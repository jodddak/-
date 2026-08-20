"""
STCO 온라인팀 광고/마케팅 성과 대시보드
==================================
매주 "STCO_주간보고서_...xlsx" 파일을 업로드하면 아래 시트들을 자동으로 인식해서
누적 저장하고, ROAS/KPI를 웹에서 바로 볼 수 있는 대시보드.

자동으로 읽는 시트:
  - "매체통합" 시트의 1) 월별 통합데이터 / 2) 매체별 현황(당월 GA비교) / 통합 주간별
  - "(SA)/(DA)/(SSP)/(브검) ○○" 형태의 매체별 요약 시트 (약 17개)
  - "GA-RAW" 시트 (소스/매체별 유입 스냅샷)

실행:
    streamlit run app.py
배포:
    README.md 참고 (GitHub + Supabase + Streamlit Community Cloud)
"""

import base64
import hashlib
import io
import json
import re
import urllib.parse
from datetime import date, datetime, timedelta

import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st
import streamlit.components.v1 as components
from PIL import Image

# 메인 타이틀 및 브라우저 탭 파비콘에 쓰는 막대그래프 아이콘(사용자가 준 PNG)을
# base64로 인라인 임베드 — 별도 이미지 파일 없이 app.py 하나로 배포되게 하기 위함.
# st.set_page_config()보다 먼저 정의해야 파비콘으로 바로 쓸 수 있다.
PAGE_TITLE_ICON_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAADAAAAAuCAIAAADhvA07AAAACXBIWXMAAA7DAAAOwwHHb6hkAAADu0lEQVRYR+2Yz47TMBDG4yTebrd/"
    "JC4sQkKIEwcQT8B78CwreAzeiTt3TtwQF1rBtk0T840nTRPHHm9IWO1ho1XqdeyZn78Ze9IqY0zykK70IcEQyyNQLCK5d8BdEksp5Z07"
    "slM5vvEv98hMTNO+NxzyRB4mLKajEGxV9vr57WuS5vpqPZvPQyuG0VzPLhZrDGAH7z7+gGhoh6TLMvX25SrP1OcPldY6y7K+cQ9QWZYw"
    "i9PAKGXC9nFaFMVBm0opsovFHEuTgjKzRD6ohnW73a5WqzTFcHfcGaiJlAVii+7o/oISUxnCUG9uIA9o0vks17mfKMWetgT7/X6xWHis"
    "JYmrEEetHkrpZGAAYoTu9CxJEGWCUonO0wud4g64vr+mV8gz/y5zbLVtn9oEyW2CRtiAkBINRAJcnwY9ge7OWBnISoT0IIWQT6RGq039"
    "LB45SxMIABjcx1yRk5qN2zvFjhy77Zb3GmUUUVghEsQKdL9XRKH7hSFvYYXoaCGJYkzRATED3eejFZqYR1SIk2jYAkePjioU2lx1/2gA"
    "14CUQzaDqDTwJHvqcJvOpFPbtTjyf1kh69eefvx3ateM1BnP+mGEQYXg7FhV5ohiLlm83e0WSy2NGPgsCITF43WiMJVo0BTb7XJFr0RT"
    "XR0gKtgKJdLGkdKnRAJBIKjVv2MICtz19TNqyDIOgT0DwSgSgup2lq1fvMYbhWCHuTHS+5IlTIw+chVieRhOnvw/aODRBbJBI5FkhRpu"
    "jteEe60DRMXdXoNe8iekwfI65xBMQ5jOSz7lM3KXXgdtEp/beFIUe8yQIzv06VkhXiju55d8OhTtHms++m2cVcrz5WEoRzPeE7JW9ti"
    "vFDVNZ+PjTLCv8XRO24+a9J85moly6ai1sd7YZ323D04FZDoaGA2f1CaBPJRTtnS0o9eJ5CmiFnGCK6pQOxa+9mSxqhcTVohkIYVweLc"
    "W7jJhiFh8B2sWBILnEtWjxK5uPKLPbe92+ytb7afCEkJmimN1eyhvD8fTX79d/NpucLID6Mun9VwzVTCK9I02Bu4pHe1qL+xnK1dd7TE"
    "lz/PnT9JXT7M/e3WB7/Y9xxj/u0ih+M3771ovai+9kHZ+sOJjGgcjLrmWNdUeHFzwgY7fNDabzeFwEJYBAPwytFwuLy8vMbfHQ78DnRXm"
    "bQ4UNGSjMAQORuF6jB4soygKeSU8ESheGspFx3GDIgNx3rTvzVrlieSyH86WUC4QP4oajdptuRjW9AMNszHpaGHbT+rnzsYegWJS/QWf"
    "cPB2ZPskBQAAAABJRU5ErkJggg=="
)


def _decode_page_icon():
    """base64 PNG를 PIL Image로 디코드해 st.set_page_config(page_icon=...)에 바로 넘긴다
    (emoji 문자열 대신 실제 이미지를 파비콘으로 쓰기 위함). 혹시 디코딩이 실패하면
    기존 이모지로 조용히 폴백해서 앱 실행 자체가 막히지 않게 한다."""
    try:
        return Image.open(io.BytesIO(base64.b64decode(PAGE_TITLE_ICON_B64)))
    except Exception:
        return "📊"


st.set_page_config(page_title="STCO 광고성과 대시보드", page_icon=_decode_page_icon(), layout="wide")

# ── 디자인 톤 (색상/버튼/표는 Toss(TDS Mobile) 스타일, 폰트는 당근마켓 SEED 시스템폰트 스택) ──
THEME_FONT_STACK = (
    "-apple-system, BlinkMacSystemFont, \"Pretendard Variable\", Pretendard, "
    "\"Apple SD Gothic Neo\", \"Malgun Gothic\", system-ui, sans-serif"
)
THEME_COLORS = {
    "primary": "#3182f6",
    "primary_hover": "#2272eb",
    "canvas": "#ffffff",
    "surface": "#f2f4f6",
    "foreground": "#191f28",
    "body": "#4e5968",
    "muted": "#8b95a1",
    "border": "#e5e8eb",
    "on_primary": "#ffffff",
    "weak_bg": "#e8f3ff",
    "weak_fg": "#1b64da",
    "danger": "#e42939",
}
px.defaults.color_discrete_sequence = ["#3182f6", "#191f28", "#8b95a1", "#1b64da"]

# 사이드바 '메뉴' 그룹 헤더 아이콘(사용자가 준 SVG, 초록 폴더+막대그래프)을 base64로 인라인 임베드.
# 별도 이미지 파일을 GitHub에 같이 올릴 필요 없이 app.py 하나로 배포되게 하기 위함.
NAV_GROUP_ICON_B64 = (
    "PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgZmlsbD0ibm9uZSI+"
    "PGcgZmlsdGVyPSJ1cmwoI2EpIj48cGF0aCBmaWxsPSIjMUY5OTU1IiBkPSJNMi42NDYgNS43N0EyLjc3IDIuNzcgMCAwIDEgNS40"
    "MTUgM2g0LjJjLjUyNCAwIC45ODguMzM1IDEuMTU0LjgzLjExLjMzMi40Mi41NTUuNzY4LjU1NWg3LjcyNGMxLjAyIDAgMS44NDYu"
    "ODI2IDEuODQ2IDEuODQ2djExLjA3N0EzLjY5IDMuNjkgMCAwIDEgMTcuNDE1IDIxSDYuMzM4YTMuNjkgMy42OSAwIDAgMS0zLjY5"
    "Mi0zLjY5MnoiLz48L2c+PGcgZmlsdGVyPSJ1cmwoI2IpIj48cGF0aCBmaWxsPSIjMUY5OTU1IiBkPSJNNC4xNjUgOS45MDljLjE4"
    "LTIuMDMyIDEuOTgtMy42NzggNC4wMi0zLjY3OGgxMC43MjljMS45OTQgMCAzLjQ4OCAxLjU3NyAzLjM3MyAzLjU2M2wtLjQyMyA3"
    "LjM1NWMtLjEyIDIuMDc2LTEuOTQzIDMuNzkzLTQuMDI3IDMuNzkzSDYuODc3Yy0yLjA0IDAtMy41NDYtMS42NDctMy4zNjUtMy42"
    "Nzh6Ii8+PC9nPjxnIGZpbHRlcj0idXJsKCNjKSI+PHJlY3Qgd2lkdGg9IjExLjA3NyIgaGVpZ2h0PSI3LjM4NSIgZmlsbD0iI0Uy"
    "RTJFMiIgcng9IjEuODQ2IiB0cmFuc2Zvcm09Im1hdHJpeCgxIDAgLS4wODg1NCAuOTk2MDcgOC4xNDUgMTAuMzg1KSIvPjwvZz48"
    "ZyBmaWxsPSIjMUY5OTU1IiBmaWx0ZXI9InVybCgjZCkiPjxwYXRoIGQ9Ik0xNC45MDIgMTUuOTAyaC0xLjM1MWwuMzI3LTMuNjc4"
    "aDEuMzUxYzEuMTI3IDAgMS43Ni42OSAxLjY1NyAxLjgzOS0uMTAyIDEuMTQ4LS44NTcgMS44MzgtMS45ODQgMS44MzhtLS40MTgt"
    "Ljc0MmguNDU0Yy42MzIgMCAxLjAwNS0uMyAxLjA3Ni0xLjA5Ny4wNzEtLjc5OC0uMjQ4LTEuMTAzLS44OTEtMS4xMDNoLS40NDR6"
    "TTEwLjQ0IDE1LjkwMmgtLjkzOGwxLjU4Mi0zLjY3OGgxLjE1MmwuOTIzIDMuNjc4aC0uOTM0bC0uMTgtLjc4OEgxMC43NnptLjU4"
    "OC0xLjQ0OGguODY3bC0uMy0xLjMyMWgtLjAzMXoiLz48L2c+PGRlZnM+PGZpbHRlciBpZD0iYSIgd2lkdGg9IjIwLjMwOCIgaGVp"
    "Z2h0PSIxOS44NDYiIHg9IjIuMTg0IiB5PSIyLjUzOCIgY29sb3ItaW50ZXJwb2xhdGlvbi1maWx0ZXJzPSJzUkdCIiBmaWx0ZXJV"
    "bml0cz0idXNlclNwYWNlT25Vc2UiPjxmZUZsb29kIGZsb29kLW9wYWNpdHk9IjAiIHJlc3VsdD0iQmFja2dyb3VuZEltYWdlRml4"
    "Ii8+PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iLjQ2MiIgZHk9Ii40NjIiLz48ZmVHYXVzc2lhbkJs"
    "dXIgc3RkRGV2aWF0aW9uPSIuNDYyIi8+PGZlQ29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBvcGVyYXRvcj0ib3V0Ii8+PGZlQ29s"
    "b3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwLjA4IDAiLz48ZmVCbGVuZCBpbjI9"
    "IkJhY2tncm91bmRJbWFnZUZpeCIgcmVzdWx0PSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzgyIi8+PGZlQmxlbmQgaW49IlNv"
    "dXJjZUdyYXBoaWMiIGluMj0iZWZmZWN0MV9kcm9wU2hhZG93Xzc2MV80MTc4MiIgcmVzdWx0PSJzaGFwZSIvPjxmZUNvbG9yTWF0"
    "cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii40NjIiIGR5PSIuNDYyIi8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlv"
    "bj0iLjIzMSIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3BlcmF0b3I9ImFyaXRobWV0aWMi"
    "Lz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMSAwIDAgMCAwIDEgMCAwIDAgMCAxIDAgMCAwIDAuMiAwIi8+PGZlQmxl"
    "bmQgaW4yPSJzaGFwZSIgcmVzdWx0PSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4MiIvPjwvZmlsdGVyPjxmaWx0ZXIgaWQ9"
    "ImIiIHdpZHRoPSIyMi4wMjciIGhlaWdodD0iMTYuNTU3IiB4PSIuNzI4IiB5PSI1LjMwOCIgY29sb3ItaW50ZXJwb2xhdGlvbi1m"
    "aWx0ZXJzPSJzUkdCIiBmaWx0ZXJVbml0cz0idXNlclNwYWNlT25Vc2UiPjxmZUZsb29kIGZsb29kLW9wYWNpdHk9IjAiIHJlc3Vs"
    "dD0iQmFja2dyb3VuZEltYWdlRml4Ii8+PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIg"
    "dmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iLTEuODQ2Ii8+"
    "PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlvbj0iLjQ2MiIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgb3BlcmF0b3I9"
    "Im91dCIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMC4yNSAwIi8+"
    "PGZlQmxlbmQgaW4yPSJCYWNrZ3JvdW5kSW1hZ2VGaXgiIHJlc3VsdD0iZWZmZWN0MV9kcm9wU2hhZG93Xzc2MV80MTc4MiIvPjxm"
    "ZUJsZW5kIGluPSJTb3VyY2VHcmFwaGljIiBpbjI9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODIiIHJlc3VsdD0ic2hhcGUi"
    "Lz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhhIiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4PSIuNDYyIiBkeT0iLjQ2MiIvPjxmZUdhdXNzaWFuQmx1"
    "ciBzdGREZXZpYXRpb249Ii4yMzEiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIGsyPSItMSIgazM9IjEiIG9wZXJhdG9y"
    "PSJhcml0aG1ldGljIi8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDEgMCAwIDAgMCAxIDAgMCAwIDAgMSAwIDAgMCAw"
    "LjIgMCIvPjxmZUJsZW5kIGluMj0ic2hhcGUiIHJlc3VsdD0iZWZmZWN0Ml9pbm5lclNoYWRvd183NjFfNDE3ODIiLz48L2ZpbHRl"
    "cj48ZmlsdGVyIGlkPSJjIiB3aWR0aD0iMTUuMTExIiBoZWlnaHQ9IjExLjA0OCIgeD0iNS4wNjMiIHk9IjkiIGNvbG9yLWludGVy"
    "cG9sYXRpb24tZmlsdGVycz0ic1JHQiIgZmlsdGVyVW5pdHM9InVzZXJTcGFjZU9uVXNlIj48ZmVGbG9vZCBmbG9vZC1vcGFjaXR5"
    "PSIwIiByZXN1bHQ9IkJhY2tncm91bmRJbWFnZUZpeCIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJo"
    "YXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVNb3JwaG9sb2d5"
    "IGluPSJTb3VyY2VBbHBoYSIgb3BlcmF0b3I9ImRpbGF0ZSIgcmFkaXVzPSIuOTIzIiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRv"
    "d183NjFfNDE3ODIiLz48ZmVPZmZzZXQgZHg9Ii0uNzM4IiBkeT0iLjQ2MiIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249"
    "Ii40NjIiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJvdXQiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9"
    "IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMDQgMCIvPjxmZUJsZW5kIGluMj0iQmFja2dyb3VuZEltYWdl"
    "Rml4IiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODIiLz48ZmVCbGVuZCBpbj0iU291cmNlR3JhcGhpYyIgaW4y"
    "PSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzgyIiByZXN1bHQ9InNoYXBlIi8+PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFs"
    "cGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIv"
    "PjxmZU9mZnNldCBkeD0iLjQ2MiIgZHk9Ii40NjIiLz48ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuMjMxIi8+PGZlQ29t"
    "cG9zaXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4"
    "IHZhbHVlcz0iMCAwIDAgMCAxIDAgMCAwIDAgMSAwIDAgMCAwIDEgMCAwIDAgMSAwIi8+PGZlQmxlbmQgaW4yPSJzaGFwZSIgcmVz"
    "dWx0PSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4MiIvPjwvZmlsdGVyPjxmaWx0ZXIgaWQ9ImQiIHdpZHRoPSI3LjM5NSIg"
    "aGVpZ2h0PSIzLjY3OCIgeD0iOS41MDIiIHk9IjEyLjIyNCIgY29sb3ItaW50ZXJwb2xhdGlvbi1maWx0ZXJzPSJzUkdCIiBmaWx0"
    "ZXJVbml0cz0idXNlclNwYWNlT25Vc2UiPjxmZUZsb29kIGZsb29kLW9wYWNpdHk9IjAiIHJlc3VsdD0iQmFja2dyb3VuZEltYWdl"
    "Rml4Ii8+PGZlQmxlbmQgaW49IlNvdXJjZUdyYXBoaWMiIGluMj0iQmFja2dyb3VuZEltYWdlRml4IiByZXN1bHQ9InNoYXBlIi8+"
    "PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldC8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlvbj0iLjIzMSIv"
    "PjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3BlcmF0b3I9ImFyaXRobWV0aWMiLz48ZmVDb2xv"
    "ck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMyAwIi8+PGZlQmxlbmQgaW4yPSJz"
    "aGFwZSIgcmVzdWx0PSJlZmZlY3QxX2lubmVyU2hhZG93Xzc2MV80MTc4MiIvPjwvZmlsdGVyPjwvZGVmcz48L3N2Zz4="
)

# 그룹별 커스텀 아이콘(사용자가 준 SVG) — 성과 리포트/운영 도구/가이드 그룹 헤더 앞에 표시.
# GA 유입 리포트 그룹은 별도로 받은 아이콘이 없어서 기존 NAV_GROUP_ICON_B64를 그대로 쓴다.
NAV_ICON_PERFORMANCE_B64 = (
    "PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgZmlsbD0ibm9u"
    "ZSI+PGcgZmlsdGVyPSJ1cmwoI2EpIj48cmVjdCB3aWR0aD0iNC43MzciIGhlaWdodD0iNC43MzciIHg9IjMiIHk9IjE2LjI2"
    "MyIgZmlsbD0idXJsKCNiKSIgcng9Ii45NDciLz48L2c+PGcgZmlsdGVyPSJ1cmwoI2MpIj48cGF0aCBmaWxsPSJ1cmwoI2Qp"
    "IiBkPSJNOS42MzIgMTQuMzY4YzAtLjUyMy40MjQtLjk0Ny45NDctLjk0N2gyLjg0MmMuNTIzIDAgLjk0Ny40MjQuOTQ3Ljk0"
    "N3Y1LjY4NWEuOTQ3Ljk0NyAwIDAgMS0uOTQ3Ljk0N0gxMC41OGEuOTQ3Ljk0NyAwIDAgMS0uOTQ3LS45NDd6Ii8+PC9nPjxn"
    "IGZpbHRlcj0idXJsKCNlKSI+PHBhdGggZmlsbD0idXJsKCNmKSIgZD0iTTE2LjI2MyAxMS41MjZjMC0uNTIzLjQyNC0uOTQ3"
    "Ljk0Ny0uOTQ3aDIuODQzYy41MjMgMCAuOTQ3LjQyNC45NDcuOTQ3djguNTI3YS45NDcuOTQ3IDAgMCAxLS45NDcuOTQ3SDE3"
    "LjIxYS45NDcuOTQ3IDAgMCAxLS45NDctLjk0N3oiLz48L2c+PGcgZmlsdGVyPSJ1cmwoI2cpIj48cGF0aCBmaWxsPSJ1cmwo"
    "I2gpIiBkPSJNMTguMTQ2IDcuNzUzQzE0Ljc2IDEyLjEzMyA3Ljc1MyAxMy4zMzQgNC4xNSAxMy40M2MtLjY0Ny4wMTctMS4x"
    "NS0uNTE3LTEuMTUtMS4xNjMgMC0uNjcxLjU0NS0xLjIxMSAxLjIxNC0xLjI2NiA1LjI1LS40MzUgMTAuMjM4LTMuNTM2IDEx"
    "LjU2OC01LjE1MWEuMDE0LjAxNCAwIDAgMCAwLS4wMThsLTEuMDY2LTEuMjhBLjk0Ny45NDcgMCAwIDEgMTUuNDQ0IDNoNC42"
    "MDljLjUyMyAwIC45NDcuNDI0Ljk0Ny45NDd2NC4zNDVjMCAuODQ0LTEuMDIgMS4yNjctMS42MTcuNjdsLTEuMjEtMS4yMTFh"
    "LjAyLjAyIDAgMCAwLS4wMjcuMDAyIi8+PC9nPjxkZWZzPjxsaW5lYXJHcmFkaWVudCBpZD0iYiIgeDE9IjUuMzY4IiB4Mj0i"
    "NS4zNjgiIHkxPSIxNi4yNjMiIHkyPSIyMSIgZ3JhZGllbnRVbml0cz0idXNlclNwYWNlT25Vc2UiPjxzdG9wIHN0b3AtY29s"
    "b3I9IiNGM0VFRUEiLz48c3RvcCBvZmZzZXQ9IjEiIHN0b3AtY29sb3I9IiNFQ0U1REYiLz48L2xpbmVhckdyYWRpZW50Pjxs"
    "aW5lYXJHcmFkaWVudCBpZD0iZCIgeDE9IjEyIiB4Mj0iMTIiIHkxPSIxMy40MjEiIHkyPSIyMSIgZ3JhZGllbnRVbml0cz0i"
    "dXNlclNwYWNlT25Vc2UiPjxzdG9wIHN0b3AtY29sb3I9IiNGM0VFRUEiLz48c3RvcCBvZmZzZXQ9IjEiIHN0b3AtY29sb3I9"
    "IiNFQ0U1REYiLz48L2xpbmVhckdyYWRpZW50PjxsaW5lYXJHcmFkaWVudCBpZD0iZiIgeDE9IjE4LjYzMiIgeDI9IjE4LjYz"
    "MiIgeTE9IjEwLjU3OSIgeTI9IjIxIiBncmFkaWVudFVuaXRzPSJ1c2VyU3BhY2VPblVzZSI+PHN0b3Agc3RvcC1jb2xvcj0i"
    "I0YzRUVFQSIvPjxzdG9wIG9mZnNldD0iMSIgc3RvcC1jb2xvcj0iI0VDRTVERiIvPjwvbGluZWFyR3JhZGllbnQ+PGxpbmVh"
    "ckdyYWRpZW50IGlkPSJoIiB4MT0iMTIiIHgyPSIxMiIgeTE9IjEzLjQzMSIgeTI9IjMiIGdyYWRpZW50VW5pdHM9InVzZXJT"
    "cGFjZU9uVXNlIj48c3RvcCBzdG9wLWNvbG9yPSIjRjk4MDQ0Ii8+PHN0b3Agb2Zmc2V0PSIxIiBzdG9wLWNvbG9yPSIjRkVB"
    "OTdGIi8+PC9saW5lYXJHcmFkaWVudD48ZmlsdGVyIGlkPSJhIiB3aWR0aD0iNi42MzIiIGhlaWdodD0iNi42MzIiIHg9IjEu"
    "ODE2IiB5PSIxNS41NTMiIGNvbG9yLWludGVycG9sYXRpb24tZmlsdGVycz0ic1JHQiIgZmlsdGVyVW5pdHM9InVzZXJTcGFj"
    "ZU9uVXNlIj48ZmVGbG9vZCBmbG9vZC1vcGFjaXR5PSIwIiByZXN1bHQ9IkJhY2tncm91bmRJbWFnZUZpeCIvPjxmZUNvbG9y"
    "TWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii0uMjM3IiBkeT0iLjIzNyIvPjxmZUdhdXNzaWFuQmx1ciBz"
    "dGREZXZpYXRpb249Ii40NzQiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJvdXQiLz48ZmVDb2xv"
    "ck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMDggMCIvPjxmZUJsZW5kIGlu"
    "Mj0iQmFja2dyb3VuZEltYWdlRml4IiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODMiLz48ZmVCbGVuZCBp"
    "bj0iU291cmNlR3JhcGhpYyIgaW4yPSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzgzIiByZXN1bHQ9InNoYXBlIi8+PGZl"
    "Q29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iLjc1OCIgZHk9Ii0uNDc0Ii8+PGZlR2F1c3NpYW5C"
    "bHVyIHN0ZERldmlhdGlvbj0iLjIzNyIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3Bl"
    "cmF0b3I9ImFyaXRobWV0aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC44NTQ2NDIgMCAwIDAgMCAwLjg1"
    "NDY0MiAwIDAgMCAwIDAuODU0NjQyIDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0ic2hhcGUiIHJlc3VsdD0iZWZmZWN0Ml9p"
    "bm5lclNoYWRvd183NjFfNDE3ODMiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhh"
    "IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4PSIuNDc0"
    "IiBkeT0iLjQ3NCIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii4yMzciLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJk"
    "QWxwaGEiIGsyPSItMSIgazM9IjEiIG9wZXJhdG9yPSJhcml0aG1ldGljIi8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAg"
    "MCAwIDAuOTAyNjA3IDAgMCAwIDAgMC44NDIwMzEgMCAwIDAgMCAwLjgwMTY0NyAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9"
    "ImVmZmVjdDJfaW5uZXJTaGFkb3dfNzYxXzQxNzgzIiByZXN1bHQ9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYxXzQxNzgzIi8+"
    "PC9maWx0ZXI+PGZpbHRlciBpZD0iYyIgd2lkdGg9IjYuNjMyIiBoZWlnaHQ9IjkuNDc0IiB4PSI4LjQ0NyIgeT0iMTIuNzEi"
    "IGNvbG9yLWludGVycG9sYXRpb24tZmlsdGVycz0ic1JHQiIgZmlsdGVyVW5pdHM9InVzZXJTcGFjZU9uVXNlIj48ZmVGbG9v"
    "ZCBmbG9vZC1vcGFjaXR5PSIwIiByZXN1bHQ9IkJhY2tncm91bmRJbWFnZUZpeCIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3Vy"
    "Y2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAg"
    "MTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii0uMjM3IiBkeT0iLjIzNyIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii40"
    "NzQiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJvdXQiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9"
    "IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMDggMCIvPjxmZUJsZW5kIGluMj0iQmFja2dyb3VuZElt"
    "YWdlRml4IiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODMiLz48ZmVCbGVuZCBpbj0iU291cmNlR3JhcGhp"
    "YyIgaW4yPSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzgzIiByZXN1bHQ9InNoYXBlIi8+PGZlQ29sb3JNYXRyaXggaW49"
    "IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iLjc1OCIgZHk9Ii0uNDc0Ii8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlv"
    "bj0iLjIzNyIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3BlcmF0b3I9ImFyaXRobWV0"
    "aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC44NTQ2NDIgMCAwIDAgMCAwLjg1NDY0MiAwIDAgMCAwIDAu"
    "ODU0NjQyIDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0ic2hhcGUiIHJlc3VsdD0iZWZmZWN0Ml9pbm5lclNoYWRvd183NjFf"
    "NDE3ODMiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhhIiB2YWx1ZXM9IjAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4PSIuNDc0IiBkeT0iLjQ3NCIvPjxm"
    "ZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii4yMzciLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIGsyPSItMSIg"
    "azM9IjEiIG9wZXJhdG9yPSJhcml0aG1ldGljIi8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAuOTAyNjA3IDAg"
    "MCAwIDAgMC44NDIwMzEgMCAwIDAgMCAwLjgwMTY0NyAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9ImVmZmVjdDJfaW5uZXJT"
    "aGFkb3dfNzYxXzQxNzgzIiByZXN1bHQ9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYxXzQxNzgzIi8+PC9maWx0ZXI+PGZpbHRl"
    "ciBpZD0iZSIgd2lkdGg9IjYuNjMyIiBoZWlnaHQ9IjEyLjMxNiIgeD0iMTUuMDc5IiB5PSI5Ljg2OCIgY29sb3ItaW50ZXJw"
    "b2xhdGlvbi1maWx0ZXJzPSJzUkdCIiBmaWx0ZXJVbml0cz0idXNlclNwYWNlT25Vc2UiPjxmZUZsb29kIGZsb29kLW9wYWNp"
    "dHk9IjAiIHJlc3VsdD0iQmFja2dyb3VuZEltYWdlRml4Ii8+PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1"
    "bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9m"
    "ZnNldCBkeD0iLS4yMzciIGR5PSIuMjM3Ii8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlvbj0iLjQ3NCIvPjxmZUNvbXBv"
    "c2l0ZSBpbjI9ImhhcmRBbHBoYSIgb3BlcmF0b3I9Im91dCIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMC4wOCAwIi8+PGZlQmxlbmQgaW4yPSJCYWNrZ3JvdW5kSW1hZ2VGaXgiIHJlc3Vs"
    "dD0iZWZmZWN0MV9kcm9wU2hhZG93Xzc2MV80MTc4MyIvPjxmZUJsZW5kIGluPSJTb3VyY2VHcmFwaGljIiBpbjI9ImVmZmVj"
    "dDFfZHJvcFNoYWRvd183NjFfNDE3ODMiIHJlc3VsdD0ic2hhcGUiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEi"
    "IHJlc3VsdD0iaGFyZEFscGhhIiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+"
    "PGZlT2Zmc2V0IGR4PSIuNzU4IiBkeT0iLS40NzQiLz48ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuMjM3Ii8+PGZl"
    "Q29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9y"
    "TWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwLjg1NDY0MiAwIDAgMCAwIDAuODU0NjQyIDAgMCAwIDAgMC44NTQ2NDIgMCAwIDAg"
    "MSAwIi8+PGZlQmxlbmQgaW4yPSJzaGFwZSIgcmVzdWx0PSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4MyIvPjxmZUNv"
    "bG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii40NzQiIGR5PSIuNDc0Ii8+PGZlR2F1c3NpYW5CbHVy"
    "IHN0ZERldmlhdGlvbj0iLjIzNyIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3BlcmF0"
    "b3I9ImFyaXRobWV0aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC45MDI2MDcgMCAwIDAgMCAwLjg0MjAz"
    "MSAwIDAgMCAwIDAuODAxNjQ3IDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0iZWZmZWN0Ml9pbm5lclNoYWRvd183NjFfNDE3"
    "ODMiIHJlc3VsdD0iZWZmZWN0M19pbm5lclNoYWRvd183NjFfNDE3ODMiLz48L2ZpbHRlcj48ZmlsdGVyIGlkPSJnIiB3aWR0"
    "aD0iMTkuODk1IiBoZWlnaHQ9IjEyLjMyNiIgeD0iMS44MTYiIHk9IjIuMjg5IiBjb2xvci1pbnRlcnBvbGF0aW9uLWZpbHRl"
    "cnM9InNSR0IiIGZpbHRlclVuaXRzPSJ1c2VyU3BhY2VPblVzZSI+PGZlRmxvb2QgZmxvb2Qtb3BhY2l0eT0iMCIgcmVzdWx0"
    "PSJCYWNrZ3JvdW5kSW1hZ2VGaXgiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhh"
    "IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4PSItLjIz"
    "NyIgZHk9Ii4yMzciLz48ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuNDc0Ii8+PGZlQ29tcG9zaXRlIGluMj0iaGFy"
    "ZEFscGhhIiBvcGVyYXRvcj0ib3V0Ii8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwLjEyIDAiLz48ZmVCbGVuZCBpbjI9IkJhY2tncm91bmRJbWFnZUZpeCIgcmVzdWx0PSJlZmZlY3QxX2Ry"
    "b3BTaGFkb3dfNzYxXzQxNzgzIi8+PGZlQmxlbmQgaW49IlNvdXJjZUdyYXBoaWMiIGluMj0iZWZmZWN0MV9kcm9wU2hhZG93"
    "Xzc2MV80MTc4MyIgcmVzdWx0PSJzaGFwZSIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJk"
    "QWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9"
    "Ii0uNDc0IiBkeT0iLS40NzQiLz48ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuMjM3Ii8+PGZlQ29tcG9zaXRlIGlu"
    "Mj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZhbHVl"
    "cz0iMCAwIDAgMCAwLjkxMzUxOCAwIDAgMCAwIDAuNTEzMTA2IDAgMCAwIDAgMC4zMTc1NTYgMCAwIDAgMSAwIi8+PGZlQmxl"
    "bmQgaW4yPSJzaGFwZSIgcmVzdWx0PSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4MyIvPjxmZUNvbG9yTWF0cml4IGlu"
    "PSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii40NzQiIGR5PSIuNDc0Ii8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlv"
    "bj0iLjIzNyIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0iMSIgb3BlcmF0b3I9ImFyaXRobWV0"
    "aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC45OTc2NTIgMCAwIDAgMCAwLjU1MzYyNCAwIDAgMCAwIDAu"
    "NDY3NTc1IDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0iZWZmZWN0Ml9pbm5lclNoYWRvd183NjFfNDE3ODMiIHJlc3VsdD0i"
    "ZWZmZWN0M19pbm5lclNoYWRvd183NjFfNDE3ODMiLz48L2ZpbHRlcj48L2RlZnM+PC9zdmc+"
)

NAV_ICON_OPERATIONS_B64 = (
    "PHN2ZyB3aWR0aD0iMjQiIGhlaWdodD0iMjQiIHZpZXdCb3g9IjAgMCAyNCAyNCIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6"
    "Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPGcgZmlsdGVyPSJ1cmwoI2ZpbHRlcjBfZGlpXzc2MV80MTc4NykiPgo8cGF0aCBk"
    "PSJNMjAuNDI1NyA2Ljc1MDM2QzIwLjkwNjggNi4yNjkyNyAyMS43MTY1IDYuNDMyOTcgMjEuODE1OCA3LjEwNjA1QzIyLjA1"
    "MjcgOC43MTIzOSAyMS41NTM0IDEwLjQwNjIgMjAuMzE2OSAxMS42NDI2QzE4Ljg4NjEgMTMuMDczNCAxNi44NDI5IDEzLjUx"
    "NjcgMTUuMDMxNCAxMi45NzQ2QzE0LjY1NzYgMTIuODYyNyAxNC4yNDI4IDEyLjkzMzMgMTMuOTY2OSAxMy4yMDkyTDYuOTkx"
    "NTYgMjAuMTg0NUM2LjIzNjc0IDIwLjkzOTIgNS4wMTI5MyAyMC45MzkzIDQuMjU4MTUgMjAuMTg0NUMzLjUwMzM2IDE5LjQy"
    "OTcgMy41MDM0MSAxOC4yMDU5IDQuMjU4MTUgMTcuNDUxMUwxMS4yMzM1IDEwLjQ3NThDMTEuNTA5MyAxMC4xOTk5IDExLjU4"
    "IDkuNzg1MDUgMTEuNDY4MSA5LjQxMTI1QzEwLjkyNiA3LjU5OTc0IDExLjM2OTIgNS41NTY2IDEyLjgwMDEgNC4xMjU3M0Mx"
    "NC4wMzY1IDIuODg5MjkgMTUuNzMwMyAyLjM4OTk4IDE3LjMzNjYgMi42MjY4OUMxOC4wMDk3IDIuNzI2MTggMTguMTczNCAz"
    "LjUzNTg2IDE3LjY5MjMgNC4wMTY5NUwxNS44OTg5IDUuODEwMzVDMTUuNTA4NCA2LjIwMDg4IDE1LjUwODQgNi44MzQwNCAx"
    "NS44OTg5IDcuMjI0NTdMMTcuMjE4MSA4LjU0Mzc2QzE3LjYwODYgOC45MzQyOSAxOC4yNDE4IDguOTM0MjkgMTguNjMyMyA4"
    "LjU0Mzc2TDIwLjQyNTcgNi43NTAzNloiIGZpbGw9IiNDNkMzQkQiLz4KPC9nPgo8ZGVmcz4KPGZpbHRlciBpZD0iZmlsdGVy"
    "MF9kaWlfNzYxXzQxNzg3IiB4PSIxLjY5MjA4IiB5PSIxLjU2OTcyIiB3aWR0aD0iMjIuMTgwOSIgaGVpZ2h0PSIyMi4xODA5"
    "IiBmaWx0ZXJVbml0cz0idXNlclNwYWNlT25Vc2UiIGNvbG9yLWludGVycG9sYXRpb24tZmlsdGVycz0ic1JHQiI+CjxmZUZs"
    "b29kIGZsb29kLW9wYWNpdHk9IjAiIHJlc3VsdD0iQmFja2dyb3VuZEltYWdlRml4Ii8+CjxmZUNvbG9yTWF0cml4IGluPSJT"
    "b3VyY2VBbHBoYSIgdHlwZT0ibWF0cml4IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEy"
    "NyAwIiByZXN1bHQ9ImhhcmRBbHBoYSIvPgo8ZmVPZmZzZXQgZHk9IjEiLz4KPGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlv"
    "bj0iMSIvPgo8ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJvdXQiLz4KPGZlQ29sb3JNYXRyaXggdHlw"
    "ZT0ibWF0cml4IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMTIgMCIvPgo8ZmVCbGVu"
    "ZCBtb2RlPSJub3JtYWwiIGluMj0iQmFja2dyb3VuZEltYWdlRml4IiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFf"
    "NDE3ODciLz4KPGZlQmxlbmQgbW9kZT0ibm9ybWFsIiBpbj0iU291cmNlR3JhcGhpYyIgaW4yPSJlZmZlY3QxX2Ryb3BTaGFk"
    "b3dfNzYxXzQxNzg3IiByZXN1bHQ9InNoYXBlIi8+CjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgdHlwZT0ibWF0"
    "cml4IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIiByZXN1bHQ9ImhhcmRBbHBo"
    "YSIvPgo8ZmVPZmZzZXQgZHg9IjEiIGR5PSItMSIvPgo8ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIwLjI1Ii8+Cjxm"
    "ZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgb3BlcmF0b3I9ImFyaXRobWV0aWMiIGsyPSItMSIgazM9IjEiLz4KPGZlQ29s"
    "b3JNYXRyaXggdHlwZT0ibWF0cml4IiB2YWx1ZXM9IjAgMCAwIDAgMC42NjI3NDUgMCAwIDAgMCAwLjYxOTYwOCAwIDAgMCAw"
    "IDAuNTgwMzkyIDAgMCAwIDEgMCIvPgo8ZmVCbGVuZCBtb2RlPSJub3JtYWwiIGluMj0ic2hhcGUiIHJlc3VsdD0iZWZmZWN0"
    "Ml9pbm5lclNoYWRvd183NjFfNDE3ODciLz4KPGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiB0eXBlPSJtYXRyaXgi"
    "IHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiIHJlc3VsdD0iaGFyZEFscGhhIi8+"
    "CjxmZU9mZnNldCBkeD0iMSIvPgo8ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIwLjEyNSIvPgo8ZmVDb21wb3NpdGUg"
    "aW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJhcml0aG1ldGljIiBrMj0iLTEiIGszPSIxIi8+CjxmZUNvbG9yTWF0cml4IHR5"
    "cGU9Im1hdHJpeCIgdmFsdWVzPSIwIDAgMCAwIDAuNTUyOTQxIDAgMCAwIDAgMC41MzMzMzMgMCAwIDAgMCAwLjUwOTgwNCAw"
    "IDAgMCAxIDAiLz4KPGZlQmxlbmQgbW9kZT0ibm9ybWFsIiBpbjI9ImVmZmVjdDJfaW5uZXJTaGFkb3dfNzYxXzQxNzg3IiBy"
    "ZXN1bHQ9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYxXzQxNzg3Ii8+CjwvZmlsdGVyPgo8L2RlZnM+Cjwvc3ZnPgo="
)

NAV_ICON_GUIDE_B64 = (
    "PHN2ZyB4bWxucz0iaHR0cDovL3d3dy53My5vcmcvMjAwMC9zdmciIHdpZHRoPSIyNCIgaGVpZ2h0PSIyNCIgZmlsbD0ibm9u"
    "ZSI+PGcgY2xpcC1wYXRoPSJ1cmwoI2EpIj48ZyBmaWx0ZXI9InVybCgjYikiPjxwYXRoIGZpbGw9IiNGRERBOEEiIGQ9Ik02"
    "LjAzNCAzLjA1MmExLjEyMiAxLjEyMiAwIDAgMSAxLjkzMiAwbDMuMzgzIDUuNzkzYy40My43MzUtLjEwNyAxLjY1NS0uOTY3"
    "IDEuNjU1SDMuNjE4Yy0uODYgMC0xLjM5Ni0uOTItLjk2Ny0xLjY1NXoiLz48L2c+PGcgZmlsdGVyPSJ1cmwoI2MpIj48cGF0"
    "aCBmaWxsPSJ1cmwoI2QpIiBkPSJNMTMuNSAzLjQ0N2MwLS41MjMuNDI0LS45NDcuOTQ3LS45NDdoNi4xMDZjLjUyMyAwIC45"
    "NDcuNDI0Ljk0Ny45NDd2Ni4xMDZhLjk0Ny45NDcgMCAwIDEtLjk0Ny45NDdoLTYuMTA2YS45NDcuOTQ3IDAgMCAxLS45NDct"
    "Ljk0N3oiLz48L2c+PGcgZmlsdGVyPSJ1cmwoI2UpIj48Y2lyY2xlIGN4PSI3IiBjeT0iMTciIHI9IjQuNSIgZmlsbD0iIzlF"
    "RTNCMyIvPjwvZz48ZyBmaWx0ZXI9InVybCgjZikiPjxjaXJjbGUgY3g9IjE3LjUiIGN5PSIxNyIgcj0iNC41IiBmaWxsPSIj"
    "OUVFM0IzIi8+PC9nPjwvZz48ZGVmcz48ZmlsdGVyIGlkPSJiIiB3aWR0aD0iMTMiIGhlaWdodD0iMTIiIHg9Ii41IiB5PSIx"
    "LjUiIGNvbG9yLWludGVycG9sYXRpb24tZmlsdGVycz0ic1JHQiIgZmlsdGVyVW5pdHM9InVzZXJTcGFjZU9uVXNlIj48ZmVG"
    "bG9vZCBmbG9vZC1vcGFjaXR5PSIwIiByZXN1bHQ9IkJhY2tncm91bmRJbWFnZUZpeCIvPjxmZUNvbG9yTWF0cml4IGluPSJT"
    "b3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMTI3IDAiLz48ZmVPZmZzZXQgZHk9IjEiLz48ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIxIi8+PGZlQ29tcG9z"
    "aXRlIGluMj0iaGFyZEFscGhhIiBvcGVyYXRvcj0ib3V0Ii8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwLjEyIDAiLz48ZmVCbGVuZCBpbjI9IkJhY2tncm91bmRJbWFnZUZpeCIgcmVzdWx0"
    "PSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzg2Ii8+PGZlQmxlbmQgaW49IlNvdXJjZUdyYXBoaWMiIGluMj0iZWZmZWN0"
    "MV9kcm9wU2hhZG93Xzc2MV80MTc4NiIgcmVzdWx0PSJzaGFwZSIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIg"
    "cmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48"
    "ZmVPZmZzZXQgZHg9IjEiIGR5PSItMSIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii41Ii8+PGZlQ29tcG9zaXRl"
    "IGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZh"
    "bHVlcz0iMCAwIDAgMCAxIDAgMCAwIDAgMC44MDQzNjQgMCAwIDAgMCAwLjM1NzE5NyAwIDAgMCAxIDAiLz48ZmVCbGVuZCBp"
    "bjI9InNoYXBlIiByZXN1bHQ9ImVmZmVjdDJfaW5uZXJTaGFkb3dfNzYxXzQxNzg2Ii8+PGZlQ29sb3JNYXRyaXggaW49IlNv"
    "dXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAg"
    "MCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iMSIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii40Ii8+PGZlQ29tcG9z"
    "aXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4"
    "IHZhbHVlcz0iMCAwIDAgMCAwLjg5OTUwOCAwIDAgMCAwIDAuNzMwODEyIDAgMCAwIDAgMC4zNDUyMjEgMCAwIDAgMSAwIi8+"
    "PGZlQmxlbmQgaW4yPSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4NiIgcmVzdWx0PSJlZmZlY3QzX2lubmVyU2hhZG93"
    "Xzc2MV80MTc4NiIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0i"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9Ii0xIiBkeT0iMSIvPjxm"
    "ZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii40Ii8+PGZlQ29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGsz"
    "PSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAxIDAgMCAwIDAgMC45"
    "MDYyMDYgMCAwIDAgMCAwLjY5MTgxOSAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYx"
    "XzQxNzg2IiByZXN1bHQ9ImVmZmVjdDRfaW5uZXJTaGFkb3dfNzYxXzQxNzg2Ii8+PC9maWx0ZXI+PGZpbHRlciBpZD0iYyIg"
    "d2lkdGg9IjEyIiBoZWlnaHQ9IjEyIiB4PSIxMS41IiB5PSIxLjUiIGNvbG9yLWludGVycG9sYXRpb24tZmlsdGVycz0ic1JH"
    "QiIgZmlsdGVyVW5pdHM9InVzZXJTcGFjZU9uVXNlIj48ZmVGbG9vZCBmbG9vZC1vcGFjaXR5PSIwIiByZXN1bHQ9IkJhY2tn"
    "cm91bmRJbWFnZUZpeCIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVl"
    "cz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHk9IjEiLz48ZmVHYXVz"
    "c2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIxIi8+PGZlQ29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBvcGVyYXRvcj0ib3V0Ii8+"
    "PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwLjEyIDAiLz48ZmVC"
    "bGVuZCBpbjI9IkJhY2tncm91bmRJbWFnZUZpeCIgcmVzdWx0PSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzg2Ii8+PGZl"
    "QmxlbmQgaW49IlNvdXJjZUdyYXBoaWMiIGluMj0iZWZmZWN0MV9kcm9wU2hhZG93Xzc2MV80MTc4NiIgcmVzdWx0PSJzaGFw"
    "ZSIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9IjEiIGR5PSItMSIvPjxmZUdhdXNzaWFu"
    "Qmx1ciBzdGREZXZpYXRpb249Ii41Ii8+PGZlQ29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVy"
    "YXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwLjg1NDY0MiAwIDAgMCAwIDAuODU0"
    "NjQyIDAgMCAwIDAgMC44NTQ2NDIgMCAwIDAgMSAwIi8+PGZlQmxlbmQgaW4yPSJzaGFwZSIgcmVzdWx0PSJlZmZlY3QyX2lu"
    "bmVyU2hhZG93Xzc2MV80MTc4NiIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIgcmVzdWx0PSJoYXJkQWxwaGEi"
    "IHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48ZmVPZmZzZXQgZHg9IjEiLz48"
    "ZmVHYXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuNCIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBr"
    "Mz0iMSIgb3BlcmF0b3I9ImFyaXRobWV0aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC45MDI2MDcgMCAw"
    "IDAgMCAwLjg0MjAzMSAwIDAgMCAwIDAuODAxNjQ3IDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0iZWZmZWN0Ml9pbm5lclNo"
    "YWRvd183NjFfNDE3ODYiIHJlc3VsdD0iZWZmZWN0M19pbm5lclNoYWRvd183NjFfNDE3ODYiLz48L2ZpbHRlcj48ZmlsdGVy"
    "IGlkPSJlIiB3aWR0aD0iMTMiIGhlaWdodD0iMTMiIHg9Ii41IiB5PSIxMS41IiBjb2xvci1pbnRlcnBvbGF0aW9uLWZpbHRl"
    "cnM9InNSR0IiIGZpbHRlclVuaXRzPSJ1c2VyU3BhY2VPblVzZSI+PGZlRmxvb2QgZmxvb2Qtb3BhY2l0eT0iMCIgcmVzdWx0"
    "PSJCYWNrZ3JvdW5kSW1hZ2VGaXgiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhh"
    "IiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR5PSIxIi8+"
    "PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlvbj0iMSIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgb3BlcmF0b3I9"
    "Im91dCIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMC4xMiAw"
    "Ii8+PGZlQmxlbmQgaW4yPSJCYWNrZ3JvdW5kSW1hZ2VGaXgiIHJlc3VsdD0iZWZmZWN0MV9kcm9wU2hhZG93Xzc2MV80MTc4"
    "NiIvPjxmZUJsZW5kIGluPSJTb3VyY2VHcmFwaGljIiBpbjI9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODYiIHJlc3Vs"
    "dD0ic2hhcGUiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFyZEFscGhhIiB2YWx1ZXM9IjAg"
    "MCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4PSIxIiBkeT0iLTEiLz48ZmVH"
    "YXVzc2lhbkJsdXIgc3RkRGV2aWF0aW9uPSIuNSIvPjxmZUNvbXBvc2l0ZSBpbjI9ImhhcmRBbHBoYSIgazI9Ii0xIiBrMz0i"
    "MSIgb3BlcmF0b3I9ImFyaXRobWV0aWMiLz48ZmVDb2xvck1hdHJpeCB2YWx1ZXM9IjAgMCAwIDAgMC41MjA3OTcgMCAwIDAg"
    "MCAwLjkzMTAyMSAwIDAgMCAwIDAuNjQ1NjQ4IDAgMCAwIDEgMCIvPjxmZUJsZW5kIGluMj0ic2hhcGUiIHJlc3VsdD0iZWZm"
    "ZWN0Ml9pbm5lclNoYWRvd183NjFfNDE3ODYiLz48ZmVDb2xvck1hdHJpeCBpbj0iU291cmNlQWxwaGEiIHJlc3VsdD0iaGFy"
    "ZEFscGhhIiB2YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDEyNyAwIi8+PGZlT2Zmc2V0IGR4"
    "PSIxIi8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlvbj0iLjQiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIGsy"
    "PSItMSIgazM9IjEiIG9wZXJhdG9yPSJhcml0aG1ldGljIi8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAuMzI0"
    "ODc0IDAgMCAwIDAgMC43Nzc1MTMgMCAwIDAgMCAwLjQ2MjYzMyAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9ImVmZmVjdDJf"
    "aW5uZXJTaGFkb3dfNzYxXzQxNzg2IiByZXN1bHQ9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYxXzQxNzg2Ii8+PGZlQ29sb3JN"
    "YXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAg"
    "MCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iLTEiIGR5PSIxIi8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlh"
    "dGlvbj0iLjQiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIGsyPSItMSIgazM9IjEiIG9wZXJhdG9yPSJhcml0aG1l"
    "dGljIi8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAuNjQ0MTU5IDAgMCAwIDAgMC45NTUzMDMgMCAwIDAgMCAw"
    "LjczODg1NSAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9ImVmZmVjdDNfaW5uZXJTaGFkb3dfNzYxXzQxNzg2IiByZXN1bHQ9"
    "ImVmZmVjdDRfaW5uZXJTaGFkb3dfNzYxXzQxNzg2Ii8+PC9maWx0ZXI+PGZpbHRlciBpZD0iZiIgd2lkdGg9IjEzIiBoZWln"
    "aHQ9IjEzIiB4PSIxMSIgeT0iMTEuNSIgY29sb3ItaW50ZXJwb2xhdGlvbi1maWx0ZXJzPSJzUkdCIiBmaWx0ZXJVbml0cz0i"
    "dXNlclNwYWNlT25Vc2UiPjxmZUZsb29kIGZsb29kLW9wYWNpdHk9IjAiIHJlc3VsdD0iQmFja2dyb3VuZEltYWdlRml4Ii8+"
    "PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeT0iMSIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZp"
    "YXRpb249IjEiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIG9wZXJhdG9yPSJvdXQiLz48ZmVDb2xvck1hdHJpeCB2"
    "YWx1ZXM9IjAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAuMTIgMCIvPjxmZUJsZW5kIGluMj0iQmFja2dy"
    "b3VuZEltYWdlRml4IiByZXN1bHQ9ImVmZmVjdDFfZHJvcFNoYWRvd183NjFfNDE3ODYiLz48ZmVCbGVuZCBpbj0iU291cmNl"
    "R3JhcGhpYyIgaW4yPSJlZmZlY3QxX2Ryb3BTaGFkb3dfNzYxXzQxNzg2IiByZXN1bHQ9InNoYXBlIi8+PGZlQ29sb3JNYXRy"
    "aXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAw"
    "IDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iMSIgZHk9Ii0xIi8+PGZlR2F1c3NpYW5CbHVyIHN0ZERldmlhdGlv"
    "bj0iLjUiLz48ZmVDb21wb3NpdGUgaW4yPSJoYXJkQWxwaGEiIGsyPSItMSIgazM9IjEiIG9wZXJhdG9yPSJhcml0aG1ldGlj"
    "Ii8+PGZlQ29sb3JNYXRyaXggdmFsdWVzPSIwIDAgMCAwIDAuNTIwNzk3IDAgMCAwIDAgMC45MzEwMjEgMCAwIDAgMCAwLjY0"
    "NTY0OCAwIDAgMCAxIDAiLz48ZmVCbGVuZCBpbjI9InNoYXBlIiByZXN1bHQ9ImVmZmVjdDJfaW5uZXJTaGFkb3dfNzYxXzQx"
    "Nzg2Ii8+PGZlQ29sb3JNYXRyaXggaW49IlNvdXJjZUFscGhhIiByZXN1bHQ9ImhhcmRBbHBoYSIgdmFsdWVzPSIwIDAgMCAw"
    "IDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAxMjcgMCIvPjxmZU9mZnNldCBkeD0iMSIvPjxmZUdhdXNzaWFuQmx1ciBz"
    "dGREZXZpYXRpb249Ii40Ii8+PGZlQ29tcG9zaXRlIGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0i"
    "YXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZhbHVlcz0iMCAwIDAgMCAwLjMyNDg3NCAwIDAgMCAwIDAuNzc3NTEzIDAg"
    "MCAwIDAgMC40NjI2MzMgMCAwIDAgMSAwIi8+PGZlQmxlbmQgaW4yPSJlZmZlY3QyX2lubmVyU2hhZG93Xzc2MV80MTc4NiIg"
    "cmVzdWx0PSJlZmZlY3QzX2lubmVyU2hhZG93Xzc2MV80MTc4NiIvPjxmZUNvbG9yTWF0cml4IGluPSJTb3VyY2VBbHBoYSIg"
    "cmVzdWx0PSJoYXJkQWxwaGEiIHZhbHVlcz0iMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMCAwIDAgMTI3IDAiLz48"
    "ZmVPZmZzZXQgZHg9Ii0xIiBkeT0iMSIvPjxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249Ii40Ii8+PGZlQ29tcG9zaXRl"
    "IGluMj0iaGFyZEFscGhhIiBrMj0iLTEiIGszPSIxIiBvcGVyYXRvcj0iYXJpdGhtZXRpYyIvPjxmZUNvbG9yTWF0cml4IHZh"
    "bHVlcz0iMCAwIDAgMCAwLjY0NDE1OSAwIDAgMCAwIDAuOTU1MzAzIDAgMCAwIDAgMC43Mzg4NTUgMCAwIDAgMSAwIi8+PGZl"
    "QmxlbmQgaW4yPSJlZmZlY3QzX2lubmVyU2hhZG93Xzc2MV80MTc4NiIgcmVzdWx0PSJlZmZlY3Q0X2lubmVyU2hhZG93Xzc2"
    "MV80MTc4NiIvPjwvZmlsdGVyPjxsaW5lYXJHcmFkaWVudCBpZD0iZCIgeDE9IjE3LjUiIHgyPSIxNy41IiB5MT0iMi41IiB5"
    "Mj0iMTAuNSIgZ3JhZGllbnRVbml0cz0idXNlclNwYWNlT25Vc2UiPjxzdG9wIHN0b3AtY29sb3I9IiNGM0VFRUEiLz48c3Rv"
    "cCBvZmZzZXQ9IjEiIHN0b3AtY29sb3I9IiNFQ0U1REYiLz48L2xpbmVhckdyYWRpZW50PjxjbGlwUGF0aCBpZD0iYSI+PHBh"
    "dGggZmlsbD0iI2ZmZiIgZD0iTTAgMGgyNHYyNEgweiIvPjwvY2xpcFBhdGg+PC9kZWZzPjwvc3ZnPg=="
)


def theme_chart(fig):
    """Plotly 차트에 테마 톤(시스템 폰트, 화이트 배경, 옅은 그리드)을 적용."""
    fig.update_layout(
        font_family=THEME_FONT_STACK,
        font_color=THEME_COLORS["foreground"],
        plot_bgcolor="#ffffff",
        paper_bgcolor="#ffffff",
        title_font_size=17,
        title_font_color=THEME_COLORS["foreground"],
        legend_title_font_color=THEME_COLORS["muted"],
        margin=dict(l=10, r=10, t=50, b=10),
    )
    # title 텍스트를 따로 지정하지 않은 차트는 title 객체에 font만 있고 text가 없어
    # 화면에 "undefined"라는 글자가 그대로 렌더링되는 경우가 있어, 빈 문자열로 명시해둔다.
    if fig.layout.title.text is None:
        fig.update_layout(title_text="")
    fig.update_xaxes(gridcolor=THEME_COLORS["surface"], zerolinecolor=THEME_COLORS["border"], linecolor=THEME_COLORS["border"])
    fig.update_yaxes(gridcolor=THEME_COLORS["surface"], zerolinecolor=THEME_COLORS["border"], linecolor=THEME_COLORS["border"])
    return fig


def kor_date_labels(dates, unit="day"):
    """date 값들을 한글 라벨(예: '8월 1일', '2026년 8월') 리스트로 변환.
    Plotly에 date 타입을 그대로 x축으로 주면 기본 영문(예: 'Aug 1')으로 표시되므로,
    차트를 그리기 전 이 라벨로 바꾸고 px.line/px.bar의 category_orders로 순서를 고정해서 쓴다."""
    if unit == "month":
        return [f"{d.year}년 {d.month}월" for d in dates]
    return [f"{d.month}월 {d.day}일" for d in dates]


def inject_theme():
    st.markdown(
        f"""
        <style>
        html, body, [class*="css"], .stApp, .stMarkdown, .stText {{
            font-family: {THEME_FONT_STACK} !important;
        }}
        .stApp, [data-testid="stAppViewContainer"] {{
            background-color: {THEME_COLORS["canvas"]};
        }}
        [data-testid="stHeader"] {{ background-color: transparent; }}
        .block-container {{ padding-top: 2rem; padding-left: 3rem; padding-right: 3rem; max-width: 100%; }}

        [data-testid="stSidebar"] {{
            background-color: {THEME_COLORS["surface"]};
            border-right: 1px solid {THEME_COLORS["border"]};
        }}

        h1, h2, h3, h4 {{
            color: {THEME_COLORS["foreground"]} !important;
            font-weight: 700 !important;
            letter-spacing: -0.01em;
        }}
        h1 {{ font-size: 32px !important; font-weight: 700 !important; }}
        h2 {{ font-size: 24px !important; font-weight: 600 !important; }}
        h3 {{ font-size: 19px !important; font-weight: 600 !important; }}
        h4 {{ font-size: 17px !important; font-weight: 600 !important; }}
        p, span, label, div {{ color: {THEME_COLORS["body"]}; }}
        .stco-daterange {{
            display:inline-block; margin:2px 0 10px;
            padding:6px 13px; border-radius:8px;
            background:#F4F2ED; border:1px solid #E3E1DC;
            font-size:15px; font-weight:700; letter-spacing:-.01em;
            color:{THEME_COLORS["body"]};
        }}
        /* 날짜 입력칸(시작일/종료일)과 기간 셀렉트도 같이 키운다 */
        [data-testid="stDateInput"] input {{ font-size:14.5px !important; font-weight:600 !important; }}
        [data-testid="stDateInput"] label, [data-testid="stSelectbox"] label {{ font-size:12.5px !important; }}
        [data-baseweb="calendar"] {{ font-size:14px !important; }}
        [data-testid="stCaptionContainer"], .stCaption, small {{
            color: {THEME_COLORS["body"]} !important;
            font-weight: 600 !important;
            /* 기본 caption이 너무 작아 화면 안내문이 안 읽힌다 */
            font-size: 13px !important;
            line-height: 1.8 !important;
        }}

        [data-testid="stMetric"] {{
            background: {THEME_COLORS["canvas"]};
            border: 1px solid {THEME_COLORS["border"]};
            border-radius: 14px;
            padding: 16px 20px;
        }}
        [data-testid="stMetricLabel"] {{ color: {THEME_COLORS["muted"]} !important; font-weight: 400 !important; }}
        [data-testid="stMetricValue"] {{ color: {THEME_COLORS["foreground"]} !important; font-weight: 700 !important; }}

        .stButton > button, .stDownloadButton > button, .stLinkButton > a {{
            background-color: {THEME_COLORS["primary"]} !important;
            color: {THEME_COLORS["on_primary"]} !important;
            border: none !important;
            border-radius: 10px !important;
            padding: 8px 20px !important;
            font-weight: 600 !important;
            font-size: 15px !important;
            box-shadow: none !important;
            transition: background-color .15s ease;
        }}
        .stButton > button:hover, .stDownloadButton > button:hover, .stLinkButton > a:hover {{
            background-color: {THEME_COLORS["primary_hover"]} !important;
            color: {THEME_COLORS["on_primary"]} !important;
        }}
        .stButton > button[kind="secondary"] {{
            background-color: {THEME_COLORS["weak_bg"]} !important;
            color: {THEME_COLORS["weak_fg"]} !important;
            border: none !important;
        }}

        [data-testid="stTabs"] button {{ color: {THEME_COLORS["muted"]}; font-weight: 600; }}
        [data-testid="stTabs"] button[aria-selected="true"] {{
            color: {THEME_COLORS["primary"]} !important;
            border-bottom-color: {THEME_COLORS["primary"]} !important;
        }}

        [data-testid="stPopover"] {{ width: fit-content !important; }}
        [data-testid="stPopover"] > div {{ width: fit-content !important; }}
        [data-testid="stPopover"] > div > button {{
            background-color: {THEME_COLORS["canvas"]} !important;
            color: {THEME_COLORS["foreground"]} !important;
            border: 1px solid {THEME_COLORS["border"]} !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
            padding: 3px 8px !important;
            font-size: 12px !important;
            white-space: nowrap !important;
            min-width: 0 !important;
            width: fit-content !important;
        }}
        [data-testid="stPopover"] > div > button:hover {{
            background-color: {THEME_COLORS["surface"]} !important;
            color: {THEME_COLORS["foreground"]} !important;
            border-color: {THEME_COLORS["primary"]} !important;
        }}
        [data-testid="stPopoverBody"] {{
            border-radius: 12px;
            border: 1px solid {THEME_COLORS["border"]};
        }}
        div[data-baseweb="popover"] {{
            z-index: 999999 !important;
        }}

        [data-baseweb="select"] > div {{
            border-radius: 8px !important;
            border-color: {THEME_COLORS["border"]} !important;
        }}
        .stTextInput > div > div, .stDateInput > div > div {{ border-radius: 8px !important; }}

        [data-testid="stFileUploader"] section {{
            border-radius: 10px;
            border: 1px dashed {THEME_COLORS["border"]};
            background: {THEME_COLORS["surface"]};
        }}

        [data-testid="stDataFrame"] {{
            border-radius: 10px;
            overflow: hidden;
            border: 1px solid {THEME_COLORS["border"]};
        }}

        hr {{ border-color: {THEME_COLORS["border"]}; }}

        /* 사이드바 '메뉴' 내비게이션 전용 스타일 (네이버 광고관리자 스타일 참고):
           그룹 헤더는 아이콘+굵은 글씨+expander 기본 화살표(펼침/접힘)만 남기고 박스/배경 제거,
           페이지 항목은 버튼 박스가 아니라 '텍스트 링크'처럼 보이게(현재 페이지만 파란 굵은 글씨). */
        div.st-key-stco_nav div[data-testid="stExpander"] {{
            border: none !important;
            box-shadow: none !important;
            background: transparent !important;
        }}
        div.st-key-stco_nav div[data-testid="stExpander"] summary {{
            font-weight: 700 !important;
            font-size: 15px !important;
            color: {THEME_COLORS["foreground"]} !important;
            padding: 8px 4px !important;
        }}
        /* 그룹명 왼쪽에 커스텀 아이콘(SVG, 이모지 대신) 표시 */
        div.st-key-stco_nav div[data-testid="stExpander"] summary p::before {{
            content: "";
            display: inline-block;
            width: 18px;
            height: 18px;
            margin-right: 6px;
            vertical-align: -4px;
            background-image: url("data:image/svg+xml;base64,{NAV_GROUP_ICON_B64}");
            background-size: contain;
            background-repeat: no-repeat;
        }}
        /* 그룹별 아이콘 지정 — render_nav()에서 그룹마다 container(key=f"navgrp_xxx")로 한 번 더
           감싸서 붙는 고유 클래스(st-key-navgrp_xxx)로 구분한다. :nth-of-type은 Streamlit이 위젯을
           각각 별도 element-container로 감싸는 구조라 전부 "1번째"로 잡혀서 못 쓴다.
           GA 유입 리포트는 별도 아이콘을 안 받아서 위 기본값(NAV_GROUP_ICON_B64) 그대로 유지. */
        div.st-key-navgrp_report div[data-testid="stExpander"] summary p::before {{
            background-image: url("data:image/svg+xml;base64,{NAV_ICON_PERFORMANCE_B64}");
        }}
        div.st-key-navgrp_ops div[data-testid="stExpander"] summary p::before {{
            background-image: url("data:image/svg+xml;base64,{NAV_ICON_OPERATIONS_B64}");
        }}
        div.st-key-navgrp_guide div[data-testid="stExpander"] summary p::before {{
            background-image: url("data:image/svg+xml;base64,{NAV_ICON_GUIDE_B64}");
        }}
        div.st-key-stco_nav div[data-testid="stExpander"] summary:hover {{
            color: {THEME_COLORS["primary"]} !important;
        }}
        div.st-key-stco_nav div[data-testid="stExpanderDetails"] {{
            padding-left: 4px !important;
        }}
        div.st-key-stco_nav .stButton > button {{
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            text-align: left !important;
            justify-content: flex-start !important;
            padding: 6px 14px !important;
            border-radius: 8px !important;
            font-weight: 400 !important;
            font-size: 14px !important;
            color: {THEME_COLORS["body"]} !important;
        }}
        div.st-key-stco_nav .stButton > button[kind="primary"] {{
            color: {THEME_COLORS["primary"]} !important;
            font-weight: 700 !important;
            background: transparent !important;
        }}
        div.st-key-stco_nav .stButton > button:hover {{
            background: {THEME_COLORS["surface"]} !important;
            color: {THEME_COLORS["primary"]} !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


inject_theme()

TABLES = {
    "weekly_overview": "weekly_overview",
    "monthly_overview": "monthly_overview",
    "daily_overview": "daily_overview",
    "channel_monthly": "channel_monthly",
    "channel_snapshot": "channel_snapshot",
    "ga_source": "ga_source",
    "creative_performance": "creative_performance",
    "channel_audience_snapshot": "channel_audience_snapshot",
    "inflow_revenue_daily": "inflow_revenue_daily",
    "ga_channel_inflow": "ga_channel_inflow",
    "agency_notes": "agency_notes",
    "channel_weekly": "channel_weekly",
    "utm_channel_map": "utm_channel_map",
    "channel_budget": "channel_budget",
    "channel_mix_budget": "channel_mix_budget",
    "ga_channel_daily": "ga_channel_daily",
    "decision_log": "decision_log",
    "ad_spend_daily": "ad_spend_daily",
    "media_master": "media_master",
    "ad_contract": "ad_contract",
    "ga_creative_daily": "ga_creative_daily",
}

# 채널 요약 시트로 취급하지 않을 시트들
SHEET_SKIP_EXACT = {"매체통합", "GA-RAW", "RD_네이버"}
SHEET_SKIP_SUBSTR = ["_data", "_date", "확인용", "소재"]

# 채널믹스 목표 비중 (STCO 퍼포먼스마케팅 인수인계서 4절 기준: 발굴 메타50/구글20/네이버20, 회수 모비온10)
CHANNEL_MIX_TARGET = {
    "메타": 0.50,
    "구글": 0.20,
    "네이버": 0.20,
    "모비온": 0.10,
}

CHANNEL_GROUP_RULES = [
    ("메타", ["메타", "페이스북", "facebook", "meta", "인스타"]),
    ("구글", ["구글", "google"]),
    ("네이버", ["네이버", "naver", "gfa", "브검", "ssp"]),
    ("모비온", ["모비온", "mobon"]),
]


def map_channel_group(channel_name: str) -> str:
    """시트/채널명을 채널믹스 목표의 4개 그룹(메타/구글/네이버/모비온)으로 매핑.
    어디에도 안 걸리면 '기타'(크리테오·에디AI·카카오모먼트 등 목표 비중 배정이 없는 매체)."""
    name = str(channel_name).lower()
    for group, keywords in CHANNEL_GROUP_RULES:
        if any(kw.lower() in name for kw in keywords):
            return group
    return "기타"


# ──────────────────────────────────────────────────────────────
# Supabase 연결 (secrets.toml 에 SUPABASE_URL / SUPABASE_KEY 필요)
# 설정이 없으면 로컬 세션 메모리로 동작 (테스트용, 새로고침 시 초기화됨)
# ──────────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase_client():
    try:
        from supabase import create_client

        return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
    except Exception:
        return None


def _local_store():
    if "local_store" not in st.session_state:
        st.session_state["local_store"] = {k: pd.DataFrame() for k in TABLES}
    return st.session_state["local_store"]


# 데이터는 하루 단위로 들어오는데 캐시 수명이 60초면, 조작할 때마다 사실상 매번 다시 읽는다.
# 15분으로 늘리고, 즉시 갱신이 필요할 땐 화면의 '지금 동기화'가 캐시를 비우게 해둔다.
@st.cache_data(ttl=900, show_spinner=False)
def load_table(name: str) -> pd.DataFrame:
    client = get_supabase_client()
    if client is None:
        return _local_store().get(name, pd.DataFrame()).copy()

    rows, page, page_size = [], 0, 1000
    try:
        while True:
            resp = (
                client.table(TABLES[name])
                .select("*")
                .range(page * page_size, page * page_size + page_size - 1)
                .execute()
            )
            chunk = resp.data or []
            rows.extend(chunk)
            if len(chunk) < page_size:
                break
            page += 1
    except Exception as e:
        # 테이블이 아직 Supabase에 없는 경우(예: creative_performance 신규 테이블 미생성) 등
        # API 에러가 나면 앱 전체가 죽지 않도록 빈 데이터로 취급하고 안내만 띄운다.
        st.sidebar.warning(f"'{TABLES.get(name, name)}' 테이블 조회 실패 — 해당 테이블이 Supabase에 없을 수 있습니다. ({e})")
        return pd.DataFrame()
    return pd.DataFrame(rows)


def delete_creative_performance_for_date(as_of_date_value):
    """오늘자 소재별 성과를 다시 저장하기 전에, 같은 날짜로 이미 저장돼있던 이전 스냅샷을
    통째로 지운다. upsert는 '새 업로드에 있는 행'만 갱신할 뿐 '새 업로드에서 빠진 행'은
    절대 지우지 않아서, 엑셀에서 시트를 숨기거나 소재를 뺐는데도 예전에 저장된 행이
    DB에 계속 남아 화면에 나오는 문제(예: 철 지난 숨김 시트 데이터)를 막기 위함이다."""
    client = get_supabase_client()
    if client is None:
        store = _local_store()
        df = store.get("creative_performance", pd.DataFrame())
        if not df.empty and "as_of_date" in df.columns:
            store["creative_performance"] = df[df["as_of_date"].astype(str) != str(as_of_date_value)]
        return
    try:
        client.table(TABLES["creative_performance"]).delete().eq(
            "as_of_date", str(as_of_date_value)
        ).execute()
    except Exception as e:
        st.sidebar.warning(f"기존 소재별 성과 스냅샷 삭제 실패(무시하고 계속 진행합니다): {e}")


def delete_channel_audience_for_date(as_of_date_value):
    """타겟팅별 성과용 채널×오디언스(신규/리타겟팅) 스냅샷도 creative_performance와 동일한 이유로
    저장 전에 같은 날짜의 이전 스냅샷을 지운다(upsert가 빠진 행을 알아서 지워주지 않아서)."""
    client = get_supabase_client()
    if client is None:
        store = _local_store()
        df = store.get("channel_audience_snapshot", pd.DataFrame())
        if not df.empty and "as_of_date" in df.columns:
            store["channel_audience_snapshot"] = df[df["as_of_date"].astype(str) != str(as_of_date_value)]
        return
    try:
        client.table(TABLES["channel_audience_snapshot"]).delete().eq(
            "as_of_date", str(as_of_date_value)
        ).execute()
    except Exception as e:
        st.sidebar.warning(f"기존 채널×오디언스 스냅샷 삭제 실패(무시하고 계속 진행합니다): {e}")


def save_table(name: str, df: pd.DataFrame, on_conflict: str, source_file: str):
    if df is None or df.empty:
        return 0
    df = df.copy()
    # 같은 업로드 안에 동일 키(예: creative_performance라면 같은 날짜+매체+소재명) 행이 중복되면
    # upsert 한 번의 요청 안에서 같은 행을 두 번 건드리게 되어 Postgres가 에러를 내므로, 저장 전에
    # 미리 정리해야 한다. 예전에는 단순 drop_duplicates(keep="last")로 나머지를 버렸는데, 같은
    # 소재가 여러 캠페인/그룹에 동시에 쓰이는 경우(예: '260630_슈즈'가 리타겟팅/프로스펙팅 캠페인에
    # 둘 다 있는 경우) 실제로는 두 행을 '합산'해야 할 실적인데 한쪽을 통째로 버려서 노출·광고비·매출이
    # 실제보다 훨씬 작게 저장되는 버그가 있었다. 그래서 중복 키가 있으면 숫자 컬럼은 합산하고,
    # 문자열/날짜 등 나머지 컬럼은 비어있지 않은 첫 값을 사용하도록 바꾼다.
    key_cols = [c.strip() for c in on_conflict.split(",")]
    if df.duplicated(subset=key_cols, keep=False).any():
        agg_map = {}
        for col in df.columns:
            if col in key_cols:
                continue
            if pd.api.types.is_numeric_dtype(df[col]):
                agg_map[col] = "sum"
            else:
                agg_map[col] = lambda s: next((v for v in s if pd.notna(v) and v != ""), None)
        df = df.groupby(key_cols, as_index=False).agg(agg_map)
    df["source_file"] = source_file
    df["uploaded_at"] = datetime.utcnow().isoformat()

    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].astype(str)
        elif df[col].dtype == object:
            df[col] = df[col].apply(lambda v: v.isoformat() if isinstance(v, (date, datetime)) else v)

    # NaN/inf는 표준 JSON에 없는 값이라 Supabase API 호출(JSON 직렬화) 자체가 통째로 실패한다
    # ("Out of range float values are not JSON compliant: nan"). 어떤 컬럼에서 오든(신규 image_url
    # 컬럼이 전부 비어 float64/NaN으로 잡히는 경우 등) 저장 직전에 전부 None으로 정리해 방지한다.
    # (float64 컬럼은 dtype을 object로 바꾸지 않으면 None이 다시 NaN으로 되돌아가므로 astype(object) 필수)
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.astype(object).where(pd.notnull(df), None)

    client = get_supabase_client()
    if client is None:
        store = _local_store()
        prev = store.get(name, pd.DataFrame())
        merged = pd.concat([prev, df], ignore_index=True)
        keys = on_conflict.split(",")
        merged = merged.drop_duplicates(subset=keys, keep="last")
        store[name] = merged
        return len(df)

    records = df.to_dict(orient="records")
    try:
        for i in range(0, len(records), 500):
            client.table(TABLES[name]).upsert(records[i : i + 500], on_conflict=on_conflict).execute()
    except Exception as e:
        # 예: creative_performance처럼 Supabase에 아직 테이블을 안 만든 경우 — 이 테이블만 건너뛰고
        # 나머지 저장(주간/월별/매체 등)은 정상 진행되도록 전체 저장 흐름을 막지 않는다.
        st.sidebar.error(f"'{TABLES.get(name, name)}' 저장 실패 — Supabase에 해당 테이블이 있는지 확인해주세요. ({e})")
        return 0
    return len(df)


# ──────────────────────────────────────────────────────────────
# 파싱 유틸
# ──────────────────────────────────────────────────────────────
def clean_col(c) -> str:
    if c is None:
        return ""
    return str(c).replace("\n", "").replace(" ", "").strip()


def match_col(columns, include_all=None, include_any=None, exclude=None):
    include_all, include_any, exclude = include_all or [], include_any or [], exclude or []
    for c in columns:
        lc = clean_col(c).lower()
        if not lc:
            continue
        if any(ex in lc for ex in exclude):
            continue
        if include_all and not all(tok in lc for tok in include_all):
            continue
        if include_any and not any(tok in lc for tok in include_any):
            continue
        return c
    return None


def metric_cols(columns):
    return dict(
        impr=match_col(columns, include_any=["노출"]),
        clicks=match_col(columns, include_any=["클릭"]),
        cost_ex=match_col(columns, include_all=["제외"], include_any=["광고비", "비용"], exclude=["마크업", "최종"]),
        cost_in=match_col(columns, include_all=["포함"], include_any=["광고비", "비용"], exclude=["마크업"]),
        signup=match_col(columns, include_any=["가입"]),
        conv=match_col(columns, include_all=["전환"], exclude=["금액", "ga", "율"]),
        rev=match_col(columns, include_any=["매출", "전환금액"], exclude=["ga", "객단가"]),
        ga_conv=match_col(columns, include_all=["ga"], include_any=["전환"]),
        ga_rev=match_col(columns, include_all=["ga"], include_any=["매출"]),
    )


def numcol(data: pd.DataFrame, c):
    if not c or c not in data.columns:
        return np.zeros(len(data))
    return pd.to_numeric(data[c], errors="coerce").fillna(0).values


def match_col_pos(headers, include_all=None, include_any=None, exclude=None):
    """match_col과 같은 규칙이지만 컬럼 '이름' 대신 헤더 리스트 안에서의 위치(정수 인덱스)를 반환한다.
    소재 단위 raw 시트는 같은 이름의 컬럼(예: 자체 전환/매출/ROAS 다음에 GA 전환/매출/ROAS가
    똑같이 '전환'/'매출'/'ROAS'라는 이름으로 한 번 더 나오는 경우)이 있어서, 이름으로 data[name]을
    하면 중복 컬럼이 DataFrame으로 잡혀 계산이 깨진다. 위치 기반으로 첫 매칭 컬럼 하나만 정확히 집는다."""
    include_all, include_any, exclude = include_all or [], include_any or [], exclude or []
    for i, c in enumerate(headers):
        lc = clean_col(c).lower()
        if not lc:
            continue
        if any(ex in lc for ex in exclude):
            continue
        if include_all and not all(tok in lc for tok in include_all):
            continue
        if include_any and not any(tok in lc for tok in include_any):
            continue
        return i
    return None


def numcol_by_pos(block: pd.DataFrame, idx):
    """위치(정수 인덱스) 기반으로 숫자 컬럼 값을 꺼낸다 — 컬럼명이 중복돼도 안전하다."""
    if idx is None or idx >= block.shape[1]:
        return np.zeros(len(block))
    return pd.to_numeric(block.iloc[:, idx], errors="coerce").fillna(0).values


def find_header_row(raw: pd.DataFrame, required=("노출수", "클릭수"), scan=10):
    for i in range(min(scan, len(raw))):
        row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
        if all(tok in row_text for tok in required):
            return i
    return None


SECTION_MARKERS = {
    "monthly": ["월별 통합데이터", "월간 데이터"],
    "channel_snap": ["매체별 현황"],
    "weekly": ["통합 주간별", "주간 데이터"],
    "daily": ["통합 일자별", "일일 데이터", "일별 데이터"],
}


def find_sections(raw: pd.DataFrame, scan_cols=(0, 1, 2)):
    """시트 안에 세로로 쌓인 여러 표(월간/주간/일별/매체별 현황 등)의 경계를 찾는다.
    각 표는 '■ 월간 데이터' 같은 제목 행 다음 줄이 헤더, 그 다음부터 다음 제목 전까지가 데이터."""
    hits = []
    for i in range(len(raw)):
        for col in scan_cols:
            if col >= raw.shape[1]:
                continue
            v = raw.iat[i, col]
            if isinstance(v, str):
                for key, tokens in SECTION_MARKERS.items():
                    if any(tok in v for tok in tokens):
                        hits.append((i, key))
                        break
    hits.sort()
    bounds = {}
    for idx, (row_i, key) in enumerate(hits):
        end = hits[idx + 1][0] if idx + 1 < len(hits) else len(raw)
        bounds.setdefault(key, (row_i, end))
    return bounds


def section_dataframe(raw: pd.DataFrame, start_row: int, end_row: int, date_tokens=("기간", "월별")):
    """섹션 제목(start_row) 바로 다음 줄을 헤더로 보고 데이터프레임 구성."""
    header_row = start_row + 1
    headers = raw.iloc[header_row].tolist()
    date_idx = None
    for i, h in enumerate(headers):
        if clean_col(h) in date_tokens:
            date_idx = i
            break
    if date_idx is None:
        return None, None
    data = raw.iloc[header_row + 1 : end_row].copy()
    data.columns = headers
    data = data[data.iloc[:, date_idx].notna()]
    return data, date_idx


def parse_monthly(raw: pd.DataFrame, bounds, today: date):
    if "monthly" not in bounds:
        return pd.DataFrame()
    data, date_idx = section_dataframe(raw, *bounds["monthly"], date_tokens=("월별", "기간"))
    if data is None or data.empty:
        return pd.DataFrame()
    m = metric_cols(list(data.columns))
    out = pd.DataFrame()
    out["report_month"] = pd.to_datetime(data.iloc[:, date_idx], errors="coerce")
    out = out[out["report_month"].notna()]
    data = data.loc[out.index]
    out["impressions"] = numcol(data, m["impr"])
    out["clicks"] = numcol(data, m["clicks"])
    out["cost_excl_vat"] = numcol(data, m["cost_ex"])
    out["cost_incl_vat"] = numcol(data, m["cost_in"])
    out["signups"] = numcol(data, m["signup"])
    out["conversions"] = numcol(data, m["conv"])
    out["revenue"] = numcol(data, m["rev"])
    out["ga_conversions"] = numcol(data, m["ga_conv"])
    out["ga_revenue"] = numcol(data, m["ga_rev"])
    out["report_month"] = out["report_month"].dt.date
    cutoff = today.replace(day=1)
    out = out[out["report_month"] <= cutoff]
    return out.reset_index(drop=True)


def _parse_weekly_section(raw: pd.DataFrame, bounds, today: date) -> pd.DataFrame:
    """'■ 주간 데이터'(또는 '통합 주간별') 섹션 하나를 파싱하는 공용 코어.
    '매체통합' 시트의 통합 주간별 표뿐 아니라, 매체 개별 시트(네이버/GFA/메타/구글/크리테오 등)에
    똑같이 있는 '■ 주간 데이터' 표에도 그대로 쓴다 — 구조가 동일하기 때문."""
    if "weekly" not in bounds:
        return pd.DataFrame()
    data, date_idx = section_dataframe(raw, *bounds["weekly"], date_tokens=("기간", "월별"))
    if data is None or data.empty:
        return pd.DataFrame()
    m = metric_cols(list(data.columns))

    year_state = {"year": None, "prev_month": None}
    rows = []
    for _, r in data.iterrows():
        label = r.iloc[date_idx]
        mm = re.search(r"\((\d{1,2})/(\d{1,2})\s*~\s*(\d{1,2})/(\d{1,2})\)", str(label))
        lead = re.search(r"^(\d{1,2})월", str(label))
        if not mm or not lead:
            continue
        month_lead = int(lead.group(1))
        if year_state["year"] is None:
            # 주간 섹션은 같은 시트의 월간 섹션 첫 달과 같은 해에서 시작한다고 본다.
            # 컬럼 순서가 시트마다 달라서(예: '매체통합'은 날짜가 1번째 컬럼, 매체 개별
            # 시트는 0번째 컬럼인 경우 등) 위치를 하드코딩하지 않고 section_dataframe으로
            # 실제 날짜 컬럼을 다시 찾아서 연도를 뽑는다.
            year_val = None
            if "monthly" in bounds:
                m_data, m_date_idx = section_dataframe(raw, *bounds["monthly"], date_tokens=("월별", "기간"))
                if m_data is not None and not m_data.empty:
                    first_date = pd.to_datetime(m_data.iloc[0, m_date_idx], errors="coerce")
                    if pd.notna(first_date):
                        year_val = first_date.year
            year_state["year"] = year_val if year_val is not None else today.year
        elif year_state["prev_month"] is not None and month_lead < year_state["prev_month"] - 6:
            year_state["year"] += 1
        year_state["prev_month"] = month_lead

        sm, sd, em, ed = map(int, mm.groups())
        s_year = year_state["year"]
        e_year = s_year if em >= sm else s_year + 1
        try:
            week_start = date(s_year, sm, sd)
            week_end = date(e_year, em, ed)
        except ValueError:
            continue

        rows.append(
            {
                "week_start": week_start,
                "week_end": week_end,
                "label": str(label).strip(),
                "impressions": float(pd.to_numeric(r.get(m["impr"]), errors="coerce") or 0) if m["impr"] else 0,
                "clicks": float(pd.to_numeric(r.get(m["clicks"]), errors="coerce") or 0) if m["clicks"] else 0,
                "cost_excl_vat": float(pd.to_numeric(r.get(m["cost_ex"]), errors="coerce") or 0) if m["cost_ex"] else 0,
                "cost_incl_vat": float(pd.to_numeric(r.get(m["cost_in"]), errors="coerce") or 0) if m["cost_in"] else 0,
                "signups": float(pd.to_numeric(r.get(m["signup"]), errors="coerce") or 0) if m["signup"] else 0,
                "conversions": float(pd.to_numeric(r.get(m["conv"]), errors="coerce") or 0) if m["conv"] else 0,
                "revenue": float(pd.to_numeric(r.get(m["rev"]), errors="coerce") or 0) if m["rev"] else 0,
            }
        )
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out[out["week_end"] <= today]
    out = out.sort_values("week_start").reset_index(drop=True)
    # parse_daily와 동일한 이유: 리포트 템플릿이 아직 안 지난 미래 주차까지 행(라벨)을 미리
    # 만들어두는 경우가 있어서, week_end<=today만으로는 다 안 걸러진다. 실제 값이 하나도 없는
    # (전부 0인) 말미 주차는 '아직 보고 안 된 주'로 보고 잘라낸다.
    metric_sum = out[["impressions", "clicks", "cost_excl_vat", "cost_incl_vat", "conversions", "revenue"]].sum(axis=1)
    nonzero_idx = metric_sum[metric_sum > 0].index
    if len(nonzero_idx):
        out = out.loc[: nonzero_idx.max()]
    return out.reset_index(drop=True)


def parse_weekly(raw: pd.DataFrame, bounds, today: date) -> pd.DataFrame:
    """'매체통합' 시트의 통합 주간별 표(전체 합산)를 파싱한다."""
    return _parse_weekly_section(raw, bounds, today)


def parse_daily(raw: pd.DataFrame, bounds, today: date):
    """'3) 통합 일자별' 표를 파싱한다. 날짜 컬럼명은 '일자', 바로 옆에 '요일' 컬럼이 있다."""
    if "daily" not in bounds:
        return pd.DataFrame()
    data, date_idx = section_dataframe(raw, *bounds["daily"], date_tokens=("일자", "기간", "월별"))
    if data is None or data.empty:
        return pd.DataFrame()
    m = metric_cols(list(data.columns))
    out = pd.DataFrame()
    out["report_date"] = pd.to_datetime(data.iloc[:, date_idx], errors="coerce")
    out = out[out["report_date"].notna()]
    data = data.loc[out.index]
    out["impressions"] = numcol(data, m["impr"])
    out["clicks"] = numcol(data, m["clicks"])
    out["cost_excl_vat"] = numcol(data, m["cost_ex"])
    out["cost_incl_vat"] = numcol(data, m["cost_in"])
    out["signups"] = numcol(data, m["signup"])
    out["conversions"] = numcol(data, m["conv"])
    out["revenue"] = numcol(data, m["rev"])
    out["report_date"] = out["report_date"].dt.date
    out = out[out["report_date"] <= today]
    out = out.sort_values("report_date").reset_index(drop=True)
    # 아직 보고되지 않은(전부 0인) 말미 날짜는 잘라낸다 (리포트 템플릿의 미래 placeholder 행)
    metric_sum = out[["impressions", "clicks", "cost_excl_vat", "cost_incl_vat", "conversions", "revenue"]].sum(axis=1)
    nonzero_idx = metric_sum[metric_sum > 0].index
    if len(nonzero_idx):
        out = out.loc[: nonzero_idx.max()]
    return out.reset_index(drop=True)


def parse_channel_snapshot(raw: pd.DataFrame, bounds, monthly_df: pd.DataFrame):
    if "channel_snap" not in bounds:
        return pd.DataFrame()
    data, date_idx_unused = section_dataframe(raw, *bounds["channel_snap"], date_tokens=("매체",))
    if data is None or data.empty:
        return pd.DataFrame()
    m = metric_cols(list(data.columns))
    channel_col = data.columns[1] if clean_col(data.columns[1]) == "매체" else data.columns[0]
    out = pd.DataFrame()
    out["channel"] = data[channel_col].astype(str)
    out = out[~out["channel"].str.contains("TOTAL", case=False, na=False)]
    data = data.loc[out.index]
    out["impressions"] = numcol(data, m["impr"])
    out["clicks"] = numcol(data, m["clicks"])
    out["cost_excl_vat"] = numcol(data, m["cost_ex"])
    out["cost_incl_vat"] = numcol(data, m["cost_in"])
    out["signups"] = numcol(data, m["signup"])
    out["conversions"] = numcol(data, m["conv"])
    out["revenue"] = numcol(data, m["rev"])
    out["ga_conversions"] = numcol(data, m["ga_conv"])
    out["ga_revenue"] = numcol(data, m["ga_rev"])
    as_of = monthly_df["report_month"].max() if len(monthly_df) else date.today().replace(day=1)
    out["as_of_month"] = as_of
    return out.reset_index(drop=True)


# 대행사가 '매체통합' 시트 하단에 자유 텍스트로 남기는 운영 메모(특이사항/코멘트)를 찾을 때
# 상위 불릿(예: "*. 7월 신규 예산 반영 완료")으로 인정할 접두어.
AGENCY_NOTE_BULLET_PREFIXES = ("*", "-", "•", "※", "▶", "①", "②", "③", "④", "⑤")
# ">"로 시작하는 줄은 바로 위 상위 불릿에 딸린 하위/부연 설명으로 취급(들여쓰기해서 표시).
AGENCY_NOTE_SUB_PREFIX = ">"


def parse_agency_notes(raw: pd.DataFrame, today: date) -> pd.DataFrame:
    """'매체통합' 시트 하단에 대행사가 정리해주는 자유 텍스트 운영 코멘트/특이사항을 추출한다.
    상단 표(월간/주간/일별/매체별 현황)는 각 행이 숫자 위주 데이터라 첫 칸이 불릿 마커로
    시작하지 않는다는 점을 이용해, 각 행의 첫 번째 비어있지 않은 문자열 셀이 '*.', '-', '>' 같은
    불릿 마커로 시작하는 행만 순서대로 모아 하나의 텍스트로 합친다. '>'로 시작하는 줄은 바로 위
    상위 항목의 하위 설명으로 보고 들여써서 표시한다."""
    if raw is None or raw.empty:
        return pd.DataFrame()

    items = []  # (level, text) — level 0=상위 불릿, 1=하위(">")
    n_rows, n_cols = raw.shape
    for i in range(n_rows):
        first_str = None
        for j in range(n_cols):
            v = raw.iat[i, j]
            if isinstance(v, str) and v.strip():
                first_str = v.strip()
                break
        if not first_str or len(first_str) < 2:
            continue
        if first_str.startswith(AGENCY_NOTE_SUB_PREFIX):
            text = first_str.lstrip(AGENCY_NOTE_SUB_PREFIX + " ").strip()
            if text:
                items.append((1, text))
        elif first_str.startswith(AGENCY_NOTE_BULLET_PREFIXES):
            # "*." "* " 등 뒤에 붙는 마침표/공백까지 같이 제거 (예: "*. 쇼핑박스..." → "쇼핑박스...")
            text = first_str.lstrip("".join(AGENCY_NOTE_BULLET_PREFIXES) + ". ").strip()
            if text:
                items.append((0, text))

    if not items:
        return pd.DataFrame()

    note_text = "\n".join(f"    - {t}" if level == 1 else f"- {t}" for level, t in items)
    return pd.DataFrame([{"as_of_date": today, "note_text": note_text}])


def discover_channel_sheets(xls: pd.ExcelFile):
    names = []
    for s in xls.sheet_names:
        if s in SHEET_SKIP_EXACT:
            continue
        low = s.lower()
        if any(p in low for p in SHEET_SKIP_SUBSTR):
            continue
        names.append(s)
    return names


def _drop_hidden_rows(raw: pd.DataFrame, hidden_rows: set | None) -> pd.DataFrame:
    """엑셀에서 행 자체를 숨겨둔(우클릭 → 행 숨기기) 행은 화면에 보이는 표와 무관한
    예전 버전/임시 데이터인 경우가 많다. 위치 기반(iloc) 파싱이 엉뚱한 행을 진짜 표로
    착각하지 않도록, 섹션을 찾기 전에 아예 원본에서 제거하고 행 번호를 다시 매긴다."""
    if not hidden_rows:
        return raw
    keep_idx = [i for i in range(len(raw)) if i not in hidden_rows]
    if len(keep_idx) == len(raw):
        return raw
    return raw.iloc[keep_idx].reset_index(drop=True)


def parse_channel_sheet(xls: pd.ExcelFile, sheet: str, today: date, hidden_rows: set | None = None):
    """매체 개별 시트(네이버/GFA/메타/구글/크리테오 등)에서 월간 데이터를 파싱한다.
    (monthly_df, weekly_df) 튜플을 반환한다 — weekly_df는 그 시트에 '■ 주간 데이터' 섹션이
    있을 때만 채워지고(대부분의 개별 매체 시트에 있음), 없으면 빈 DataFrame이다.
    hidden_rows: 엑셀에서 숨겨진 행 번호(raw의 0-indexed 위치) 집합 — 있으면 파싱 전에 미리
    제거한다. 시트 안에 예전 버전 표가 통째로 숨겨진 채 남아있어 잘못 잡히는 걸 막기 위함."""
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)
    raw = _drop_hidden_rows(raw, hidden_rows)
    bounds = find_sections(raw)
    if "monthly" in bounds:
        data, date_idx = section_dataframe(raw, *bounds["monthly"], date_tokens=("기간", "월별"))
    else:
        # '■ 월간 데이터' 같은 섹션 제목이 없는 단순 시트는 기존 방식으로 폴백
        hdr = find_header_row(raw)
        if hdr is None:
            return None, pd.DataFrame()
        headers = raw.iloc[hdr].tolist()
        date_idx = next((i for i, h in enumerate(headers) if clean_col(h) in ("기간", "월별")), None)
        if date_idx is None:
            return None, pd.DataFrame()
        data = raw.iloc[hdr + 1 :].copy()
        data.columns = headers
        data = data[data.iloc[:, date_idx].notna()]

    weekly_out = pd.DataFrame()
    if "weekly" in bounds:
        weekly_out = _parse_weekly_section(raw, bounds, today)
        if not weekly_out.empty:
            weekly_out["channel"] = sheet

    if data is None or data.empty:
        return None, weekly_out
    m = metric_cols(list(data.columns))
    out = pd.DataFrame()
    out["report_month"] = pd.to_datetime(data.iloc[:, date_idx], errors="coerce")
    out = out[out["report_month"].notna()]
    data = data.loc[out.index]
    out["impressions"] = numcol(data, m["impr"])
    out["clicks"] = numcol(data, m["clicks"])
    out["cost_excl_vat"] = numcol(data, m["cost_ex"])
    out["cost_incl_vat"] = numcol(data, m["cost_in"])
    out["signups"] = numcol(data, m["signup"])
    out["conversions"] = numcol(data, m["conv"])
    out["revenue"] = numcol(data, m["rev"])
    out["report_month"] = out["report_month"].dt.date
    out = out[out["report_month"] <= today.replace(day=1)]
    out["channel"] = sheet
    return out.reset_index(drop=True), weekly_out


def parse_ga_raw(xls: pd.ExcelFile, today: date):
    if "GA-RAW" not in xls.sheet_names:
        return pd.DataFrame()
    raw = pd.read_excel(xls, sheet_name="GA-RAW")
    raw.columns = [clean_col(c) for c in raw.columns]
    rename = {
        "매체": "source_medium",
        "사용자": "users",
        "신규방문자": "new_users",
        "세션": "sessions",
        "이탈률": "bounce_rate",
        "세션당페이지수": "pages_per_session",
        "평균세션시간": "avg_session_duration",
        "전자상거래전환율": "ecommerce_cvr",
        "거래수": "transactions",
        "수익": "revenue",
    }
    raw = raw.rename(columns=rename)
    keep = [c for c in rename.values() if c in raw.columns]
    out = raw[keep].dropna(subset=["source_medium"]).copy()
    for c in keep:
        if c != "source_medium":
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
    out["as_of_date"] = today
    return out.reset_index(drop=True)


INFLOW_REVENUE_SHEET = "일별 GA,어드민 지표 비교"

# (엑셀 열 위치 0-based, 우리 컬럼명, 퍼센트 변환 배수) — 이름이 아니라 '위치'로 읽는 이유:
# 시트 헤더가 4행이고 그 위에 병합된 그룹 헤더(GA/자체 분석 툴/매체 기준/GA 기준)가 있어서
# "결제수·CVR·AOV·매출액·ROAS"가 두 번씩 반복되는데, pandas로 이름 기준 read_excel을 하면
# 뒤에 나온 컬럼이 앞 컬럼을 덮어써버려 데이터가 사라진다. 열 위치 고정이 훨씬 안전하다.
INFLOW_REVENUE_COL_MAP = [
    (0, "report_date", None),
    (1, "users", None),
    (2, "new_users", None),
    (3, "returning_users", None),
    (4, "signups", None),
    (5, "bounce_rate", 100),
    # 6번(인덱스6, 숫자 초 단위)은 실제로는 대부분 비어있고(372행 중 31행만 값 존재),
    # 7번(인덱스7, "3분 26초" 텍스트)이 거의 매일 채워져 있어서 이쪽을 파싱해서 쓴다 — 아래
    # avg_session_duration 처리 참고. 6번은 그래서 COL_MAP에서 건너뛴다.
    (8, "pageviews", None),
    (9, "admin_orders_total", None),
    (10, "admin_orders_real", None),
    (11, "admin_qty", None),
    (12, "admin_revenue", None),          # 회사 내부(어드민) 기준 매출
    (13, "price_rate", 100),
    (14, "admin_aov", None),
    (15, "admin_returns", None),
    (16, "return_rate", 100),
    # 17번(인덱스17)은 "매체 TOTAL" 아래 날짜 중복 컬럼이라 건너뜀
    (18, "impressions", None),
    (19, "clicks", None),
    (20, "ctr", 100),
    (21, "cpc", None),
    (22, "cost_incl_vat", None),
    (23, "conversions", None),            # 매체 리포트 기준 결제수
    (24, "cvr", 100),
    (25, "aov", None),
    (26, "revenue", None),                # 매체 리포트(보고서) 기준 매출
    (27, "roas", 100),
    (28, "ga_conversions", None),
    (29, "ga_cvr", 100),
    (30, "ga_aov", None),
    (31, "ga_revenue", None),             # GA 기준 매출
    (32, "ga_roas", 100),
    (33, "new_paying_customers", None),
    (34, "cac", None),
    # 35~39(인덱스)는 퍼널(1.유입~5.구매) 라벨 텍스트 칸이라 실데이터가 아니어서 제외
]


def parse_inflow_revenue_sheet(xls: pd.ExcelFile):
    """형이 별도로 정리해서 준 '일별 GA,어드민 지표 비교' 파일 전용 파서.
    5~11행에는 월별 요약행이 먼저 있고 그 아래(12행~)부터 실제 일별 데이터가 시작되는데,
    요약행은 날짜 칸이 '2026년 1월' 같은 문자열이라 datetime 파싱이 실패하는 것으로 자연스럽게
    걸러낸다. 아직 값이 없는 미래 날짜(연말까지 미리 만들어둔 빈 행)는 방문자수가 비어있는 것으로
    걸러낸다."""
    if INFLOW_REVENUE_SHEET not in xls.sheet_names:
        return pd.DataFrame()
    raw = pd.read_excel(xls, sheet_name=INFLOW_REVENUE_SHEET, header=None, skiprows=4)
    if raw.empty:
        return pd.DataFrame()

    out = pd.DataFrame(index=raw.index)
    for idx, name, scale in INFLOW_REVENUE_COL_MAP:
        if idx >= raw.shape[1]:
            out[name] = pd.NA
            continue
        col = raw.iloc[:, idx]
        if name == "report_date":
            out[name] = col
            continue
        col = pd.to_numeric(col, errors="coerce")
        if scale:
            col = col * scale
        out[name] = col

    # 평균 체류시간 — "3분 26초" 같은 텍스트(8번째 열, 인덱스7)를 초 단위 숫자로 변환.
    if raw.shape[1] > 7:
        dur_text = raw.iloc[:, 7].astype(str)
        dur_match = dur_text.str.extract(r"(?:(\d+)\s*분)?\s*(?:(\d+)\s*초)?")
        minutes = pd.to_numeric(dur_match[0], errors="coerce").fillna(0)
        seconds = pd.to_numeric(dur_match[1], errors="coerce").fillna(0)
        parsed = minutes * 60 + seconds
        parsed[~dur_text.str.contains("분|초", na=False)] = pd.NA
        out["avg_session_duration"] = parsed
    else:
        out["avg_session_duration"] = pd.NA

    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce")
    out = out.dropna(subset=["report_date", "users"]).copy()
    out["report_date"] = out["report_date"].dt.date
    return out.reset_index(drop=True)


# GA 내보내기 화면/버전마다 컬럼 순서·개수가 달라질 수 있어(뒤에 빈 컬럼이나 메모가 붙기도
# 하고, '거래'/'세션수' 같은 컬럼이 추가/누락되기도 함) 위치가 아니라 헤더 '이름'으로 찾는다.
# 후보를 여러 개 둔 건 GA4 내보내기 화면 언어/버전에 따라 라벨이 조금씩 달라질 수 있어서다.
GA_CHANNEL_INFLOW_HEADER_CANDIDATES = {
    "report_date": ["날짜"],
    "source_medium": ["세션 소스/매체", "세션 소스 / 매체", "소스/매체", "소스 / 매체"],
    "users": ["총 사용자", "사용자"],
    "new_users": ["새 사용자 수", "새 사용자"],
    "returning_users": ["재 사용자", "재방문 사용자", "재방문자"],
    "bounce_rate": ["이탈률"],
    "pageviews": ["조회수"],
    "avg_session_duration": ["평균 세션 시간", "평균 참여 시간", "평균 체류 시간"],
    "conversions": ["거래", "구매", "전환수"],
    "revenue": ["총수익", "수익", "매출"],
}
GA_CHANNEL_INFLOW_PCT_FIELDS = {"bounce_rate"}  # 0~1 비율로 들어오면 %로 쓰기 위해 100을 곱함


def _match_header_index(columns: list, candidates: list) -> int:
    """컬럼명 리스트에서 후보 이름과 '정확히' 일치하는 첫 컬럼의 위치를 찾는다. 정확히 일치하는
    것만 인정해서(부분일치 X) '조회수'와 '세션당 조회수'처럼 한쪽이 다른 쪽을 포함하는 헤더가
    섞여 있어도 엉뚱한 컬럼을 잡지 않는다."""
    cols_clean = [str(c).strip() for c in columns]
    for cand in candidates:
        for i, c in enumerate(cols_clean):
            if c == cand:
                return i
    return None


def parse_ga_channel_inflow_sheet(xls: pd.ExcelFile, channel_map: dict = None):
    """'GA 세션소스 매체&일자별 데이터' 같은 GA 원본 내보내기 파일 전용 파서.
    날짜×세션 소스/매체 단위로 하루 단위 데이터가 들어있다(하루에 소스/매체별로 한 행씩).
    '매체'(채널 그룹핑, 예: (SA) 네이버/구글(P-MAX) 등)는 UTM 소스/매체 원본에는 없는 값이라,
    별도로 학습해둔 utm_channel_map(build_utm_channel_lookup) 매핑을 channel_map으로 받아
    source_medium 문자열로 채운다(대소문자·공백 차이는 무시하고 매칭). 매핑에 없는 값은
    None으로 남겨 화면에서 '미매핑'으로 구분할 수 있게 한다."""
    if not xls.sheet_names:
        return pd.DataFrame()
    sheet = xls.sheet_names[0]
    raw = pd.read_excel(xls, sheet_name=sheet, header=0)
    if raw.empty:
        return pd.DataFrame()

    columns = list(raw.columns)
    idx_map = {
        name: _match_header_index(columns, candidates)
        for name, candidates in GA_CHANNEL_INFLOW_HEADER_CANDIDATES.items()
    }
    # 날짜/소스매체 컬럼을 못 찾으면 이 파일은 GA 채널 유입 파일이 아니라고 보고 빈 결과를 낸다
    # (같은 업로더에서 다른 파서들도 같이 시도되므로, 여기서 조용히 빠지는 게 맞다).
    if idx_map.get("report_date") is None or idx_map.get("source_medium") is None:
        return pd.DataFrame()

    out = pd.DataFrame(index=raw.index)
    for name, idx in idx_map.items():
        if idx is None:
            out[name] = pd.NA
            continue
        col = raw.iloc[:, idx]
        if name in ("report_date", "source_medium"):
            out[name] = col
            continue
        col = pd.to_numeric(col, errors="coerce")
        if name in GA_CHANNEL_INFLOW_PCT_FIELDS:
            col = col * 100
        out[name] = col

    # 날짜가 "2026. 1. 1." 같은 텍스트 포맷이라 고정 포맷으로 우선 파싱하고, 혹시 다른 형식이
    # 섞여 들어와도 죽지 않도록 실패한 행만 일반 파싱으로 한 번 더 시도한다.
    parsed = pd.to_datetime(out["report_date"], errors="coerce", format="%Y. %m. %d.")
    fallback = parsed.isna()
    if fallback.any():
        parsed.loc[fallback] = pd.to_datetime(out.loc[fallback, "report_date"], errors="coerce")
    out["report_date"] = parsed

    out = out.dropna(subset=["report_date", "source_medium"]).copy()
    out["report_date"] = out["report_date"].dt.date
    if channel_map:
        out["channel"] = out["source_medium"].map(
            lambda sm: channel_map.get(str(sm).strip().lower())
        )
    else:
        out["channel"] = None
    return out.reset_index(drop=True)


def parse_utm_channel_map(xls: pd.ExcelFile) -> pd.DataFrame:
    """'STCO_UTM 리스트' 같은 UTM 소스/매체 ↔ 보고서 매체명 매핑표 파일을 파싱한다.
    헤더 위치가 고정돼 있지 않을 수 있어(형이 셀을 밀거나 시트를 늘릴 수 있으니), '매체명'과
    '소스'/'매체'가 같이 있는 행을 헤더로 찾고, 그 아래 두 컬럼(매체명/소스·매체)이 둘 다
    채워진 행만 데이터로 읽는다. '비고' 컬럼이 있으면 참고용으로 같이 저장한다."""
    rows = []
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        if raw.empty:
            continue
        header_row, chan_col, src_col, note_col = None, None, None, None
        for r in range(min(10, len(raw))):
            row_vals = raw.iloc[r].tolist()
            c_idx = next((i for i, v in enumerate(row_vals) if "매체명" in str(v)), None)
            s_idx = next((i for i, v in enumerate(row_vals) if "소스" in str(v) and "매체" in str(v)), None)
            if c_idx is not None and s_idx is not None:
                header_row, chan_col, src_col = r, c_idx, s_idx
                note_col = next((i for i, v in enumerate(row_vals) if "비고" in str(v)), None)
                break
        if header_row is None:
            continue
        for r in range(header_row + 1, len(raw)):
            channel = raw.iloc[r, chan_col] if chan_col < raw.shape[1] else None
            source_medium = raw.iloc[r, src_col] if src_col < raw.shape[1] else None
            if pd.isna(channel) or pd.isna(source_medium) or not str(channel).strip() or not str(source_medium).strip():
                continue
            note = raw.iloc[r, note_col] if note_col is not None and note_col < raw.shape[1] else None
            rows.append({
                "source_medium": str(source_medium).strip(),
                "channel": str(channel).strip(),
                "note": str(note).strip() if pd.notna(note) else None,
            })
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    # 같은 소스/매체가 여러 시트/행에 중복돼 있으면 마지막(가장 아래) 값을 채택한다.
    out = out.drop_duplicates(subset=["source_medium"], keep="last").reset_index(drop=True)
    return out


# '26년 매체별 채널 믹스' 같은 연간 채널 예산 파일 전용 파서. '◆26년 월별 예산 정리'(자사몰
# 예산 박스형 표, parse_channel_budget_sheet)와는 다른 파일이다 — 이쪽은 전체 매체 집행
# 예산(자사몰 한정 아님)을 채널 1행 = 연간합계 + 1~12월로 단순하게 정리한 표다.
# '매체'(합계행)·법인카드정산·촬영샘플·잔여비용은 매체 집행이 아니라서 채널 믹스에서 제외한다.
CHANNEL_MIX_NON_MEDIA_ROWS = {"매체", "법인카드정산", "촬영샘플", "잔여비용"}


def parse_channel_mix_sheet(xls: pd.ExcelFile, source_name: str = "") -> pd.DataFrame:
    """B열=채널명, C열=연간 합계, D~O열=1~12월인 고정 양식을 위치 기준으로 읽는다.
    'TOTAL' 라벨과 그 오른쪽에 1~12가 순서대로 나열된 헤더 행을 찾아서, 그 아래 채널명이
    있는 행만 데이터로 삼는다. 연도는 파일명(예: '26년 매체별 채널 믹스.xlsx')에서 'NN년'
    패턴을 찾아 20NN으로 추정하고, 못 찾으면 오늘 날짜 기준 연도를 쓴다."""
    sheet = xls.sheet_names[0]
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)
    if raw.empty:
        return pd.DataFrame()

    header_row = None
    for r in range(min(6, len(raw))):
        row_vals = raw.iloc[r].tolist()
        if not any(str(v).strip().upper() == "TOTAL" for v in row_vals if pd.notna(v)):
            continue
        nums = [v for v in row_vals if isinstance(v, (int, float)) and pd.notna(v)]
        if len(nums) >= 12 and [int(x) for x in nums[:12]] == list(range(1, 13)):
            header_row = r
            break
    if header_row is None:
        return pd.DataFrame()

    year_guess = date.today().year
    ym = re.search(r"(\d{2})년", source_name)
    if ym:
        year_guess = 2000 + int(ym.group(1))

    rows = []
    for r in range(header_row + 1, len(raw)):
        if raw.shape[1] <= 2:
            break
        name = raw.iloc[r, 1]
        if pd.isna(name):
            continue
        name = str(name).strip()
        if not name or name in CHANNEL_MIX_NON_MEDIA_ROWS or len(name) > 20:
            continue
        total_val = pd.to_numeric(raw.iloc[r, 2], errors="coerce")
        if pd.isna(total_val):
            continue
        for month in range(1, 13):
            col_idx = 2 + month
            val = pd.to_numeric(raw.iloc[r, col_idx], errors="coerce") if raw.shape[1] > col_idx else 0
            rows.append({
                "channel": name, "year": year_guess, "month": month,
                "budget": float(val) if pd.notna(val) else 0.0,
            })
    return pd.DataFrame(rows)


BUDGET_SCOPE_TITLES = {
    "자사몰": "자사몰 현황",
    # "온라인사업팀": "온라인사업팀 현황",  # 형 요청으로 우선 제외 — 자사몰만 본다
    # "외부몰": "외부몰 현황",  # 구조가 더 복잡해서(직입점/벤더 중첩) 다음 단계에서 추가 예정
}
# 실제로 시트에 등장하는 모든 섹션 제목(파싱 대상이 아닌 것도 포함) — 섹션 경계(어디까지가
# '자사몰 현황' 구간인지)를 정확히 자르기 위해 필요하다. 하나의 시트 안에 온라인사업팀/자사몰/
# 외부몰 현황이 위아래로 이어져 있을 수 있어서, 자사몰 구간이 다음 섹션(외부몰 현황) 내용까지
# 삼켜버리지 않도록 전체 목록으로 경계를 잡는다.
ALL_BUDGET_SECTION_TITLES = ["온라인사업팀 현황", "자사몰 현황", "외부몰 현황"]
BUDGET_SUBSECTION_DETAIL = "매체 세부내역"
# '매체 세부내역' 라벨을 못 찾았을 때, 요약 행(예상매출/실매출/매출 달성률/광고비/비율/전년
# 대비 등 — 파일마다 문구가 조금씩 다름)이 채널로 잘못 잡히는 걸 막기 위해, 요약 행을
# "제외"하는 방식 대신 실제 매체명으로 알려진 키워드를 "포함"하는지로 판단한다(화이트리스트가
# 훨씬 안전 — 요약 행 문구는 파일마다 계속 달라졌지만 매체명은 안정적이었다).
BUDGET_CHANNEL_NAME_KEYWORDS = [
    "네이버", "메타", "구글", "크리테오", "GFA", "카카오", "모비온", "AEDI", "틱톡", "당근",
    "쿠팡", "무신사", "G마켓", "바이럴", "토스", "하프클럽", "LF몰", "신규 매체", "신규매체",
    "법인카드", "촬영", "잔여", "기타 비용", "기타비용", "웹뜰", "AGENCY", "에이전시",
]
BUDGET_CHANNEL_NAME_RE = re.compile("|".join(re.escape(k) for k in BUDGET_CHANNEL_NAME_KEYWORDS), re.IGNORECASE)
# 화이트리스트를 통과해도, 이 키워드가 들어있으면 무조건 요약 행으로 보고 제외한다(이중 안전장치
# — '매출 달성율'처럼 표기가 달라 화이트리스트/기존 블랙리스트 둘 다 놓친 사례가 실제로 있었음).
BUDGET_HARD_EXCLUDE_RE = re.compile(r"매출|달성|증감|대비|비율|목표|누계|광고비|광고선전비")


def _find_budget_month_columns(raw: pd.DataFrame, start_row: int = 0, end_row: int | None = None):
    """[start_row, end_row) 범위 안에서 '1월'~'12월'이 순서대로 나열된 헤더 행을 찾아
    (헤더 행 번호, {월번호: 컬럼위치})를 반환한다. 10개월 이상 잡히는 첫 행을 헤더로 인정한다
    (연말 등 일부 컬럼이 빠져도 허용).

    섹션(온라인사업팀 현황/자사몰 현황/외부몰 현황 등)마다 표가 따로 그려져 있어서 헤더 행 위치와
    컬럼 배치가 섹션마다 다를 수 있다 — 그래서 시트 전체에서 한 번만 찾지 않고, 각 섹션 구간
    안에서 지역적으로(local) 찾는다."""
    end_row = len(raw) if end_row is None else min(end_row, len(raw))
    for r in range(max(start_row, 0), end_row):
        month_cols = {}
        for i, v in enumerate(raw.iloc[r].tolist()):
            month_num = None
            s = str(v).strip()
            m = re.match(r"^(\d{1,2})\s*월$", s)
            if m:
                month_num = int(m.group(1))
            elif isinstance(v, (int, float)) and not isinstance(v, bool) and pd.notna(v):
                # 셀 값이 텍스트 '1월'이 아니라 숫자 1(서식만 'N월'로 표시되는 경우)일 수 있어서,
                # 정수 1~12 값도 후보로 받아들인다.
                if float(v) == int(v) and 1 <= int(v) <= 12:
                    month_num = int(v)
            if month_num is not None and 1 <= month_num <= 12:
                month_cols[month_num] = i
        if len(month_cols) >= 10:
            # 숫자형 월(1~12)로만 잡힌 경우, 우연히 다른 데이터 행이 걸리는 걸 막기 위해
            # 컬럼 순서대로 월이 1씩 증가하는지(오름차순 연속) 확인한다.
            ordered = sorted(month_cols.items(), key=lambda kv: kv[1])
            seq_ok = all(b[0] - a[0] == 1 for a, b in zip(ordered, ordered[1:]))
            if seq_ok:
                return r, month_cols
    return None, {}


def _diagnose_budget_sheet(xls: pd.ExcelFile) -> str:
    """예산 파일 파싱이 실패했을 때, 형이 화면을 캡쳐해서 보내주면 바로 원인을 알 수 있도록
    시트별로 무엇을 찾았는지/못 찾았는지 자세히 보여주는 진단 리포트를 만든다."""
    lines = [f"전체 시트 목록({len(xls.sheet_names)}개): {xls.sheet_names}", ""]
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        lines.append(f"### 시트: '{sheet}' ({raw.shape[0]}행 x {raw.shape[1]}열)")
        if raw.empty:
            lines.append("- (빈 시트)")
            continue

        title_scan_cols = min(3, raw.shape[1])
        title_positions = []  # [(row, title), ...] 등장 순서대로 — 파서와 동일한 방식(왼쪽 컬럼만)
        for r in range(len(raw)):
            row_str = raw.iloc[r, :title_scan_cols].astype(str)
            for t in ALL_BUDGET_SECTION_TITLES:
                if row_str.str.contains(t, na=False, regex=False).any():
                    title_positions.append((r, t))
        title_hits = {}
        for r, t in title_positions:
            title_hits.setdefault(t, []).append(r)
        if title_hits:
            for t, hit_rows in title_hits.items():
                lines.append(f"- '{t}' 텍스트 발견(맨 왼쪽 {title_scan_cols}개 컬럼만 확인): {hit_rows[:5]}행" + (" 등" if len(hit_rows) > 5 else ""))
        else:
            lines.append("- '온라인사업팀 현황'/'자사몰 현황'/'외부몰 현황' 텍스트를 이 시트 왼쪽 컬럼에서 못 찾음")

        # 실제 파서가 판단할 [start, end) 구간과, 그 범위 안에 '26년'/'25년'/'매체 세부내역'이
        # 실제로 들어오는지를 직접 계산해서 보여준다 — 이게 어긋나면 구간이 너무 일찍 끊긴 것
        # (파싱 실패의 가장 흔한 원인).
        for scope_check, title_check in BUDGET_SCOPE_TITLES.items():
            starts = title_hits.get(title_check, [])
            if not starts:
                continue
            s = starts[0]
            later = sorted(r for r, _ in title_positions if r > s)
            e = later[0] if later else len(raw)
            year_in = [r for r in range(len(raw)) if s <= r < e and raw.iloc[r].astype(str).str.contains("26년|25년", na=False, regex=True).any()]
            detail_in = [r for r in range(len(raw)) if s <= r < e and raw.iloc[r].astype(str).str.contains(BUDGET_SUBSECTION_DETAIL, na=False, regex=False).any()]
            lines.append(
                f"- [파서 판정] '{title_check}' 구간 = {s}~{e - 1}행 / 이 구간 안 '26·25년' 발견: "
                f"{year_in if year_in else '없음'} / 이 구간 안 '매체 세부내역' 발견: "
                f"{detail_in if detail_in else '없음 ← 이게 없으면 매체 데이터가 0개로 나옵니다'}"
            )

        all_month_headers = []
        r = 0
        while r < len(raw):
            hr, mc = _find_budget_month_columns(raw, start_row=r, end_row=len(raw))
            if hr is None:
                break
            all_month_headers.append((hr, len(mc), min(mc.values()), max(mc.values())))
            r = hr + 1
        if all_month_headers:
            lines.append(f"- '1월~12월' 헤더 후보 {len(all_month_headers)}개: " + ", ".join(
                f"{hr}행(월 {n}개, 컬럼 {c0}~{c1})" for hr, n, c0, c1 in all_month_headers[:8]
            ))
        else:
            lines.append("- '1월'~'12월'이 10개 이상 나열된 헤더 행을 이 시트에서 못 찾음")

        year_hits = [r for r in range(len(raw)) if raw.iloc[r].astype(str).str.contains("26년|25년", na=False, regex=True).any()]
        lines.append(f"- '26년'/'25년' 텍스트 발견 행: {year_hits[:10]}" if year_hits else "- '26년'/'25년' 텍스트를 못 찾음")

        detail_hits = [r for r in range(len(raw)) if raw.iloc[r].astype(str).str.contains(BUDGET_SUBSECTION_DETAIL, na=False, regex=False).any()]
        lines.append(f"- '매체 세부내역' 텍스트 발견 행: {detail_hits[:10]}" if detail_hits else "- '매체 세부내역' 텍스트를 못 찾음")

        # '매체 세부내역'이라는 라벨 문구에 기대지 않고, 실제 매체명(네이버/메타/구글 등)이
        # 어디에 있는지 직접 찾는다 — 라벨 문구가 실제 파일과 다를 수 있어서, 채널명 자체를
        # 찾는 게 더 확실하다.
        channel_keywords = ["네이버", "메타", "구글", "크리테오", "GFA", "카카오", "모비온", "AEDI", "틱톡", "당근"]
        channel_hits = []
        for r in range(len(raw)):
            row_str = raw.iloc[r].astype(str)
            for kw in channel_keywords:
                if row_str.str.contains(kw, na=False, regex=False).any():
                    channel_hits.append((r, kw))
                    break
        if channel_hits:
            lines.append(
                "- 매체명 키워드(네이버/메타/구글/크리테오/GFA/카카오/모비온 등) 직접 검색 결과: "
                + ", ".join(f"{r}행({kw})" for r, kw in channel_hits[:20])
                + (" 등" if len(channel_hits) > 20 else "")
            )
        else:
            lines.append("- 매체명 키워드(네이버/메타/구글 등)를 이 시트 전체에서 하나도 못 찾음")

        preview_cols = min(18, raw.shape[1])

        def dump_rows(row_start, row_end):
            out_lines = []
            for r in range(max(0, row_start), min(len(raw), row_end)):
                cells = []
                for c in range(preview_cols):
                    v = raw.iat[r, c]
                    cells.append("·" if pd.isna(v) or str(v).strip() == "" else str(v).strip())
                out_lines.append(f"  {r:>3}행: " + " | ".join(cells))
            return out_lines

        # '자사몰 현황' 구간이 있으면 그 주변을 컬럼 위치가 그대로 보이게(빈 칸 생략 없이)
        # 찍어서, 헤더 행과 데이터 행의 컬럼이 실제로 어떻게 어긋나는지 한눈에 보이게 한다.
        zasamall_rows = title_hits.get("자사몰 현황", [])
        if zasamall_rows:
            r0 = zasamall_rows[0]
            lines.append(f"- '자사몰 현황' 주변 원본 셀 (col0~col{preview_cols - 1}, {max(0, r0 - 1)}~{min(len(raw), r0 + 14) - 1}행, 빈칸도 그대로 표시):")
            lines.extend(dump_rows(r0 - 1, r0 + 14))
        else:
            lines.append(f"- 앞부분 원본 셀 (col0~col{preview_cols - 1}, 0~{min(8, len(raw)) - 1}행, 빈칸도 그대로 표시):")
            lines.extend(dump_rows(0, 8))

        # '매체 세부내역' 텍스트가 자사몰 구간 밖(예: 훨씬 아래쪽)에서 발견되는 경우가 있어서,
        # 그 위치 주변도 별도로 그대로 찍어준다 — 진짜 매체별 예산표가 어떤 모양인지 보기 위해.
        if detail_hits:
            for dh in detail_hits[:2]:
                lines.append(f"- '매체 세부내역'(행 {dh}) 주변 원본 셀 (col0~col{preview_cols - 1}, {max(0, dh - 2)}~{min(len(raw), dh + 15) - 1}행, 빈칸도 그대로 표시):")
                lines.extend(dump_rows(dh - 2, dh + 15))
        lines.append("")
    return "\n".join(lines)


BUDGET_YEAR_LABELS = ["26년", "25년"]  # 파싱할 연도(등장 순서와 무관하게 둘 다 찾는다)
# channel 컬럼에 실제 매체명 대신 넣는 특수 표시값 — 광고비(TOTAL)와 구분해서
# 매출 지표도 같은 테이블/스키마에 같이 저장하기 위한 것.
BUDGET_SENTINEL_EXPECTED_REVENUE = "__예상매출__"
BUDGET_SENTINEL_ACTUAL_REVENUE = "__실매출__"


def parse_channel_budget_sheet(xls: pd.ExcelFile) -> pd.DataFrame:
    """'◆26년 월별 예산 정리' 파일에서 '자사몰 현황' 섹션의 매체별 월간 예산(광고비), 예상매출,
    실매출을 26년/25년 둘 다 파싱한다. 형이 캡쳐해준 화면 구조를 보고 만든 버전 — 한 시트 안에
    온라인사업팀 현황/자사몰 현황/외부몰 현황이 위아래로 이어져 있고 섹션마다 표 레이아웃(헤더 행
    위치, 라벨 컬럼 개수)이 다를 수 있다고 보고, '1월'~'12월' 헤더를 시트 전체에서 한 번만 찾지
    않고 섹션(자사몰 현황) 구간 안에서 지역적으로 찾는다.

    형이 보내준 진단/실제 화면으로 확인된 실제 파일 특성을 반영했다:
    1) '1월'~'12월' 헤더 셀이 문자열 '1월'이 아니라 숫자 1(서식만 'N월'로 표시)로 저장돼 있어서,
       숫자형 월(1~12, 컬럼 순서대로 1씩 증가)도 헤더로 인식한다.
    2) 화면엔 '(단위: 천원, +VAT)'로 표시되지만, 셀에 저장된 실제 값은 이미 원 단위였다 — 1000을
       곱하지 않고 원본 값을 그대로 저장한다.
    3) 광고비 합계 행 라벨이 파일 버전마다 다르다('광고선전비 (1월~12월)' / '실제 광고비
       (1월~6월)'+'잔여 광고비 (7월~12월)' 등) — 특정 라벨 문구에 의존하지 않도록, 광고비
       TOTAL은 라벨 매칭 대신 '매체 세부내역' 채널별 합계를 그대로 더해서 계산한다.
    channel='TOTAL'은 광고비 합계(매체 합산), channel='__예상매출__'/'__실매출__'은 매출
    지표를 담는다.
    """
    rows = []
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        if raw.empty:
            continue

        # 시트 안에 등장하는 모든 섹션 제목의 위치를 먼저 전부 찾아둔다 (온라인사업팀 현황이
        # 자사몰 현황보다 앞에 있을 수 있고, 그 위쪽에 다른 요약 표가 더 있을 수도 있어서 —
        # 헤더/컬럼 위치를 시트 전체에서 한 번만 찾지 않고 섹션별로 다시 찾기 위한 기준점).
        # 왼쪽 몇 개 컬럼(라벨 자리)만 본다 — 전체 행을 다 보면 '매체 세부내역'의 채널명이나
        # 코멘트 셀에 우연히 '외부몰' 같은 단어가 들어있을 때 그걸 새 섹션 제목으로 잘못 인식해서
        # 구간이 너무 일찍 끊길 수 있다.
        title_scan_cols = min(3, raw.shape[1])
        title_positions = []  # [(row_idx, title), ...] 등장 순서대로
        for r in range(len(raw)):
            row_str = raw.iloc[r, :title_scan_cols].astype(str)
            for t in ALL_BUDGET_SECTION_TITLES:
                if row_str.str.contains(t, na=False, regex=False).any():
                    title_positions.append((r, t))

        for scope, scope_title in BUDGET_SCOPE_TITLES.items():
            scope_starts = [r for r, t in title_positions if t == scope_title]
            if not scope_starts:
                continue
            start = scope_starts[0]
            # 다음 섹션 제목(어떤 제목이든)이 나오는 행 전까지가 이 scope 구간.
            later_titles = [r for r, t in title_positions if r > start]
            end = later_titles[0] if later_titles else len(raw)

            # 이 섹션 구간 안에서 지역적으로 '1월'~'12월' 헤더 행/컬럼 위치를 찾는다 — 섹션마다
            # 표 레이아웃(라벨 컬럼 개수 등)이 다를 수 있어 전역 헤더를 재사용하면 엉뚱한
            # 컬럼에서 값을 읽게 된다. 26년/25년 블록은 보통 같은 헤더(같은 컬럼 배치)를 공유한다.
            header_row, month_cols = _find_budget_month_columns(raw, start_row=start, end_row=end)
            if header_row is None or not month_cols:
                continue
            label_col_end = min(month_cols.values())

            # 헤더 행에서 'TOTAL'/'연 목표' 컬럼과 '당월 누계' 컬럼 위치를 찾는다 — 이 값들은
            # 월별 합계로 다시 계산하지 않고 원본 셀 값을 그대로 쓴다(월별로 아직 안 나뉜 금액이
            # TOTAL에만 잡혀있는 항목이 있어서, 월별 합계로 대체하면 값이 달라진다).
            total_col = mtd_col = None
            for c in range(label_col_end):
                v = str(raw.iat[header_row, c]).strip()
                if v in ("TOTAL", "연 목표", "연목표"):
                    total_col = c
                elif v in ("당월 누계", "당월누계"):
                    mtd_col = c

            # 연도 블록 시작 위치를 전부 찾고(26년, 25년 등장 순서대로) 각 블록의 끝을 다음 연도
            # 블록 시작(또는 섹션 끝)으로 잡는다.
            year_hits = []
            for yl in BUDGET_YEAR_LABELS:
                hits = [
                    r for r in range(start, end)
                    if raw.iloc[r].astype(str).str.contains(yl, na=False, regex=False).any()
                ]
                if hits:
                    year_hits.append((hits[0], yl))
            year_hits.sort(key=lambda x: x[0])

            for i, (y_start, yl) in enumerate(year_hits):
                y_end = year_hits[i + 1][0] if i + 1 < len(year_hits) else end
                year_num = 2000 + int(re.sub(r"\D", "", yl))

                # 예상매출/실매출 행 — 광고비와 달리 라벨이 안정적으로 '예상매출'/'실매출' 문구를
                # 쓰고 있어서 그대로 매칭한다.
                for label, sentinel in [
                    ("예상매출", BUDGET_SENTINEL_EXPECTED_REVENUE),
                    ("실매출", BUDGET_SENTINEL_ACTUAL_REVENUE),
                ]:
                    idxs = [
                        r for r in range(y_start, y_end)
                        if raw.iloc[r].astype(str).str.contains(label, na=False, regex=False).any()
                    ]
                    if not idxs:
                        continue
                    r = idxs[0]
                    for m, c in month_cols.items():
                        v = pd.to_numeric(raw.iat[r, c], errors="coerce")
                        if pd.notna(v):
                            rows.append({"scope": scope, "channel": sentinel, "year": year_num, "month": m, "budget_cost": float(v)})
                    # 원본 'TOTAL'/'당월 누계' 셀 값도 그대로 저장한다(month=0/-1로 구분) — 월별
                    # 합계와 다를 수 있어서, 화면에는 계산값 대신 이 원본값을 우선 쓴다.
                    for pseudo_m, col in ((0, total_col), (-1, mtd_col)):
                        if col is None:
                            continue
                        v = pd.to_numeric(raw.iat[r, col], errors="coerce")
                        if pd.notna(v):
                            rows.append({"scope": scope, "channel": sentinel, "year": year_num, "month": pseudo_m, "budget_cost": float(v)})

                # 매체 세부내역: 라벨 행 다음부터 y_end까지, 각 행의 첫 문자열 셀(라벨 컬럼 범위
                # 안)을 매체명으로 쓴다. 숫자로 시작하는 값/빈 라벨/전부 빈 월별 값인 행은 건너뛴다.
                detail_start_candidates = [
                    r for r in range(y_start, y_end)
                    if raw.iloc[r].astype(str).str.contains(BUDGET_SUBSECTION_DETAIL, na=False, regex=False).any()
                ]
                if detail_start_candidates:
                    channel_row_range = range(detail_start_candidates[0] + 1, y_end)
                else:
                    # '매체 세부내역' 라벨 문구를 못 찾은 경우 — 파일마다 이 라벨 문구가 없거나
                    # 다를 수 있어서, y_start부터 훑되 아래에서 매체명 화이트리스트로 다시 거른다.
                    channel_row_range = range(y_start, y_end)

                for r in channel_row_range:
                    label = None
                    for c in range(label_col_end):
                        v = raw.iat[r, c]
                        if isinstance(v, str) and v.strip() and not re.match(r"^\d", v.strip()):
                            label = v.strip()
                    if not label:
                        continue
                    # '매체 세부내역' 라벨 아래에 있어도, 매출/달성/증감/비율 같은 요약성 단어가
                    # 들어간 라벨은 무조건 제외한다(라벨이 있든 없든 항상 적용 — '매체 세부내역'
                    # 구간 안에 요약 행이 섞여 들어온 사례가 실제로 있었음).
                    if BUDGET_HARD_EXCLUDE_RE.search(label):
                        continue
                    if not detail_start_candidates:
                        # '매체 세부내역' 라벨이 없을 때는 요약 행을 "제외"하는 대신, 알려진
                        # 매체명 키워드를 "포함"하는 라벨만 채널로 인정한다 — 요약 행 문구는
                        # 파일마다 계속 달라서 블랙리스트로는 다 못 걸렀다(예: '매출 달성률'이
                        # 새 채널처럼 잡히는 사고가 실제로 있었음). 화이트리스트가 훨씬 안전하다.
                        if not BUDGET_CHANNEL_NAME_RE.search(label):
                            continue
                    for m, c in month_cols.items():
                        v = pd.to_numeric(raw.iat[r, c], errors="coerce")
                        if pd.notna(v):
                            rows.append({"scope": scope, "channel": label, "year": year_num, "month": m, "budget_cost": float(v)})
                    for pseudo_m, col in ((0, total_col), (-1, mtd_col)):
                        if col is None:
                            continue
                        v = pd.to_numeric(raw.iat[r, col], errors="coerce")
                        if pd.notna(v):
                            rows.append({"scope": scope, "channel": label, "year": year_num, "month": pseudo_m, "budget_cost": float(v)})

    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out = out.groupby(["scope", "channel", "year", "month"], as_index=False)["budget_cost"].sum()

    # 광고비 TOTAL = 매체 세부내역(실제 채널명 행)의 월별 합계. 요약 행 라벨 문구에 기대지 않기
    # 위해 이렇게 계산한다.
    sentinels = {BUDGET_SENTINEL_EXPECTED_REVENUE, BUDGET_SENTINEL_ACTUAL_REVENUE}
    ch_only = out[~out["channel"].isin(sentinels)]
    if not ch_only.empty:
        total = ch_only.groupby(["scope", "year", "month"], as_index=False)["budget_cost"].sum()
        total["channel"] = "TOTAL"
        out = pd.concat([out, total], ignore_index=True)
    return out


def build_utm_channel_lookup(utm_map: pd.DataFrame) -> dict:
    """utm_channel_map 테이블(source_medium/channel)을 대소문자·공백 차이에 안전한
    lookup dict로 바꾼다 — GA 원본 값이 'Naver / cpc'처럼 대소문자가 섞여 올 수 있어서."""
    if utm_map is None or utm_map.empty:
        return {}
    return {
        str(sm).strip().lower(): ch
        for sm, ch in zip(utm_map["source_medium"], utm_map["channel"])
        if pd.notna(sm) and pd.notna(ch)
    }


def _ga_daily_agg(ga_channel_inflow: pd.DataFrame) -> pd.DataFrame:
    """ga_channel_inflow(일자×소스/매체, UTM 매핑 완료)를 날짜 단위로 합산해 ga_conversions/
    ga_revenue를 만든다. '일자별 누적' 표에 GA 컬럼을 붙일 때 쓴다. 우리가 실제로 운영 중인
    매체(UTM 매핑표에 있는 소스/매체)만 합산한다 — 자연유입/direct/referral/이메일 등은
    광고 성과가 아니라서 여기 섞이면 자체 ROAS랑 비교가 안 맞게 된다."""
    if ga_channel_inflow is None or ga_channel_inflow.empty:
        return pd.DataFrame(columns=["report_date", "ga_conversions", "ga_revenue"])
    g = ga_channel_inflow.copy()
    g = g[g["channel"].notna()]
    if g.empty:
        return pd.DataFrame(columns=["report_date", "ga_conversions", "ga_revenue"])
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    return g.groupby("report_date", as_index=False).agg(
        ga_conversions=("conversions", "sum"), ga_revenue=("revenue", "sum")
    )


def _ga_daily_agg_all(ga_channel_inflow: pd.DataFrame) -> pd.DataFrame:
    """_ga_daily_agg와 반대로, 매체 매핑 여부와 상관없이(자연유입/direct/referral/이메일 등
    포함) 날짜 단위로 전부 합산한다. '유입·매출 비교' 페이지의 GA-매출은 어드민(사이트 전체)
    매출과 비교하는 용도라, 광고 채널만 걸러내면 안 되고 GA가 보는 전체를 그대로 써야 한다."""
    if ga_channel_inflow is None or ga_channel_inflow.empty:
        return pd.DataFrame(columns=["report_date", "ga_conversions_all", "ga_revenue_all"])
    g = ga_channel_inflow.copy()
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    return g.groupby("report_date", as_index=False).agg(
        ga_conversions_all=("conversions", "sum"), ga_revenue_all=("revenue", "sum")
    )


def _ga_weekly_agg(ga_channel_inflow: pd.DataFrame, weekly: pd.DataFrame) -> pd.DataFrame:
    """'주간별 누적' 표의 각 주(week_start~week_end) 구간에 맞춰 ga_channel_inflow를
    합산한다. 리포트 원본 주간 섹션엔 GA 컬럼이 아예 없어서, 이 방식으로만 채울 수 있다.
    _ga_daily_agg와 마찬가지로 UTM 매핑된(=우리가 운영 중인) 매체만 합산한다."""
    if ga_channel_inflow is None or ga_channel_inflow.empty or weekly is None or weekly.empty:
        return pd.DataFrame(columns=["week_start", "ga_conversions", "ga_revenue"])
    g = ga_channel_inflow.copy()
    g = g[g["channel"].notna()]
    if g.empty:
        return pd.DataFrame(columns=["week_start", "ga_conversions", "ga_revenue"])
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    rows = []
    for _, w in weekly.iterrows():
        mask = (g["report_date"] >= w["week_start"]) & (g["report_date"] <= w["week_end"])
        rows.append({
            "week_start": w["week_start"],
            "ga_conversions": g.loc[mask, "conversions"].sum(),
            "ga_revenue": g.loc[mask, "revenue"].sum(),
        })
    return pd.DataFrame(rows)


def _ga_channel_agg(ga_channel_inflow: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """선택된 기간 안에서, UTM 매핑이 된(channel이 채워진) 행만 매체 단위로 합산한다.
    '매체별 성과' 표에 (기간 선택에 맞춰 갱신되는) GA 컬럼을 붙일 때 쓴다."""
    if ga_channel_inflow is None or ga_channel_inflow.empty:
        return pd.DataFrame(columns=["channel", "ga_conversions", "ga_revenue"])
    g = ga_channel_inflow.copy()
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end) & g["channel"].notna()]
    if g.empty:
        return pd.DataFrame(columns=["channel", "ga_conversions", "ga_revenue"])
    return g.groupby("channel", as_index=False).agg(
        ga_conversions=("conversions", "sum"), ga_revenue=("revenue", "sum")
    )


# ──────────────────────────────────────────────────────────────
# 채널 퍼널 리포트 (신규 — 신규 고객 발굴 / 매출 확보 2목적축) 전용 집계 헬퍼
# ──────────────────────────────────────────────────────────────
def classify_ga_bucket(row) -> str:
    """ga_channel_inflow 한 행(세션 소스/매체)을 광고/자연유입/기타로 나눈다.
    '매체'(channel) 컬럼이 채워져 있으면(=UTM 매핑된 유료 매체) 광고로 보고,
    그 외에는 source_medium 문자열(organic/referral/direct 등)로 자연유입/기타를 구분한다."""
    if pd.notna(row.get("channel")):
        return "광고"
    sm = str(row.get("source_medium", "")).lower()
    if "organic" in sm or "referral" in sm or "direct" in sm:
        return "자연유입"
    return "기타"


def _funnel_from_audience(audience: pd.DataFrame, start: date, end: date, audience_type: str) -> pd.DataFrame:
    """channel_audience_snapshot(캠페인/그룹명 기준 신규·리타겟팅 분류, 매체 리포트 원본)에서
    기간 내 채널×오디언스타입 '최신 스냅샷'만 채널 단위로 합산한다. render_targeting_performance_page
    와 동일한 dedup 로직(같은 달 여러 번 업로드해도 최신 값만 사용)을 그대로 따른다."""
    cols = ["channel", "impressions", "clicks", "cost_incl_vat", "signups", "conversions", "revenue"]
    if audience is None or audience.empty:
        return pd.DataFrame(columns=cols)
    a = audience.copy()
    a["as_of_date"] = pd.to_datetime(a["as_of_date"]).dt.date
    a = a[(a["as_of_date"] >= start) & (a["as_of_date"] <= end) & (a["audience_type"] == audience_type)]
    if a.empty:
        return pd.DataFrame(columns=cols)
    a = a.sort_values("as_of_date").drop_duplicates(subset=["channel", "audience_type"], keep="last")
    return a.groupby("channel", as_index=False).agg(
        impressions=("impressions", "sum"), clicks=("clicks", "sum"),
        cost_incl_vat=("cost_incl_vat", "sum"), signups=("signups", "sum"),
        conversions=("conversions", "sum"), revenue=("revenue", "sum"),
    )


def _ga_visits_by_channel(ga_channel_inflow: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """ga_channel_inflow(UTM 매핑 완료)에서 기간 내 채널 단위 신규/재방문 세션수를 합산한다.
    매출 확보 퍼널의 '재방문' 단계, 신규 발굴 퍼널의 '방문(신규유입)' 단계에 쓴다."""
    cols = ["channel", "new_users", "returning_users"]
    if ga_channel_inflow is None or ga_channel_inflow.empty:
        return pd.DataFrame(columns=cols)
    g = ga_channel_inflow.copy()
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end) & g["channel"].notna()]
    if g.empty:
        return pd.DataFrame(columns=cols)
    return g.groupby("channel", as_index=False).agg(
        new_users=("new_users", "sum"), returning_users=("returning_users", "sum")
    )


def _channel_spend_total(channels_weekly: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """channel_weekly(주간 매체 리포트, 신규/리타겟 구분 없는 채널 전체 광고비)에서 기간과
    겹치는 주(week_start~week_end)만 채널 단위로 합산한다. 채널 믹스(연예산) 대비 '실제
    집행비중'을 계산할 때, 오디언스타입 한쪽(신규/리타겟)만이 아니라 채널 전체 지출을 써야
    26년 채널믹스 파일(전체 매체 집행 기준)과 스코프가 맞는다."""
    cols = ["channel", "cost_incl_vat"]
    if channels_weekly is None or channels_weekly.empty:
        return pd.DataFrame(columns=cols)
    w = channels_weekly.copy()
    w["week_start"] = pd.to_datetime(w["week_start"]).dt.date
    w["week_end"] = pd.to_datetime(w["week_end"]).dt.date
    w = w[(w["week_start"] <= end) & (w["week_end"] >= start)]
    if w.empty:
        return pd.DataFrame(columns=cols)
    # 주간 광고비를 그대로 더하면 기간을 8/1~8/11로 잘라도 그 주 '전체' 광고비가 들어가 과다 집계된다.
    # 선택 기간과 겹치는 날짜 수만큼만 비례 배분한다.
    def _prorate(r):
        w_days = (r["week_end"] - r["week_start"]).days + 1
        ov = (min(r["week_end"], end) - max(r["week_start"], start)).days + 1
        if w_days <= 0 or ov <= 0:
            return 0.0
        return float(r["cost_incl_vat"]) * min(ov, w_days) / w_days
    w = w.copy()
    w["cost_incl_vat"] = w.apply(_prorate, axis=1)
    return w.groupby("channel", as_index=False).agg(cost_incl_vat=("cost_incl_vat", "sum"))


def _channel_mix_ratio(mix: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """channel_mix_budget(26년 채널 믹스 예산)에서 선택 기간과 겹치는 연/월만 채널 단위로
    합산해 연예산비중(%)을 계산한다."""
    cols = ["channel", "budget", "budget_ratio"]
    if mix is None or mix.empty:
        return pd.DataFrame(columns=cols)
    m = mix.copy()
    period_months = pd.period_range(start, end, freq="M")
    keys = {(p.year, p.month) for p in period_months}
    m = m[m.apply(lambda r: (int(r["year"]), int(r["month"])) in keys, axis=1)]
    if m.empty:
        return pd.DataFrame(columns=cols)
    agg = m.groupby("channel", as_index=False).agg(budget=("budget", "sum"))
    total = agg["budget"].sum()
    agg["budget_ratio"] = np.where(total > 0, agg["budget"] / total * 100, 0)
    return agg


CREATIVE_NAME_HINTS = ["소재성과", "소재별성과"]  # 예: 페이스북_소재성과요약


def _infer_channel_from_sheet(sheet: str) -> str:
    """'(DA) 구글_실적최대화_data' → '구글', '페이스북_소재성과요약' → '페이스북' 처럼
    시트명 접두어에서 매체명만 뽑아낸다."""
    name = re.sub(r"^\([^)]*\)\s*", "", sheet)  # 앞의 "(DA) " 같은 괄호 접두어 제거
    name = name.split("_")[0].strip()
    return name or sheet


CREATIVE_SCAN_ROWS = 100  # 실제 파일에서 소재 상세표 헤더가 50~60번째 행 근처에 있는 경우가 있어 넉넉하게 잡음


def find_creative_sheets(xls: pd.ExcelFile):
    """소재 단위 데이터가 있는 시트를 찾는다.
    시트명 패턴(예: '_data' 접미사)에 의존하면 '(DA) GFA_data(자사몰)'처럼
    '_data'가 끝이 아니라 중간에 오는 등 실제 파일의 이름 규칙이 제각각이라 놓치기 쉽다.
    그래서 이름은 참고만 하고, 실제로 각 시트 안에 '광고소재' 컬럼이 있는지
    내용 기준으로 직접 확인한다 (숨겨진/그룹화된 행이 있어도 pandas는 다 읽으므로 문제 없음).
    """
    found = []
    for s in xls.sheet_names:
        if any(hint in s for hint in CREATIVE_NAME_HINTS):
            found.append(s)
            continue
        try:
            probe = pd.read_excel(xls, sheet_name=s, header=None, nrows=CREATIVE_SCAN_ROWS)
        except Exception:
            continue
        if find_header_row(probe, required=("광고소재",), scan=CREATIVE_SCAN_ROWS) is not None:
            found.append(s)
    return found


def _find_last_matching_row(raw: pd.DataFrame, required, scan=CREATIVE_SCAN_ROWS):
    """find_header_row와 달리 '가장 먼저' 매칭되는 행이 아니라 '가장 마지막으로' 매칭되는 행을 찾는다.
    실제 파일은 한 시트 안에 캠페인/그룹 집계표 → (일부 채널은) 소재 요약 집계표 → 진짜 소재별
    raw data 표 순서로 여러 표가 쌓여 있고, 우리가 원하는 건 항상 가장 마지막(가장 상세한) 표라서
    노출수·클릭수만 있는 앞쪽 집계표를 잘못 집지 않으려면 '마지막 매칭'을 써야 한다."""
    last = None
    for i in range(min(scan, len(raw))):
        row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
        if all(tok in row_text for tok in required):
            last = i
    return last


# 소재별 성과 화면에는 '현재 실제로 운영 중인' 매체만 노출한다 (당근마켓/카카오모먼트/DV360/
# 구글에즈/MOBON/에디AI/네이버 SA·SSP 등은 현재 미운영이라 제외).
# GFA는 캠페인명에 붙는 '_PC'/'_MO' 접미사로 기기별(네이버 GFA PC / 네이버 GFA MO)로 쪼갠다.
CREATIVE_CHANNEL_WHITELIST_MAP = {
    "페이스북": "메타",
    "구글": "구글(P-MAX)",  # '(DA) 구글_실적최대화_data' 시트 = Performance Max
    "크리테오": "크리테오",
}


def _map_creative_channel(inferred_channel: str, campaign_series: pd.Series) -> pd.Series:
    """행별 매체탭 라벨을 계산. 현재 미운영 매체는 None을 반환해 화면/저장에서 제외되도록 한다."""
    idx = campaign_series.index
    if inferred_channel in CREATIVE_CHANNEL_WHITELIST_MAP:
        label = CREATIVE_CHANNEL_WHITELIST_MAP[inferred_channel]
        return pd.Series([label] * len(idx), index=idx, dtype=object)
    if inferred_channel == "GFA":
        camp = campaign_series.astype(str)
        is_pc = camp.str.contains("_PC", case=False, na=False)
        is_mo = camp.str.contains("_MO", case=False, na=False)
        result = pd.Series([None] * len(idx), index=idx, dtype=object)
        result[is_pc] = "네이버 GFA PC"
        result[~is_pc & is_mo] = "네이버 GFA MO"
        return result
    return pd.Series([None] * len(idx), index=idx, dtype=object)


# ──────────────────────────────────────────────────────────────
# 채널 그룹(타겟팅 오디언스) 단위 파싱 — 타겟팅별 성과용 (신규)
# 소재 상세표보다 앞에 나오는 '캠페인 분류 | 그룹 분류 | 노출수 ...' 집계표에서
# 그룹(광고그룹/타겟팅 오디언스) 단위로 신규/리타겟팅을 분류해 저장한다.
# ──────────────────────────────────────────────────────────────

# 사용자가 정리해준 '260804_광고 매체 그룹별 타겟팅 분류.xlsx' 기준 매핑.
# 키: (원본채널, 그룹명) 또는 [그룹 분류 컬럼이 없는 채널은] (원본채널, 캠페인분류명) → "신규"/"리타겟팅"
AUDIENCE_TYPE_MAP = {
    # 메타
    ("페이스북", "의류관심타겟"): "신규",
    ("페이스북", "패션관심타겟"): "신규",
    ("페이스북", "방문자180일"): "리타겟팅",
    # 네이버 GFA (자사몰 + 애드부스트)
    ("GFA", "패션관심타겟"): "신규",
    ("GFA", "쇼핑관심타겟"): "신규",
    ("GFA", "방문자180일"): "리타겟팅",
    ("GFA", "3050연령타겟"): "신규",
    ("GFA", "3050남성타겟"): "신규",
    # 크리테오 (그룹 분류 컬럼이 없어 캠페인 분류 자체가 그룹 역할)
    ("크리테오", "다이나믹_신규"): "신규",
    ("크리테오", "스태틱_신규"): "신규",
    ("크리테오", "다이나믹_리텐션"): "리타겟팅",
    ("크리테오", "스태틱_리텐션"): "리타겟팅",
    # 구글 실적최대화 (그룹 분류 없음 → 캠페인분류 전체가 신규)
    ("구글", "Pmax_온라인팀"): "신규",
}


def classify_audience_type(origin_channel: str, label) -> str:
    """(원본채널, 그룹명/캠페인분류명) → '신규' / '리타겟팅' / '미분류'.
    1) 사용자가 정리해준 매핑표를 먼저 조회
    2) 못 찾으면 이름에 흔히 쓰이는 키워드로 최대한 추정
    3) 그래도 애매하면 '미분류'로 남겨서 대시보드에서 바로 눈에 띄게 한다 (임의로 단정하지 않음).
    새 캠페인/그룹이 추가되면 위 AUDIENCE_TYPE_MAP에 항목을 추가해주면 된다."""
    name = str(label).strip()
    key = (origin_channel, name)
    if key in AUDIENCE_TYPE_MAP:
        return AUDIENCE_TYPE_MAP[key]
    if "재방문자제외" in name or "신규" in name or "관심타겟" in name:
        return "신규"
    if any(kw in name for kw in ["방문자", "리텐션", "리타겟"]):
        return "리타겟팅"
    return "미분류"


def parse_channel_group_sheet(xls: pd.ExcelFile, sheet: str, today: date):
    """소재 상세표보다 앞에 나오는 '캠페인 분류 | 그룹 분류 | 노출수 ...' 집계표를 파싱해서
    그룹(타겟팅 오디언스) 단위 성과 + 신규/리타겟팅 분류를 반환한다.
    그룹 분류 컬럼이 없는 채널(크리테오)은 캠페인 분류를 그룹처럼 취급한다."""
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)
    hdr = find_header_row(raw, required=("노출수", "클릭수"), scan=30)
    if hdr is None:
        return None

    headers = [clean_col(h) for h in raw.iloc[hdr].tolist()]
    campaign_idx = match_col_pos(headers, include_any=["캠페인"])
    group_idx = match_col_pos(headers, include_any=["그룹"])
    if campaign_idx is None:
        return None
    label_idx = group_idx if group_idx is not None else campaign_idx

    # 이 집계표 바로 아래에는 컬럼 배치가 다른 '소재 상세표'(캠페인|그룹|광고소재|...)가 또 나오는데,
    # 헤더 텍스트가 똑같이 '노출수'+'클릭수'를 포함해서 그대로 이어 읽으면 소재명이 그룹명 자리에
    # 잘못 섞여 들어간다. 그래서 다음 헤더가 나오는 지점 '직전'까지만 이 표의 데이터로 본다.
    next_hdr = None
    for i in range(hdr + 1, len(raw)):
        row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
        if "노출수" in row_text and "클릭수" in row_text:
            next_hdr = i
            break
    body = raw.iloc[hdr + 1 : next_hdr] if next_hdr is not None else raw.iloc[hdr + 1 :]
    if body.empty:
        return None

    label_series = body.iloc[:, label_idx].astype(str).str.strip()
    keep_mask = body.iloc[:, label_idx].notna() & (label_series != "") & (label_series.str.lower() != "nan")
    keep_mask &= ~label_series.str.contains("합계|TOTAL|총계", case=False, na=False)
    body = body[keep_mask]
    label_series = label_series[keep_mask]
    if body.empty:
        return None

    impr_idx = match_col_pos(headers, include_any=["노출"])
    clicks_idx = match_col_pos(headers, include_any=["클릭"])
    cost_ex_idx = match_col_pos(headers, include_all=["제외"], include_any=["광고비", "비용"], exclude=["포함"])
    cost_in_idx = match_col_pos(headers, include_all=["포함"], include_any=["광고비", "비용"], exclude=["제외"])
    if cost_ex_idx is None and cost_in_idx is None:
        cost_generic_idx = match_col_pos(headers, include_any=["광고비", "비용"])
        cost_ex_idx = cost_in_idx = cost_generic_idx
    signup_idx = match_col_pos(headers, include_any=["가입"])
    conv_idx = match_col_pos(headers, include_all=["전환"], exclude=["금액", "ga", "율"])
    rev_idx = match_col_pos(headers, include_any=["매출", "전환금액"], exclude=["ga", "객단가"])

    campaign_series = (
        body.iloc[:, campaign_idx].astype(str).str.strip() if campaign_idx is not None else label_series
    )

    origin_channel = _infer_channel_from_sheet(sheet)

    out = pd.DataFrame()
    out["channel"] = _map_creative_channel(origin_channel, campaign_series).values
    out["audience_type"] = [classify_audience_type(origin_channel, v) for v in label_series]
    out["impressions"] = numcol_by_pos(body, impr_idx)
    out["clicks"] = numcol_by_pos(body, clicks_idx)
    out["cost_excl_vat"] = numcol_by_pos(body, cost_ex_idx)
    out["cost_incl_vat"] = numcol_by_pos(body, cost_in_idx)
    out["signups"] = numcol_by_pos(body, signup_idx)
    out["conversions"] = numcol_by_pos(body, conv_idx)
    out["revenue"] = numcol_by_pos(body, rev_idx)
    out["as_of_date"] = today
    # 현재 미운영 매체(채널 매핑이 None인 행)는 제외
    out = out[out["channel"].notna()]
    if out.empty:
        return None
    out = out[
        (out["impressions"] > 0) | (out["clicks"] > 0) | (out["cost_incl_vat"] > 0)
        | (out["conversions"] > 0) | (out["revenue"] > 0)
    ]
    return out.reset_index(drop=True)


# 네이버 검색광고/쇼핑검색광고/브랜드검색광고 — 사용자 지시에 따라 캠페인/그룹 세부 분류 없이
# "채널 전체 = 리타겟팅"으로 우선 분류한다(신규 유입보다는 이미 브랜드/검색 의도가 있는 사용자가
# 대상이라는 판단). GFA/메타/크리테오처럼 그룹명 기준 신규·리타겟팅을 나누지 않고 시트 전체를
# 채널 합계 1행으로 요약한다.
#
# 데이터 소스는 "(SA)/(SSP)/(브검) 네이버_data"(원본 raw 표)가 아니라 "_data"가 없는 요약 시트의
# "■ 일일 데이터 → 당월 총합" 행을 쓴다. 두 시트가 같은 채널인데도 서로 다른 숫자를 담고 있는
# 경우가 있고(예: 브랜드검색광고는 _data 시트에 광고비가 0으로 비어있음), 사용자가 실제 주간
# 보고서에서 보는 숫자는 요약 시트의 "당월 총합"이라고 확인해줘서 그 쪽을 공식 소스로 쓴다.
NAVER_SEARCH_SHEET_CHANNEL_MAP = {
    "(SA)": "네이버 검색광고",
    "(SSP)": "네이버 쇼핑검색광고",
    "(브검)": "네이버 브랜드검색광고",
}


def _naver_search_channel_from_sheet(sheet: str):
    """요약 시트("_data"가 붙지 않은)만 대상으로 한다 — 원본 raw 시트는 제외."""
    if sheet.endswith("_data"):
        return None
    for prefix, label in NAVER_SEARCH_SHEET_CHANNEL_MAP.items():
        if sheet.startswith(prefix):
            return label
    return None


def parse_naver_search_sheet(xls: pd.ExcelFile, sheet: str, today: date):
    """(SA)/(SSP)/(브검) 네이버 요약 시트의 "■ 일일 데이터 → 당월 총합" 행 하나를
    채널 합계로 반환한다."""
    channel_label = _naver_search_channel_from_sheet(sheet)
    if channel_label is None:
        return None
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)

    total_row = None
    for i in range(len(raw)):
        row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
        if "당월" in row_text and "총합" in row_text:
            total_row = i
            break
    if total_row is None:
        return None

    hdr = None
    for i in range(total_row - 1, -1, -1):
        row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
        if "노출수" in row_text and "클릭수" in row_text:
            hdr = i
            break
    if hdr is None:
        return None

    headers = [clean_col(h) for h in raw.iloc[hdr].tolist()]
    impr_idx = match_col_pos(headers, include_any=["노출"])
    clicks_idx = match_col_pos(headers, include_any=["클릭"])
    # 일부 시트는 "광고비(VAT제외)"처럼 '제외'가 명시돼 있고, 일부는 그냥 "광고비"(제외 의미)만
    # 있고 옆에 "광고비(VAT포함)"가 따로 있어 '제외' 키워드 없이도 구분해야 한다.
    cost_ex_idx = match_col_pos(headers, include_any=["광고비", "비용"], exclude=["포함"])
    cost_in_idx = match_col_pos(headers, include_all=["포함"], include_any=["광고비", "비용"])
    signup_idx = match_col_pos(headers, include_any=["가입"])
    conv_idx = match_col_pos(headers, include_all=["전환"], exclude=["금액", "ga", "율"])
    rev_idx = match_col_pos(headers, include_any=["매출", "전환금액"], exclude=["ga", "객단가"])

    row = raw.iloc[total_row]

    def _val(idx):
        if idx is None or idx >= len(row):
            return 0.0
        v = pd.to_numeric(row.iloc[idx], errors="coerce")
        return 0.0 if pd.isna(v) else float(v)

    out = pd.DataFrame([{
        "channel": channel_label,
        "audience_type": "리타겟팅",
        "impressions": _val(impr_idx),
        "clicks": _val(clicks_idx),
        "cost_excl_vat": _val(cost_ex_idx),
        "cost_incl_vat": _val(cost_in_idx),
        "signups": _val(signup_idx),
        "conversions": _val(conv_idx),
        "revenue": _val(rev_idx),
        "as_of_date": today,
    }])
    return out


def parse_creative_sheet(xls: pd.ExcelFile, sheet: str, today: date, hidden_rows: set = None):
    """소재 단위 시트 하나를 파싱. 시트마다 컬럼 구성이 조금씩 달라
    '소재명/광고소재/행 레이블' 컬럼과 노출/클릭/광고비/전환/매출 컬럼을 유연하게 매칭한다.
    hidden_rows: 엑셀에서 '행 숨기기'로 숨긴 행 번호(0-indexed) 집합. 해당 소재는 더 이상
    안 쓰는 것으로 보고 집계에서 제외한다(다시 보이게 하면 다음 업로드부터 자동으로 복귀)."""
    raw = pd.read_excel(xls, sheet_name=sheet, header=None)

    # 1순위: 노출수·클릭수·소재(광고소재/소재명 등)가 다 있는 '진짜 소재 상세표' 헤더 중 가장 마지막 것.
    # 컬럼명이 '광고소재'가 아니라 그냥 '소재'/'소재명'인 매체도 있어 '소재'로만 느슨하게 확인한다.
    hdr = _find_last_matching_row(raw, required=("노출수", "클릭수", "소재"))
    if hdr is None:
        # 2순위: '소재' 대신 '행 레이블' 같은 이름을 쓰는 피벗 요약 시트 (예: 페이스북_소재성과요약)
        hdr = find_header_row(raw, required=("노출수", "클릭수"), scan=CREATIVE_SCAN_ROWS)
    if hdr is None:
        return None

    headers = [clean_col(h) for h in raw.iloc[hdr].tolist()]
    creative_idx = None
    for cand in (["소재명"], ["광고소재"], ["행레이블"], ["소재"]):
        creative_idx = match_col_pos(headers, include_any=cand)
        if creative_idx is not None:
            break
    if creative_idx is None:
        return None

    body = raw.iloc[hdr + 1:]
    if body.empty:
        return None

    # 엑셀에서 '행 숨기기'로 숨겨둔 소재 행은 더 이상 안 쓰는 것으로 보고 제외
    if hidden_rows:
        body = body[~body.index.isin(hidden_rows)]
        if body.empty:
            return None

    # '총 합계'/TOTAL 행(캠페인·그룹 칸에 라벨이 오기도 함)과, 그 아래 붙는 '소재 이미지'
    # 미리보기 같은 성격이 다른 표는 소재명 칸 자체가 비어있거나(합계 행) 다른 컬럼 위치에
    # 있어서(이미지 표) 아래 notna/공백 필터에서 자연스럽게 걸러진다 — 별도 절단 없이 그대로 둔다.
    creative_series = body.iloc[:, creative_idx].astype(str).str.strip()
    keep_mask = body.iloc[:, creative_idx].notna() & (creative_series != "") & (creative_series.str.lower() != "nan")
    keep_mask &= ~creative_series.str.contains("합계|TOTAL|총계", case=False, na=False)
    body = body[keep_mask]
    creative_series = creative_series[keep_mask]
    if body.empty:
        return None

    impr_idx = match_col_pos(headers, include_any=["노출"])
    clicks_idx = match_col_pos(headers, include_any=["클릭"])
    cost_ex_idx = match_col_pos(headers, include_all=["제외"], include_any=["광고비", "비용"], exclude=["포함"])
    cost_in_idx = match_col_pos(headers, include_all=["포함"], include_any=["광고비", "비용"], exclude=["제외"])
    if cost_ex_idx is None and cost_in_idx is None:
        # VAT 구분 없이 '광고비' 한 컬럼만 있는 시트 대응 (포함/제외를 같은 값으로 취급)
        cost_generic_idx = match_col_pos(headers, include_any=["광고비", "비용"])
        cost_ex_idx = cost_in_idx = cost_generic_idx
    conv_idx = match_col_pos(headers, include_all=["전환"], exclude=["금액", "ga", "율"])
    rev_idx = match_col_pos(headers, include_any=["매출", "전환금액"], exclude=["ga", "객단가"])
    campaign_idx = match_col_pos(headers, include_any=["캠페인"])
    campaign_series = (
        body.iloc[:, campaign_idx] if campaign_idx is not None
        else pd.Series([""] * len(body), index=body.index)
    )

    out = pd.DataFrame()
    out["creative"] = creative_series.values
    out["channel"] = _map_creative_channel(_infer_channel_from_sheet(sheet), campaign_series).values
    out["impressions"] = numcol_by_pos(body, impr_idx)
    out["clicks"] = numcol_by_pos(body, clicks_idx)
    out["cost_excl_vat"] = numcol_by_pos(body, cost_ex_idx)
    out["cost_incl_vat"] = numcol_by_pos(body, cost_in_idx)
    out["conversions"] = numcol_by_pos(body, conv_idx)
    out["revenue"] = numcol_by_pos(body, rev_idx)
    out["as_of_date"] = today
    # 현재 미운영 매체(채널 매핑이 None인 행)는 소재별 성과에서 제외
    out = out[out["channel"].notna()]
    if out.empty:
        return None
    out = out[
        (out["impressions"] > 0) | (out["clicks"] > 0) | (out["cost_incl_vat"] > 0)
        | (out["conversions"] > 0) | (out["revenue"] > 0)
    ]
    return out.reset_index(drop=True)


# ──────────────────────────────────────────────────────────────
# 소재 이미지 추출/업로드 (신규)
# 각 채널 시트 끝에 실제로 embedded 되어 있는 '소재명 | 이미지' 표에서
# 이미지를 직접 추출해 Supabase Storage에 업로드하고, (원본채널, 소재명) → 공개 URL로 매핑한다.
# 소재 단위 성과 파싱(parse_creative_sheet)과는 별도 경로 — 실패해도 성과 데이터 저장에는 영향 없음.
# ──────────────────────────────────────────────────────────────
CREATIVE_IMAGE_BUCKET = "creative-images"

# 소재별 성과 화면의 매체탭(기기 분리 후) → 이미지가 들어있는 원본 채널명 역매핑
TAB_TO_ORIGIN_CHANNEL = {
    "메타": "페이스북",
    "구글(P-MAX)": "구글",
    "크리테오": "크리테오",
    "네이버 GFA PC": "GFA",
    "네이버 GFA MO": "GFA",
}

# Storage 저장 경로(폴더명)는 한글이 섞이면 일부 클라이언트/버전에서 URL 인코딩 문제로
# 업로드가 통째로 실패할 수 있어, 폴더명은 반드시 ASCII로 고정한다.
ORIGIN_CHANNEL_FOLDER = {
    "페이스북": "facebook",
    "구글": "google",
    "크리테오": "criteo",
    "GFA": "gfa",
}


def _creative_image_key(name) -> str:
    """소재명 표기 차이를 무시하고 매칭하기 위한 정규화 키.
    - 줄바꿈/공백 제거
    - 크리테오처럼 성과표의 소재명 끝에 '(크리에이티브ID)'가 붙지만 이미지 캡션표에는
      없는 경우가 있어, 끝에 붙은 '(숫자)'는 제거해서 매칭한다 ('(수정)'처럼 숫자가 아닌
      괄호 표기는 그대로 남겨 다른 소재와 구분되도록 유지)."""
    s = re.sub(r"\s+", "", str(name)).strip()
    s = re.sub(r"\(\d+\)$", "", s)
    return s


def extract_creative_images(file, sheets_and_channels) -> dict:
    """업로드된 엑셀 원본에서 시트별로 embedded 이미지를 찾아
    {(원본채널, 정규화된 소재명): (이미지bytes, 확장자)} 형태로 반환한다.
    시트 안에 '소재명(또는 광고소재/소재)'과 '이미지'가 같이 있는 헤더 행을 찾고,
    그 아래 각 행에 앵커된 이미지를 같은 행의 소재명 값과 매칭한다."""
    images = {}
    try:
        file.seek(0)
    except Exception:
        pass
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        return images
    finally:
        try:
            file.seek(0)
        except Exception:
            pass

    for sheet, origin_channel in sheets_and_channels:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        imgs = getattr(ws, "_images", [])
        if not imgs:
            continue
        try:
            raw = pd.read_excel(file, sheet_name=sheet, header=None)
        except Exception:
            continue
        finally:
            try:
                file.seek(0)
            except Exception:
                pass

        # '소재명 | 이미지' 헤더 행(진짜 성과 상세표와는 별개, 항상 시트 맨 끝 쪽에 있음) 탐색
        hdr = None
        for i in range(len(raw)):
            row_text = " ".join(str(x) for x in raw.iloc[i].tolist())
            if "이미지" in row_text and any(tok in row_text for tok in ("소재명", "광고소재", "소재")):
                hdr = i
        if hdr is None:
            continue

        headers = [clean_col(h) for h in raw.iloc[hdr].tolist()]
        name_idx = None
        for cand in (["소재명"], ["광고소재"], ["소재"]):
            name_idx = match_col_pos(headers, include_any=cand)
            if name_idx is not None:
                break
        if name_idx is None:
            continue

        for img in imgs:
            try:
                row_i = img.anchor._from.row  # 0-indexed, header=None으로 읽은 raw의 행 인덱스와 동일
                if row_i <= hdr or row_i >= len(raw):
                    continue
                name = str(raw.iloc[row_i, name_idx]).replace("\n", "").strip()
                if not name or name.lower() == "nan":
                    continue
                data = img._data()
                ext = (img.format or "png").lower()
                images[(origin_channel, _creative_image_key(name))] = (data, ext)
            except Exception:
                continue

    return images


def _safe_storage_name(name_key: str) -> str:
    """Supabase Storage의 객체 키 검증(isValidKey)은 영문/숫자/일부 기호만 허용하고
    한글 등 non-ASCII 문자는 URL 인코딩 여부와 무관하게 무조건 거부한다(서버가 디코딩한
    키 자체를 검사하기 때문에 percent-encoding으로는 우회되지 않음).
    그래서 한글은 아예 제거하고, 대신 원본 소재명의 해시를 붙여 서로 다른 소재끼리
    충돌하지 않게 한다(같은 소재는 항상 같은 해시 → upsert로 정상적으로 덮어써짐)."""
    ascii_part = re.sub(r"[^0-9A-Za-z_\-]", "", name_key)[:60]
    digest = hashlib.md5(name_key.encode("utf-8")).hexdigest()[:10]
    base = ascii_part or "creative"
    return f"{base}_{digest}"


def upload_creative_images(images: dict):
    """추출된 이미지를 Supabase Storage(버킷: creative-images)에 업로드하고
    ({(원본채널, 정규화된 소재명): 공개 URL} 딕셔너리, 에러 메시지 목록)을 반환한다.
    Supabase 미연결이거나 버킷이 없는 등 실패 시 조용히 건너뛴다(성과 저장 자체는 막지 않음)."""
    client = get_supabase_client()
    if client is None or not images:
        return {}, []
    urls = {}
    errors = []
    for (origin_channel, name_key), (data, ext) in images.items():
        folder = ORIGIN_CHANNEL_FOLDER.get(origin_channel, "etc")
        path = f"{folder}/{_safe_storage_name(name_key)}.{ext}"
        try:
            client.storage.from_(CREATIVE_IMAGE_BUCKET).upload(
                path, data, {"content-type": f"image/{ext}", "upsert": "true"}
            )
            urls[(origin_channel, name_key)] = client.storage.from_(CREATIVE_IMAGE_BUCKET).get_public_url(path)
        except Exception as e:
            if len(errors) < 3:  # 화면에 너무 길게 쌓이지 않도록 앞의 몇 개만 보관
                errors.append(f"{origin_channel}/{name_key}: {e}")
            continue
    return urls, errors


def attach_creative_images(creatives: pd.DataFrame, image_urls: dict) -> pd.DataFrame:
    """소재별 성과 데이터프레임(channel=탭 라벨, creative=소재명)에 image_url 컬럼을 붙인다."""
    if creatives is None or creatives.empty or not image_urls:
        return creatives
    creatives = creatives.copy()
    creatives["image_url"] = creatives.apply(
        lambda r: image_urls.get(
            (TAB_TO_ORIGIN_CHANNEL.get(r["channel"], r["channel"]), _creative_image_key(r["creative"]))
        ),
        axis=1,
    )
    return creatives


def backfill_missing_creative_images(creatives: pd.DataFrame) -> pd.DataFrame:
    """이번 주 엑셀에 이미지가 안 박혀있거나 인식이 안 돼서 image_url이 비는 소재는,
    과거에 저장된 같은 매체+소재명의 최근 이미지 URL이 있으면 그대로 이어붙인다.
    Storage 경로가 소재명 해시로 고정돼 있어 파일 자체는 계속 남아있는데, 이번 주
    스냅샷 행에는 그 URL을 다시 채워주지 않으면 이미지가 빈 걸로 보이는 문제가 있었음
    (2026-08 발견). 소재명이 같으면 매체가 같은 것으로 보고 가장 최근 저장분의 URL을 쓴다."""
    if creatives is None or creatives.empty:
        return creatives
    creatives = creatives.copy()
    if "image_url" not in creatives.columns:
        creatives["image_url"] = None

    missing_mask = creatives["image_url"].isna() | (creatives["image_url"].astype(str) == "")
    if not missing_mask.any():
        return creatives

    prev = load_table("creative_performance")
    if prev.empty or "image_url" not in prev.columns:
        return creatives
    prev = prev[prev["image_url"].notna() & (prev["image_url"].astype(str) != "")]
    if prev.empty:
        return creatives
    if "as_of_date" in prev.columns:
        prev = prev.sort_values("as_of_date")
    lookup = prev.drop_duplicates(subset=["channel", "creative"], keep="last").set_index(
        ["channel", "creative"]
    )["image_url"]

    def _fill(row):
        cur = row["image_url"]
        if pd.notna(cur) and str(cur) != "":
            return cur
        return lookup.get((row["channel"], row["creative"]))

    creatives.loc[missing_mask, "image_url"] = creatives.loc[missing_mask].apply(_fill, axis=1)
    return creatives


def get_hidden_rows(wb, sheet: str) -> set:
    """엑셀에서 마우스 우클릭 → '행 숨기기'로 숨긴 행의 번호(0-indexed, header=None으로 읽은
    raw DataFrame의 행 인덱스와 동일)를 반환한다. 소재 하나를 '이제 안 씀' 표시로 숨겨두면
    소재별 성과 집계에서 자동으로 제외하기 위한 용도 — 숨김을 풀면 다음 업로드부터 다시 잡힌다."""
    if wb is None or sheet not in wb.sheetnames:
        return set()
    ws = wb[sheet]
    hidden = set()
    for row_idx, dim in ws.row_dimensions.items():
        if dim.hidden:
            hidden.add(row_idx - 1)  # openpyxl 행 번호는 1-indexed
    return hidden


def get_visible_sheets(wb) -> set:
    """워크북 탭 자체가 숨겨진(우클릭 → 시트 숨기기) 시트 이름을 제외한, 눈에 보이는 시트 이름 집합.
    현재 안 쓰는 매체나 철 지난 시트를 통째로 탭 숨김 처리해두는 게 이 대시보드 운영자의
    실제 워크플로우라, '시트가 숨겨져 있으면 데이터도 안 가져온다'가 가장 직접적인 신호다.
    (wb가 없으면 None을 반환해 — 필터링 자체를 건너뛰고 기존처럼 전부 읽도록 — 안전하게 폴백한다.)"""
    if wb is None:
        return None
    return {name for name in wb.sheetnames if wb[name].sheet_state == "visible"}


def parse_workbook(file, today: date):
    xls = pd.ExcelFile(file)
    try:
        file.seek(0)
    except Exception:
        pass
    try:
        wb_for_hidden = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        wb_for_hidden = None
    finally:
        try:
            file.seek(0)
        except Exception:
            pass
    visible_sheets = get_visible_sheets(wb_for_hidden)
    result = {
        "weekly": pd.DataFrame(),
        "monthly": pd.DataFrame(),
        "daily": pd.DataFrame(),
        "channel_snapshot": pd.DataFrame(),
        "channels": pd.DataFrame(),
        "channels_weekly": pd.DataFrame(),
        "ga": pd.DataFrame(),
        "creatives": pd.DataFrame(),
        "channel_audience": pd.DataFrame(),
        "channel_sheets_found": [],
        "channel_sheets_parsed": [],
        "creative_sheets_found": [],
        "agency_notes": pd.DataFrame(),
    }
    if "매체통합" in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name="매체통합", header=None)
        raw = _drop_hidden_rows(raw, get_hidden_rows(wb_for_hidden, "매체통합"))
        bounds = find_sections(raw)
        result["monthly"] = parse_monthly(raw, bounds, today)
        result["weekly"] = parse_weekly(raw, bounds, today)
        result["daily"] = parse_daily(raw, bounds, today)
        result["channel_snapshot"] = parse_channel_snapshot(raw, bounds, result["monthly"])
        result["agency_notes"] = parse_agency_notes(raw, today)

    chan_frames = []
    chan_weekly_frames = []
    channel_sheet_candidates = discover_channel_sheets(xls)
    if visible_sheets is not None:
        channel_sheet_candidates = [s for s in channel_sheet_candidates if s in visible_sheets]
    for s in channel_sheet_candidates:
        result["channel_sheets_found"].append(s)
        df, weekly_df = parse_channel_sheet(xls, s, today, hidden_rows=get_hidden_rows(wb_for_hidden, s))
        if df is not None and len(df):
            chan_frames.append(df)
            result["channel_sheets_parsed"].append(s)
        if weekly_df is not None and len(weekly_df):
            chan_weekly_frames.append(weekly_df)
    result["channels_weekly"] = (
        pd.concat(chan_weekly_frames, ignore_index=True) if chan_weekly_frames else pd.DataFrame()
    )
    if chan_frames:
        result["channels"] = pd.concat(chan_frames, ignore_index=True)

    creative_frames = []
    creative_sheets = find_creative_sheets(xls)
    if visible_sheets is not None:
        creative_sheets = [s for s in creative_sheets if s in visible_sheets]
    for s in creative_sheets:
        hidden_rows = get_hidden_rows(wb_for_hidden, s)
        df = parse_creative_sheet(xls, s, today, hidden_rows=hidden_rows)
        if df is not None and len(df):
            creative_frames.append(df)
    result["creatives"] = pd.concat(creative_frames, ignore_index=True) if creative_frames else pd.DataFrame()
    result["creative_sheets_found"] = creative_sheets

    # 타겟팅별 성과(신규 타겟팅 vs 리타겟팅)용 그룹(타겟팅 오디언스) 단위 데이터.
    # 소재 파싱과 같은 시트 목록을 쓰되, 소재 상세표가 아니라 그 앞의 그룹 집계표를 읽는다.
    audience_frames = []
    for s in creative_sheets:
        df = parse_channel_group_sheet(xls, s, today)
        if df is not None and len(df):
            audience_frames.append(df)
    # 네이버 검색광고/쇼핑검색광고/브랜드검색광고 — (SA)/(SSP)/(브검) 시트는 '광고소재' 컬럼이
    # 없어 creative_sheets 목록에 안 잡히므로 별도로 찾아서 리타겟팅 채널로 추가한다.
    naver_search_sheets = [s for s in xls.sheet_names if _naver_search_channel_from_sheet(s)]
    for s in naver_search_sheets:
        df = parse_naver_search_sheet(xls, s, today)
        if df is not None and len(df):
            audience_frames.append(df)
    result["channel_audience"] = pd.concat(audience_frames, ignore_index=True) if audience_frames else pd.DataFrame()

    result["ga"] = parse_ga_raw(xls, today)
    return result


# ──────────────────────────────────────────────────────────────
# KPI
# ──────────────────────────────────────────────────────────────
def add_kpis(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    cost_in = df["cost_incl_vat"] if "cost_incl_vat" in df else 0
    cost_ex = df["cost_excl_vat"] if "cost_excl_vat" in df else 0
    df["ctr"] = np.where(df.get("impressions", 0) > 0, df["clicks"] / df["impressions"] * 100, 0)
    df["cpc"] = np.where(df.get("clicks", 0) > 0, cost_in / df["clicks"], 0)
    df["cpa"] = np.where(df.get("conversions", 0) > 0, cost_ex / df["conversions"], 0)
    df["cvr"] = np.where(df.get("clicks", 0) > 0, df["conversions"] / df["clicks"] * 100, 0)
    df["roas"] = np.where(cost_in > 0, df["revenue"] / cost_in * 100, 0)
    df["aov"] = np.where(df.get("conversions", 0) > 0, df["revenue"] / df["conversions"], 0)
    if "ga_revenue" in df.columns:
        df["ga_roas"] = np.where(cost_in > 0, df["ga_revenue"] / cost_in * 100, 0)
    return df


def kpi_cards(df: pd.DataFrame):
    cost = df["cost_incl_vat"].sum()
    revenue = df["revenue"].sum()
    conv = df["conversions"].sum()
    roas = (revenue / cost * 100) if cost > 0 else 0
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("총 광고비 (VAT포함)", f"{cost:,.0f} 원")
    c2.metric("총 매출", f"{revenue:,.0f} 원")
    c3.metric("총 전환수", f"{conv:,.0f} 건")
    c4.metric("ROAS", f"{roas:,.1f} %")


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="data")
    return buf.getvalue()


# 화면/엑셀에 표시할 때 쓰는 한글 컬럼명
KOR_COLS = {
    "channel": "매체",
    "creative": "소재명",
    "creative_image": "소재",
    "image_url": "소재이미지URL",
    "impressions": "노출수",
    "clicks": "클릭수",
    "cost_excl_vat": "광고비(VAT제외)",
    "cost_incl_vat": "광고비(VAT+)",
    "signups": "회원가입",
    "signup_rate": "가입율(%)",
    "audience_type": "구분(신규/리타겟팅)",
    "conversions": "전환수",
    "revenue": "매출",
    "ctr": "CTR(%)",
    "cpc": "CPC",
    "cpa": "CPA",
    "cvr": "CVR(%)",
    "roas": "ROAS(%)",
    "aov": "객단가",
    "ga_conversions": "GA-전환수",
    "ga_revenue": "GA-매출",
    "ga_roas": "GA-ROAS(%)",
    "report_month": "월",
    "report_date": "일자",
    "weekday": "요일",
    "week_start": "주 시작일",
    "week_end": "주 종료일",
    "label": "기간",
    "week_no": "주차",
    "week_range": "기간(월~일)",
    "as_of_month": "기준월",
    "as_of_date": "기준일",
    "source_medium": "소스/매체",
    "users": "사용자",
    "new_users": "신규방문자",
    "sessions": "세션",
    "bounce_rate": "이탈률(%)",
    "pages_per_session": "세션당 페이지수",
    "avg_session_duration": "평균 세션시간(초)",
    "ecommerce_cvr": "전자상거래 전환율(%)",
    "transactions": "거래수",
    # ↓ 유입·매출 비교(일별 GA·어드민 지표 비교) 전용 컬럼
    "returning_users": "재방문자",
    "pageviews": "페이지뷰",
    "admin_orders_total": "총 결제건수(어드민)",
    "admin_orders_real": "실 결제건수(어드민)",
    "admin_qty": "실 상품수량(어드민)",
    "admin_revenue": "어드민 매출(회사 내부 기준)",
    "price_rate": "판가율(%)",
    "admin_aov": "평균 구매금액(어드민)",
    "admin_returns": "반품건수",
    "return_rate": "반품율(%)",
    "ga_cvr": "GA-CVR(%)",
    "ga_aov": "GA-객단가",
    "new_paying_customers": "신규결제고객수",
    "cac": "CAC(고객획득비용)",
    "report_gap_pct": "보고서 매출 격차(%)",
    "ga_gap_pct": "GA 매출 격차(%)",
}


def korify(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns=KOR_COLS)


# 종합 대시보드(주간 추이)만 예전처럼 전체 프리셋을 쓰고, 나머지 페이지는 아래 축약판을 쓴다
# (형이 "오늘/어제/지난주/지난달/이번달/직접선택 6개만 남기고 줄여달라"고 요청 — 종합 대시보드는 제외).
DATE_PRESETS_FULL = [
    "오늘", "어제", "이번주", "지난주",
    "최근 7일(오늘 포함)", "최근 7일(오늘 제외)",
    "이번달", "지난달",
    "최근 30일(오늘 포함)", "최근 30일(오늘 제외)",
]
DATE_PRESETS_SHORT = ["오늘", "어제", "지난주", "지난달", "이번달"]


def _preset_to_range(name: str, min_d: date, max_d: date):
    """프리셋 이름 → (start, end). '오늘' 기준일은 실제 오늘 날짜이되, 데이터 범위 밖이면 잘라낸다."""
    today = date.today()
    if name == "오늘":
        s, e = today, today
    elif name == "어제":
        s = e = today - timedelta(days=1)
    elif name == "이번주":
        s, e = today - timedelta(days=today.weekday()), today
    elif name == "지난주":
        this_mon = today - timedelta(days=today.weekday())
        s = this_mon - timedelta(days=7)
        e = s + timedelta(days=6)
    elif name == "최근 7일(오늘 포함)":
        s, e = today - timedelta(days=6), today
    elif name == "최근 7일(오늘 제외)":
        s, e = today - timedelta(days=7), today - timedelta(days=1)
    elif name == "이번달":
        s, e = today.replace(day=1), today
    elif name == "지난달":
        last_prev = today.replace(day=1) - timedelta(days=1)
        s, e = last_prev.replace(day=1), last_prev
    elif name == "최근 30일(오늘 포함)":
        s, e = today - timedelta(days=29), today
    elif name == "최근 30일(오늘 제외)":
        s, e = today - timedelta(days=30), today - timedelta(days=1)
    else:
        s, e = min_d, max_d
    s = max(min_d, min(s, max_d))
    e = min(max_d, max(e, min_d))
    if s > e:
        s = e
    return s, e


DATE_PERIOD_OPTIONS_FULL = DATE_PRESETS_FULL + ["전체", "직접선택"]
DATE_PERIOD_OPTIONS = DATE_PRESETS_SHORT + ["직접선택"]


def period_filter(min_d: date, max_d: date, key: str, default_preset: str = "이번달", options: list = None):
    """날짜 프리셋 버튼 목록(누적 표의 preset_button_picker와 동일한 방식) + 직접선택 달력.
    버튼 아래에 실제로 적용된 날짜범위를 항상 캡션으로 보여준다. 반환값은 (start, end).
    options를 안 주면 축약판(오늘/어제/지난주/지난달/이번달/직접선택) 6개만 쓴다."""
    preset = preset_button_picker(options or DATE_PERIOD_OPTIONS, key=f"{key}_dateperiod", default=default_preset)

    if preset == "전체":
        start, end = min_d, max_d
    elif preset == "직접선택":
        narrow_col, _spacer = st.columns([3, 9])
        with narrow_col:
            date_range = st.date_input(
                "기간 직접 선택", value=(min_d, max_d), min_value=min_d, max_value=max_d, key=f"{key}_manual",
            )
        if isinstance(date_range, tuple) and len(date_range) == 2:
            start, end = date_range
        else:
            start, end = min_d, max_d
    else:
        start, end = _preset_to_range(preset, min_d, max_d)

    # 지금 어느 기간을 보고 있는지는 화면에서 가장 자주 확인하는 정보다.
    # 기본 caption은 너무 작아서 전용 스타일로 크게 뽑는다.
    st.markdown(
        f'<div class="stco-daterange">📆 {start:%Y-%m-%d} ~ {end:%Y-%m-%d}</div>',
        unsafe_allow_html=True,
    )
    return start, end


def preset_button_picker(options: list, key: str, default: str, label_prefix: str = "📅"):
    """연도/기간 프리셋을 달력 버튼 + 팝오버 패널 방식으로 고르는 UI (누적 표용).
    st.selectbox 대신 목업과 같은 '버튼 → 패널' 스타일을 쓰되, 옵션 내용은 그대로(연도 단위) 유지."""
    sel_key = f"{key}_preset_sel"
    if sel_key not in st.session_state or st.session_state[sel_key] not in options:
        st.session_state[sel_key] = default if default in options else options[0]

    current = st.session_state[sel_key]
    with st.popover(f"{label_prefix} {current}", use_container_width=False):
        cols = st.columns(2)
        for i, opt in enumerate(options):
            c = cols[i % 2]
            btn_type = "primary" if opt == current else "secondary"
            if c.button(opt, key=f"{key}_preset_opt_{i}", use_container_width=True, type=btn_type):
                st.session_state[sel_key] = opt
                st.rerun()
    return st.session_state[sel_key]


PAGE_SIZE_OPTIONS = [20, 50, 100, 200]

# 표시 포맷: 정수+콤마(돈/카운트), 소수점 2자리 %(CTR·CVR류), 소수점 0자리 %(ROAS류)
MONEY_COLS = {
    "impressions", "clicks", "signups", "conversions", "cost_excl_vat", "cost_incl_vat",
    "cpc", "cpa", "revenue", "aov", "ga_conversions", "ga_revenue",
    "users", "new_users", "sessions", "transactions",
    "returning_users", "pageviews", "admin_orders_total", "admin_orders_real", "admin_qty",
    "admin_revenue", "admin_aov", "admin_returns", "ga_aov", "new_paying_customers", "cac",
}
PCT2_COLS = {"ctr", "cvr", "bounce_rate", "ecommerce_cvr", "signup_rate", "price_rate", "return_rate", "ga_cvr"}
PCT0_COLS = {"roas", "ga_roas"}


KOR_WEEKDAY = ["월", "화", "수", "목", "금", "토", "일"]


def format_display(df: pd.DataFrame) -> pd.DataFrame:
    """화면/엑셀에 보여줄 때 쓰는 최종 포맷팅 (콤마, 소수점 자리수, 날짜 형식)."""
    df = df.copy()
    for c in df.columns:
        if c == "report_month":
            df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m")
        elif c == "report_date":
            dt = pd.to_datetime(df[c])
            df[c] = dt.dt.strftime("%Y-%m-%d") + " (" + dt.dt.dayofweek.map(lambda i: KOR_WEEKDAY[int(i)]) + ")"
        elif c in ("week_start", "week_end", "as_of_month", "as_of_date"):
            df[c] = pd.to_datetime(df[c]).dt.strftime("%Y-%m-%d")
        elif c in MONEY_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce").map(lambda v: f"{v:,.0f}" if pd.notna(v) else "")
        elif c in PCT2_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce").map(lambda v: f"{v:.2f}%" if pd.notna(v) else "")
        elif c in PCT0_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce").map(lambda v: f"{v:.0f}%" if pd.notna(v) else "")
    return df


KOR_COLS_REV = {v: k for k, v in KOR_COLS.items()}


def render_html_table(table: pd.DataFrame, raw: pd.DataFrame = None, raw_label_map: dict = None):
    """pandas Styler(jinja2 의존) 없이 순수 HTML로 표를 그린다.
    ▲(상승)는 빨간색, ▼(하락)는 파란색 글씨로 표시하고, 인덱스는 표시하지 않는다.
    raw(원본 숫자 컬럼, TOTAL 행 제외 · table의 데이터 행과 같은 순서)를 넘기면 raw에 있는
    숫자 컬럼에 한해 헤더를 클릭해 오름차순/내림차순 정렬할 수 있다(TOTAL 행은 항상 고정).
    raw_label_map(화면표시 헤더 → 원본 컬럼명)을 따로 주면, 공용 KOR_COLS 대신 이 표 전용
    라벨(예: "매체-매출")로 정렬 매칭을 한다 — 페이지마다 다른 이름을 쓰고 싶을 때 사용."""
    if table.empty:
        st.caption("데이터가 아직 없습니다.")
        return

    cols = list(table.columns)

    sortable_keys = {}
    if raw is not None:
        for c in cols:
            orig = (raw_label_map or {}).get(c) or KOR_COLS_REV.get(c)
            if orig and orig in raw.columns and pd.api.types.is_numeric_dtype(raw[orig]):
                sortable_keys[c] = orig

    if sortable_keys:
        thead = "".join(
            f'<th class="stco-sortable" data-key="{sortable_keys[c]}">{c} <span class="stco-sort-arrow">⇅</span></th>'
            if c in sortable_keys else f"<th>{c}</th>"
            for c in cols
        )
    else:
        thead = "".join(f"<th>{c}</th>" for c in cols)

    # TOTAL 행이 맨 위/맨 아래 중 어느 쪽에 있는지에 따라, 정렬 후에도 같은 위치에 고정한다.
    total_pin = "top" if (not table.empty and str(table.iloc[0][cols[0]]).strip() == "TOTAL") else "bottom"

    row_htmls = []
    raw_idx = 0  # table에서 TOTAL 행을 제외한 순번 == raw의 행 순번(raw가 있을 때)
    for _, row in table.iterrows():
        first_text = str(row[cols[0]]).strip()
        is_total = first_text == "TOTAL"
        is_label_row = is_total or first_text.endswith("대비")
        # TOTAL/증감 행에서 앞의 두 컬럼이 라벨용(둘 다 텍스트)이고 두 번째 칸이 비어있으면
        # 두 칸을 하나로 합쳐서(colspan) 가운데 정렬로 보여준다 (예: 주차 + 기간 컬럼).
        merge_first_two = False
        if is_label_row and len(cols) > 1:
            second_val = row[cols[1]]
            second_text = "" if pd.isna(second_val) else str(second_val).strip()
            merge_first_two = second_text == ""

        cells = []
        skip_next = False
        for i, c in enumerate(cols):
            if skip_next:
                skip_next = False
                continue
            val = row[c]
            text = "" if pd.isna(val) else str(val)
            style = ""
            colspan = ""
            if merge_first_two and i == 0:
                colspan = ' colspan="2"'
                style = "text-align:center;"
                skip_next = True
            if text.startswith("▲"):
                style += "color:#d93025;"
            elif text.startswith("▼"):
                style += "color:#1a73e8;"
            data_attr = ""
            if c in sortable_keys and not is_total and raw is not None and raw_idx < len(raw):
                sort_col = sortable_keys[c]
                raw_val = raw.iloc[raw_idx][sort_col]
                if pd.notna(raw_val):
                    data_attr = f' data-value="{float(raw_val)}"'
            cells.append(f'<td{colspan} style="{style}"{data_attr}>{text}</td>')
        row_class = ' class="stco-total-row"' if is_total else ""
        row_htmls.append(f"<tr{row_class}>{''.join(cells)}</tr>")
        if not is_total:
            raw_idx += 1

    table_style = f"""
    <style>
    .stco-table-wrap {{
        overflow-x:auto; border:1px solid {THEME_COLORS["border"]}; border-radius:10px; background:{THEME_COLORS["canvas"]};
        margin-bottom:18px;
    }}
    .stco-table {{
        width:100%; table-layout:auto; border-collapse:collapse; font-size:14px;
        font-family: {THEME_FONT_STACK};
    }}
    .stco-table th {{
        background:{THEME_COLORS["surface"]}; color:{THEME_COLORS["muted"]}; font-weight:600; padding:8px 12px;
        text-align:right; border-bottom:1px solid {THEME_COLORS["border"]}; white-space:nowrap;
    }}
    .stco-table th:first-child {{ text-align:left; border-top-left-radius:10px; width:1%; }}
    .stco-table th:last-child {{ border-top-right-radius:10px; }}
    .stco-table td {{
        padding:8px 12px; text-align:right; color:{THEME_COLORS["foreground"]};
        border-bottom:1px solid {THEME_COLORS["border"]}; white-space:nowrap;
    }}
    .stco-table td:first-child {{ text-align:left; }}
    .stco-table tr:last-child td {{ border-bottom:none; }}
    .stco-table tr:hover td {{ background:{THEME_COLORS["surface"]}; }}
    .stco-table tr.stco-total-row td {{
        background:{THEME_COLORS["surface"]}; color:{THEME_COLORS["foreground"]};
        font-weight:700; border-top:2px solid {THEME_COLORS["border"]};
    }}
    .stco-table tr.stco-total-row:hover td {{ background:{THEME_COLORS["surface"]}; }}
    .stco-sortable {{ cursor:pointer; user-select:none; }}
    .stco-sortable:hover {{ color:{THEME_COLORS["foreground"]}; }}
    .stco-sort-arrow {{ font-size:11px; margin-left:2px; }}
    .stco-sortable.stco-sort-active {{ color:{THEME_COLORS["foreground"]}; }}
    .stco-sortable.stco-sort-active .stco-sort-arrow {{ font-weight:700; }}
    </style>
    """

    if not sortable_keys:
        # 정렬 기능이 필요 없는 표는 기존대로 st.markdown으로 그린다(고정 높이 없이 자연스럽게 흐름).
        # 주의: 문자열 내용이 4칸 이상 들여쓰기 되면 마크다운이 "코드 블록"으로 인식해 HTML을
        # 렌더링하지 않고 태그를 그대로 텍스트로 보여준다 — 그래서 내용은 항상 왼쪽에 맞춰 쓴다.
        html = (
            table_style
            + '<div class="stco-table-wrap">'
            + '<table class="stco-table">'
            + f"<thead><tr>{thead}</tr></thead>"
            + f"<tbody>{''.join(row_htmls)}</tbody>"
            + "</table>"
            + "</div>"
        )
        st.markdown(html, unsafe_allow_html=True)
        return

    # 정렬 기능이 있는 표는 st.markdown이 <script>를 실행하지 않기 때문에 iframe(components.html)으로 그린다.
    n_rows = len(table)
    iframe_height = min(max(90 + n_rows * 41, 150), 900)
    table_id = f"stco-table-{abs(hash(tuple(cols))) % 100000}"
    html = f"""
    <html><head>{table_style}</head>
    <body style="margin:0; font-family:{THEME_FONT_STACK};">
    <div class="stco-table-wrap">
    <table class="stco-table" id="{table_id}">
      <thead><tr>{thead}</tr></thead>
      <tbody>{''.join(row_htmls)}</tbody>
    </table>
    </div>
    <script>
    (function() {{
        var table = document.getElementById("{table_id}");
        var tbody = table.querySelector("tbody");
        var headers = table.querySelectorAll("th.stco-sortable");
        var pinTop = {str(total_pin == "top").lower()};
        var state = {{ key: null, dir: 1 }};

        headers.forEach(function(th) {{
            th.addEventListener("click", function() {{
                var key = th.getAttribute("data-key");
                var colIndex = Array.prototype.indexOf.call(th.parentNode.children, th);
                if (state.key === key) {{
                    state.dir = -state.dir;
                }} else {{
                    state.key = key;
                    state.dir = -1;  // 처음 클릭하면 내림차순(큰 값 먼저)
                }}
                headers.forEach(function(h) {{
                    h.classList.remove("stco-sort-active");
                    h.querySelector(".stco-sort-arrow").textContent = "⇅";
                }});
                th.classList.add("stco-sort-active");
                th.querySelector(".stco-sort-arrow").textContent = state.dir === 1 ? "▲" : "▼";

                var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
                var totalRows = rows.filter(function(r) {{ return r.classList.contains("stco-total-row"); }});
                var dataRows = rows.filter(function(r) {{ return !r.classList.contains("stco-total-row"); }});

                dataRows.sort(function(a, b) {{
                    var av = parseFloat(a.children[colIndex].getAttribute("data-value"));
                    var bv = parseFloat(b.children[colIndex].getAttribute("data-value"));
                    var aNum = isNaN(av) ? -Infinity : av;
                    var bNum = isNaN(bv) ? -Infinity : bv;
                    return (aNum - bNum) * state.dir;
                }});

                dataRows.forEach(function(r) {{ tbody.appendChild(r); }});
                if (pinTop) {{
                    for (var i = totalRows.length - 1; i >= 0; i--) {{
                        tbody.insertBefore(totalRows[i], tbody.firstChild);
                    }}
                }} else {{
                    totalRows.forEach(function(r) {{ tbody.appendChild(r); }});
                }}
            }});
        }});
    }})();
    </script>
    </body></html>
    """
    components.html(html, height=iframe_height, scrolling=True)


# 소재별 성과 표에서 헤더 클릭으로 정렬 가능한 컬럼: 화면표시(한글) 헤더 → 정렬 기준이 되는 원본 컬럼 키.
# st.markdown은 보안상 <script>를 실행하지 않아서(react-markdown이 그냥 무시함), 클릭 정렬처럼
# JS가 필요한 표는 반드시 st.components.v1.html(iframe)로 렌더링해야 한다.
CREATIVE_SORT_COLS = {
    "노출수": "impressions",
    "클릭수": "clicks",
    "CTR(%)": "ctr",
    "CPC": "cpc",
    "전환수": "conversions",
    "ROAS(%)": "roas",
}


def render_sortable_creative_table(show: pd.DataFrame, raw: pd.DataFrame):
    """소재별 성과 표 전용 렌더러. render_html_table과 모양은 같지만, 노출/클릭/CTR/CPC/전환/ROAS
    헤더를 클릭하면 오름차순↔내림차순으로 재정렬된다 (TOTAL 행은 항상 맨 위 고정).
    show: korify+format_display를 거친 화면표시용 표 (맨 위 TOTAL 행 포함, 문자열 포맷).
    raw : TOTAL 행을 제외한, show의 데이터 행과 같은 순서의 원본(raw, 숫자) 값 표
          (정렬 기준값을 문자열이 아니라 실제 숫자로 비교하기 위해 필요)."""
    if show.empty:
        st.caption("데이터가 아직 없습니다.")
        return

    cols = list(show.columns)
    thead_cells = []
    for c in cols:
        sort_key = CREATIVE_SORT_COLS.get(c)
        if sort_key:
            thead_cells.append(
                f'<th class="stco-sortable" data-key="{sort_key}">{c} <span class="stco-sort-arrow">⇅</span></th>'
            )
        else:
            thead_cells.append(f"<th>{c}</th>")
    thead = "".join(thead_cells)

    row_htmls = []
    raw_idx = 0  # show에서 TOTAL 행(맨 위 1개)을 제외한 순번 == raw의 행 순번
    for _, row in show.iterrows():
        first_text = str(row[cols[0]]).strip()
        is_total = first_text == "TOTAL"
        cells = []
        for c in cols:
            val = row[c]
            text = "" if pd.isna(val) else str(val)
            sort_key = CREATIVE_SORT_COLS.get(c)
            data_attr = ""
            if sort_key and not is_total and raw_idx < len(raw) and sort_key in raw.columns:
                raw_val = raw.iloc[raw_idx][sort_key]
                if pd.notna(raw_val):
                    data_attr = f' data-value="{float(raw_val)}"'
            cells.append(f"<td{data_attr}>{text}</td>")
        row_class = ' class="stco-total-row"' if is_total else ""
        row_htmls.append(f"<tr{row_class}>{''.join(cells)}</tr>")
        if not is_total:
            raw_idx += 1

    n_data_rows = max(len(show) - 1, 0)
    iframe_height = min(max(110 + n_data_rows * 156, 220), 900)

    html = f"""
    <style>
    body {{ margin:0; font-family:{THEME_FONT_STACK}; }}
    .stco-table-wrap {{
        overflow:auto; border:1px solid {THEME_COLORS["border"]}; border-radius:10px; background:{THEME_COLORS["canvas"]};
    }}
    .stco-table {{ width:auto; max-width:100%; border-collapse:collapse; font-size:14px; }}
    .stco-table th {{
        background:{THEME_COLORS["surface"]}; color:{THEME_COLORS["muted"]}; font-weight:600; padding:8px 14px;
        text-align:right; border-bottom:1px solid {THEME_COLORS["border"]}; white-space:nowrap;
        position:sticky; top:0; z-index:1;
    }}
    .stco-table th:first-child {{ text-align:left; border-top-left-radius:10px; }}
    .stco-table th:last-child {{ border-top-right-radius:10px; }}
    .stco-table td {{
        padding:8px 14px; text-align:right; color:{THEME_COLORS["foreground"]};
        border-bottom:1px solid {THEME_COLORS["border"]}; white-space:nowrap;
    }}
    .stco-table td:first-child {{ text-align:left; }}
    .stco-table tr:last-child td {{ border-bottom:none; }}
    .stco-table tr:hover td {{ background:{THEME_COLORS["surface"]}; }}
    .stco-table tr.stco-total-row td {{
        background:{THEME_COLORS["surface"]}; color:{THEME_COLORS["foreground"]};
        font-weight:700; border-top:2px solid {THEME_COLORS["border"]};
    }}
    .stco-sortable {{ cursor:pointer; user-select:none; }}
    .stco-sortable:hover {{ color:{THEME_COLORS["foreground"]}; }}
    .stco-sort-arrow {{ font-size:11px; margin-left:2px; }}
    .stco-sortable.stco-sort-active {{ color:{THEME_COLORS["foreground"]}; }}
    .stco-sortable.stco-sort-active .stco-sort-arrow {{ font-weight:700; }}
    </style>
    <div class="stco-table-wrap">
    <table class="stco-table" id="stco-creative-table">
      <thead><tr>{thead}</tr></thead>
      <tbody>{''.join(row_htmls)}</tbody>
    </table>
    </div>
    <script>
    (function() {{
        var table = document.getElementById("stco-creative-table");
        var tbody = table.querySelector("tbody");
        var headers = table.querySelectorAll("th.stco-sortable");
        var state = {{ key: null, dir: 1 }};

        headers.forEach(function(th) {{
            th.addEventListener("click", function() {{
                var key = th.getAttribute("data-key");
                var colIndex = Array.prototype.indexOf.call(th.parentNode.children, th);
                if (state.key === key) {{
                    state.dir = -state.dir;
                }} else {{
                    state.key = key;
                    state.dir = -1;  // 처음 클릭하면 내림차순(큰 값 먼저)
                }}
                headers.forEach(function(h) {{
                    h.classList.remove("stco-sort-active");
                    h.querySelector(".stco-sort-arrow").textContent = "⇅";
                }});
                th.classList.add("stco-sort-active");
                th.querySelector(".stco-sort-arrow").textContent = state.dir === 1 ? "▲" : "▼";

                var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
                var totalRows = rows.filter(function(r) {{ return r.classList.contains("stco-total-row"); }});
                var dataRows = rows.filter(function(r) {{ return !r.classList.contains("stco-total-row"); }});

                dataRows.sort(function(a, b) {{
                    var av = parseFloat(a.children[colIndex].getAttribute("data-value"));
                    var bv = parseFloat(b.children[colIndex].getAttribute("data-value"));
                    var aNum = isNaN(av) ? -Infinity : av;
                    var bNum = isNaN(bv) ? -Infinity : bv;
                    return (aNum - bNum) * state.dir;
                }});

                totalRows.forEach(function(r) {{ tbody.appendChild(r); }});
                dataRows.forEach(function(r) {{ tbody.appendChild(r); }});
                for (var i = totalRows.length - 1; i >= 0; i--) {{
                    tbody.insertBefore(totalRows[i], tbody.firstChild);
                }}
            }});
        }});
    }})();
    </script>
    """
    components.html(html, height=iframe_height, scrolling=True)


def pct_change_row(d_full: pd.DataFrame, latest_pos: int, numeric_cols: list, label_col: str, label_text: str = "전기간 대비"):
    """d_full(전체 정렬 데이터)에서 latest_pos 위치의 행과 바로 이전 행을 비교해 증감율 행을 만든다."""
    if latest_pos <= 0 or latest_pos >= len(d_full):
        return None
    latest, prev = d_full.iloc[latest_pos], d_full.iloc[latest_pos - 1]
    row = {label_col: label_text}
    for c in numeric_cols:
        if c not in d_full.columns:
            continue
        pv, lv = prev.get(c), latest.get(c)
        if pd.isna(pv) or pv in (0, None):
            row[c] = "-"
            continue
        change = (lv - pv) / abs(pv) * 100
        arrow = "▲" if change >= 0 else "▼"
        row[c] = f"{arrow}{change:+.1f}%"
    return row


def build_total_row(view_raw: pd.DataFrame, display_cols: list, label_col: str, label_text: str = "TOTAL"):
    """현재 화면에 표시 중인 원본(raw) 행들을 합산해 TOTAL 행을 만든다.
    노출/클릭/비용/전환/매출 등은 단순 합산하고, CTR·CPC·CPA·CVR·ROAS·객단가·GA-ROAS 같은
    비율/단가 지표는 합산된 값 기준으로 다시 계산한다 (개별 행 비율의 평균이 아님)."""
    if view_raw is None or view_raw.empty:
        return None

    def s(col):
        return pd.to_numeric(view_raw[col], errors="coerce").sum() if col in view_raw.columns else 0

    imp, clk = s("impressions"), s("clicks")
    cost_ex, cost_in = s("cost_excl_vat"), s("cost_incl_vat")
    signups, conv, revenue = s("signups"), s("conversions"), s("revenue")
    ga_conv, ga_rev = s("ga_conversions"), s("ga_revenue")

    raw = {
        "impressions": imp, "clicks": clk, "cost_excl_vat": cost_ex, "cost_incl_vat": cost_in,
        "signups": signups, "conversions": conv, "revenue": revenue,
        "ga_conversions": ga_conv, "ga_revenue": ga_rev,
        "ctr": (clk / imp * 100) if imp else 0,
        "cpc": (cost_in / clk) if clk else 0,
        "cpa": (cost_ex / conv) if conv else 0,
        "cvr": (conv / clk * 100) if clk else 0,
        "roas": (revenue / cost_in * 100) if cost_in else 0,
        "aov": (revenue / conv) if conv else 0,
        "ga_roas": (ga_rev / cost_in * 100) if cost_in else 0,
    }
    row = {label_col: label_text}
    for c in display_cols:
        if c == label_col:
            continue
        if c not in raw:
            row[c] = ""
            continue
        v = raw[c]
        if c in MONEY_COLS:
            row[c] = f"{v:,.0f}"
        elif c in PCT2_COLS:
            row[c] = f"{v:.2f}%"
        elif c in PCT0_COLS:
            row[c] = f"{v:.0f}%"
        else:
            row[c] = v
    return row


def build_year_options(date_series: pd.Series):
    years = sorted({d.year for d in pd.to_datetime(date_series).dropna()})
    return years


def render_pager(total_pages: int, key: str) -> int:
    """« 1 2 3 » 형태의 페이지 버튼. 처음 볼 땐 최신 데이터가 있는 마지막 페이지부터 보여준다."""
    state_key = f"{key}_pagenum"
    if state_key not in st.session_state:
        st.session_state[state_key] = total_pages
    cur = min(max(st.session_state[state_key], 1), total_pages)

    window = 5
    start_p = max(1, cur - window // 2)
    end_p = min(total_pages, start_p + window - 1)
    start_p = max(1, end_p - window + 1)

    spacer, pager_area = st.columns([3, 4])
    with pager_area:
        n_buttons = end_p - start_p + 3
        btn_cols = st.columns(n_buttons)
        if btn_cols[0].button("«", key=f"{key}_prev", disabled=cur <= 1, use_container_width=True):
            cur = max(1, cur - 1)
        for i, p in enumerate(range(start_p, end_p + 1)):
            if btn_cols[i + 1].button(
                str(p), key=f"{key}_p{p}", type=("primary" if p == cur else "secondary"), use_container_width=True
            ):
                cur = p
        if btn_cols[-1].button("»", key=f"{key}_next", disabled=cur >= total_pages, use_container_width=True):
            cur = min(total_pages, cur + 1)
    st.session_state[state_key] = cur
    return cur


def render_cumulative_table(df: pd.DataFrame, date_col: str, show_cols: list, numeric_cols: list,
                             title: str, key: str, mode: str):
    """월별/주간/일자별 누적 표.
    mode: 'month' → 기본값 당해년도(1월~최신월) / 'week' → 기본값 최근 5주 / 'day' → 기본값 이번달
    프리셋: (week·day는) 기본, 연도별, 전체, 직접선택
    """
    st.markdown(f"#### {title}")
    if df is None or df.empty:
        st.caption("데이터가 아직 없습니다.")
        return

    d = df.sort_values(date_col).reset_index(drop=True)
    d[date_col] = pd.to_datetime(d[date_col])
    years = build_year_options(d[date_col])
    year_labels = [f"{y}년" for y in years]

    if mode == "month":
        options = year_labels + ["전체", "직접선택"]
        this_year = date.today().year
        default_label = f"{this_year}년" if this_year in years else (year_labels[-1] if year_labels else "전체")
    else:
        options = ["기본"] + year_labels + ["전체", "직접선택"]
        default_label = "기본"

    preset = preset_button_picker(options, key=f"{key}_periodpicker", default=default_label)

    need_pagination = True
    if preset == "기본":
        if mode == "week":
            view_all = d.tail(5)
        else:  # day
            cur_month = pd.Timestamp(date.today().replace(day=1))
            view_all = d[d[date_col] >= cur_month]
            if view_all.empty:
                view_all = d.tail(31)
        need_pagination = False
    elif preset == "전체":
        view_all = d
    elif preset == "직접선택":
        min_d, max_d = d[date_col].min().date(), d[date_col].max().date()
        narrow_col, _spacer = st.columns([3, 9])
        with narrow_col:
            date_range = st.date_input("기간 직접 선택", value=(min_d, max_d), min_value=min_d, max_value=max_d, key=f"{key}_manual")
        start, end = date_range if isinstance(date_range, tuple) and len(date_range) == 2 else (min_d, max_d)
        view_all = d[(d[date_col].dt.date >= start) & (d[date_col].dt.date <= end)]
    else:  # "YYYY년"
        y = int(preset.replace("년", ""))
        view_all = d[d[date_col].dt.year == y]

    total = len(view_all)
    if need_pagination and total > PAGE_SIZE_OPTIONS[0]:
        narrow_ps_col, _ps_spacer = st.columns([2, 10])
        with narrow_ps_col:
            page_size = st.selectbox("페이지당 표시", PAGE_SIZE_OPTIONS, index=1, key=f"{key}_{preset}_pagesize")
        total_pages = max(1, -(-total // page_size))
        page = render_pager(total_pages, key=f"{key}_{preset}") if total_pages > 1 else 1
        start_i, end_i = (page - 1) * page_size, page * page_size
        view = view_all.iloc[start_i:end_i]
        show_change_row = page == total_pages
    else:
        view = view_all
        show_change_row = True

    display_cols = [c for c in show_cols if c in view.columns]
    table = format_display(view[display_cols])  # 먼저 숫자/날짜 포맷 적용 (증감율 행은 이미 문자열이라 따로 붙임)

    change_label = {"month": "전월 대비", "week": "전주 대비", "day": "전일 대비"}.get(mode, "전기간 대비")
    has_change_row = False
    if show_change_row and len(view):
        latest_pos = view.index[-1]
        change_row = pct_change_row(d, latest_pos, numeric_cols, display_cols[0], label_text=change_label)
        if change_row:
            table = pd.concat([table, pd.DataFrame([change_row])], ignore_index=True)
            has_change_row = True

    total_row = build_total_row(view[display_cols], display_cols, display_cols[0], label_text="TOTAL")
    if total_row:
        table = pd.concat([table, pd.DataFrame([total_row])], ignore_index=True)

    render_html_table(korify(table))
    st.download_button(
        f"⬇️ 엑셀 다운로드 ({title})",
        data=to_excel_bytes(korify(format_display(view[display_cols]))),
        file_name=f"{key}.xlsx",
        key=f"{key}_{preset}_dl",
    )


# ──────────────────────────────────────────────────────────────
# 업로드 패널
# ──────────────────────────────────────────────────────────────
# 매체 리포트(소재별 일일 성과 등)에서 찾을 컬럼 이름 후보.
# 매체마다 헤더 문구가 조금씩 달라서 넉넉하게 잡는다.
MEDIA_REPORT_COLS = {
    "date": ["기간", "날짜", "일자", "일별", "date", "day"],
    "cost": ["총비용", "광고비", "비용", "소진액", "지출", "cost", "spend"],
    # 완전일치를 먼저 보므로 '노출'이 '노출 기간'(시간대별 표)에 잘못 붙지 않는다.
    "imp": ["노출수", "노출", "impression", "imp"],
    "click": ["클릭수", "클릭", "click", "clk"],
    "campaign": ["캠페인 이름", "캠페인명", "캠페인", "campaign"],
}


def _mr_pick(cols, keys):
    """헤더 목록에서 후보 키워드에 맞는 컬럼을 하나 고른다(정확히 같은 이름 우선)."""
    low = {c: str(c).strip().lower().replace(" ", "") for c in cols}
    for k in keys:
        kk = k.lower().replace(" ", "")
        for c in cols:
            if low[c] == kk:
                return c
    for k in keys:
        kk = k.lower().replace(" ", "")
        for c in cols:
            if kk in low[c]:
                return c
    return None


def _mr_is_header(cells) -> bool:
    """이 행이 표의 머리글인지 판단한다.

    '노출'이라는 글자가 본문에 우연히 들어간 경우와 구분해야 해서, 지표 이름과
    **완전히 같은** 셀이 2개 이상일 때만 머리글로 본다.
    """
    vals = {str(v).strip().lower().replace(" ", "") for v in cells}
    hit = 0
    for key in ("date", "imp", "click", "cost"):
        if any(k.lower().replace(" ", "") in vals for k in MEDIA_REPORT_COLS[key]):
            hit += 1
    return hit >= 2


def _mr_read(file) -> pd.DataFrame:
    """csv/xlsx 어느 쪽이든 읽고, 머리글이 몇 줄 아래 있어도 찾아낸다."""
    name = (getattr(file, "name", "") or "").lower()
    file.seek(0)
    if name.endswith(".csv"):
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                file.seek(0)
                raw = pd.read_csv(file, encoding=enc, header=None, dtype=str)
                break
            except Exception:
                raw = None
        if raw is None:
            raise RuntimeError("CSV를 읽지 못했습니다(인코딩 확인 필요).")
    else:
        raw = pd.read_excel(file, header=None, dtype=str)

    # 보장형 매체(맨즈탭 등) 리포트에는 비용 열이 아예 없고 노출·클릭만 있는 경우가 있다.
    # 그래서 '날짜 + (비용 또는 노출 또는 클릭)'이면 머리글로 인정한다.
    hdr = None
    for i in range(min(len(raw), 30)):
        cells = list(raw.iloc[i])
        if _mr_pick(cells, MEDIA_REPORT_COLS["date"]) is None:
            continue
        if any(_mr_pick(cells, MEDIA_REPORT_COLS[k]) is not None
               for k in ("cost", "imp", "click")):
            hdr = i
            break
    if hdr is None:
        raise RuntimeError(
            "날짜 열을 못 찾았습니다. 파일에 '기간/날짜/일자' 같은 머리글과 "
            "'총비용/노출수/클릭수' 중 하나 이상이 있어야 합니다.")

    # 한 시트에 표가 여러 개 쌓여 있는 파일이 있다(맨즈탭 TOTAL 시트: 전체 / 대표소재 / 추가소재).
    # 아래 표들은 위 표를 쪼갠 것이라 같이 더하면 값이 배로 부푼다. 다음 머리글 직전에서 끊는다.
    end = len(raw)
    for i in range(hdr + 1, len(raw)):
        if _mr_is_header(list(raw.iloc[i])):
            end = i
            break

    df = raw.iloc[hdr + 1:end].copy()
    df.columns = [str(v).strip() for v in raw.iloc[hdr]]
    return df.reset_index(drop=True)


def _mr_dates(s: pd.Series) -> pd.Series:
    """매체마다 제각각인 날짜 표기를 하나로 맞춘다.

    실제로 만난 것들:
      · '2026.08.11.'      — GFA, 끝에 점
      · '2026-07-06 (월)'  — 맨즈탭, 요일 괄호
      · '20260706'         — GA raw, 구분자 없음
    """
    t = s.astype(str).str.strip()
    t = t.str.replace(r"\s*\([^)]*\)\s*$", "", regex=True)   # 끝의 (월) 제거
    t = t.str.rstrip(".")
    out = pd.to_datetime(t, errors="coerce")
    miss = out.isna() & t.str.fullmatch(r"\d{8}")
    if miss.any():
        out.loc[miss] = pd.to_datetime(t[miss], format="%Y%m%d", errors="coerce")
    return out.dt.date


def _mr_channel_of(campaign: str, filename: str) -> str:
    """캠페인명·파일명으로 어느 매체인지 가른다.

    GFA 계정 안에 GFA(웹사이트 전환)와 애드부스트(ADVoost 쇼핑)가 같이 있고,
    맨즈탭은 자사몰/외부몰로 갈라 보기 때문에 이름으로 구분해야 한다.
    """
    blob = f"{campaign} {filename}".lower()
    ext = ("외부몰" in blob) or ("스마트스토어" in blob)
    if any(k in blob for k in ("advoost", "adboost", "애드부스트", "ad voost")):
        return "네이버 애드부스트"
    if any(k in blob for k in ("맨즈탭", "맨즈텝", "manstab", "n.box", "핫아이템")):
        return "네이버 맨즈탭_외부몰" if ext else "네이버 맨즈탭_자사몰"
    if "gfa" in blob or "자사몰" in blob or "전환" in blob:
        return "네이버 GFA"
    return "네이버 GFA"


def parse_media_report_file(file, vat_included: bool = True) -> pd.DataFrame:
    """매체 관리자에서 내려받은 '일별/소재별 성과' 파일을 일별 광고비로 접는다.

    소재·광고그룹 단위로 여러 줄이 와도 날짜×매체로 합산한다.
    네이버 계열 리포트(GFA·맨즈탭)의 '총비용'은 VAT가 포함된 금액이라 기본값을 True로 둔다.
    VAT 별도로 내려오는 매체 파일이면 체크를 풀어 ×1.1을 적용한다.
    """
    df = _mr_read(file)
    fname = getattr(file, "name", "") or ""
    cols = list(df.columns)
    c_date = _mr_pick(cols, MEDIA_REPORT_COLS["date"])
    c_cost = _mr_pick(cols, MEDIA_REPORT_COLS["cost"])   # 없을 수 있다(보장형)
    c_imp = _mr_pick(cols, MEDIA_REPORT_COLS["imp"])
    c_clk = _mr_pick(cols, MEDIA_REPORT_COLS["click"])
    c_cmp = _mr_pick(cols, MEDIA_REPORT_COLS["campaign"])

    def num(s):
        return pd.to_numeric(
            s.astype(str).str.replace(r"[^0-9.\-]", "", regex=True), errors="coerce").fillna(0)

    out = pd.DataFrame({
        "report_date": _mr_dates(df[c_date]),
        "cost_incl_vat": num(df[c_cost]) if c_cost else 0.0,
        "impressions": num(df[c_imp]) if c_imp else 0,
        "clicks": num(df[c_clk]) if c_clk else 0,
        "channel": (df[c_cmp].astype(str) if c_cmp else pd.Series([""] * len(df))).map(
            lambda v: _mr_channel_of(v, fname)),
    })
    out = out.dropna(subset=["report_date"])
    if not vat_included:
        out["cost_incl_vat"] = out["cost_incl_vat"] * 1.1
    out = out.groupby(["report_date", "channel"], as_index=False).agg(
        cost_incl_vat=("cost_incl_vat", "sum"),
        impressions=("impressions", "sum"), clicks=("clicks", "sum"))
    return out[out[["cost_incl_vat", "impressions", "clicks"]].sum(axis=1) > 0]


def detect_upload_kind(file) -> str:
    """올린 엑셀이 네 종류 중 무엇인지 판단한다.

    파일명으로 먼저 보고(가장 확실), 안 되면 각 파서를 실제로 돌려서 데이터가 나오는 쪽을 고른다.
    각 파서에는 이미 시트명·필수컬럼 가드가 있어서 엉뚱한 파일에는 빈 결과를 낸다.
    마지막까지 못 고르면 '주간 리포트'로 본다(가장 자주 올리는 파일).
    """
    name = (getattr(file, "name", "") or "")
    flat = name.replace(" ", "").lower()
    if "주간보고서" in flat:
        return "weekly"
    if any(k in flat for k in ("맨즈탭", "맨즈텝", "소재별", "일일성과", "result.csv")):
        return "media_report"
    if "채널믹스" in flat:
        return "mix"
    if "예산" in flat and "믹스" not in flat:
        return "budget"

    try:
        file.seek(0)
        xls = pd.ExcelFile(file)
    except Exception:
        try:
            if not parse_media_report_file(file).empty:
                return "media_report"
        except Exception:
            pass
        return "weekly"

    sheets = " ".join(str(s) for s in xls.sheet_names)
    if "매체통합" in sheets:
        return "weekly"

    try:
        if not parse_media_report_file(file).empty:
            return "media_report"
    except Exception:
        pass

    for kind, fn in (
        ("mix", lambda: parse_channel_mix_sheet(xls, source_name=name)),
        ("budget", lambda: parse_channel_budget_sheet(xls)),
        ("ga", lambda: parse_inflow_revenue_sheet(xls)),
        ("ga", lambda: parse_utm_channel_map(xls)),
    ):
        try:
            if not fn().empty:
                return kind
        except Exception:
            continue
    return "weekly"


def render_upload_panel():
    st.sidebar.header("⚙️ 데이터 관리")
    client = get_supabase_client()
    st.sidebar.caption(f"저장소: {'Supabase (Postgres)' if client else '로컬 세션 (테스트용, 새로고침 시 초기화)'}")

    # 업로더를 4개 두면 사이드바가 너무 길어져서 하나로 합쳤다. 파일을 넣으면 종류를 스스로
    # 알아내고, 잘못 잡히면 아래 셀렉트로 직접 고를 수 있게 한다.
    up_file = st.sidebar.file_uploader(
        "데이터 파일 업로드", type=["xlsx", "xls", "csv"], key="unified_uploader",
        help="주간 리포트 · GA 유입 데이터 · 연간 예산 · 채널 믹스 · 매체 리포트(GFA·맨즈탭 등) — "
             "아무거나 넣으면 종류를 알아서 판단합니다.",
    )

    KIND_LABELS = {
        "weekly": "① 주간 리포트 (STCO_주간보고서)",
        "ga": "② GA 유입 데이터 (유입·매출 / 매체별 유입 / UTM 매핑표)",
        "budget": "③ 연간 예산 (◆26년 월별 예산 정리)",
        "mix": "④ 채널 믹스 (26년 매체별 채널 믹스)",
        "media_report": "⑤ 매체 리포트 (GFA·맨즈탭 등 일별 성과)",
    }
    kind = None
    if up_file is not None:
        auto = detect_upload_kind(up_file)
        opts = ["자동 인식: " + KIND_LABELS[auto]] + [v for k, v in KIND_LABELS.items()]
        pick = st.sidebar.selectbox("파일 종류", opts, index=0, key="unified_kind")
        if pick.startswith("자동 인식"):
            kind = auto
        else:
            kind = next(k for k, v in KIND_LABELS.items() if v == pick)

    file = up_file if kind == "weekly" else None
    ga_file = up_file if kind == "ga" else None
    budget_file = up_file if kind == "budget" else None
    mix_file = up_file if kind == "mix" else None
    mr_file = up_file if kind == "media_report" else None

    if file is not None:
        today = date.today()
        with st.sidebar.status("파일 분석 중...", expanded=True) as status:
            result = parse_workbook(file, today)
            st.write(f"📅 월별 통합데이터: {len(result['monthly'])}개월")
            st.write(f"📆 통합 주간별: {len(result['weekly'])}주")
            st.write(f"🗓️ 통합 일자별: {len(result['daily'])}일")
            st.write(f"🏷️ 당월 매체별 스냅샷: {len(result['channel_snapshot'])}개 매체")
            st.write(f"📊 매체별 시트 인식: {len(result['channel_sheets_parsed'])}/{len(result['channel_sheets_found'])}개")
            cw = result.get("channels_weekly", pd.DataFrame())
            st.write(f"📆 매체별 주간 데이터 인식: {len(cw)}행 (매체 {cw['channel'].nunique() if not cw.empty else 0}종)")
            st.write(f"🔎 GA 유입경로: {len(result['ga'])}건")
            st.write(
                f"🎨 소재별 성과 인식: {len(result.get('creatives', []))}행 "
                f"({', '.join(result.get('creative_sheets_found', [])) or '해당 시트 없음'})"
            )
            agency_notes_df = result.get("agency_notes", pd.DataFrame())
            st.write(
                f"📝 매체통합 시트 하단 운영 메모: "
                f"{'인식됨' if not agency_notes_df.empty else '인식 안 됨(없거나 형식 미확인)'}"
            )
            audience_df = result.get("channel_audience", pd.DataFrame())
            unclassified_n = (
                int((audience_df["audience_type"] == "미분류").sum()) if not audience_df.empty else 0
            )
            st.write(f"🧭 타겟팅별 성과(신규/리타겟팅) 그룹 인식: {len(audience_df)}행")
            if unclassified_n:
                st.warning(
                    f"'미분류'로 남은 그룹 {unclassified_n}행이 있습니다 — 새 캠페인/그룹이 추가된 것일 수 있어요. "
                    "app.py의 AUDIENCE_TYPE_MAP에 분류를 추가해주세요."
                )
            missing = set(result["channel_sheets_found"]) - set(result["channel_sheets_parsed"])
            if missing:
                st.warning(f"인식 실패한 매체 시트: {', '.join(missing)}")
            status.update(label="분석 완료", state="complete")

        if st.sidebar.button("💾 전체 저장하기", type="primary"):
            creatives_df = result.get("creatives", pd.DataFrame())
            if not creatives_df.empty:
                with st.sidebar.status("🖼️ 소재 이미지 추출/업로드 중...", expanded=False):
                    sheets_and_channels = [
                        (s, _infer_channel_from_sheet(s)) for s in result.get("creative_sheets_found", [])
                    ]
                    images = extract_creative_images(file, sheets_and_channels)
                    image_urls, image_errors = upload_creative_images(images)
                    creatives_df = attach_creative_images(creatives_df, image_urls)
                    if images:
                        st.write(f"소재 이미지 {len(images)}개 인식, {len(image_urls)}개 업로드 성공")
                        if image_errors:
                            st.warning("업로드 실패 예시:\n" + "\n".join(image_errors))
                    before_missing = int(
                        creatives_df["image_url"].isna().sum() if "image_url" in creatives_df.columns
                        else len(creatives_df)
                    )
                    creatives_df = backfill_missing_creative_images(creatives_df)
                    after_missing = int(creatives_df["image_url"].isna().sum())
                    backfilled_n = before_missing - after_missing
                    if backfilled_n > 0:
                        st.write(f"🔁 이번 파일엔 이미지가 없어 이전 저장분 이미지로 채운 소재: {backfilled_n}개")

            n1 = save_table("weekly_overview", result["weekly"], "week_start", file.name)
            n2 = save_table("monthly_overview", result["monthly"], "report_month", file.name)
            n3 = save_table("channel_monthly", result["channels"], "report_month,channel", file.name)
            n3w = save_table("channel_weekly", result.get("channels_weekly", pd.DataFrame()), "channel,week_start", file.name)
            n4 = save_table("channel_snapshot", result["channel_snapshot"], "as_of_month,channel", file.name)
            n5 = save_table("ga_source", result["ga"], "as_of_date,source_medium", file.name)
            n6 = save_table("daily_overview", result["daily"], "report_date", file.name)
            delete_creative_performance_for_date(today)
            n7 = save_table(
                "creative_performance", creatives_df,
                "as_of_date,channel,creative", file.name,
            )
            delete_channel_audience_for_date(today)
            n8 = save_table(
                "channel_audience_snapshot", result.get("channel_audience", pd.DataFrame()),
                "as_of_date,channel,audience_type", file.name,
            )
            n9 = save_table(
                "agency_notes", result.get("agency_notes", pd.DataFrame()),
                "as_of_date", file.name,
            )
            st.cache_data.clear()
            st.sidebar.success(
                f"저장 완료! 주간 {n1} · 월별 {n2} · 일자별 {n6} · 매체(월) {n3} · 매체(주간) {n3w} · 매체(당월) {n4} · "
                f"GA {n5}건 · 소재 {n7}건 · 퍼널(신규/리타겟) {n8}건 · 운영 메모 {n9}건"
            )
            st.rerun()

    if ga_file is not None:
        with st.sidebar.status("파일 분석 중...", expanded=True) as status2:
            # 세 파일은 시트 구조가 서로 달라서, 하나의 업로더에서 세 파서를 다 시도해보고
            # 실제로 데이터가 나온 쪽만 채택한다(맞지 않는 파서는 항상 빈 결과를 내도록 이미
            # 각 파서 안에 시트명/필수컬럼 가드가 있어서 안전하게 겸용 가능).
            ga_xls = pd.ExcelFile(ga_file)
            inflow_df = parse_inflow_revenue_sheet(ga_xls)
            utm_map_df = parse_utm_channel_map(ga_xls)
            # 이번에 새로 올라온 매핑표 + 이미 저장돼있는 매핑표를 합쳐서 '매체' 채널 그룹핑에 쓴다
            # (새 파일에 없는 소스/매체는 기존에 저장된 매핑을 그대로 쓰기 위함).
            existing_utm_map = load_table("utm_channel_map")
            combined_utm_map = pd.concat([existing_utm_map, utm_map_df], ignore_index=True) if not utm_map_df.empty else existing_utm_map
            channel_lookup = build_utm_channel_lookup(combined_utm_map)
            ga_channel_df = parse_ga_channel_inflow_sheet(ga_xls, channel_map=channel_lookup)

            if inflow_df.empty and ga_channel_df.empty and utm_map_df.empty:
                st.warning(
                    "인식 가능한 데이터를 찾지 못했습니다. '일별 GA,어드민 지표 비교' 시트가 있는 "
                    "유입·매출 비교 파일이거나, 날짜/세션 소스/매체 컬럼이 있는 GA 매체별 유입 경로 "
                    "파일, 또는 '매체명'/'소스 / 매체' 컬럼이 있는 UTM 매핑표 파일인지 확인해주세요."
                )
            if not inflow_df.empty:
                st.write(
                    f"🚶 일별 유입·매출 데이터 인식: {len(inflow_df)}일 "
                    f"({inflow_df['report_date'].min()} ~ {inflow_df['report_date'].max()})"
                )
            if not utm_map_df.empty:
                st.write(f"🗺️ UTM 소스/매체 매핑 인식: {len(utm_map_df)}행")
            if not ga_channel_df.empty:
                mapped_n = ga_channel_df["channel"].notna().sum()
                st.write(
                    f"🔎 GA 매체별 유입 경로 인식: {len(ga_channel_df)}행 "
                    f"({ga_channel_df['report_date'].min()} ~ {ga_channel_df['report_date'].max()}, "
                    f"소스/매체 {ga_channel_df['source_medium'].nunique()}종, 매체 매핑 {mapped_n}행)"
                )
                if mapped_n == 0:
                    st.caption("※ '매체'(채널 그룹핑) 컬럼이 아직 비어있습니다 — UTM 매핑표를 같이/먼저 올리면 채워집니다.")
                elif mapped_n < len(ga_channel_df):
                    # source_medium에 문자열이 아닌 값(숫자로 잘못 인식된 셀 등)이 섞여 있으면
                    # sorted()/join()이 타입 비교 에러로 죽을 수 있어, 비교 전에 전부 문자열로 맞춘다.
                    unmapped = sorted(
                        ga_channel_df.loc[ga_channel_df["channel"].isna(), "source_medium"]
                        .dropna().astype(str).unique()
                    )
                    st.caption(f"※ 매핑표에 없는 소스/매체 {len(unmapped)}종은 미매핑 상태입니다: {', '.join(unmapped[:10])}{' 외' if len(unmapped) > 10 else ''}")
            status2.update(label="분석 완료", state="complete")

        if st.sidebar.button("💾 GA 유입 데이터 저장하기", type="primary", key="ga_combined_save_btn"):
            saved_msgs = []
            if not inflow_df.empty:
                n_inflow = save_table("inflow_revenue_daily", inflow_df, "report_date", ga_file.name)
                saved_msgs.append(f"유입·매출 비교 {n_inflow}일")
            if not utm_map_df.empty:
                n_utm = save_table("utm_channel_map", utm_map_df, "source_medium", ga_file.name)
                saved_msgs.append(f"UTM 매핑 {n_utm}행")
            if not ga_channel_df.empty:
                n_ga_channel = save_table(
                    "ga_channel_inflow", ga_channel_df, "report_date,source_medium", ga_file.name
                )
                saved_msgs.append(f"GA 매체별 유입 경로 {n_ga_channel}행")
            st.cache_data.clear()
            st.sidebar.success("저장 완료! " + (" · ".join(saved_msgs) if saved_msgs else "저장할 데이터가 없습니다."))
            st.rerun()

    if budget_file is not None:
        with st.sidebar.status("파일 분석 중...", expanded=True) as status3:
            budget_xls = pd.ExcelFile(budget_file)
            budget_df = parse_channel_budget_sheet(budget_xls)
            sentinels = {BUDGET_SENTINEL_EXPECTED_REVENUE, BUDGET_SENTINEL_ACTUAL_REVENUE}
            if budget_df.empty:
                st.warning(
                    "인식 가능한 예산 데이터를 찾지 못했습니다. '자사몰 현황' 섹션, '26년' 블록, "
                    "'매체 세부내역'이 있는 파일인지 확인해주세요."
                )
            else:
                summary_bits = []
                for scope in budget_df["scope"].unique():
                    bsub = budget_df[budget_df["scope"] == scope]
                    for year in sorted(bsub["year"].unique(), reverse=True):
                        n_ch = bsub[(bsub["year"] == year) & (~bsub["channel"].isin(sentinels | {"TOTAL"}))]["channel"].nunique()
                        summary_bits.append(f"{scope}·{year}년 매체 {n_ch}개")
                st.write("💰 예산 인식: " + " · ".join(summary_bits))
                if (budget_df[~budget_df["channel"].isin(sentinels | {"TOTAL"})].empty):
                    st.warning(
                        "매체 세부내역(개별 매체별 예산)은 하나도 못 찾았습니다 — 매출 지표만 "
                        "인식됐습니다. 아래 진단 정보를 펼쳐서 캡쳐해서 보내주면 원인을 바로 알 수 "
                        "있습니다."
                    )
            with st.expander("🔍 진단 정보 보기 (매체가 0개거나 이상하면 이 내용을 캡쳐해서 보내주세요)"):
                st.text(_diagnose_budget_sheet(budget_xls))
            status3.update(label="분석 완료", state="complete")

        if st.sidebar.button("💾 예산 데이터 저장하기", type="primary", key="budget_save_btn"):
            if budget_df.empty:
                st.sidebar.warning("저장할 예산 데이터가 없습니다.")
            else:
                n_budget = save_table("channel_budget", budget_df, "scope,channel,year,month", budget_file.name)
                st.cache_data.clear()
                st.sidebar.success(f"저장 완료! 예산 {n_budget}행")
                st.rerun()

    if mix_file is not None:
        with st.sidebar.status("파일 분석 중...", expanded=True) as status4:
            mix_xls = pd.ExcelFile(mix_file)
            mix_df = parse_channel_mix_sheet(mix_xls, source_name=mix_file.name)
            if mix_df.empty:
                st.warning(
                    "인식 가능한 채널 믹스 데이터를 찾지 못했습니다. 'TOTAL' + 1~12월 헤더가 있는 "
                    "채널별 예산 표 파일인지 확인해주세요."
                )
            else:
                yr = mix_df["year"].iloc[0]
                st.write(f"📊 채널 믹스 인식: {yr}년 · 매체 {mix_df['channel'].nunique()}개")
            status4.update(label="분석 완료", state="complete")

        if st.sidebar.button("💾 채널 믹스 저장하기", type="primary", key="channel_mix_save_btn"):
            if mix_df.empty:
                st.sidebar.warning("저장할 채널 믹스 데이터가 없습니다.")
            else:
                n_mix = save_table("channel_mix_budget", mix_df, "channel,year,month", mix_file.name)
                st.cache_data.clear()
                st.sidebar.success(f"저장 완료! 채널 믹스 {n_mix}행")
                st.rerun()

    if mr_file is not None:
        st.sidebar.caption(
            "매체 관리자에서 받은 일별/소재별 성과 파일입니다. 날짜×매체로 합쳐 광고비·노출·클릭을 "
            "저장하고, 예산 일할값 대신 이 실집행액을 씁니다."
        )
        drop_cost = st.sidebar.checkbox(
            "광고비는 무시하고 노출·클릭만 가져오기", value=False, key="mr_nocost",
            help="맨즈탭처럼 정액 계약 매체는 리포트의 광고비가 '계약금액 ÷ 리포트 일수'로 계산돼 "
                 "부분 기간 파일에서는 부풀 수 있습니다. 광고비는 '정액 계약' 패널로 넣으세요.",
        )
        vat_in = st.sidebar.checkbox(
            "이 파일 금액이 이미 VAT 포함", value=True, key="mr_vat",
            help="체크 안 하면 VAT를 더해(×1.1) 다른 매체와 단위를 맞춥니다. "
                 "매체 관리자 화면 금액과 대조해서 정하세요.",
        )
        try:
            mr_df = parse_media_report_file(mr_file, vat_included=vat_in)
        except Exception as e:
            mr_df = pd.DataFrame()
            st.sidebar.warning(f"읽지 못했습니다: {e}")
        if not mr_df.empty and drop_cost:
            mr_df = mr_df.assign(cost_incl_vat=0.0)
        if not mr_df.empty:
            chans = sorted(mr_df["channel"].unique())
            st.sidebar.write(
                f"📅 {mr_df['report_date'].min()} ~ {mr_df['report_date'].max()} · "
                f"{len(mr_df)}행"
            )
            for ch in chans:
                sub = mr_df[mr_df["channel"] == ch]
                st.sidebar.write(
                    f"• **{ch}** — 광고비 {sub['cost_incl_vat'].sum():,.0f}원 · "
                    f"노출 {sub['impressions'].sum():,.0f} · 클릭 {sub['clicks'].sum():,.0f}"
                )
            # 캠페인명으로 매체를 잘못 짚었을 때 손으로 바꿀 수 있게 한다
            fix = st.sidebar.selectbox(
                "매체 지정 (자동 인식이 틀렸을 때만)",
                ["자동 인식 그대로"] + MANUAL_SPEND_CHANNELS, key="mr_fix")
            if fix != "자동 인식 그대로":
                mr_df = mr_df.assign(channel=fix).groupby(
                    ["report_date", "channel"], as_index=False).sum(numeric_only=True)

            if st.sidebar.button("💾 매체 리포트 저장하기", type="primary", key="mr_save_btn"):
                save_df = mr_df.copy()
                save_df["source"] = "manual"
                n = save_table("ad_spend_daily", save_df,
                               "report_date,channel,source", mr_file.name)
                st.cache_data.clear()
                st.sidebar.success(f"저장 완료! {n}행")
                st.rerun()

    st.sidebar.markdown("---")
    wk = load_table("weekly_overview")
    st.sidebar.metric("누적 주간 데이터", f"{len(wk):,} 주")
    if st.sidebar.button("🔄 새로고침 (캐시 비우기)"):
        st.cache_data.clear()
        st.rerun()


# ──────────────────────────────────────────────────────────────
# 예산 현황 (신규, 1차 버전) — '◆26년 월별 예산 정리' 파일 기반, 자사몰만
# ──────────────────────────────────────────────────────────────
def _budget_month_series(b: pd.DataFrame, year: int, channel: str) -> dict:
    """b(scope='자사몰'로 이미 필터된 budget df)에서 특정 연도·channel의 {월: 값} dict를 만든다
    (1~12월만, month=0/-1인 원본 TOTAL/당월누계 값은 제외). 값이 없는 월은 None."""
    vals = {m: None for m in range(1, 13)}
    sub = b[(b["year"] == year) & (b["channel"] == channel) & (b["month"] >= 1) & (b["month"] <= 12)]
    for _, r in sub.iterrows():
        vals[int(r["month"])] = float(r["budget_cost"])
    return vals


def _budget_raw_total_mtd(b: pd.DataFrame, year: int, channel: str) -> tuple:
    """파일에 원본 'TOTAL'/'당월 누계' 셀 값이 있으면(month=0/-1로 저장됨) 그대로 돌려준다.
    없으면 (None, None) — 이 경우 호출 쪽에서 월별 합계로 계산한다. 월별로 아직 안 나뉜 금액이
    TOTAL 셀에만 있는 항목(예: 촬영샘플처럼 월별은 0인데 연간 합계만 있는 경우)이 있어서,
    가능하면 원본 값을 우선 쓴다."""
    sub = b[(b["year"] == year) & (b["channel"] == channel) & (b["month"].isin([0, -1]))]
    total = mtd = None
    for _, r in sub.iterrows():
        if r["month"] == 0:
            total = float(r["budget_cost"])
        elif r["month"] == -1:
            mtd = float(r["budget_cost"])
    return total, mtd


def _budget_total_mtd(vals: dict, year: int, today, raw_total=None, raw_mtd=None) -> tuple:
    """연간 합계와 '당월 누계'를 계산한다. raw_total/raw_mtd(파일 원본 셀 값)가 주어지면 그걸
    우선 쓰고, 없을 때만 월별 값을 더해서 계산한다(오늘 기준, 해당 연도가 올해면 이번 달까지,
    과거 연도면 12월까지 값이 있는 월만)."""
    present = {m: v for m, v in vals.items() if v is not None}
    computed_total = sum(present.values()) if present else None
    cutoff = today.month if year == today.year else 12
    mtd_vals = [v for m, v in present.items() if m <= cutoff]
    computed_mtd = sum(mtd_vals) if mtd_vals else None
    total = raw_total if raw_total is not None else computed_total
    mtd = raw_mtd if raw_mtd is not None else computed_mtd
    return total, mtd


def _budget_ratio_series(num: dict, den: dict) -> dict:
    out = {}
    for m in range(1, 13):
        n, d = num.get(m), den.get(m)
        out[m] = (n / d * 100) if (n is not None and d not in (None, 0)) else None
    return out


def _budget_yoy_series(cur: dict, prev: dict) -> dict:
    out = {}
    for m in range(1, 13):
        c, p = cur.get(m), prev.get(m)
        out[m] = ((c - p) / p * 100) if (c is not None and p not in (None, 0)) else None
    return out


def _budget_fmt_money(v):
    return f"{v:,.0f}" if v is not None else "-"


def _budget_fmt_pct(v):
    return f"{v:.1f}%" if v is not None else "-"


def _budget_fmt_pct_change(v):
    if v is None:
        return "-"
    arrow = "▲" if v >= 0 else "▼"
    return f"{arrow}{abs(v):.1f}%"


BUDGET_TABLE_CSS = f"""
<style>
.stco-budget-wrap {{ overflow-x:auto; border:1px solid {THEME_COLORS["border"]}; border-radius:10px;
  background:{THEME_COLORS["canvas"]}; margin-bottom:12px; }}
.stco-budget-table {{ width:100%; border-collapse:collapse; font-size:13px;
  font-family:{THEME_FONT_STACK}; }}
.stco-budget-table th {{ background:{THEME_COLORS["surface"]}; color:{THEME_COLORS["muted"]};
  font-weight:600; padding:6px 10px; text-align:right; border-bottom:1px solid {THEME_COLORS["border"]};
  white-space:nowrap; }}
.stco-budget-table th:nth-child(-n+3) {{ text-align:center; }}
.stco-budget-table td {{ padding:6px 10px; text-align:right; color:{THEME_COLORS["foreground"]};
  border:1px solid {THEME_COLORS["border"]}; white-space:nowrap; }}
.stco-budget-table td:nth-child(-n+3) {{ text-align:center; color:{THEME_COLORS["body"]}; font-weight:600; }}
.stco-budget-table tr:hover td {{ background:{THEME_COLORS["surface"]}; }}
.stco-budget-strong {{ background:#fff8e1; font-weight:700; color:{THEME_COLORS["foreground"]} !important; }}
</style>
"""


def _render_budget_box_table(scope_label: str, b_scope: pd.DataFrame, years: list, today) -> dict:
    """scope_label(예: '자사몰')에 해당하는 예산 데이터를, 원본 엑셀 표처럼 연도/그룹이
    rowspan으로 병합된 박스형 HTML 표로 그린다. 실매출·광고비 행은 원본의 노란 강조와 비슷하게
    배경색을 넣는다. 진단용으로 연도별 {채널수} dict를 반환한다."""

    def series(year, channel):
        return _budget_month_series(b_scope, year, channel)

    def money_cells(vals, year, channel, strong=False):
        raw_total, raw_mtd = _budget_raw_total_mtd(b_scope, year, channel)
        total, mtd = _budget_total_mtd(vals, year, today, raw_total, raw_mtd)
        cls = ' class="stco-budget-strong"' if strong else ""
        cells = f"<td{cls}>{_budget_fmt_money(total)}</td><td{cls}>{_budget_fmt_money(mtd)}</td>"
        for m in range(1, 13):
            cells += f"<td{cls}>{_budget_fmt_money(vals.get(m))}</td>"
        return cells

    def pct_cells(vals, fmt_fn=_budget_fmt_pct):
        cells = "<td>-</td><td>-</td>"
        for m in range(1, 13):
            text = fmt_fn(vals.get(m))
            style = ""
            if text.startswith("▲"):
                style = ' style="color:#d93025;"'
            elif text.startswith("▼"):
                style = ' style="color:#1a73e8;"'
            cells += f"<td{style}>{text}</td>"
        return cells

    month_headers = "".join(f"<th>{m}월</th>" for m in range(1, 13))
    html = [
        '<div class="stco-budget-wrap"><table class="stco-budget-table"><thead><tr>'
        f"<th>연도</th><th>구분</th><th>항목</th><th>합계(연간)</th><th>당월 누계</th>{month_headers}"
        "</tr></thead><tbody>"
    ]

    revenue_by_year, cost_by_year, channel_counts = {}, {}, {}
    for idx, year in enumerate(years):
        rev_e = series(year, BUDGET_SENTINEL_EXPECTED_REVENUE)
        rev_a = series(year, BUDGET_SENTINEL_ACTUAL_REVENUE)
        cost = series(year, "TOTAL")
        revenue_by_year[year] = rev_a
        cost_by_year[year] = cost
        achieve = _budget_ratio_series(rev_a, rev_e)
        cost_ratio = _budget_ratio_series(cost, rev_a)

        yoy_rows = []
        if idx + 1 < len(years):
            prev_year = years[idx + 1]
            prev_rev = revenue_by_year.get(prev_year) or series(prev_year, BUDGET_SENTINEL_ACTUAL_REVENUE)
            prev_cost = cost_by_year.get(prev_year) or series(prev_year, "TOTAL")
            yoy_rows = [
                ("전년 대비 매출 증감율", _budget_yoy_series(rev_a, prev_rev)),
                ("전년 대비 광고비 증감율", _budget_yoy_series(cost, prev_cost)),
            ]

        chs = b_scope[
            (b_scope["year"] == year)
            & (~b_scope["channel"].isin(["TOTAL", BUDGET_SENTINEL_EXPECTED_REVENUE, BUDGET_SENTINEL_ACTUAL_REVENUE]))
        ]["channel"].unique().tolist()
        chs_sorted = sorted(chs, key=lambda c: -(sum(v for v in series(year, c).values() if v is not None) or 0))
        channel_counts[year] = len(chs_sorted)

        sales_rowspan = 4 + len(yoy_rows)
        detail_rowspan = max(len(chs_sorted), 1)
        year_rowspan = sales_rowspan + detail_rowspan

        html.append(
            f'<tr><td rowspan="{year_rowspan}">{year}년</td>'
            f'<td rowspan="{sales_rowspan}">매출/광고비</td>'
            f"<td>실매출</td>{money_cells(rev_a, year, BUDGET_SENTINEL_ACTUAL_REVENUE, strong=True)}</tr>"
        )
        html.append(f"<tr><td>매출 달성률</td>{pct_cells(achieve)}</tr>")
        html.append(f"<tr><td>광고비</td>{money_cells(cost, year, 'TOTAL', strong=True)}</tr>")
        html.append(f"<tr><td>광고비 비율</td>{pct_cells(cost_ratio)}</tr>")
        for label, yoy_vals in yoy_rows:
            html.append(f"<tr><td>{label}</td>{pct_cells(yoy_vals, _budget_fmt_pct_change)}</tr>")

        if chs_sorted:
            html.append(
                f'<tr><td rowspan="{detail_rowspan}">매체 세부내역</td>'
                f"<td>{chs_sorted[0]}</td>{money_cells(series(year, chs_sorted[0]), year, chs_sorted[0])}</tr>"
            )
            for ch in chs_sorted[1:]:
                html.append(f"<tr><td>{ch}</td>{money_cells(series(year, ch), year, ch)}</tr>")
        else:
            html.append(
                f'<tr><td rowspan="{detail_rowspan}">매체 세부내역</td>'
                f'<td colspan="14">이 연도는 매체 세부내역을 못 찾았습니다</td></tr>'
            )

    html.append("</tbody></table></div>")
    st.markdown(BUDGET_TABLE_CSS + "".join(html), unsafe_allow_html=True)
    return channel_counts


def render_budget_page(monthly: pd.DataFrame, budget: pd.DataFrame):
    st.subheader("💰 예산 현황")
    st.caption(
        "'◆26년 월별 예산 정리' 파일의 '온라인사업팀 현황'·'자사몰 현황' 섹션을 원본 엑셀 표와 "
        "같은 형태(연도/구분/항목 병합, 26년·25년 비교, 매체 세부내역)로 보여줍니다. 원본 표시는 "
        "천원 단위지만 여기 숫자는 원 단위입니다. 광고비 합계는 '매체 세부내역' 채널별 합으로 "
        "계산합니다(요약 행 라벨이 파일 버전마다 달라서, 라벨 대신 실제 매체별 값을 더한 값입니다) "
        "— 원본 표의 합계와 다르면 알려주세요."
    )
    if budget is None or budget.empty:
        st.info("아직 예산 데이터가 없습니다. 왼쪽 사이드바 '③ 연간 예산 파일 업로드'에서 파일을 올려주세요.")
        return

    today = datetime.now()
    scopes_present = [s for s in BUDGET_SCOPE_TITLES if s in budget["scope"].unique()]
    if not scopes_present:
        st.info("예산 데이터를 찾지 못했습니다. 왼쪽 사이드바 업로드 화면의 진단 정보를 확인해주세요.")
        return

    for scope in scopes_present:
        b = budget[budget["scope"] == scope].copy()
        if b.empty:
            continue
        years = sorted(b["year"].unique(), reverse=True)
        st.markdown(f"##### {BUDGET_SCOPE_TITLES[scope]}")
        channel_counts = _render_budget_box_table(scope, b, years, today)
        diag = " · ".join(f"{y}년 매체 {n}개" for y, n in channel_counts.items())
        st.caption(f"매체 세부내역 인식: {diag}")
        st.markdown("---")

    st.caption(
        "'매체 세부내역' 매체명은 원본 파일 표기 그대로이며, 다른 페이지(매체별 성과 등)의 리포트 "
        "매체명과 아직 자동으로 매칭되지 않았습니다."
    )


# ──────────────────────────────────────────────────────────────
# 채널믹스 목표 대비 (신규) — "매체별 성과" 페이지 안에서 호출
# ──────────────────────────────────────────────────────────────
def render_channel_mix(fc: pd.DataFrame):
    st.markdown("---")
    st.markdown("### 채널믹스 목표 대비 (발굴 메타50·구글20·네이버20 / 회수 모비온10)")
    st.caption(
        "목표 비중은 「STCO 퍼포먼스마케팅 인수인계서」 4절 기준. "
        "'기타'는 목표 비중 배정이 없는 매체(크리테오 등)입니다."
    )
    if fc is None or fc.empty:
        st.info("데이터가 아직 없습니다.")
        return

    grp = fc.copy()
    grp["채널그룹"] = grp["channel"].apply(map_channel_group)
    by_group = grp.groupby("채널그룹", as_index=False)["cost_incl_vat"].sum()
    total_cost = by_group["cost_incl_vat"].sum()
    if total_cost <= 0:
        st.info("광고비 데이터가 없습니다.")
        return

    all_groups = list(CHANNEL_MIX_TARGET.keys())
    if (by_group["채널그룹"] == "기타").any():
        all_groups = all_groups + ["기타"]
    by_group = (
        by_group.set_index("채널그룹").reindex(all_groups, fill_value=0).reset_index()
    )

    by_group["실제비중(%)"] = by_group["cost_incl_vat"] / total_cost * 100
    by_group["목표비중(%)"] = by_group["채널그룹"].map(lambda g: CHANNEL_MIX_TARGET.get(g, 0) * 100)
    by_group["차이(%p)"] = by_group["실제비중(%)"] - by_group["목표비중(%)"]

    chart_df = by_group.melt(
        id_vars="채널그룹", value_vars=["목표비중(%)", "실제비중(%)"], var_name="구분", value_name="비중(%)"
    )
    fig = px.bar(
        chart_df, x="채널그룹", y="비중(%)", color="구분", barmode="group", text_auto=".1f",
        title="채널그룹별 목표 vs 실제 예산 비중",
    )
    st.plotly_chart(theme_chart(fig), use_container_width=True)

    show = by_group.copy()
    show["광고비(원)"] = show["cost_incl_vat"].map(lambda v: f"{v:,.0f}")
    show["실제비중(%)"] = show["실제비중(%)"].map(lambda v: f"{v:.1f}%")
    show["목표비중(%)"] = show["목표비중(%)"].map(lambda v: f"{v:.1f}%")
    show["차이(%p)"] = show["차이(%p)"].map(lambda v: f"{v:+.1f}%p")
    render_html_table(show[["채널그룹", "광고비(원)", "목표비중(%)", "실제비중(%)", "차이(%p)"]])

    over = by_group[(by_group["채널그룹"] != "기타") & (by_group["차이(%p)"] > 5)]
    under = by_group[(by_group["채널그룹"] != "기타") & (by_group["차이(%p)"] < -5)]
    if len(over) or len(under):
        msgs = []
        for _, r in over.iterrows():
            msgs.append(f"- **{r['채널그룹']}** 목표 대비 +{r['차이(%p)']:.1f}%p 초과 집행 중 — 단계적 축소(10~15% 내) 검토")
        for _, r in under.iterrows():
            msgs.append(f"- **{r['채널그룹']}** 목표 대비 {r['차이(%p)']:.1f}%p 미달 — 단계적 증액 검토")
        st.markdown("\n".join(msgs))


# ──────────────────────────────────────────────────────────────
# 소재별 성과 (신규 페이지)
# ──────────────────────────────────────────────────────────────
# 현재 실제로 운영 중인 매체 탭만 고정 순서로 노출 (TOTAL이 맨 왼쪽)
CREATIVE_TABS = ["TOTAL", "네이버 GFA PC", "네이버 GFA MO", "메타", "구글(P-MAX)", "크리테오"]


def _render_creative_table(fc: pd.DataFrame, channel_name: str = None):
    """선택된 탭(매체)의 필터링된 소재 데이터로 집계 테이블 + 판정 + 다운로드 버튼을 렌더링.
    channel_name이 주어지면(=TOTAL 탭이 아니면) 표 위에 그 매체의 우수/부진 소재를 짚어주는
    코멘트를 한 줄 붙인다 — TOTAL 탭엔 매체가 섞여 있어 우수/부진 비교가 의미가 없어 생략한다."""
    if fc.empty:
        st.info("선택한 기간/매체에 데이터가 없습니다.")
        return

    has_image = "image_url" in fc.columns

    # 주간 리포트가 '해당 월 1일~업로드일까지 누적' 형태라, 같은 달 안에 여러 번(매주) 업로드하면
    # 소재마다 as_of_date가 다른 여러 스냅샷이 쌓인다. 이걸 그대로 sum하면 매주 누적치를 또 더해서
    # 몇 배로 부풀려지므로, 기간 내 '가장 최근 스냅샷' 한 건만 남기고(이미 그 시점까지의 누적값이라
    # 합산이 아니라 대체) 집계한다.
    fc_latest = fc.sort_values("as_of_date").drop_duplicates(subset=["channel", "creative"], keep="last")

    agg_kwargs = dict(
        impressions=("impressions", "sum"), clicks=("clicks", "sum"),
        cost_excl_vat=("cost_excl_vat", "sum"), cost_incl_vat=("cost_incl_vat", "sum"),
        conversions=("conversions", "sum"), revenue=("revenue", "sum"),
    )
    if has_image:
        agg_kwargs["image_url"] = ("image_url", "first")
    agg = fc_latest.groupby(["channel", "creative"], as_index=False).agg(**agg_kwargs)
    agg = add_kpis(agg)

    MIN_SPEND = 50000  # 표본 기준: 광고비 5만원 미만은 판단 보류 (performance-marketing-analysis 스킬 기본값)
    total_cost, total_rev = agg["cost_incl_vat"].sum(), agg["revenue"].sum()
    account_avg_roas = (total_rev / total_cost * 100) if total_cost else 0

    def judge(row):
        if row["cost_incl_vat"] < MIN_SPEND:
            return "판단 보류(표본 부족)"
        if account_avg_roas <= 0:
            return "판단 보류"
        ratio = row["roas"] / account_avg_roas
        if ratio >= 1.2:
            return "우수"
        if ratio <= 0.7:
            return "부진"
        return "평균 수준"

    agg["판정"] = agg.apply(judge, axis=1)
    agg = agg.sort_values("cost_incl_vat", ascending=False)

    st.caption(f"계정 평균 ROAS(선택 기간): {account_avg_roas:,.0f}% · 광고비 {MIN_SPEND:,}원 미만은 표본 부족으로 판단 보류 처리")

    if channel_name:
        judged = agg[agg["판정"] != "판단 보류(표본 부족)"]
        n_good = int((judged["판정"] == "우수").sum())
        n_mid = int((judged["판정"] == "평균 수준").sum())
        n_bad = int((judged["판정"] == "부진").sum())
        n_hold = len(agg) - len(judged)
        if len(judged) >= 2:
            best = judged.loc[judged["roas"].idxmax()]
            worst = judged.loc[judged["roas"].idxmin()]
            lines = [
                f"{channel_name} 소재 {len(judged)}개 중 우수 {n_good}개 · 평균 수준 {n_mid}개 · 부진 {n_bad}개입니다"
                + (f" (표본 부족 {n_hold}개는 판단 보류)." if n_hold else "."),
                f"<b>{best['creative']}</b>가 광고비 {best['cost_incl_vat']:,.0f}원으로 ROAS {best['roas']:,.0f}%를 기록해 가장 우수했고, "
                f"<b>{worst['creative']}</b>는 광고비 {worst['cost_incl_vat']:,.0f}원 대비 ROAS {worst['roas']:,.0f}%로 "
                f"계정 평균({account_avg_roas:,.0f}%) 대비 크게 낮아 가장 부진했습니다.",
            ]
            if n_bad:
                lines.append(_ops_next_action(f"부진 소재({n_bad}개)는 소재 교체 또는 예산 축소를 검토하는 것을 권장합니다."))
            st.markdown("<br>".join(lines), unsafe_allow_html=True)
        elif len(judged) == 1:
            st.markdown(
                f"{channel_name}은 판단 가능한(표본 충분) 소재가 <b>{judged.iloc[0]['creative']}</b> 1개뿐이라, "
                "우수/부진 비교는 소재가 더 쌓이면 확인하겠습니다.",
                unsafe_allow_html=True,
            )
        else:
            st.caption(f"{channel_name}은 아직 판단 가능한(표본 충분) 소재가 없습니다.")

    display_cols = ["channel", "creative"]
    if has_image:
        # render_html_table은 셀 값을 그대로 <td>에 넣으므로 <img> 태그 문자열이 실제 썸네일로 렌더링된다.
        agg["creative_image"] = agg["image_url"].map(
            lambda u: (
                f'<img src="{u}" style="height:140px;width:140px;border-radius:8px;object-fit:cover;">'
                if isinstance(u, str) and u else ""
            )
        )
        display_cols.append("creative_image")
    display_cols += ["impressions", "clicks", "ctr", "cpc", "cost_incl_vat",
                      "conversions", "cvr", "cpa", "revenue", "roas", "판정"]
    display_cols = [c for c in display_cols if c in agg.columns]

    show = format_display(agg[display_cols])

    # TOTAL 행: 선택된 탭/기간 전체 합계를 표 맨 위(헤더 바로 아래)에 음영 처리해서 보여준다.
    # render_html_table은 첫 컬럼이 정확히 "TOTAL"이면 자동으로 강조 스타일(음영/굵게)을 입힌다.
    total_base = pd.DataFrame([{
        "impressions": agg["impressions"].sum(), "clicks": agg["clicks"].sum(),
        "cost_excl_vat": agg["cost_excl_vat"].sum(), "cost_incl_vat": agg["cost_incl_vat"].sum(),
        "conversions": agg["conversions"].sum(), "revenue": agg["revenue"].sum(),
    }])
    total_kpis = add_kpis(total_base).iloc[0]
    total_row = {c: "" for c in display_cols}
    total_row["channel"] = "TOTAL"
    for c in ("impressions", "clicks", "ctr", "cpc", "cost_incl_vat", "conversions", "cvr", "cpa", "revenue", "roas"):
        if c in display_cols:
            total_row[c] = total_kpis[c]
    total_show = format_display(pd.DataFrame([total_row])[display_cols])
    show = pd.concat([total_show, show], ignore_index=True)

    render_sortable_creative_table(korify(show), agg)

    # 엑셀 다운로드에는 이미지 썸네일(HTML) 대신 원본 URL을 남긴다.
    dl_cols = [c for c in display_cols if c != "creative_image"]
    if has_image and "image_url" in agg.columns and "image_url" not in dl_cols:
        dl_cols.insert(dl_cols.index("creative") + 1, "image_url")
    st.download_button(
        "⬇️ 엑셀 다운로드 (소재별 성과)",
        data=to_excel_bytes(korify(format_display(agg[dl_cols]))),
        file_name="creative_performance.xlsx",
        key=f"dl_creative_{fc['channel'].iloc[0] if fc['channel'].nunique() == 1 else 'total'}",
    )


def render_creative_performance(creatives: pd.DataFrame):
    st.subheader("🎨 소재별 성과")
    st.caption("현재 운영 중인 매체(네이버 GFA PC/MO, 메타, 구글 P-MAX, 크리테오)의 소재 단위 데이터만 표시됩니다.")
    if creatives is None or creatives.empty:
        st.info(
            "소재별 데이터가 아직 없습니다. 엑셀에 '○○_소재성과요약' 시트 또는 "
            "'광고소재' 컬럼이 있는 '_data' 시트가 있는지 확인해주세요."
        )
        return

    creatives = creatives.copy()
    creatives["as_of_date"] = pd.to_datetime(creatives["as_of_date"]).dt.date
    min_d, max_d = creatives["as_of_date"].min(), creatives["as_of_date"].max()
    start, end = period_filter(min_d, max_d, key="creative")
    fc = creatives[(creatives["as_of_date"] >= start) & (creatives["as_of_date"] <= end)]
    # 예전 업로드분에 남아있을 수 있는 미운영 매체(당근마켓 등) 잔여 행은 TOTAL에서도 제외
    fc = fc[fc["channel"].isin(CREATIVE_TABS[1:])]

    tabs = st.tabs(CREATIVE_TABS)
    for tab_widget, tab_name in zip(tabs, CREATIVE_TABS):
        with tab_widget:
            tab_fc = fc if tab_name == "TOTAL" else fc[fc["channel"] == tab_name]
            _render_creative_table(tab_fc, channel_name=None if tab_name == "TOTAL" else tab_name)


# ──────────────────────────────────────────────────────────────
# 사이드바 그룹 네비게이션 (신규 — st.tabs() 대체)
# ──────────────────────────────────────────────────────────────
NAV_GROUPS = {
    "GA 유입 리포트": ["채널 퍼널 리포트", "채널 성과", "GA 소재별 성과", "예산 관리",
                    "GA 매체별 유입 경로", "GA4 라이브 리포트", "유입·매출 비교"],
    "성과 리포트": ["종합 대시보드", "매체별 성과", "타겟팅별 성과", "소재별 성과", "예산 현황"],
    "운영 코멘트": ["운영 코멘트"],
    "운영 도구": ["UTM 빌더", "소재 로그"],
    "가이드": ["가이드"],
}

# 그룹별로 다른 CSS 클래스(st-key-navgrp_xxx)를 붙이기 위한 영문 키. GA 유입 리포트/운영 코멘트는
# 별도 아이콘을 안 받아서 키를 안 주고(None) 공통 기본 아이콘(NAV_GROUP_ICON_B64)을 그대로 쓴다.
NAV_GROUP_KEYS = {
    "성과 리포트": "report",
    "운영 코멘트": None,
    "GA 유입 리포트": None,
    "운영 도구": "ops",
    "가이드": "guide",
}

# 아직 실제 데이터/로직이 없는 페이지들 — main()의 페이지 분기에서 이 목록에 있으면
# render_coming_soon()으로 "준비 중" 안내만 보여준다. 나중에 진짜 렌더 함수가 생기면
# main()에 elif 분기를 추가하고 여기서 이름을 지워주면 된다.
NAV_PAGES_COMING_SOON = {"UTM 빌더", "소재 로그", "가이드"}


def render_coming_soon(page_name: str):
    st.subheader(page_name)
    st.info(f"'{page_name}' 페이지는 아직 준비 중입니다. 데이터 연동이 되면 채워질 예정이에요.")


def render_nav() -> str:
    st.sidebar.markdown("---")
    # 매일 처음 보는 화면이 '채널 퍼널 리포트'라서 기본 페이지를 여기로 둔다.
    default_page = NAV_GROUPS["GA 유입 리포트"][0]
    if "nav_page" not in st.session_state:
        st.session_state["nav_page"] = default_page
    # key="stco_nav"로 감싸면 Streamlit이 이 블록에 "st-key-stco_nav" 클래스를 붙여주는데,
    # inject_theme()의 CSS가 그 클래스 안의 expander/button만 골라 '텍스트 링크' 스타일로 바꾼다
    # (사이드바 위쪽의 업로드/저장 버튼 등 다른 버튼들은 기존 스타일 그대로 유지됨).
    # container(key=...)는 비교적 최신 Streamlit에서만 지원하므로, 혹시 배포 환경 버전이 낮아
    # TypeError가 나더라도 내비게이션 자체는 동작하도록 안전하게 폴백한다(이 경우 스타일만 예전
    # 버튼 박스 모양으로 보임).
    try:
        nav_box = st.sidebar.container(key="stco_nav")
    except TypeError:
        nav_box = st.sidebar.container()
    with nav_box:
        for group, pages in NAV_GROUPS.items():
            # 아이콘은 텍스트가 아니라 CSS(summary p::before)로 붙인다 — NAV_GROUP_ICON_B64 참고.
            # 그룹이 여러 개가 되면 전부 펼쳐두면 사이드바가 너무 길어지니, 현재 선택된 페이지가
            # 속한 그룹만 자동으로 펼치고 나머지는 접어둔다.
            is_active_group = st.session_state["nav_page"] in pages
            # 그룹별로 다른 아이콘을 쓰려면 CSS에서 그룹을 구분할 수 있어야 하는데, Streamlit은
            # 위젯마다 개별 element-container로 감싸버려서 :nth-of-type으로는 구분이 안 된다
            # (모든 expander가 "자기 부모 안에서는 1번째"라 전부 같은 규칙에 걸림). 그래서 그룹마다
            # container(key=...)로 한 번 더 감싸 고유 CSS 클래스(st-key-{key})를 붙여준다.
            group_key = NAV_GROUP_KEYS.get(group)
            try:
                group_box = st.container(key=f"navgrp_{group_key}") if group_key else st.container()
            except TypeError:
                group_box = st.container()
            with group_box, st.expander(group, expanded=is_active_group):
                for p in pages:
                    is_current = st.session_state["nav_page"] == p
                    if st.button(
                        p, key=f"nav_{p}", use_container_width=True,
                        type="primary" if is_current else "secondary",
                    ):
                        st.session_state["nav_page"] = p
                        st.rerun()
    return st.session_state["nav_page"]


# ──────────────────────────────────────────────────────────────
# 페이지별 렌더 함수 (예전 tab1~tab4의 내용을 그대로 옮김, 로직 변경 없음)
# ──────────────────────────────────────────────────────────────
def render_overview_page(
    weekly: pd.DataFrame, monthly: pd.DataFrame, daily: pd.DataFrame,
    channels: pd.DataFrame = None, channels_weekly: pd.DataFrame = None,
    ga_channel_inflow: pd.DataFrame = None,
):
    weekly = _drop_trailing_zero_weeks(weekly)
    if weekly is None:
        weekly = pd.DataFrame()
    if not weekly.empty:
        st.subheader("🔎 기간 필터 (주간 기준)")
        min_d, max_d = weekly["week_start"].min(), weekly["week_end"].max()
        start, end = period_filter(min_d, max_d, key="weekly", options=DATE_PERIOD_OPTIONS_FULL)
        fw = weekly[(weekly["week_start"] >= start) & (weekly["week_start"] <= end)]
        fw = add_kpis(fw).sort_values("week_start")

        kpi_cards(fw)
        st.markdown("### 주간 추이")
        if fw.empty:
            st.caption("선택한 기간(주간 기준)에 해당하는 데이터가 없습니다.")
        else:
            week_label_order = kor_date_labels(fw["week_start"], "day")
            fw = fw.assign(주=week_label_order)
            c1, c2 = st.columns(2)
            with c1:
                chart_df = fw.rename(columns={"cost_incl_vat": "광고비(VAT포함)", "revenue": "매출"})
                fig = px.bar(
                    chart_df, x="주", y=["광고비(VAT포함)", "매출"], barmode="group",
                    title="주간 비용(VAT포함) vs 매출",
                    labels={"value": "금액(원)", "variable": "구분"},
                    category_orders={"주": week_label_order},
                )
                fig.update_yaxes(tickformat=",.0f")
                fig.for_each_trace(
                    lambda t: t.update(
                        hovertemplate=f"구분={t.name}<br>주 시작일=%{{x}}<br>금액(원)=%{{y:,.0f}}원<extra></extra>"
                    )
                )
                st.plotly_chart(theme_chart(fig), use_container_width=True)
            with c2:
                fig2 = px.line(
                    fw, x="주", y="roas", markers=True, title="주간 ROAS 추이 (%)",
                    labels={"roas": "ROAS(%)"},
                    category_orders={"주": week_label_order},
                )
                fig2.update_yaxes(tickformat=",.0f", ticksuffix="%")
                fig2.update_traces(hovertemplate="주=%{x}<br>ROAS=%{y:,.0f}%<extra></extra>")
                st.plotly_chart(theme_chart(fig2), use_container_width=True)

        # 플레이스홀더로 미리 만들어둔 빈 월(광고비 0원)까지 그리면 옛날~올해가 다 이어진 밋밋한
        # 0선이 껴서 정작 보고 싶은 최근 구간이 눌려 보인다 — 실제로 집행된(광고비>0) 달만 그린다.
        monthly_real = (
            monthly[monthly["cost_incl_vat"] > 0]
            if not monthly.empty and "cost_incl_vat" in monthly.columns
            else monthly
        )
        if not monthly_real.empty:
            st.markdown("---")
            st.markdown("### 월별 GA-ROAS vs 플랫폼 ROAS")
            fm_chart = add_kpis(monthly_real).sort_values("report_month").rename(
                columns={"roas": "플랫폼 ROAS", "ga_roas": "GA ROAS"}
            )
            month_label_order = kor_date_labels(fm_chart["report_month"], "month")
            fm_chart = fm_chart.assign(월=month_label_order)
            fig3 = px.line(
                fm_chart, x="월", y=["플랫폼 ROAS", "GA ROAS"], markers=True,
                labels={"value": "ROAS(%)", "variable": "기준"},
                title="플랫폼 리포팅 ROAS vs GA 기준 ROAS",
                category_orders={"월": month_label_order},
            )
            fig3.update_yaxes(tickformat=",.0f", ticksuffix="%")
            fig3.for_each_trace(
                lambda t: t.update(hovertemplate=f"기준={t.name}<br>월=%{{x}}<br>ROAS=%{{y:,.0f}}%<extra></extra>")
            )
            st.plotly_chart(theme_chart(fig3), use_container_width=True)
            st.caption("* 광고비가 집행된 달만 표시합니다 (미집행 빈 달 제외). GA-매출/GA-ROAS는 쇼핑검색 및 GFA 외부몰 데이터가 미집계될 수 있습니다 (원본 시트 주석 기준).")

        st.markdown("---")
        st.markdown("## 📚 누적 데이터")
        st.caption("기본은 최근 데이터만 보여주고, '이전 데이터 더 보기'를 켜면 10/30/50/100/200개 단위로 넘겨볼 수 있어요.")

        month_show_cols = ["report_month", "impressions", "clicks", "ctr", "cpc", "cost_excl_vat", "cost_incl_vat",
                            "signups", "cpa", "conversions", "cvr", "revenue", "roas", "aov",
                            "ga_conversions", "ga_revenue", "ga_roas"]
        month_numeric_cols = [c for c in month_show_cols if c != "report_month"]
        render_cumulative_table(
            add_kpis(monthly) if not monthly.empty else monthly,
            date_col="report_month", show_cols=month_show_cols, numeric_cols=month_numeric_cols,
            title="1) 월별 누적", key="monthly_cum", mode="month",
        )
        render_ops_comment_monthly(monthly, channels=channels)

        wk = weekly.copy()
        if not wk.empty:
            wk["week_no"] = wk["label"].astype(str).str.replace(r"\s*\(.*\)\s*$", "", regex=True).str.strip()
            wk["week_range"] = wk.apply(lambda r: f"{r['week_start']:%Y-%m-%d}~{r['week_end']:%Y-%m-%d}", axis=1)
            # 리포트 원본 주간 섹션엔 GA 컬럼이 없어서(월별 섹션에만 있음), GA 매체별 유입 경로
            # (UTM 매핑 완료분)를 주 단위로 직접 합산해 붙인다 — 업로드가 안 됐으면 조용히 생략.
            ga_wk = _ga_weekly_agg(ga_channel_inflow, wk)
            if not ga_wk.empty:
                wk = wk.merge(ga_wk, on="week_start", how="left")
                wk["ga_conversions"] = wk["ga_conversions"].fillna(0)
                wk["ga_revenue"] = wk["ga_revenue"].fillna(0)
                wk["ga_roas"] = np.where(wk["cost_incl_vat"] > 0, wk["ga_revenue"] / wk["cost_incl_vat"] * 100, 0)
        week_show_cols = ["week_range", "week_no", "impressions", "clicks", "ctr", "cpc",
                           "cost_excl_vat", "cost_incl_vat", "signups", "cpa", "conversions", "cvr", "revenue", "roas", "aov"]
        if not wk.empty and "ga_conversions" in wk.columns:
            week_show_cols += ["ga_conversions", "ga_revenue", "ga_roas"]
        week_numeric_cols = [c for c in week_show_cols if c not in ("week_no", "week_range")]
        render_cumulative_table(
            add_kpis(wk) if not wk.empty else wk,
            date_col="week_start", show_cols=week_show_cols, numeric_cols=week_numeric_cols,
            title="2) 주간별 누적", key="weekly_cum", mode="week",
        )
        if ga_channel_inflow is None or ga_channel_inflow.empty:
            st.caption("※ 'GA 매체별 유입 경로' 데이터를 업로드하면 GA-전환수/GA-매출/GA-ROAS 컬럼이 여기에도 채워집니다.")
        render_ops_comment_weekly(weekly, channels_weekly=channels_weekly)

        dy = daily.copy()
        if not dy.empty:
            ga_dy = _ga_daily_agg(ga_channel_inflow)
            if not ga_dy.empty:
                dy = dy.merge(ga_dy, on="report_date", how="left")
                dy["ga_conversions"] = dy["ga_conversions"].fillna(0)
                dy["ga_revenue"] = dy["ga_revenue"].fillna(0)
                dy["ga_roas"] = np.where(dy["cost_incl_vat"] > 0, dy["ga_revenue"] / dy["cost_incl_vat"] * 100, 0)
        day_show_cols = ["report_date", "impressions", "clicks", "ctr", "cpc", "cost_excl_vat", "cost_incl_vat",
                          "signups", "cpa", "conversions", "cvr", "revenue", "roas", "aov"]
        if not dy.empty and "ga_conversions" in dy.columns:
            day_show_cols += ["ga_conversions", "ga_revenue", "ga_roas"]
        day_numeric_cols = [c for c in day_show_cols if c != "report_date"]
        render_cumulative_table(
            add_kpis(dy) if not dy.empty else dy,
            date_col="report_date", show_cols=day_show_cols, numeric_cols=day_numeric_cols,
            title="3) 일자별 누적", key="daily_cum", mode="day",
        )
    else:
        st.info("주간 데이터가 아직 없습니다.")


def render_channel_page(channels: pd.DataFrame, snapshot: pd.DataFrame, ga_channel_inflow: pd.DataFrame = None):
    if not channels.empty:
        channels["report_month"] = pd.to_datetime(channels["report_month"]).dt.date
        st.subheader("🔎 기간 필터 (월별 기준)")
        min_m = channels["report_month"].min()
        max_m = (pd.Timestamp(channels["report_month"].max()) + pd.offsets.MonthEnd(0)).date()
        mstart, mend = period_filter(min_m, max_m, key="channel")
        fc = channels[(channels["report_month"] >= mstart) & (channels["report_month"] <= mend)]

        by_channel = (
            fc.groupby("channel", as_index=False)
            .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
                 cost_excl_vat=("cost_excl_vat", "sum"), cost_incl_vat=("cost_incl_vat", "sum"),
                 conversions=("conversions", "sum"), revenue=("revenue", "sum"))
        )
        # 리포트 원본 매체별 GA 컬럼은 '당월 스냅샷' 기준이라 아래 별도 표로만 보여주고, 여기
        # 메인 표에는 UTM 매핑된 GA 매체별 유입 경로 데이터를 선택 기간(mstart~mend) 그대로
        # 합산해서 붙인다 — 기간을 바꾸면 이 GA 컬럼도 같이 갱신된다.
        ga_by_channel = _ga_channel_agg(ga_channel_inflow, mstart, mend)
        if not ga_by_channel.empty:
            by_channel = by_channel.merge(ga_by_channel, on="channel", how="left")
            by_channel["ga_conversions"] = by_channel["ga_conversions"].fillna(0)
            by_channel["ga_revenue"] = by_channel["ga_revenue"].fillna(0)
        by_channel = add_kpis(by_channel).sort_values("cost_incl_vat", ascending=False)
        if "ga_revenue" in by_channel.columns:
            by_channel["ga_roas"] = np.where(
                by_channel["cost_incl_vat"] > 0, by_channel["ga_revenue"] / by_channel["cost_incl_vat"] * 100, 0
            )

        by_channel_chart = by_channel.copy()
        by_channel_chart["roas_label"] = by_channel_chart["roas"].map(lambda v: f"{v:.0f}%")
        fig = px.bar(
            by_channel_chart, x="channel", y="roas", title="매체별 ROAS (%, 선택 기간 합산)", text="roas_label",
            labels={"channel": "매체", "roas": "ROAS(%)"},
        )
        fig.update_yaxes(tickformat=",.0f", ticksuffix="%")
        fig.update_traces(hovertemplate="매체=%{x}<br>ROAS=%{y:,.0f}%<extra></extra>")
        st.plotly_chart(theme_chart(fig), use_container_width=True)

        # 광고비(VAT제외)는 CPA 계산에만 쓰고 화면/엑셀에는 광고비(VAT+)만 노출 — 컬럼이 많아
        # 표 폭이 들쭉날쭉해지는 것도 줄어든다. 컬럼 순서는 종합 대시보드/타겟팅별 성과와 동일하게
        # 노출수 → 클릭수 → CTR → CPC → 광고비 순으로 맞춘다.
        BC_COL_ORDER = ["channel", "impressions", "clicks", "ctr", "cpc", "cost_incl_vat",
                        "cpa", "conversions", "cvr", "revenue", "roas", "aov",
                        "ga_conversions", "ga_revenue", "ga_roas"]
        bc_cols = [c for c in BC_COL_ORDER if c in by_channel.columns]
        bc_cols += [c for c in by_channel.columns if c not in bc_cols and c != "cost_excl_vat"]
        bc_table = format_display(by_channel[bc_cols])
        if "ga_revenue" not in by_channel.columns:
            st.caption("※ 'GA 매체별 유입 경로' 데이터를 업로드하면(UTM 매핑 필요) 매체별 GA-전환수/GA-매출/GA-ROAS도 여기 표시됩니다.")
        # CPA는 광고비(VAT제외) 기준으로 계산되는데, bc_cols에서 그 컬럼을 표시에서 뺐다고
        # TOTAL 합산용 원본까지 잘라버리면 cost_excl_vat이 0 취급돼 CPA가 0으로 나온다.
        # 합산은 by_channel 전체 컬럼 기준으로 하고, 화면에 낼 항목만 bc_cols로 고른다.
        bc_total = build_total_row(by_channel, bc_cols, "channel", label_text="TOTAL")
        if bc_total:
            bc_table = pd.concat([bc_table, pd.DataFrame([bc_total])], ignore_index=True)
        render_html_table(korify(bc_table), raw=by_channel[bc_cols])
        st.download_button(
            "⬇️ 엑셀 다운로드 (매체별·월별)",
            data=to_excel_bytes(korify(format_display(by_channel[bc_cols]))), file_name="channel_performance.xlsx",
        )
        render_ops_comment_channel_narrative(
            by_channel, "roas",
            footnote="※ 선택 기간 합산·자체 ROAS 기준입니다. GA-ROAS 비교는 아래 '당월 매체별 GA 비교' 참고.",
        )

        render_channel_mix(fc)  # ← 신규: 채널믹스 목표 대비
    else:
        st.info("매체별 데이터가 아직 없습니다.")

    if not snapshot.empty:
        st.markdown("---")
        st.markdown("### 당월 매체별 GA 비교 (최신 스냅샷)")
        latest_month = snapshot["as_of_month"].max()
        snap_latest = add_kpis(snapshot[snapshot["as_of_month"] == latest_month])
        st.caption(f"기준월: {latest_month}")
        cols = ["channel", "impressions", "clicks", "cost_incl_vat", "conversions", "revenue", "roas", "ga_conversions", "ga_revenue", "ga_roas"]
        cols = [c for c in cols if c in snap_latest.columns]
        st.dataframe(korify(format_display(snap_latest[cols].sort_values("cost_incl_vat", ascending=False))), use_container_width=True, hide_index=True)


# "리타겟팅 매체(자사몰)"이 아니라 "리타겟팅 매체(네이버 스토어)"로 따로 빼는 채널.
# 네이버 쇼핑검색광고는 클릭 도착 페이지가 자사몰이 아니라 네이버 스토어(스마트스토어)라 구분한다.
TARGETING_STORE_CHANNELS = {"네이버 쇼핑검색광고"}

# 사용자가 정리해준 리포트 순서(비용순 정렬이 아니라 매체 관례상 고정 순서) — 목록에 없는
# 채널은 뒤에 광고비 내림차순으로 붙는다.
TARGETING_NEW_CHANNEL_ORDER = ["네이버 GFA PC", "네이버 GFA MO", "메타", "구글(P-MAX)", "네이버 맨즈탭"]
TARGETING_RETARGET_OWN_CHANNEL_ORDER = ["네이버 검색광고", "네이버 브랜드검색광고", "네이버 GFA PC", "네이버 GFA MO", "메타", "크리테오"]

TARGETING_CORE_COLS = ["channel", "impressions", "clicks", "ctr", "cpc", "cost_incl_vat",
                        "signups", "signup_rate", "conversions", "revenue", "roas"]


def _targeting_order_channels(df: pd.DataFrame, order_list: list) -> pd.DataFrame:
    df = df.copy()
    df["_ord"] = df["channel"].apply(lambda c: order_list.index(c) if c in order_list else len(order_list) + 1)
    return df.sort_values(["_ord", "cost_incl_vat"], ascending=[True, False]).drop(columns="_ord")


def _ops_targeting_summary_comment(audience_summary: pd.DataFrame, total_row: pd.Series, start: date, end: date) -> str:
    """타겟팅별 성과 ① TOTAL 비교 표 아래에 쓰는 코멘트. 신규/리타겟팅 어느 쪽이 매출·ROAS를
    견인했는지 짚어준다. 표가 이미 선택된 기간(period_filter)을 기준으로 집계돼 있어, 코멘트도
    같은 기간을 그대로 따른다 — '지난주'로 보고 싶으면 위 기간 선택에서 '지난주'를 고르면 된다."""
    status = _ops_kpi_status(total_row["roas"])
    bits = [
        f"**{start:%Y-%m-%d}~{end:%Y-%m-%d} 기준 전체 ROAS {total_row['roas']:,.0f}%로 {status}**입니다 "
        f"(KPI {OPS_KPI_ROAS_LOW}~{OPS_KPI_ROAS_HIGH}%)."
    ]
    new_r = audience_summary[audience_summary["channel"] == "신규 타겟팅"]
    re_r = audience_summary[audience_summary["channel"] == "리타겟팅"]
    if not new_r.empty and not re_r.empty and total_row["revenue"]:
        n, r = new_r.iloc[0], re_r.iloc[0]
        if r["roas"] >= n["roas"]:
            lead_label, lead, other_label, other = "리타겟팅", r, "신규 타겟팅", n
        else:
            lead_label, lead, other_label, other = "신규 타겟팅", n, "리타겟팅", r
        rev_share = lead["revenue"] / total_row["revenue"] * 100
        ratio_txt = f"{(lead['roas'] / other['roas']):.1f}배 높습니다." if other["roas"] else "훨씬 높습니다."
        bits.append(
            f"{lead_label}이 매출의 {rev_share:,.0f}%를 차지하며 계정 성과를 견인했고, "
            f"ROAS도 {lead['roas']:,.0f}%로 {other_label}({other['roas']:,.0f}%) 대비 {ratio_txt}"
        )
    return " ".join(bits)


def _targeting_total_row(df: pd.DataFrame, label: str = "TOTAL") -> pd.DataFrame:
    row = {
        "channel": label,
        "impressions": df["impressions"].sum(), "clicks": df["clicks"].sum(),
        "cost_incl_vat": df["cost_incl_vat"].sum(), "signups": df["signups"].sum(),
        "conversions": df["conversions"].sum(), "revenue": df["revenue"].sum(),
    }
    row["ctr"] = (row["clicks"] / row["impressions"] * 100) if row["impressions"] else 0
    row["cpc"] = (row["cost_incl_vat"] / row["clicks"]) if row["clicks"] else 0
    row["signup_rate"] = (row["signups"] / row["clicks"] * 100) if row["clicks"] else 0
    row["roas"] = (row["revenue"] / row["cost_incl_vat"] * 100) if row["cost_incl_vat"] else 0
    return pd.DataFrame([row])[TARGETING_CORE_COLS]


def render_targeting_performance_page(audience: pd.DataFrame, creatives_fallback: pd.DataFrame = None):
    """타겟팅별 성과 — 신규 타겟팅 vs 리타겟팅을 서로 겹치지 않게(mutually exclusive) 나눠서 본다.
    ① TOTAL 비교(TOTAL = 신규 타겟팅 + 리타겟팅) → ② 매체별 성과(신규 타겟팅 매체(자사몰) /
    리타겟팅 매체(자사몰) / 리타겟팅 매체(네이버 스토어)) → ③ 매체별 ROAS 비교, 3단 구성.
    '발굴/회수/수확' 역할 구분이나 CAC 중심 판정은 쓰지 않는다 — 대행사 운영 + 매체 예산이
    이미 고정된 구조에는 그 프레임이 안 맞는다는 판단에 따른 것."""
    if audience.empty and (creatives_fallback is None or creatives_fallback.empty):
        st.info(
            "아직 데이터가 없습니다. 주간 리포트를 업로드하면 캠페인 그룹명을 기준으로 "
            "신규/리타겟팅이 자동 분류돼 채워집니다."
        )
        return

    audience = audience.copy()
    if not audience.empty:
        audience["as_of_date"] = pd.to_datetime(audience["as_of_date"]).dt.date
        min_d, max_d = audience["as_of_date"].min(), audience["as_of_date"].max()
    else:
        cf = creatives_fallback.copy()
        cf["as_of_date"] = pd.to_datetime(cf["as_of_date"]).dt.date
        min_d, max_d = cf["as_of_date"].min(), cf["as_of_date"].max()
    start, end = period_filter(min_d, max_d, key="targeting")
    fa = audience[(audience["as_of_date"] >= start) & (audience["as_of_date"] <= end)] if not audience.empty else audience
    # 소재별 성과와 같은 이유로, 같은 달 안에 여러 번 업로드해도 채널×오디언스별 '최신 스냅샷'만 사용
    # (누적 리포트를 그대로 더하면 몇 배로 부풀려짐).
    if not fa.empty:
        fa = fa.sort_values("as_of_date").drop_duplicates(subset=["channel", "audience_type"], keep="last")

    # 구글(P-MAX)은 리포트 원본에 그룹(오디언스)별 집계 자체가 비어있어(전부 0) 위 표에는 안 잡힌다.
    # 사용자가 확인해준 매핑상 PMax는 항상 전체 신규이므로, 이미 검증된 소재별 성과(creative_performance)
    # 합계를 '신규'로 채워 넣는다 (회원가입 수는 이 표에서 안 잡히는 값이라 0으로 남는다).
    if creatives_fallback is not None and not creatives_fallback.empty:
        already_has_google = (not fa.empty) and (fa["channel"] == "구글(P-MAX)").any()
        if not already_has_google:
            g = creatives_fallback.copy()
            g["as_of_date"] = pd.to_datetime(g["as_of_date"]).dt.date
            g = g[(g["channel"] == "구글(P-MAX)") & (g["as_of_date"] >= start) & (g["as_of_date"] <= end)]
            if not g.empty:
                g_latest = g.sort_values("as_of_date").drop_duplicates(subset=["channel", "creative"], keep="last")
                synth = pd.DataFrame([{
                    "channel": "구글(P-MAX)", "audience_type": "신규",
                    "impressions": g_latest["impressions"].sum(), "clicks": g_latest["clicks"].sum(),
                    "cost_excl_vat": g_latest["cost_excl_vat"].sum(), "cost_incl_vat": g_latest["cost_incl_vat"].sum(),
                    "signups": 0, "conversions": g_latest["conversions"].sum(), "revenue": g_latest["revenue"].sum(),
                }])
                fa = pd.concat([fa, synth], ignore_index=True) if not fa.empty else synth

    if fa.empty:
        st.info("선택한 기간에 데이터가 없습니다.")
        return

    unclassified = fa[fa["audience_type"] == "미분류"]
    if not unclassified.empty:
        st.warning(
            f"'미분류'로 남은 그룹이 {len(unclassified)}건 있습니다(신규 캠페인/그룹이 추가됐을 수 있어요). "
            "신규·리타겟팅 어느 쪽에도 안 잡히고 아래 표에 별도로만 표시됩니다."
        )

    agg = (
        fa.groupby(["channel", "audience_type"], as_index=False)
        .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
             cost_excl_vat=("cost_excl_vat", "sum"), cost_incl_vat=("cost_incl_vat", "sum"),
             signups=("signups", "sum"), conversions=("conversions", "sum"), revenue=("revenue", "sum"))
    )
    agg = add_kpis(agg)
    agg["signup_rate"] = np.where(agg["clicks"] > 0, agg["signups"] / agg["clicks"] * 100, 0)

    # 네이버 맨즈탭 — 아직 별도 리포트 연동 전이라 데이터가 없다. 항목은 항상 노출하되 0으로 표시
    # (나중에 전용 리포트를 업로드받으면 이 자리에 실제 값이 채워지게 될 예정).
    if "네이버 맨즈탭" not in agg["channel"].values:
        placeholder = pd.DataFrame([{
            "channel": "네이버 맨즈탭", "audience_type": "신규",
            "impressions": 0.0, "clicks": 0.0, "cost_excl_vat": 0.0, "cost_incl_vat": 0.0,
            "signups": 0.0, "conversions": 0.0, "revenue": 0.0,
            "ctr": 0.0, "cpc": 0.0, "cpa": 0.0, "cvr": 0.0, "roas": 0.0, "aov": 0.0, "signup_rate": 0.0,
        }])
        agg = pd.concat([agg, placeholder], ignore_index=True)

    # ── ① 타겟팅별 성과 TOTAL 비교 (TOTAL = 신규 타겟팅 + 리타겟팅) ──
    st.markdown("##### ① 타겟팅별 성과 TOTAL 비교")
    AUDIENCE_ORDER = ["신규", "리타겟팅", "미분류"]
    AUDIENCE_LABEL = {"신규": "신규 타겟팅", "리타겟팅": "리타겟팅", "미분류": "미분류"}
    audience_summary = (
        agg.groupby("audience_type", as_index=False)
        .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
             cost_excl_vat=("cost_excl_vat", "sum"), cost_incl_vat=("cost_incl_vat", "sum"),
             signups=("signups", "sum"), conversions=("conversions", "sum"), revenue=("revenue", "sum"))
    )
    audience_summary = add_kpis(audience_summary)
    audience_summary["signup_rate"] = np.where(
        audience_summary["clicks"] > 0, audience_summary["signups"] / audience_summary["clicks"] * 100, 0
    )
    audience_summary["_order"] = audience_summary["audience_type"].apply(
        lambda x: AUDIENCE_ORDER.index(x) if x in AUDIENCE_ORDER else 99
    )
    audience_summary = audience_summary.sort_values("_order")
    audience_summary = audience_summary.rename(columns={"audience_type": "channel"})
    audience_summary["channel"] = audience_summary["channel"].map(AUDIENCE_LABEL).fillna(audience_summary["channel"])
    show_top = format_display(audience_summary[TARGETING_CORE_COLS])
    total_top = format_display(_targeting_total_row(audience_summary, "TOTAL"))
    # ②의 신규/리타겟팅 매체 표와 같은 폭으로 보이도록 동일하게 2단 컬럼의 절반 폭에 맞춘다
    # (표 하나만 있다고 전체 페이지 폭으로 늘어나면 컬럼 간 간격이 과하게 벌어져 보인다).
    top_col, _spacer_col = st.columns(2)
    with top_col:
        render_html_table(korify(pd.concat([total_top, show_top], ignore_index=True)))
    st.caption("TOTAL = 신규 타겟팅 + 리타겟팅 (겹치지 않는 완전 분리 기준)")
    total_row_raw = _targeting_total_row(audience_summary, "TOTAL").iloc[0]
    st.markdown(_ops_targeting_summary_comment(audience_summary, total_row_raw, start, end), unsafe_allow_html=True)

    st.markdown("##### ② 타겟팅별 매체별 성과 비교")
    new_df = agg[agg["audience_type"].isin(["신규", "미분류"])]
    retarget_df = agg[agg["audience_type"] == "리타겟팅"]
    retarget_store_df = retarget_df[retarget_df["channel"].isin(TARGETING_STORE_CHANNELS)]
    retarget_own_df = retarget_df[~retarget_df["channel"].isin(TARGETING_STORE_CHANNELS)]

    col_new, col_re = st.columns(2)

    with col_new:
        st.markdown("###### 1) 신규 타겟팅 매체 (자사몰)")
        if new_df.empty:
            st.info("선택 기간에 신규 분류 데이터가 없습니다.")
        else:
            new_ordered = _targeting_order_channels(new_df, TARGETING_NEW_CHANNEL_ORDER)
            show_new = format_display(new_ordered[TARGETING_CORE_COLS])
            total_new = format_display(_targeting_total_row(new_ordered, "TOTAL"))
            render_html_table(korify(pd.concat([total_new, show_new], ignore_index=True)))
            st.download_button(
                "⬇️ 엑셀 다운로드 (신규 타겟팅 매체)",
                data=to_excel_bytes(korify(new_ordered[TARGETING_CORE_COLS])),
                file_name="targeting_new.xlsx",
            )

    with col_re:
        st.markdown("###### 2-1) 리타겟팅 매체 (자사몰)")
        if retarget_df.empty:
            st.info("선택 기간에 리타겟팅 분류 데이터가 없습니다.")
        else:
            own_ordered = _targeting_order_channels(retarget_own_df, TARGETING_RETARGET_OWN_CHANNEL_ORDER)
            show_own = format_display(own_ordered[TARGETING_CORE_COLS])
            # TOTAL은 자사몰만이 아니라 리타겟팅 전체(자사몰 + 네이버 스토어) 합계 — 사용자가
            # 확인해준 대로 ①의 TOTAL 비교 '리타겟팅' 행과 정확히 일치해야 한다.
            total_re = format_display(_targeting_total_row(retarget_df, "TOTAL"))
            render_html_table(korify(pd.concat([total_re, show_own], ignore_index=True)))

        st.markdown("###### 2-2) 리타겟팅 매체 (네이버 스토어)")
        if retarget_store_df.empty:
            st.info("아직 데이터가 없습니다.")
        else:
            store_ordered = retarget_store_df.sort_values("cost_incl_vat", ascending=False)
            render_html_table(korify(format_display(store_ordered[TARGETING_CORE_COLS])))

        if not retarget_df.empty:
            retarget_all = pd.concat([retarget_own_df, retarget_store_df], ignore_index=True)
            st.download_button(
                "⬇️ 엑셀 다운로드 (리타겟팅 매체, 자사몰+네이버 스토어)",
                data=to_excel_bytes(korify(retarget_all[["channel"] + TARGETING_CORE_COLS[1:]])),
                file_name="targeting_retarget.xlsx",
            )

    st.markdown("##### ③ 매체별 ROAS 성과 비교")
    chart_new, chart_re = st.columns(2)
    with chart_new:
        st.markdown(
            f"<div style='font-size:16px;font-weight:700;color:{THEME_COLORS['foreground']};margin-bottom:6px;'>신규 타겟팅 매체별 ROAS(%)</div>",
            unsafe_allow_html=True,
        )
        if not new_df.empty:
            new_chart_df = new_df.sort_values("roas", ascending=False)
            new_chart_df["roas_label"] = new_chart_df["roas"].map(lambda v: f"{v:.0f}%")
            fig_new = px.bar(
                new_chart_df, x="channel", y="roas", text="roas_label",
                labels={"channel": "매체", "roas": "ROAS(%)"},
            )
            fig_new.update_yaxes(tickformat=",.0f", ticksuffix="%")
            fig_new.update_traces(hovertemplate="매체=%{x}<br>ROAS=%{y:,.0f}%<extra></extra>")
            st.plotly_chart(theme_chart(fig_new), use_container_width=True)
    with chart_re:
        st.markdown(
            f"<div style='font-size:16px;font-weight:700;color:{THEME_COLORS['foreground']};margin-bottom:6px;'>리타겟팅 매체별 ROAS(%) (자사몰 + 네이버 스토어)</div>",
            unsafe_allow_html=True,
        )
        if not retarget_df.empty:
            re_chart_df = retarget_df.sort_values("roas", ascending=False)
            re_chart_df["roas_label"] = re_chart_df["roas"].map(lambda v: f"{v:.0f}%")
            fig_re = px.bar(
                re_chart_df, x="channel", y="roas", text="roas_label",
                labels={"channel": "매체", "roas": "ROAS(%)"},
            )
            fig_re.update_yaxes(tickformat=",.0f", ticksuffix="%")
            fig_re.update_traces(hovertemplate="매체=%{x}<br>ROAS=%{y:,.0f}%<extra></extra>")
            st.plotly_chart(theme_chart(fig_re), use_container_width=True)


def render_ga_page(ga: pd.DataFrame):
    if not ga.empty:
        ga["as_of_date"] = pd.to_datetime(ga["as_of_date"]).dt.date
        st.subheader("🔎 기간 필터")
        min_g, max_g = ga["as_of_date"].min(), ga["as_of_date"].max()
        gstart, gend = period_filter(min_g, max_g, key="ga")
        g_in_range = ga[(ga["as_of_date"] >= gstart) & (ga["as_of_date"] <= gend)]
        if g_in_range.empty:
            st.info("선택한 기간에 해당하는 GA 스냅샷이 없습니다.")
        else:
            latest = g_in_range["as_of_date"].max()
            g = g_in_range[g_in_range["as_of_date"] == latest].sort_values("revenue", ascending=False)
            st.caption(f"기준일: {latest} (선택 기간 내 가장 최신 업로드 스냅샷)")
            st.dataframe(korify(format_display(g)), use_container_width=True, hide_index=True)
            st.download_button("⬇️ 엑셀 다운로드 (GA 유입경로)", data=to_excel_bytes(korify(format_display(g))), file_name="ga_source.xlsx")
    else:
        st.info("GA 유입경로 데이터가 아직 없습니다.")


def render_inflow_revenue_page(df: pd.DataFrame, ga_channel_inflow: pd.DataFrame = None):
    """유입·매출 비교 — 형이 별도로 정리해서 준 '일별 GA·어드민 지표 비교' 파일 기반.
    채널별이 아니라 사이트 전체 일별 합산 기준. ① 방문자 추이(GA 총/신규) ② 매출 비교
    (어드민=회사 내부 기준 vs GA 기준, 어드민 대비 격차%)로 구성.
    '매체-매출'(광고 리포트 기준)은 이 비교에서 뺐다 — 어드민 매출은 사이트 전체 매출이고
    GA 매출도 자연유입 포함 전체인데, 매체-매출은 광고 채널 것만 걷은 거라 셋을 나란히 비교하면
    서로 다른 걸 비교하는 셈이 된다(형 확인). GA-매출은 원본 시트에 채워져 있으면 그 값을 쓰고,
    비어있으면(0/NaN) 'GA 매체별 유입 경로'(UTM 매핑 전 원본 전체 합산)로 채운다."""
    if df.empty:
        st.info(
            "아직 데이터가 없습니다. 왼쪽 사이드바 '② GA 유입 데이터 업로드'에서 "
            "'일별 GA,어드민 지표 비교' 시트가 있는 파일을 올려주세요."
        )
        return

    df = df.copy()
    df["report_date"] = pd.to_datetime(df["report_date"]).dt.date
    ga_all = _ga_daily_agg_all(ga_channel_inflow)
    if not ga_all.empty:
        # 원본 시트 값이 전부 0/NaN이면(int64) 실수 GA 매출을 대입할 때 타입 경고가 나므로,
        # 대입 전에 float으로 미리 맞춰둔다.
        df["ga_revenue"] = pd.to_numeric(df["ga_revenue"], errors="coerce").astype(float)
        df = df.merge(ga_all, on="report_date", how="left")
        needs_fill = df["ga_revenue"].isna() | (df["ga_revenue"] == 0)
        df.loc[needs_fill, "ga_revenue"] = df.loc[needs_fill, "ga_revenue_all"]
        df = df.drop(columns=["ga_revenue_all", "ga_conversions_all"], errors="ignore")
        # GA-ROAS(=GA매출/광고비)는 여기서 안 쓴다 — GA-매출은 자연유입 포함 사이트 전체 매출인데
        # 광고비는 매체(광고)에만 든 비용이라, 둘을 나누면 범위가 안 맞는 숫자가 나온다(형 확인).
    st.subheader("🔎 기간 필터")
    min_d, max_d = df["report_date"].min(), df["report_date"].max()
    start, end = period_filter(min_d, max_d, key="inflow", default_preset="이번달")
    fd = df[(df["report_date"] >= start) & (df["report_date"] <= end)].sort_values("report_date")

    if fd.empty:
        st.info("선택한 기간에 데이터가 없습니다.")
        return

    # ── ① 방문자 추이 (GA 기준) ──
    st.markdown("##### ① 방문자 추이 (GA 기준)")
    users_sum = fd["users"].sum()
    new_users_sum = fd["new_users"].sum()
    returning_users_sum = fd["returning_users"].sum()
    new_ratio = (new_users_sum / users_sum * 100) if users_sum else 0
    returning_ratio = (returning_users_sum / users_sum * 100) if users_sum else 0
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("총 방문자 합계", f"{users_sum:,.0f} 명")
    c2.metric("신규 방문자 합계", f"{new_users_sum:,.0f} 명")
    c3.metric("재방문자 합계", f"{returning_users_sum:,.0f} 명")
    c4.metric("신규 방문자 비중", f"{new_ratio:.1f} %")
    c5.metric("재방문자 비중", f"{returning_ratio:.1f} %")
    _cap = ("총 방문자 = 신규 방문자 + 재방문자 (GA4 방문수 기준). "
            "GA4 보고서의 '총 사용자'는 중복을 뺀 사람 수라 이보다 작고, 신규+재방문과도 "
            "맞지 않습니다 — 매체별로 나눌 수 없는 지표라 대시보드에서는 쓰지 않습니다.")
    if "unknown_sessions" in fd.columns:
        _unk = float(pd.to_numeric(fd["unknown_sessions"], errors="coerce").fillna(0).sum())
        if _unk > 0:
            _cap += (f" · GA4가 신규/재방문을 판정하지 못한 방문 {_unk:,.0f}건"
                     f"({_unk / users_sum * 100:.1f}%)은 같은 날·같은 소스의 판정된 비율대로 "
                     "나눠 담았습니다.")
    st.caption(_cap)

    day_label_order = kor_date_labels(fd["report_date"], "day")
    fd = fd.assign(일자=day_label_order)
    visit_chart_df = fd.melt(
        id_vars=["일자"], value_vars=["users", "new_users"],
        var_name="구분", value_name="방문자수",
    )
    visit_chart_df["구분"] = visit_chart_df["구분"].map({"users": "총 방문자", "new_users": "신규 방문자"})
    fig_visit = px.line(
        visit_chart_df, x="일자", y="방문자수", color="구분", markers=True,
        category_orders={"일자": day_label_order},
    )
    fig_visit.update_yaxes(tickformat=",.0f")
    fig_visit.for_each_trace(
        lambda t: t.update(hovertemplate=f"구분={t.name}<br>일자=%{{x}}<br>방문자수=%{{y:,.0f}}명<extra></extra>")
    )
    st.plotly_chart(theme_chart(fig_visit), use_container_width=True)

    # ── ② 매출 비교 (어드민 vs GA) ──
    # '보고서(매체 리포트) 기준 매출'은 광고 채널에서 걷힌 매출만이라, 사이트 전체 매출인
    # 어드민·GA와 나란히 비교하면 서로 다른 범위를 비교하는 셈이라 이 표에서는 뺐다(형 확인).
    # 매체 리포트 기준 매출·ROAS는 '종합 대시보드'/'매체별 성과'에서 따로 확인할 수 있다.
    st.markdown("##### ② 매출 비교")
    st.caption("어드민 = 회사 내부(카페24 등 백엔드) 기준, 사이트 전체 매출 · GA = GA가 집계한 사이트 전체 매출(자연유입·direct 포함)")
    admin_sum = fd["admin_revenue"].sum()
    ga_sum = fd["ga_revenue"].sum()
    cost_sum = fd["cost_incl_vat"].sum()
    gap_ga = ((ga_sum - admin_sum) / admin_sum * 100) if admin_sum else 0

    r1, r3 = st.columns(2)
    r1.metric("어드민 매출(회사 내부 기준)", f"{admin_sum:,.0f} 원")
    r3.metric("GA 기준 매출", f"{ga_sum:,.0f} 원", f"{gap_ga:+.1f}% vs 어드민")
    st.caption("격차(%)는 어드민(회사 내부 기준) 매출 대비 초과/미달 비율입니다.")

    revenue_chart_df = fd.melt(
        id_vars=["일자"], value_vars=["admin_revenue", "ga_revenue"],
        var_name="구분", value_name="매출",
    )
    revenue_chart_df["구분"] = revenue_chart_df["구분"].map({
        "admin_revenue": "어드민(회사 내부)", "ga_revenue": "GA 기준",
    })
    fig_rev = px.line(
        revenue_chart_df, x="일자", y="매출", color="구분", markers=True,
        category_orders={"일자": day_label_order},
    )
    fig_rev.update_yaxes(tickformat=",.0f")
    fig_rev.for_each_trace(
        lambda t: t.update(hovertemplate=f"구분={t.name}<br>일자=%{{x}}<br>매출=%{{y:,.0f}}원<extra></extra>")
    )
    st.plotly_chart(theme_chart(fig_rev), use_container_width=True)

    # ── 일자별 상세 표 (정렬 가능) — 형이 정리한 엑셀 양식(사용자→신규방문자→재방문자→이탈률→
    # 페이지뷰→평균체류시간→어드민/매체/GA 매출→매체/GA-ROAS→어드민 대비 비교) 그대로 맞춘다.
    # "매출"/"ROAS"/"어드민 매출" 같은 이름은 다른 페이지에서도 공용으로 쓰는 라벨이라 KOR_COLS를
    # 바꾸면 거기까지 다 바뀌어버리므로, 이 표에서만 쓸 라벨을 따로 둔다.
    # GA-ROAS(=GA매출/광고비)는 이 표에서 뺐다 — GA-매출은 자연유입 포함 사이트 전체 매출인데
    # 광고비는 매체(광고)에만 쓴 비용이라 나누면 범위가 안 맞는 숫자가 나온다(형 확인). 매체 기준
    # ROAS는 종합 대시보드·매체별 성과에서 보는 걸로 하고, 여기는 어드민 vs GA 매출 비교만 한다.
    INFLOW_DETAIL_LABELS = {
        "report_date": "일자",
        "users": "총 방문자",
        "new_users": "신규방문자",
        "returning_users": "재방문자",
        "bounce_rate": "이탈률",
        "pageviews": "페이지뷰",
        "avg_session_duration": "평균 체류시간",
        "admin_revenue": "어드민-매출",
        "ga_revenue": "GA-매출",
        "ga_gap_pct": "어드민-GA 매출 비교",
    }
    INFLOW_DETAIL_LABELS_REV = {v: k for k, v in INFLOW_DETAIL_LABELS.items()}

    detail_cols = [
        "report_date", "users", "new_users", "returning_users", "bounce_rate", "pageviews",
        "avg_session_duration", "admin_revenue", "ga_revenue",
    ]
    detail = fd[detail_cols].copy()
    detail["ga_gap_pct"] = np.where(
        detail["admin_revenue"] > 0, (detail["ga_revenue"] - detail["admin_revenue"]) / detail["admin_revenue"] * 100, 0
    )
    final_cols = detail_cols + ["ga_gap_pct"]

    def _fmt_duration(v):
        # 초 단위 숫자를 "3분 26초" 형태로 되돌려 보여준다 (원본 엑셀 표기와 맞춤).
        if pd.isna(v):
            return "-"
        v = int(round(v))
        return f"{v // 60}분 {v % 60}초"

    show = format_display(detail[detail_cols])
    show["avg_session_duration"] = detail["avg_session_duration"].map(_fmt_duration)

    def _fmt_gap(v):
        # 매체 리포트/GA 데이터가 아직 안 올라온 최근 며칠은 NaN이라 "-"로 표시 (0%로 오해하지 않도록).
        if pd.isna(v):
            return "-"
        return f"{'▲' if v >= 0 else '▼'}{v:+.1f}%"

    show["ga_gap_pct"] = detail["ga_gap_pct"].map(_fmt_gap)

    returning_users_sum = fd["returning_users"].sum()  # (위 ①에서도 계산하지만 이 표 TOTAL 행에도 필요)
    total_row = {
        "report_date": "TOTAL",
        "users": f"{users_sum:,.0f}",
        "new_users": f"{new_users_sum:,.0f}",
        "returning_users": f"{returning_users_sum:,.0f}",
        "bounce_rate": f"{fd['bounce_rate'].mean():.2f}%" if fd["bounce_rate"].notna().any() else "-",
        "pageviews": f"{fd['pageviews'].sum():,.0f}",
        "avg_session_duration": _fmt_duration(fd["avg_session_duration"].mean()),
        "admin_revenue": f"{admin_sum:,.0f}",
        "ga_revenue": f"{ga_sum:,.0f}",
        "ga_gap_pct": f"{'▲' if gap_ga >= 0 else '▼'}{gap_ga:+.1f}%",
    }
    table = pd.concat([pd.DataFrame([total_row])[final_cols], show[final_cols]], ignore_index=True)
    table = table.rename(columns=INFLOW_DETAIL_LABELS)
    raw_numeric_cols = [
        "users", "new_users", "returning_users", "bounce_rate", "pageviews", "avg_session_duration",
        "admin_revenue", "ga_revenue",
    ]
    render_html_table(table, raw=detail[raw_numeric_cols], raw_label_map=INFLOW_DETAIL_LABELS_REV)

    # 이 페이지는 ROAS(성과) 판단용이 아니라 '어드민 장부 매출 vs GA 추적 매출'이 서로 맞게
    # 잡히고 있는지 보는 데이터 정합성 체크용이라, KPI 상태 표시 대신 격차 크기만 짚어준다.
    st.markdown("#### 💬 코멘트")
    comment_body = [
        f"총 방문자 {users_sum:,.0f}명(신규 비중 {new_ratio:.1f}%), 어드민 매출 대비 GA 매출 격차는 {_ops_fmt_pct(gap_ga)}입니다."
    ]
    st.markdown(" ".join(comment_body), unsafe_allow_html=True)
    if abs(gap_ga) >= 20:
        st.markdown(
            _ops_next_action("격차가 큰 편이니 GA 전자상거래 추적 설정, 어드민 매출 집계 범위(취소·환불 반영 여부)를 확인해보는 것을 권장합니다."),
            unsafe_allow_html=True,
        )

    dl_df = format_display(detail[detail_cols])
    dl_df["avg_session_duration"] = detail["avg_session_duration"].map(_fmt_duration)
    st.download_button(
        "⬇️ 엑셀 다운로드 (유입·매출 비교)",
        data=to_excel_bytes(dl_df.rename(columns=INFLOW_DETAIL_LABELS)),
        file_name="inflow_revenue_daily.xlsx",
    )


GA_CHANNEL_LABELS = {
    "source_medium": "세션 소스/매체",
    "channel": "매체",
    "users": "총 방문자",
    "new_users": "신규 방문자",
    "returning_users": "재 방문자",
    "bounce_rate": "이탈률",
    "pageviews": "조회수",
    "avg_session_duration": "평균 체류 시간",
    "conversions": "구매",
    "revenue": "매출",
}
GA_CHANNEL_LABELS_REV = {v: k for k, v in GA_CHANNEL_LABELS.items()}
GA_CHANNEL_DETAIL_COLS = [
    "source_medium", "channel", "users", "new_users", "returning_users",
    "bounce_rate", "pageviews", "avg_session_duration", "conversions", "revenue",
]
# ③ TOP 10 탭에서 쓰는 (컬럼명, 탭 라벨) 순서 — 형이 요청한 순서(총방문자→신규→재방문→구매→매출) 그대로.
GA_CHANNEL_TOP_METRICS = [
    ("users", "총 방문자"),
    ("new_users", "신규 방문자"),
    ("returning_users", "재방문자"),
    ("conversions", "구매"),
    ("revenue", "매출"),
]


def _fmt_duration_padded(v):
    # "00분 00초" 형태로 0을 채워서 보여준다 (형이 요청한 포맷).
    if pd.isna(v):
        return "-"
    v = int(round(v))
    return f"{v // 60:02d}분 {v % 60:02d}초"


def render_ga_channel_inflow_page(df: pd.DataFrame):
    """GA 매체별 유입 경로 — 세션 소스/매체 단위 일별 데이터를 기간 합산해서 보여준다.
    '매체'(채널 그룹핑, 예: 네이버 검색광고/구글(P-MAX) 등)는 형이 UTM 매핑을 완료해서
    다시 올리기 전까지는 비어있는 게 정상이다."""
    if df.empty:
        st.info(
            "아직 데이터가 없습니다. 왼쪽 사이드바 '② GA 유입 데이터 업로드'에서 "
            "파일을 올려주세요."
        )
        return

    df = df.copy()
    df["report_date"] = pd.to_datetime(df["report_date"]).dt.date
    st.subheader("🔎 기간 필터")
    min_d, max_d = df["report_date"].min(), df["report_date"].max()
    start, end = period_filter(min_d, max_d, key="ga_channel", default_preset="이번달")
    fd = df[(df["report_date"] >= start) & (df["report_date"] <= end)]

    if fd.empty:
        st.info("선택한 기간에 데이터가 없습니다.")
        return

    if fd["channel"].notna().any():
        st.caption("※ '매체' 컬럼은 UTM 매핑이 반영된 소스/매체만 채워져 있습니다.")
    else:
        st.caption("※ '매체'(채널 그룹핑) 매핑은 아직 반영 전이라 전부 비어있습니다 — 매핑 완료 파일을 다시 올리면 채워집니다.")

    # ── ① 방문자 추이 (GA 기준) — '유입·매출 비교' 페이지와 동일한 카드 구성 ──
    st.markdown("##### ① 방문자 추이 (GA 기준)")
    users_sum = fd["users"].sum()
    new_users_sum = fd["new_users"].sum()
    returning_users_sum = fd["returning_users"].sum()
    new_ratio = (new_users_sum / users_sum * 100) if users_sum else 0
    returning_ratio = (returning_users_sum / users_sum * 100) if users_sum else 0
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("총 방문자 합계", f"{users_sum:,.0f} 명")
    c2.metric("신규 방문자 합계", f"{new_users_sum:,.0f} 명")
    c3.metric("재방문자 합계", f"{returning_users_sum:,.0f} 명")
    c4.metric("신규 방문자 비중", f"{new_ratio:.1f} %")
    c5.metric("재방문자 비중", f"{returning_ratio:.1f} %")
    _cap = ("총 방문자 = 신규 방문자 + 재방문자 (GA4 방문수 기준). "
            "GA4 보고서의 '총 사용자'는 중복을 뺀 사람 수라 이보다 작고, 신규+재방문과도 "
            "맞지 않습니다 — 매체별로 나눌 수 없는 지표라 대시보드에서는 쓰지 않습니다.")
    if "unknown_sessions" in fd.columns:
        _unk = float(pd.to_numeric(fd["unknown_sessions"], errors="coerce").fillna(0).sum())
        if _unk > 0:
            _cap += (f" · GA4가 신규/재방문을 판정하지 못한 방문 {_unk:,.0f}건"
                     f"({_unk / users_sum * 100:.1f}%)은 같은 날·같은 소스의 판정된 비율대로 "
                     "나눠 담았습니다.")
    st.caption(_cap)

    # 소스/매체별로 나뉜 데이터를 날짜 기준으로 다시 합쳐서(전체 소스/매체 합산) 일별 추이를 그린다.
    daily_trend = fd.groupby("report_date", as_index=False).agg(
        users=("users", "sum"), new_users=("new_users", "sum")
    ).sort_values("report_date")
    date_label_order = kor_date_labels(daily_trend["report_date"], "day")
    daily_trend["일자"] = date_label_order
    visit_chart_df = daily_trend.melt(
        id_vars=["일자"], value_vars=["users", "new_users"],
        var_name="구분", value_name="방문자수",
    )
    visit_chart_df["구분"] = visit_chart_df["구분"].map({"users": "총 방문자", "new_users": "신규 방문자"})
    fig_visit = px.line(
        visit_chart_df, x="일자", y="방문자수", color="구분", markers=True,
        category_orders={"일자": date_label_order},
    )
    fig_visit.update_yaxes(tickformat=",.0f")
    fig_visit.for_each_trace(
        lambda t: t.update(hovertemplate=f"구분={t.name}<br>일자=%{{x}}<br>방문자수=%{{y:,.0f}}명<extra></extra>")
    )
    st.plotly_chart(theme_chart(fig_visit), use_container_width=True)

    # ── ② 구매·매출 합계 ──
    st.markdown("##### ② 구매·매출 합계")
    conv_sum = fd["conversions"].sum()
    rev_sum = fd["revenue"].sum()
    r1, r2 = st.columns(2)
    r1.metric("구매 합계", f"{conv_sum:,.0f} 건")
    r2.metric("매출 합계", f"{rev_sum:,.0f} 원")

    # 선택 기간 안에서 소스/매체 단위로 합산(방문자·조회수·구매·매출은 합, 이탈률·체류시간은 평균).
    agg = (
        fd.groupby("source_medium", as_index=False)
        .agg(
            channel=("channel", "first"),
            users=("users", "sum"),
            new_users=("new_users", "sum"),
            returning_users=("returning_users", "sum"),
            bounce_rate=("bounce_rate", "mean"),
            pageviews=("pageviews", "sum"),
            avg_session_duration=("avg_session_duration", "mean"),
            conversions=("conversions", "sum"),
            revenue=("revenue", "sum"),
        )
        .sort_values("users", ascending=False)
        .reset_index(drop=True)
    )

    st.markdown("##### ③ 소스/매체별 TOP 10")
    metric_tabs = st.tabs([label for _, label in GA_CHANNEL_TOP_METRICS])
    for tab_widget, (metric_col, metric_label) in zip(metric_tabs, GA_CHANNEL_TOP_METRICS):
        with tab_widget:
            top10 = agg.sort_values(metric_col, ascending=False).head(10)
            fig_top = px.bar(
                top10, x="source_medium", y=metric_col,
                labels={"source_medium": "세션 소스/매체", metric_col: metric_label},
            )
            fig_top.update_yaxes(tickformat=",.0f")
            unit = "원" if metric_col == "revenue" else ("건" if metric_col == "conversions" else "명")
            fig_top.update_traces(
                hovertemplate=f"세션 소스/매체=%{{x}}<br>{metric_label}=%{{y:,.0f}}{unit}<extra></extra>"
            )
            st.plotly_chart(theme_chart(fig_top), use_container_width=True, key=f"ga_channel_top_{metric_col}")

    st.markdown("##### ④ 소스/매체별 상세 표")
    st.caption("TOTAL 행은 현재 페이지가 아니라 선택한 기간 전체(모든 소스/매체) 기준 합계입니다.")

    total = len(agg)
    narrow_ps_col, _ps_spacer = st.columns([2, 10])
    with narrow_ps_col:
        page_size = st.selectbox("페이지당 표시", PAGE_SIZE_OPTIONS, index=1, key="ga_channel_pagesize")
    total_pages = max(1, -(-total // page_size))
    page = render_pager(total_pages, key="ga_channel_pager") if total_pages > 1 else 1
    start_i, end_i = (page - 1) * page_size, page * page_size
    view = agg.iloc[start_i:end_i].copy()

    show = format_display(view[GA_CHANNEL_DETAIL_COLS])
    show["avg_session_duration"] = view["avg_session_duration"].map(_fmt_duration_padded)
    show["channel"] = view["channel"].fillna("미분류")

    total_row = {
        "source_medium": "TOTAL",
        "channel": "",
        "users": f"{users_sum:,.0f}",
        "new_users": f"{new_users_sum:,.0f}",
        "returning_users": f"{fd['returning_users'].sum():,.0f}",
        "bounce_rate": f"{fd['bounce_rate'].mean():.2f}%" if fd["bounce_rate"].notna().any() else "-",
        "pageviews": f"{fd['pageviews'].sum():,.0f}",
        "avg_session_duration": _fmt_duration_padded(fd["avg_session_duration"].mean()),
        "conversions": f"{conv_sum:,.0f}",
        "revenue": f"{rev_sum:,.0f}",
    }
    table = pd.concat([pd.DataFrame([total_row])[GA_CHANNEL_DETAIL_COLS], show[GA_CHANNEL_DETAIL_COLS]], ignore_index=True)
    table = table.rename(columns=GA_CHANNEL_LABELS)
    raw_numeric_cols = [
        "users", "new_users", "returning_users", "bounce_rate", "pageviews",
        "avg_session_duration", "conversions", "revenue",
    ]
    render_html_table(table, raw=view[raw_numeric_cols], raw_label_map=GA_CHANNEL_LABELS_REV)

    st.markdown("#### 💬 코멘트")
    comment_body = [
        f"선택 기간 총 방문자 {users_sum:,.0f}명(신규 비중 {new_ratio:.1f}%, 재방문 비중 {returning_ratio:.1f}%), "
        f"구매 {conv_sum:,.0f}건, 매출 {rev_sum:,.0f}원입니다."
    ]
    if not agg.empty:
        top1 = agg.iloc[0]
        comment_body.append(f"가장 유입이 많은 소스/매체는 {top1['source_medium']}({top1['users']:,.0f}명)입니다.")
    st.markdown(" ".join(comment_body))
    if not agg.empty:
        st.markdown(
            _ops_next_action(
                f"{top1['source_medium']} 유입이 다음 기간에도 유지되는지 확인하고, "
                "유입 대비 구매(전환)가 낮은 소스/매체가 있으면 랜딩페이지·타겟팅 점검을 권장합니다."
            ),
            unsafe_allow_html=True,
        )
    if agg["channel"].isna().all() if not agg.empty else False:
        st.caption("※ '매체'(채널 그룹핑) 매핑 전이라 소스/매체 단위로만 코멘트했습니다.")

    dl_df = format_display(agg[GA_CHANNEL_DETAIL_COLS])
    dl_df["avg_session_duration"] = agg["avg_session_duration"].map(_fmt_duration_padded)
    dl_df["channel"] = agg["channel"].fillna("미분류")
    st.download_button(
        "⬇️ 엑셀 다운로드 (GA 매체별 유입 경로)",
        data=to_excel_bytes(dl_df.rename(columns=GA_CHANNEL_LABELS)),
        file_name="ga_channel_inflow.xlsx",
    )


# ──────────────────────────────────────────────────────────────
# GA4 Data API 자동 연동 (신규) — 엑셀 수동 업로드 없이 매일 자동으로 끌어온다
# ──────────────────────────────────────────────────────────────
# 필요한 것 (Streamlit Secrets):
#   [gcp_service_account]  ← 서비스 계정 JSON 내용 그대로
#   GA4_PROPERTY_ID = "123456789"
# requirements.txt 에 google-analytics-data 추가 필요.
#
# 스코프 주의: sessionSourceMedium은 '세션' 범위, totalUsers/newVsReturning은 '사용자' 범위라
# 둘을 섞으면 합계가 미묘하게 안 맞을 수 있다(구글 이슈로도 보고된 사항). 그래서 사용자수와
# 세션수를 둘 다 받아 저장해두고, 화면에는 기존 표들과 단위를 맞추기 위해 사용자수를 쓰되
# 필요하면 세션 기준으로 언제든 바꿔볼 수 있게 한다.
GA4_DIMENSIONS = ["date", "sessionSourceMedium", "newVsReturning"]
# 소재별 성과용. utm_campaign(캠페인) / utm_content(소재)까지 쪼갠다. 차원이 늘수록 행이
# 곱으로 불어나므로(날짜×소스×캠페인×소재×신규재방문) 채널용과 테이블을 나눠 저장한다.
GA4_CREATIVE_DIMENSIONS = ["date", "sessionSourceMedium", "sessionCampaignName",
                           "sessionManualAdContent", "newVsReturning"]
GA4_METRICS = ["totalUsers", "sessions", "transactions", "purchaseRevenue"]
GA4_LOOKBACK_DAYS = 30      # 최초 연동 시 끌어올 기간
GA4_RESYNC_TAIL_DAYS = 3    # GA는 하루이틀 뒤 값이 보정되므로 최근 며칠은 매번 다시 받아 덮어쓴다


@st.cache_resource(show_spinner=False)
def _build_ga4_client(sa_json: str):
    """실제 클라이언트 생성. 캐시 키를 서비스 계정 JSON 문자열로 잡아서, Secrets를 고치면
    캐시 키가 바뀌어 자동으로 다시 시도된다. 실패 시 예외를 그대로 올려보내는 게 핵심 —
    st.cache_resource는 예외를 캐시하지 않으므로 '한 번 실패하면 Secrets를 고쳐도 계속
    옛날 에러가 뜨는' 문제가 안 생긴다. (예전에 실패값을 캐시해서 실제로 이 버그가 있었다.)"""
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.oauth2 import service_account

    info = json.loads(sa_json)
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/analytics.readonly"]
    )
    return BetaAnalyticsDataClient(credentials=creds)


def get_ga4_client():
    """(클라이언트, 에러메시지)를 돌려준다. 실패는 캐시되지 않는다."""
    try:
        import google.analytics.data_v1beta  # noqa: F401
        import google.oauth2.service_account  # noqa: F401
    except Exception:
        return None, "google-analytics-data 패키지가 없습니다. requirements.txt에 추가해주세요."
    try:
        info = dict(st.secrets["gcp_service_account"])
    except Exception:
        return None, "Secrets에 [gcp_service_account]가 없습니다."
    try:
        return _build_ga4_client(json.dumps(info, sort_keys=True)), None
    except Exception as e:
        return None, f"서비스 계정 인증 실패: {e}"


def _ga4_property_id():
    try:
        return str(st.secrets["GA4_PROPERTY_ID"]).strip()
    except Exception:
        return None


def fetch_ga4_channel_daily(start: date, end: date, channel_map: dict = None) -> pd.DataFrame:
    """GA4에서 [날짜 × 세션 소스/매체 × 신규·재방문] 단위로 사용자수·세션수·구매·구매수익을 받아온다.
    utm_channel_map 매핑이 있으면 '매체'(channel) 컬럼까지 채워서 돌려준다."""
    client, err = get_ga4_client()
    prop = _ga4_property_id()
    if client is None or not prop:
        return pd.DataFrame()

    from google.analytics.data_v1beta.types import (
        DateRange, Dimension, Metric, RunReportRequest,
    )

    rows_out, offset, page_size = [], 0, 100_000
    while True:
        req = RunReportRequest(
            property=f"properties/{prop}",
            date_ranges=[DateRange(start_date=str(start), end_date=str(end))],
            dimensions=[Dimension(name=d) for d in GA4_DIMENSIONS],
            metrics=[Metric(name=m) for m in GA4_METRICS],
            limit=page_size,
            offset=offset,
        )
        resp = client.run_report(req)
        for r in resp.rows:
            dv = [d.value for d in r.dimension_values]
            mv = [m.value for m in r.metric_values]
            rows_out.append({
                "report_date": dv[0], "source_medium": dv[1], "user_type": dv[2],
                "users": mv[0], "sessions": mv[1], "conversions": mv[2], "revenue": mv[3],
            })
        offset += page_size
        if offset >= getattr(resp, "row_count", 0) or not resp.rows:
            break

    if not rows_out:
        return pd.DataFrame()

    # 회원가입은 GA4 이벤트로 따로 받는다. 같은 요청에 못 넣는 이유는 eventCount가
    # '모든 이벤트'를 세기 때문 — eventName으로 걸러야 가입만 나온다.
    # 이벤트 이름이 회사마다 달라서 Secrets(GA4_SIGNUP_EVENT)로 바꿀 수 있게 뒀다.
    signup_rows = []
    try:
        from google.analytics.data_v1beta.types import Filter, FilterExpression
        ev = str(_secrets_get("GA4_SIGNUP_EVENT", "sign_up")).strip()
        offset = 0
        while True:
            req = RunReportRequest(
                property=f"properties/{prop}",
                date_ranges=[DateRange(start_date=str(start), end_date=str(end))],
                dimensions=[Dimension(name=d) for d in GA4_DIMENSIONS],
                metrics=[Metric(name="eventCount")],
                dimension_filter=FilterExpression(
                    filter=Filter(field_name="eventName",
                                  string_filter=Filter.StringFilter(value=ev))),
                limit=page_size, offset=offset,
            )
            resp = client.run_report(req)
            for r in resp.rows:
                dv = [d.value for d in r.dimension_values]
                signup_rows.append({"report_date": dv[0], "source_medium": dv[1],
                                    "user_type": dv[2], "signups": r.metric_values[0].value})
            offset += page_size
            if offset >= getattr(resp, "row_count", 0) or not resp.rows:
                break
    except Exception:
        # 이벤트 이름이 없거나 권한이 모자라도 나머지 지표는 살려야 한다
        signup_rows = []

    out = pd.DataFrame(rows_out)
    if signup_rows:
        s = pd.DataFrame(signup_rows)
        out = out.merge(s, on=["report_date", "source_medium", "user_type"], how="left")
    if "signups" not in out.columns:
        out["signups"] = 0
    out["report_date"] = pd.to_datetime(out["report_date"], format="%Y%m%d", errors="coerce")
    out = out.dropna(subset=["report_date"])
    out["report_date"] = out["report_date"].dt.date
    for c in ["users", "sessions", "conversions", "revenue", "signups"]:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
    # GA4의 newVsReturning은 'new'/'returning', 값이 없으면 빈 문자열로 온다.
    out["user_type"] = out["user_type"].map(
        lambda v: "신규" if str(v).lower().startswith("new") else ("재방문" if str(v).lower().startswith("return") else "미상")
    )
    if channel_map:
        out["channel"] = out["source_medium"].map(lambda sm: channel_map.get(str(sm).strip().lower()))
    else:
        out["channel"] = None
    return out.reset_index(drop=True)


def fetch_ga4_creative_daily(start: date, end: date, channel_map: dict = None) -> pd.DataFrame:
    """GA4에서 [날짜 × 소스/매체 × 캠페인 × 소재 × 신규·재방문] 단위로 받아온다.

    소재(utm_content)는 광고 랜딩 URL에 utm_content가 박혀 있어야만 채워진다. 안 박혀 있으면
    GA4가 '(not set)'을 주는데, 그걸 그대로 두면 소재별 표가 (not set) 한 줄로 뭉개진다.
    그래서 '(미설정)'으로 바꿔 표시하고, 화면에서 비율을 같이 보여줘 UTM을 고칠 근거로 쓴다.
    """
    client, err = get_ga4_client()
    prop = _ga4_property_id()
    if client is None or not prop:
        return pd.DataFrame()

    from google.analytics.data_v1beta.types import (
        DateRange, Dimension, Metric, RunReportRequest,
    )

    def _run(metrics, dim_filter=None, out_key=None):
        rows, offset, page = [], 0, 100_000
        while True:
            kw = dict(
                property=f"properties/{prop}",
                date_ranges=[DateRange(start_date=str(start), end_date=str(end))],
                dimensions=[Dimension(name=d) for d in GA4_CREATIVE_DIMENSIONS],
                metrics=[Metric(name=m) for m in metrics],
                limit=page, offset=offset,
            )
            if dim_filter is not None:
                kw["dimension_filter"] = dim_filter
            resp = client.run_report(RunReportRequest(**kw))
            for r in resp.rows:
                dv = [d.value for d in r.dimension_values]
                mv = [m.value for m in r.metric_values]
                rec = {"report_date": dv[0], "source_medium": dv[1],
                       "campaign": dv[2], "creative": dv[3], "user_type": dv[4]}
                if out_key:
                    rec[out_key] = mv[0]
                else:
                    rec.update({"sessions": mv[0], "conversions": mv[1], "revenue": mv[2]})
                rows.append(rec)
            offset += page
            if offset >= getattr(resp, "row_count", 0) or not resp.rows:
                break
        return rows

    base = _run(["sessions", "transactions", "purchaseRevenue"])
    if not base:
        return pd.DataFrame()

    signup_rows = []
    try:
        from google.analytics.data_v1beta.types import Filter, FilterExpression
        ev = str(_secrets_get("GA4_SIGNUP_EVENT", "sign_up")).strip()
        signup_rows = _run(
            ["eventCount"],
            dim_filter=FilterExpression(
                filter=Filter(field_name="eventName",
                              string_filter=Filter.StringFilter(value=ev))),
            out_key="signups",
        )
    except Exception:
        signup_rows = []

    out = pd.DataFrame(base)
    keys = ["report_date", "source_medium", "campaign", "creative", "user_type"]
    if signup_rows:
        out = out.merge(pd.DataFrame(signup_rows), on=keys, how="left")
    if "signups" not in out.columns:
        out["signups"] = 0

    out["report_date"] = pd.to_datetime(out["report_date"], format="%Y%m%d", errors="coerce")
    out = out.dropna(subset=["report_date"])
    out["report_date"] = out["report_date"].dt.date
    for c in ["sessions", "conversions", "revenue", "signups"]:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
    out["user_type"] = out["user_type"].map(
        lambda v: "신규" if str(v).lower().startswith("new")
        else ("재방문" if str(v).lower().startswith("return") else "미상")
    )
    for c in ("campaign", "creative"):
        out[c] = out[c].map(
            lambda v: "(미설정)" if str(v).strip().lower() in
            ("(not set)", "(not provided)", "", "nan", "none") else str(v).strip()
        )
    if channel_map:
        out["channel"] = out["source_medium"].map(
            lambda sm: channel_map.get(str(sm).strip().lower()))
    else:
        out["channel"] = None
    return out.reset_index(drop=True)


def sync_ga4_creative_daily(existing: pd.DataFrame, channel_map: dict, force_full: bool = False):
    """소재별 GA 데이터 동기화. 채널용과 같은 규칙(어제까지, 최근 며칠 재수집)으로 돈다."""
    client, err = get_ga4_client()
    if client is None:
        return 0, None, None, err
    end = date.today() - timedelta(days=1)
    if force_full or existing is None or existing.empty or "report_date" not in existing.columns:
        start = end - timedelta(days=GA4_LOOKBACK_DAYS - 1)
    else:
        last = pd.to_datetime(existing["report_date"]).max().date()
        start = min(last - timedelta(days=GA4_RESYNC_TAIL_DAYS - 1), end)
    if start > end:
        return 0, None, None, None
    try:
        df = fetch_ga4_creative_daily(start, end, channel_map)
    except Exception as e:
        return 0, None, None, f"GA4 조회 실패: {e}"
    if df.empty:
        return 0, start, end, "GA4에서 받아온 소재 데이터가 0행입니다"
    n = save_table("ga_creative_daily", df,
                   "report_date,source_medium,campaign,creative,user_type", "GA4 API")
    if not n:
        return 0, start, end, (
            "GA4에서 데이터는 받았지만 저장에 실패했습니다 — "
            "ga_creative_daily 테이블이 없을 수 있습니다(배포 SQL 실행 필요)"
        )
    return n, start, end, None


def sync_ga4_channel_daily(existing: pd.DataFrame, channel_map: dict, force_full: bool = False):
    """저장된 마지막 날짜 이후(+최근 며칠 보정분)만 GA4에서 받아 upsert한다.
    (받은 행수, 시작일, 종료일, 에러메시지) 를 돌려준다."""
    client, err = get_ga4_client()
    if client is None:
        return 0, None, None, err
    today = date.today()
    end = today - timedelta(days=1)          # 어제까지 (오늘은 아직 집계 중이라 제외)
    if force_full or existing is None or existing.empty or "report_date" not in existing.columns:
        start = end - timedelta(days=GA4_LOOKBACK_DAYS - 1)
    else:
        last = pd.to_datetime(existing["report_date"]).max().date()
        start = min(last - timedelta(days=GA4_RESYNC_TAIL_DAYS - 1), end)
    if start > end:
        return 0, None, None, None
    try:
        df = fetch_ga4_channel_daily(start, end, channel_map)
    except Exception as e:
        return 0, None, None, f"GA4 조회 실패: {e}"
    if df.empty:
        return 0, start, end, "GA4에서 받아온 데이터가 0행입니다 (속성ID/뷰어 권한 확인 필요)"
    n = save_table("ga_channel_daily", df, "report_date,source_medium,user_type", "GA4 API")
    if not n:
        return 0, start, end, (
            "GA4에서 데이터는 받았지만 Supabase 저장에 실패했습니다 — "
            "ga_channel_daily 테이블이 없을 수 있습니다(ga4_tables.sql 실행 필요)"
        )
    return n, start, end, None


B64_CHARS = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=")


def diagnose_ga4_setup() -> str:
    """GA4 자동 연동이 어느 단계에서 막혔는지 한 화면에서 보여준다.
    비밀값(private_key 등)은 절대 출력하지 않고 '형식이 맞는지'만 검사한다.
    예산 파일 진단 패널(_diagnose_budget_sheet)과 같은 용도 — 막히면 이걸 먼저 펼쳐본다."""
    L = []
    ok = lambda s: L.append(f"✅ {s}")
    ng = lambda s: L.append(f"❌ {s}")
    info = lambda s: L.append(f"   ↳ {s}")

    # 1) 라이브러리
    try:
        import google.analytics.data_v1beta  # noqa: F401
        ok("1. google-analytics-data 설치됨")
    except Exception as e:
        ng(f"1. google-analytics-data 없음 → requirements.txt에 추가 후 재배포 필요 ({e})")
        return "\n".join(L)
    try:
        import google.oauth2.service_account  # noqa: F401
        ok("2. google-auth 설치됨")
    except Exception as e:
        ng(f"2. google-auth 없음 → requirements.txt에 추가 필요 ({e})")
        return "\n".join(L)

    # 3) Secrets — 값은 안 보여주고 존재/형식만
    try:
        sa = dict(st.secrets["gcp_service_account"])
        ok("3. Secrets [gcp_service_account] 있음")
    except Exception:
        ng("3. Secrets에 [gcp_service_account] 섹션이 없음 → 셋업 가이드 5단계 확인")
        return "\n".join(L)

    need = ["type", "project_id", "private_key_id", "private_key", "client_email", "client_id", "token_uri"]
    missing = [k for k in need if not str(sa.get(k, "")).strip()]
    if missing:
        ng(f"4. 빠진 항목: {', '.join(missing)}")
    else:
        ok("4. 필수 항목 모두 채워짐")
    # 예시값 검사는 '부분 포함'으로 본다 — 'stco-dashboard@xxx.iam.gserviceaccount.com' 처럼
    # 예시 문자열이 값 중간에 박혀 있는 경우를 전체 일치 검사로는 놓쳤다(실제로 겪은 사례).
    ph_marks = ["여기에", "xxx", "<JSON", "<GA4", "아주 긴 문자열"]
    placeholders = []
    for k, v in sa.items():
        if k == "private_key" or not isinstance(v, str):
            continue
        if any(mk in v for mk in ph_marks):
            placeholders.append(k)
    if placeholders:
        ng(f"5. 예시값이 그대로 남아있음: {', '.join(placeholders)} → JSON 파일의 실제 값으로 교체 필요")
        for k in placeholders:
            info(f"{k} = {sa.get(k)}")
    else:
        ok("5. 예시값(여기에-값 / xxx 등) 남은 것 없음")

    pk = str(sa.get("private_key", ""))
    if not pk.startswith("-----BEGIN"):
        ng("6. private_key가 '-----BEGIN' 으로 시작하지 않음")
    elif "END PRIVATE KEY" not in pk:
        ng("6. private_key에 '-----END PRIVATE KEY-----' 가 없음(값이 잘렸을 수 있음)")
    elif "\n" not in pk:
        ng("6. private_key에 줄바꿈(\\n)이 없음 → JSON 원본의 \\n 을 지우지 말 것")
    else:
        # 진짜 RSA 2048 키는 본문만 1,600자 이상이다. 그보다 짧으면 예시값(...)이 남은 것.
        body = pk.split("-----BEGIN PRIVATE KEY-----")[-1].split("-----END")[0].strip()
        body_lines = [ln for ln in body.split(chr(10)) if ln.strip()]
        # 어디에 이상한 문자가 섞였는지 '위치'로 정확히 짚는다 — 추측으로 안내하지 않기 위함.
        # 키 내용 자체는 출력하지 않고, 줄 번호와 문제 문자만 보여준다.
        bad_spots = []
        for li, ln in enumerate(body_lines, start=1):
            for ci, chx in enumerate(ln.strip(), start=1):
                if chx not in B64_CHARS:
                    bad_spots.append((li, ci, chx))
                    break
        clean_body = "".join(c for c in body if c in B64_CHARS)
        if len(body) < 500:
            ng(f"6. private_key 본문이 너무 짧습니다({len(body)}자) → 실제 키가 안 들어갔습니다. "
               "JSON 파일의 private_key 값 전체를 복사해주세요")
        elif bad_spots:
            ng(f"6. private_key 안에 base64가 아닌 문자가 {len(bad_spots)}곳 있습니다")
            for li, ci, chx in bad_spots[:5]:
                info(f"본문 {li}번째 줄의 {ci}번째 글자: '{chx}' (코드 {ord(chx)})")
            info("→ 손으로 고치지 마시고, JSON의 private_key 값을 통째로 다시 복사해 붙여넣으세요")
        else:
            # 정상 RSA 2048 키의 base64 본문은 1,624자 안팎이다. 크게 벗어나면 뭔가 섞인 것.
            note = ""
            if not (1500 <= len(clean_body) <= 1800):
                note = f" ⚠ 보통 1,624자인데 {len(clean_body)}자입니다 — 값이 잘렸거나 덧붙었을 수 있습니다"
            ok(f"6. private_key 형식 정상 (base64 {len(clean_body)}자, 줄 {len(body_lines)}개){note}")
    info(f"서비스 계정 이메일: {sa.get('client_email', '(없음)')}")
    info("↑ 이 이메일이 GA4 [관리 → 속성 액세스 관리]에 '뷰어'로 추가돼 있어야 합니다")

    # 7) 속성 ID
    prop = _ga4_property_id()
    if not prop:
        ng("7. Secrets에 GA4_PROPERTY_ID 없음")
        return "\n".join(L)
    if not prop.isdigit():
        ng(f"7. GA4_PROPERTY_ID가 숫자가 아님: '{prop}' (G-XXXX 측정ID가 아니라 숫자 속성ID여야 함)")
    else:
        ok(f"7. GA4_PROPERTY_ID = {prop}")

    # 8) 인증
    client, err = get_ga4_client()
    if client is None:
        ng(f"8. 인증 실패: {err}")
        return "\n".join(L)
    ok("8. 서비스 계정 인증 성공")

    # 9) 실제 API 호출
    try:
        test_end = date.today() - timedelta(days=1)
        test_start = test_end - timedelta(days=2)
        df = fetch_ga4_channel_daily(test_start, test_end)
        if df.empty:
            ng(f"9. API는 응답했지만 데이터 0행 ({test_start}~{test_end})")
            info("→ 속성ID가 다른 사이트이거나, 서비스 계정이 아직 뷰어로 추가 안 됐을 수 있습니다")
        else:
            ok(f"9. GA4 조회 성공: {len(df):,}행 ({test_start}~{test_end})")
            info(f"소스/매체 예시: {', '.join(df['source_medium'].dropna().unique()[:3])}")
            info(f"사용자수 합계 {df['users'].sum():,.0f} · 구매 {df['conversions'].sum():,.0f}건")
    except Exception as e:
        msg = str(e)
        ng(f"9. GA4 조회 실패: {msg[:300]}")
        if "PERMISSION_DENIED" in msg or "403" in msg:
            info("→ 서비스 계정이 GA4 속성에 뷰어로 추가되지 않았습니다 (셋업 가이드 4단계)")
        elif "404" in msg or "NOT_FOUND" in msg:
            info("→ GA4_PROPERTY_ID가 잘못됐습니다 (GA4 관리 → 속성 설정의 숫자 ID)")
        return "\n".join(L)

    # 10) Supabase 테이블
    sb = get_supabase_client()
    if sb is None:
        ng("10. Supabase 연결 없음 (로컬 메모리 모드로 동작 중 — 새로고침하면 사라집니다)")
    else:
        try:
            sb.table("ga_channel_daily").select("report_date").limit(1).execute()
            ok("10. Supabase ga_channel_daily 테이블 접근 가능")
        except Exception as e:
            ng(f"10. ga_channel_daily 테이블이 없거나 접근 불가 → ga4_tables.sql 실행 필요 ({str(e)[:200]})")
        try:
            sb.table("decision_log").select("decided_on").limit(1).execute()
            ok("11. Supabase decision_log 테이블 접근 가능")
        except Exception as e:
            ng(f"11. decision_log 테이블 없음 → ga4_tables.sql 실행 필요 ({str(e)[:160]})")

    return "\n".join(L)


def ga_inflow_source(ga_daily: pd.DataFrame, excel_inflow: pd.DataFrame) -> pd.DataFrame:
    """GA 유입 화면들이 쓸 원본 한 벌. GA4 API로 받은 게 있으면 그걸 쓰고, 없을 때만
    예전 엑셀 업로드분으로 내려간다. 화면마다 다른 원본을 보면 같은 기간인데 숫자가 달라져서
    (형이 겪은 그 문제) 반드시 한 군데서 정한다."""
    if ga_daily is not None and not ga_daily.empty:
        try:
            shaped = ga4_daily_to_inflow_shape(ga_daily)
            if shaped is not None and not shaped.empty:
                return shaped
        except Exception:
            pass
    return excel_inflow if excel_inflow is not None else pd.DataFrame()


def ga4_daily_to_inflow_shape(ga_daily: pd.DataFrame) -> pd.DataFrame:
    """ga_channel_daily(신규/재방문이 행으로 쪼개져 있음)를 기존 ga_channel_inflow와 같은
    모양(날짜×소스매체 1행, new_users/returning_users 컬럼)으로 바꾼다. 이렇게 해두면 기존
    집계 함수들(_ga_visits_by_channel, classify_ga_bucket 등)을 그대로 재사용할 수 있다."""
    cols = ["report_date", "source_medium", "channel", "users", "new_users",
            "returning_users", "sessions", "new_sessions", "ret_sessions",
            "signups", "conversions", "revenue",
            "new_signups", "new_conv", "new_rev", "ret_conv", "ret_rev"]
    if ga_daily is None or ga_daily.empty:
        return pd.DataFrame(columns=cols)
    g = ga_daily.copy()
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    for c in ["users", "sessions", "conversions", "revenue", "signups"]:
        g[c] = pd.to_numeric(g.get(c), errors="coerce").fillna(0).astype(float)
    # 목적축(신규 발굴 / 매출 확보)마다 봐야 할 숫자가 다르다. 사용자만 쪼개고 구매·매출을
    # 합쳐버리면 두 탭이 똑같은 표가 되어버리므로, 구매·매출·가입도 유형별로 나눠 들고 간다.
    is_new = g["user_type"] == "신규"
    is_ret = g["user_type"] == "재방문"
    g["new_users"] = np.where(is_new, g["users"], 0)
    g["returning_users"] = np.where(is_ret, g["users"], 0)
    # 방문수(세션)도 신규/재방문으로 갈라둔다. 사용자는 날짜별로 더하면 같은 사람이 여러 번
    # 세어져 GA4 보고서와 안 맞지만, 세션은 더해도 되는 지표라 보고서와 정확히 일치한다.
    g["new_sessions"] = np.where(is_new, g["sessions"], 0)
    g["ret_sessions"] = np.where(is_ret, g["sessions"], 0)
    g["new_signups"] = np.where(is_new, g["signups"], 0)
    g["new_conv"] = np.where(is_new, g["conversions"], 0)
    g["new_rev"] = np.where(is_new, g["revenue"], 0)
    g["ret_conv"] = np.where(is_ret, g["conversions"], 0)
    g["ret_rev"] = np.where(is_ret, g["revenue"], 0)
    out = g.groupby(["report_date", "source_medium"], as_index=False).agg(
        channel=("channel", "first"), users=("users", "sum"), new_users=("new_users", "sum"),
        returning_users=("returning_users", "sum"), sessions=("sessions", "sum"),
        new_sessions=("new_sessions", "sum"), ret_sessions=("ret_sessions", "sum"),
        signups=("signups", "sum"),
        conversions=("conversions", "sum"), revenue=("revenue", "sum"),
        new_signups=("new_signups", "sum"),
        new_conv=("new_conv", "sum"), new_rev=("new_rev", "sum"),
        ret_conv=("ret_conv", "sum"), ret_rev=("ret_rev", "sum"),
    )

    # ── 화면에 세우는 '방문자'는 세션(방문수)이다 ──────────────────────────────
    # GA4의 사용자 수는 기간 전체에서 중복을 뺀 값이라, 날짜·매체별로 쪼개 저장한 걸 더하면
    # 같은 사람이 여러 번 세어져 GA4 보고서와 안 맞는다(더할 수 없는 지표). 게다가
    # 총 사용자 ≠ 신규 + 재방문이라 표의 세로 계산도 안 맞는다.
    # 세션은 더해도 되고 총 방문 = 신규 + 재방문이 정확히 성립하며, 광고 클릭 1회 = 방문 1회라
    # 광고비와 단위도 맞는다. 사람 수는 people_* 로 남겨둔다(필요할 때 꺼내 쓸 수 있게).
    out["people"] = out["users"]
    out["new_people"] = out["new_users"]
    out["ret_people"] = out["returning_users"]
    has_sess = out["sessions"] > 0
    out["users"] = np.where(has_sess, out["sessions"], out["users"])
    out["new_users"] = np.where(has_sess, out["new_sessions"], out["new_users"])
    out["returning_users"] = np.where(has_sess, out["ret_sessions"], out["returning_users"])
    # ── 신규/재방문이 안 붙은 '미상' 세션을 배분한다 ─────────────────────────
    # GA4는 동의 거부·쿠키 삭제·앱↔웹 전환 같은 이유로 일부 세션에 newVsReturning을 못 붙인다.
    # 그대로 두면 총 방문자 ≠ 신규 + 재방문이 되어 표의 세로 계산이 깨진다(형이 잡은 그 문제).
    # 같은 날·같은 소스에서 판정된 세션의 신규:재방문 비율로 나눠 담아 합을 맞춘다.
    # 배분 전 원래 값은 *_known 으로, 미상 규모는 unknown_sessions 로 남겨 검증할 수 있게 한다.
    out["new_known"] = out["new_users"]
    out["ret_known"] = out["returning_users"]
    known = out["new_users"] + out["returning_users"]
    out["unknown_sessions"] = (out["users"] - known).clip(lower=0)
    resid = out["unknown_sessions"]
    sh_known = (out["new_users"] / known).replace([np.inf, -np.inf], np.nan)
    sh_people = (out["new_people"] / out["people"]).replace([np.inf, -np.inf], np.nan)
    share_new = sh_known.fillna(sh_people).fillna(0.0).clip(0, 1)
    out["new_users"] = out["new_users"] + resid * share_new
    out["returning_users"] = out["returning_users"] + resid * (1 - share_new)

    # 구매·매출도 같은 이유로 유형별 합이 전체보다 작을 수 있다. 같은 비율로 맞춰준다.
    for whole, a, b in (("conversions", "new_conv", "ret_conv"),
                        ("revenue", "new_rev", "ret_rev")):
        r = (out[whole] - out[a] - out[b]).clip(lower=0)
        out[a] = out[a] + r * share_new
        out[b] = out[b] + r * (1 - share_new)

    # GA 매체별 유입 경로 화면이 쓰는 칸들. GA4 API에는 없어서 비워둔다(0이 아니라 '-'로 보이게).
    for c in ("bounce_rate", "avg_session_duration", "pageviews"):
        if c not in out.columns:
            out[c] = np.nan
    return out[cols + ["people", "new_people", "ret_people",
                       "new_known", "ret_known", "unknown_sessions",
                       "bounce_rate", "avg_session_duration", "pageviews"]]


# ──────────────────────────────────────────────────────────────
# 채널 퍼널 리포트 V4 — 디자인/집계 상수
# ──────────────────────────────────────────────────────────────
# 채널명은 소스마다 표기가 다르다(채널믹스 파일 '네이버 맨즈탭' / 매체 리포트 시트 '(DA) 네이버_맨즈탭'
# / GA 매핑 '네이버맨즈' 등). 계획(예산) 대비 실제 집행을 한 줄에 놓으려면 반드시 한 이름으로 모아야
# 해서, 대표 채널명 → 매칭 키워드 목록으로 정규화한다. 위에서부터 먼저 걸리는 규칙을 쓰므로
# 더 구체적인 규칙(네이버 맨즈탭)을 일반 규칙(네이버 검색광고)보다 앞에 둔다.
FUNNEL_CANON_RULES = [
    ("네이버 맨즈탭", ["맨즈탭", "맨즈", "mens"]),
    ("네이버 브랜드검색광고", ["브랜드검색", "브검"]),
    ("네이버 쇼핑검색광고", ["쇼핑검색", "ssp"]),
    ("네이버 검색광고", ["네이버 검색", "네이버검색", "(sa)", "네이버sa"]),
    ("네이버 애드부스트", ["애드부스트", "adboost", "advoost", "ad voost"]),
    ("네이버 트렌드픽", ["트렌드픽", "trendpick", "trend pick"]),
    ("네이버 GFA", ["gfa"]),
    ("메타", ["메타", "페이스북", "facebook", "meta", "인스타", "instagram"]),
    ("구글", ["구글", "google", "p-max", "pmax", "실적최대화", "demand"]),
    ("카카오", ["카카오", "kakao", "플친"]),
    ("모비온", ["모비온", "mobon"]),
    ("크리테오", ["크리테오", "criteo"]),
    ("AEDI", ["aedi", "에디"]),
]

# 퍼널 단계별 전환율 벤치마크(%). 실행 트래커의 '벤치마크(목표)' 칸과 같은 역할로, 이 값보다
# 낮은 단계를 병목(⚠)으로 표시한다. 베이스라인이 쌓이면 이 숫자만 바꾸면 된다.
FUNNEL_BENCHMARK_NEW = {"클릭": 1.0, "방문": 80.0, "가입": 2.0, "첫구매": 25.0}
FUNNEL_BENCHMARK_RETURN = {"재클릭": 1.0, "재방문": 80.0, "재구매": 3.0}
FUNNEL_MIX_DRIFT_PP = 5.0        # 계획 대비 집행 비중이 이만큼(%p) 벌어지면 '예산 이탈'로 표시
FUNNEL_MIN_SPEND = 100_000       # 이 미만 광고비 채널은 표본이 작아 '판단 보류'로 둔다

FUNNEL_V4_CSS = """
<style>
.fv4-wrap { font-family: -apple-system, BlinkMacSystemFont, "Pretendard", "Segoe UI", sans-serif; }
.fv4-banner {
  background:#17170f; border-radius:14px; padding:22px 26px; display:flex; gap:26px;
  align-items:stretch; flex-wrap:wrap; margin-bottom:14px;
}
.fv4-banner-lead { display:flex; gap:16px; align-items:flex-start; min-width:280px; flex:1 1 300px; }
.fv4-count {
  width:34px; height:34px; border-radius:50%; background:#c8f231; color:#17170f;
  font-weight:800; font-size:16px; display:flex; align-items:center; justify-content:center; flex:none;
}
.fv4-banner-eyebrow { color:#8f8f80; font-size:13.5px; letter-spacing:.04em; margin-bottom:6px; }
.fv4-banner-title { color:#fdfdf7; font-size:21px; font-weight:800; line-height:1.42; }
.fv4-signals { display:flex; gap:30px; flex:2 1 520px; flex-wrap:wrap; }
.fv4-signal { min-width:180px; flex:1 1 180px; }
.fv4-bk-cap{font-size:14px;color:#6E747C;margin:16px 0 9px;line-height:1.75}
.fv4-bk{margin-bottom:4px}
.fv4-bk-row td{background:#FBFAF7;font-size:14px;font-weight:600}
.fv4-bk-row td.l{text-align:left}
.fv4-bk-sub{font-size:13.5px;color:#7A8088;font-weight:400;margin-left:8px}
.fv4-bk-share{display:inline-block;padding:2px 7px;border-radius:5px;
  background:#EAF7D9;color:#3F6B12;font-size:13px;font-weight:700}
.fv4-bk-split{font-size:13px;font-weight:700;color:#6E747C;margin:18px 0 6px;letter-spacing:.02em}
.fv4-bk-split-row td{background:#FFF !important;color:#6E747C;font-size:13.5px;font-weight:700;
  letter-spacing:.06em;padding-top:16px !important;border-bottom:1px solid #E3E1DC}
.fv4-sum-row td{background:#14181F !important;color:#F3F1EC !important;font-weight:800;
  border-bottom:none}
.fv4-sum-row td.l{color:#D9F27E !important;letter-spacing:.06em}
.fv4-sum-row .fv4-bk-share{background:rgba(217,242,126,.18);color:#D9F27E}
.fv4-sec{font-size:13.5px;font-weight:800;color:#14181F;margin:22px 0 6px;letter-spacing:-.01em}
.fv4-ar{margin-left:5px;color:#C2C6CC;font-size:13px;font-weight:400}
.fv4-tbl th:hover{background:#EDEAE3}
.fv4-ind{padding-left:14px;display:inline-block}
.fv4-na{color:#C2C6CC}
.fv4-mixpanel-wide{background:#FFF;border:1px solid #E3E1DC;border-radius:14px;padding:22px 26px;margin-top:16px}
.fv4-mix-amt-big{font-size:34px;font-weight:800;letter-spacing:-.03em;color:#14181F;margin:4px 0 2px}
.fv4-mix-ch{font-size:14px;font-weight:700;color:#14181F}
.fv4-mix-val{font-size:13.5px;color:#6E747C;font-weight:600}
.fv4-mix-val b{margin-left:10px;color:#14181F;font-weight:800}
.fv4-mixpanel-wide .fv4-mix-row{margin-top:12px}
.fv4-mixpanel-wide .fv4-mix-track{height:7px;background:#EFEDE8;border-radius:4px;margin-top:5px;overflow:hidden}
.fv4-mixpanel-wide .fv4-mix-track i{display:block;height:100%;background:#8B7BE8;border-radius:4px}
.fv4-mixpanel-wide .fv4-mix-list{margin-top:16px}
.cp-rec-det{color:#3F5210;font-size:14.5px}
.fv4-chip {
  display:inline-block; font-size:13px; font-weight:700; padding:3px 9px; border-radius:5px; margin-bottom:9px;
}
.fv4-chip.bad  { background:#fadadd; color:#a3172b; }
.fv4-chip.good { background:#d9f5cf; color:#1f6b2c; }
.fv4-chip.warn { background:#f7edc4; color:#7a5c14; }
.fv4-chip.hold { background:#ececdf; color:#6d6d5d; }
.fv4-signal-title { color:#fdfdf7; font-size:14px; font-weight:700; margin-bottom:5px; }
.fv4-signal-sub { color:#96968a; font-size:13.5px; line-height:1.5; }

.fv4-kpis { display:flex; border:1px solid #e6e4da; border-radius:12px; overflow:hidden; background:#fffef9; margin-bottom:26px; flex-wrap:wrap; }
.fv4-kpi { flex:1 1 170px; padding:17px 20px; border-right:1px solid #eceadf; }
.fv4-kpi:last-child { border-right:none; }
.fv4-kpi-label { color:#8a8a7c; font-size:13.5px; margin-bottom:9px; }
.fv4-kpi-value { color:#17170f; font-size:26px; font-weight:800; letter-spacing:-.02em; display:flex; align-items:baseline; gap:8px; }
.fv4-kpi-value.money { font-size:21px; letter-spacing:-.03em; flex-wrap:wrap; gap:6px; }
.fv4-chg-note { margin-top:10px; padding-top:9px; border-top:1px dashed #DEDCCF;
                color:#8a8a7c; font-size:14px; line-height:1.6; }
.fv4-chg-p   { color:#8a8a7c; font-weight:500; }
.fv4-chg-tag { display:inline-block; margin-left:6px; padding:2px 9px; border-radius:999px;
               background:#EFEEE6; color:#7A7A6E; font-size:13px; font-weight:700; }
.fv4-chg-tag.warn { background:#FBE9E7; color:#B03A2E; }
.fv4-chg-dot.flat { background:#C9C7BA; }
.gc-sub { display:block; margin-top:3px; color:#8a8a7c; font-size:13px; font-weight:500; }
.fv4-kpi-delta { font-size:13.5px; font-weight:700; }
.fv4-up   { color:#c0392b; }
.fv4-down { color:#2563c9; }
.fv4-kpi-sub { color:#7A7A6E; font-size:14px; margin-top:7px; }
.fv4-stack { display:flex; height:9px; border-radius:5px; overflow:hidden; margin:12px 0 8px; background:#eceadf; }
.fv4-stack i { display:block; height:100%; }

.fv4-eyebrow { color:#85857a; font-size:13.5px; font-weight:700; letter-spacing:.14em; margin-bottom:6px; }
.fv4-h2 { color:#17170f; font-size:23px; font-weight:800; margin:0 0 4px 0; }

.fv4-card { border:1px solid #e6e4da; border-radius:12px; background:#fffef9; padding:22px 24px; margin-bottom:16px; }
.fv4-badge-dark { display:inline-block; background:#17170f; color:#fdfdf7; font-size:13px; font-weight:700;
  letter-spacing:.1em; padding:4px 9px; border-radius:4px; margin-bottom:11px; }
.fv4-card-title { color:#17170f; font-size:18px; font-weight:800; margin-bottom:4px; }
.fv4-card-sub { color:#8a8a7c; font-size:14px; margin-bottom:18px; }

.fv4-funnel { display:flex; align-items:stretch; background:#f0efe6; border-radius:10px; padding:6px; }
.fv4-stage { flex:1 1 0; text-align:center; padding:18px 6px; border-radius:8px; }
.fv4-stage.hot { background:#fdeaea; box-shadow:inset 0 0 0 1px #f3cccc; }
.fv4-stage-v { color:#17170f; font-size:19px; font-weight:800; letter-spacing:-.03em;
               white-space:nowrap; }
.fv4-stage-l { color:#8a8a7c; font-size:13.5px; margin-top:6px; }
.fv4-conv { flex:0 0 62px; display:flex; align-items:center; justify-content:center; color:#a3a396; font-size:13.5px; font-weight:600; }
.fv4-conv.low { color:#c0392b; }

.fv4-tbl { width:100%; border-collapse:collapse; font-size:14px; margin-top:18px; }
.fv4-tbl th { color:#8a8a7c; font-weight:600; font-size:13.5px; text-align:right; padding:10px 12px;
  border-bottom:1px solid #e6e4da; background:#f7f6ef; white-space:nowrap; }
.fv4-tbl th:first-child, .fv4-tbl th:nth-child(2) { text-align:left; }
.fv4-tbl td { color:#26261c; text-align:right; padding:13px 12px; border-bottom:1px solid #efeee4; white-space:nowrap; }
.fv4-tbl td:first-child { text-align:left; font-weight:700; }
.fv4-tbl td:nth-child(2) { text-align:left; }
.fv4-tbl tr:hover td { background:#faf9f2; }
.fv4-mix-bar { width:120px; height:6px; border-radius:3px; background:#e6e4da; position:relative; overflow:hidden; }
.fv4-mix-bar b { position:absolute; left:0; top:0; height:100%; display:block; border-radius:3px; }
.fv4-mix-plan { background:#17170f; }
.fv4-mix-act  { background:#c8f231; }
.fv4-mix-lbl { color:#9a9a8c; font-size:13px; margin-top:5px; }

.fv4-bottom { display:flex; gap:16px; flex-wrap:wrap; }
.fv4-mixpanel { flex:2 1 420px; border:1px solid #e6e4da; border-radius:12px; background:#fffef9; padding:24px; display:flex; gap:30px; flex-wrap:wrap; }
.fv4-mix-total { flex:0 0 190px; }
.fv4-mix-amt { color:#17170f; font-size:31px; font-weight:800; letter-spacing:-.03em; margin:8px 0 6px; }
.fv4-mix-list { flex:1 1 240px; }
.fv4-mix-row { margin-bottom:11px; }
.fv4-mix-row-top { display:flex; justify-content:space-between; font-size:13.5px; margin-bottom:5px; }
.fv4-mix-row-top span:first-child { color:#26261c; font-weight:600; }
.fv4-mix-row-top span:last-child { color:#26261c; font-weight:700; }
.fv4-mix-track { height:7px; border-radius:4px; background:#eceadf; overflow:hidden; }
.fv4-mix-track i { display:block; height:100%; background:#6b5ce7; border-radius:4px; }

.fv4-nba { flex:1 1 300px; background:#ccf344; border-radius:12px; padding:24px; }
.fv4-nba-eyebrow { color:#5a6b1c; font-size:13px; font-weight:700; letter-spacing:.12em; margin-bottom:6px; }
.fv4-nba-title { color:#17170f; font-size:19px; font-weight:800; margin-bottom:16px; }
.fv4-nba-item { border-top:1px solid rgba(23,23,15,.16); padding:13px 0; display:flex; gap:12px; }
.fv4-nba-no { color:#5a6b1c; font-size:13.5px; font-weight:800; flex:none; padding-top:1px; }
.fv4-nba-h { color:#17170f; font-size:14px; font-weight:700; margin-bottom:3px; }
.fv4-nba-s { color:#4b5a1a; font-size:13.5px; line-height:1.5; }
.fv4-chg { border:1px solid #e6e4da; border-left:3px solid #17170f; border-radius:10px; background:#fffef9;
  padding:15px 18px; margin-bottom:14px; }
.fv4-chg-h { color:#8a8a7c; font-size:13px; font-weight:700; letter-spacing:.1em; margin-bottom:10px; }
.fv4-chg-item { color:#26261c; font-size:14.5px; padding:4px 0; display:flex; align-items:center; gap:9px; }
.fv4-chg-dot { width:6px; height:6px; border-radius:50%; flex:none; }
.fv4-chg-dot.down { background:#c0392b; }
.fv4-chg-dot.up { background:#2f8f46; }
.fv4-rev { border:1px solid #e6e4da; border-radius:12px; background:#fffef9; padding:20px 24px; margin-top:16px; }
.fv4-rev-item { border-top:1px solid #efeee4; padding:12px 0; }
.fv4-rev-item:first-of-type { border-top:none; }
.fv4-rev-h { color:#17170f; font-size:14.5px; font-weight:700; }
.fv4-rev-s { color:#8a8a7c; font-size:13.5px; margin-top:3px; }
.fv4-foot { display:flex; justify-content:space-between; color:#a3a396; font-size:13.5px; padding:14px 2px 0; flex-wrap:wrap; gap:8px; }
</style>
"""


def _v4_canon_channel(name) -> str:
    """어떤 소스에서 온 채널명이든 대표 채널명 하나로 모은다(계획 예산 ↔ 실제 집행 매칭용)."""
    raw = str(name).strip()
    low = raw.lower()
    for canon, keywords in FUNNEL_CANON_RULES:
        if any(kw.lower() in low for kw in keywords):
            return canon
    return raw


def _v4_num(v, unit="") -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    return f"{v:,.0f}{unit}"


def _v4_compact(v) -> str:
    """8,400,000 → 8.4M / 176,000 → 176K (퍼널 단계 숫자용)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    v = float(v)
    if abs(v) >= 1_000_000:
        return f"{v / 1_000_000:.1f}M"
    if abs(v) >= 10_000:
        return f"{v / 1_000:.0f}K"
    return f"{v:,.0f}"


def _v4_money_short(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "-"
    v = float(v)
    if abs(v) >= 100_000_000:
        return f"₩{v / 100_000_000:.1f}억"
    if abs(v) >= 1_000_000:
        return f"₩{v / 1_000_000:.1f}M"
    return f"₩{v:,.0f}"


def _v4_delta_html(cur, prev) -> str:
    """직전 동일 길이 기간 대비 증감률. 표(render_html_table)와 같은 색 규칙(▲빨강/▼파랑)."""
    if not prev or prev == 0 or pd.isna(prev):
        return ""
    pct = (cur - prev) / prev * 100
    cls = "fv4-up" if pct >= 0 else "fv4-down"
    arrow = "↗" if pct >= 0 else "↘"
    return f'<span class="fv4-kpi-delta {cls}">{arrow} {abs(pct):.1f}%</span>'


def _v4_verdict(roas: float, spend: float) -> tuple:
    """(뱃지 라벨, 뱃지 클래스) — 표본이 작으면 단정하지 않고 '판단 보류'로 둔다
    (performance-marketing-analysis 스킬의 소표본 처리 원칙)."""
    if spend < FUNNEL_MIN_SPEND:
        return "판단 보류", "hold"
    if roas >= OPS_KPI_ROAS_HIGH:
        return "증액 검토", "good"
    if roas < OPS_KPI_ROAS_LOW:
        return "효율 미달", "bad"
    return "관찰", "warn"


def _v4_ga_rows(ga_inflow, start: date, end: date):
    """GA4 유입을 (대분류 소계, 광고 매체별) 두 벌로 접는다.

    위 소계표와 아래 매체표가 같은 원본·같은 방식이어야 '광고' 줄 합계와 매체 합계가 맞는다.
    신규/재방문을 둘 다 담아두고, 어느 걸 앞에 세울지는 그리는 쪽에서 정한다.
    """
    cols = ["key", "users", "new", "ret", "signup",
            "new_conv", "new_rev", "ret_conv", "ret_rev", "conv", "rev"]
    empty = pd.DataFrame(columns=cols)
    if ga_inflow is None or ga_inflow.empty:
        return empty, empty
    g = ga_inflow.copy()
    g["report_date"] = pd.to_datetime(g["report_date"], errors="coerce").dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end)]
    if g.empty:
        return empty, empty

    need = ("users", "new_users", "returning_users",
            "signups", "conversions", "revenue",
            "new_signups", "new_conv", "new_rev", "ret_conv", "ret_rev")
    for c in need:
        if c not in g.columns:
            g[c] = 0.0
        g[c] = pd.to_numeric(g[c], errors="coerce").fillna(0).astype(float)

    # 엑셀로 올린 옛 데이터에는 유형별 구매·매출이 없다(전체 값만 있음).
    # 그럴 때만 사용자 비율로 나눠 근사치를 쓴다. GA4 API로 받은 데이터는 실제 값이라 손대지 않는다.
    # 컬럼 일부만 .loc으로 덮으면 pandas가 dtype 충돌로 막으므로, 컬럼 전체를 한 번에 바꾼다.
    split_sum = g[["new_conv", "ret_conv", "new_rev", "ret_rev"]].sum(axis=1)
    miss = (split_sum == 0) & ((g["conversions"] + g["revenue"]) > 0)
    if bool(miss.any()):
        share = (g["new_users"] / g["users"]).replace([np.inf, -np.inf], np.nan).fillna(0)
        g["new_conv"] = np.where(miss, g["conversions"] * share, g["new_conv"])
        g["ret_conv"] = np.where(miss, g["conversions"] * (1 - share), g["ret_conv"])
        g["new_rev"] = np.where(miss, g["revenue"] * share, g["new_rev"])
        g["ret_rev"] = np.where(miss, g["revenue"] * (1 - share), g["ret_rev"])
    # 화면에 세우는 '방문' 지표는 세션이다. 사용자는 날짜·매체별로 쪼개 저장한 걸 더하면
    # 같은 사람이 중복으로 세어져 GA4 보고서와 안 맞는다(더할 수 없는 지표). 세션은 더해도
    # 되고 광고 클릭과 단위가 같아서 클릭→방문 전환율도 계산된다.
    # 옛 엑셀 업로드분처럼 세션이 아예 없는 데이터만 사용자로 대체한다.
    g["_bucket"] = g.apply(classify_ga_bucket, axis=1)

    def fold(df, key):
        out = df.groupby(key, as_index=False).agg(
            users=("users", "sum"), new=("new_users", "sum"),
            ret=("returning_users", "sum"),
            signup=("new_signups", "sum") if "new_signups" in df.columns else ("signups", "sum"),
            new_conv=("new_conv", "sum"), new_rev=("new_rev", "sum"),
            ret_conv=("ret_conv", "sum"), ret_rev=("ret_rev", "sum"),
            conv=("conversions", "sum"), rev=("revenue", "sum"))
        return out.rename(columns={key: "key"})

    buckets = fold(g, "_bucket")
    ads = g[g["_bucket"] == "광고"].copy()
    ads["channel"] = ads["channel"].map(_v4_canon_channel)
    media = fold(ads, "channel").sort_values("users", ascending=False)
    return buckets, media


def _v4_actions_html(media, spend_map, is_new, start, end) -> str:
    """퍼널 화면의 NEXT BEST ACTION. 목적축(신규 발굴 / 매출 확보)에 맞춰 매체를 전부 훑는다.

    예전에는 상단 '오늘의 액션 신호' 3개만 재활용해서 매체 한둘만 나왔다.
    판단은 매체별로 내리는 거라, 모든 매체를 판정하고 그 이유를 같이 적는다.
    """
    if media is None or media.empty:
        return ""
    axis = "신규 고객 발굴" if is_new else "매출 확보"
    rev_k, conv_k, base_k = (("new_rev", "new_conv", "new") if is_new
                             else ("ret_rev", "ret_conv", "ret"))
    up, keep, down, hold, skip = [], [], [], [], []
    for _, m in media.iterrows():
        ch = m["key"]
        cost = float(spend_map.get(ch, {}).get("cost_incl_vat", 0) or 0)
        rev = float(m[rev_k] or 0)
        conv = float(m[conv_k] or 0)
        base = float(m[base_k] or 0)
        roas = (rev / cost * 100) if cost > 0 else None
        rec = (ch, cost, rev, conv, base, roas)
        if cost <= 0 and rev > 0:
            skip.append(rec)
        elif cost <= 0:
            continue
        elif cost < CP_MIN_SPEND_FOR_JUDGE:
            hold.append(rec)
        elif roas is None:
            skip.append(rec)
        elif roas >= OPS_KPI_ROAS_HIGH:
            up.append(rec)
        elif roas >= OPS_KPI_ROAS_LOW:
            keep.append(rec)
        else:
            down.append(rec)
    up.sort(key=lambda x: -(x[5] or 0))
    down.sort(key=lambda x: (x[5] or 0))

    if up and down:
        head = f"<b>{up[0][0]}</b>를 늘리고 <b>{down[0][0]}</b>를 줄이세요."
    elif up:
        head = f"<b>{up[0][0]}</b>에 예산을 더 태울 여지가 있습니다."
    elif down:
        head = f"<b>{down[0][0]}</b>를 손봐야 합니다."
    elif keep:
        head = "지금은 손댈 매체가 없습니다."
    else:
        head = "아직 판단할 만한 데이터가 없습니다."
    sub = (f"{axis} 기준 · 판단 가능 {len(up)+len(keep)+len(down)}개 · "
           f"보류 {len(hold)}개 · 제외 {len(skip)}개")

    kind_lbl = "첫구매" if is_new else "재구매"
    base_lbl = "신규" if is_new else "재방문"

    def li(rec, tag, cls, act):
        ch, cost, rev, conv, base, roas = rec
        cac = (cost / conv) if conv else None
        detail = (f'{base_lbl} {base:,.0f}명 → {kind_lbl} {conv:,.0f}건'
                  + (f' · {kind_lbl} CAC \u20a9{cac:,.0f}' if cac else ''))
        roas_s = f'{roas:,.0f}%' if roas else '—'
        return (f'<li><span class="cp-tag {cls}">{tag}</span> <b>{ch}</b> — '
                f'{axis} ROAS {roas_s} · 집행 \u20a9{cost:,.0f}<br>'
                f'<span class="cp-rec-det">{detail} → {act}</span></li>')

    items = []
    for r in up:
        move = r[1] * CP_STEP
        items.append(li(r, "증액 검토", "up",
                        f'예산 +{CP_STEP*100:.0f}% (\u20a9{move:,.0f}) 후 48시간 관찰'))
    for r in down:
        move = r[1] * CP_STEP
        items.append(li(r, "감액·점검", "down",
                        f'예산 -{CP_STEP*100:.0f}% (\u20a9{move:,.0f}) 또는 소재·타겟 점검 먼저'))
    for r in keep:
        items.append(li(r, "유지", "keep", "현 수준 유지"))
    for ch, cost, rev, conv, base, roas in hold:
        items.append(f'<li><span class="cp-tag hold">판단 보류</span> <b>{ch}</b> — '
                     f'집행 \u20a9{cost:,.0f}로 표본이 작습니다<br>'
                     f'<span class="cp-rec-det">{CP_MIN_SPEND_FOR_JUDGE:,}원 넘을 때까지 판단 유보</span></li>')
    for ch, cost, rev, conv, base, roas in skip:
        why = (f'매출 \u20a9{rev:,.0f}은 잡히는데 광고비가 없어 ROAS 계산 불가'
               if cost <= 0 else '판단 근거 부족')
        items.append(f'<li><span class="cp-tag skip">판단 제외</span> <b>{ch}</b> — '
                     f'<span class="cp-rec-det">{why}</span></li>')

    return (
        '<div class="cp-wrap"><div class="cp-rec">'
        '<div class="cp-rec-eyebrow">NEXT BEST ACTION</div>'
        f'<div class="cp-rec-head">{head}</div>'
        f'<div class="cp-rec-sub">{sub}</div>'
        f'<ul class="cp-rec-list">{"".join(items)}</ul>'
        f'<div class="cp-rec-foot">기준: {axis} ROAS 목표 {OPS_KPI_ROAS_LOW:.0f}~{OPS_KPI_ROAS_HIGH:.0f}% · '
        f'집행 {CP_MIN_SPEND_FOR_JUDGE:,}원 미만은 판단 보류 · '
        f'예산 이동은 한 번에 {CP_STEP*100:.0f}%씩 · 조회 기간 {start} ~ {end}</div>'
        '</div></div>')


def _v4_row_new(name, r, cost, extra_cls=""):
    """신규 고객 발굴 표의 한 줄.
    핵심 지표: 신규가입 수 · 신규가입 CAC · 신규 첫구매 CAC · 신규 ROAS."""
    users = float(r["users"] or 0)
    new = float(r["new"] or 0)
    sign = float(r.get("signup", 0) or 0)
    conv = float(r.get("new_conv", 0) or 0)
    rev = float(r.get("new_rev", 0) or 0)
    share = (new / users * 100) if users else 0
    sign_rate = (sign / new * 100) if new else 0
    buy_rate = (conv / new * 100) if new else 0
    cac_sign = (cost / sign) if sign else None
    cac_buy = (cost / conv) if conv else None
    roas = (rev / cost * 100) if cost > 0 else None
    NA = '<span class="fv4-na">\u2014</span>'
    roas_txt = f"{roas:,.0f}%" if roas else NA
    n = lambda v: NA if v is None else f'\u20a9{v:,.0f}'
    return (
        f'<tr class="{extra_cls}"><td class="l">{name}</td>'
        f'<td data-v="{users:.0f}">{_v4_num(users)}</td>'
        f'<td data-v="{new:.0f}">{_v4_num(new)}</td>'
        f'<td data-v="{share:.2f}"><span class="fv4-bk-share">{share:.1f}%</span></td>'
        f'<td data-v="{sign:.0f}">{_v4_num(sign)}</td>'
        f'<td data-v="{sign_rate:.3f}">{sign_rate:.2f}%</td>'
        f'<td data-v="{conv:.0f}">{_v4_num(conv)}</td>'
        f'<td data-v="{buy_rate:.3f}">{buy_rate:.2f}%</td>'
        f'<td data-v="{rev:.0f}">\u20a9{rev:,.0f}</td>'
        f'<td data-v="{cost:.0f}">\u20a9{cost:,.0f}</td>'
        f'<td data-v="{cac_sign if cac_sign else -1:.0f}">{n(cac_sign)}</td>'
        f'<td data-v="{cac_buy if cac_buy else -1:.0f}">{n(cac_buy)}</td>'
        f'<td data-v="{roas if roas else -1:.2f}">{roas_txt}</td>')


def _v4_row_ret(name, r, cost, extra_cls=""):
    """매출 확보 표의 한 줄.
    핵심 지표: 재구매 건수 · 재구매 매출 · 재구매 ROAS · 방문→재구매 전환율."""
    users = float(r["users"] or 0)
    ret = float(r["ret"] or 0)
    new = float(r["new"] or 0)
    conv = float(r.get("ret_conv", 0) or 0)
    rev = float(r.get("ret_rev", 0) or 0)
    share = (ret / users * 100) if users else 0
    buy_rate = (conv / ret * 100) if ret else 0
    aov = (rev / conv) if conv else None
    cac = (cost / conv) if conv else None
    roas = (rev / cost * 100) if cost > 0 else None
    NA = '<span class="fv4-na">\u2014</span>'
    roas_txt = f"{roas:,.0f}%" if roas else NA
    n = lambda v: NA if v is None else f'\u20a9{v:,.0f}'
    return (
        f'<tr class="{extra_cls}"><td class="l">{name}</td>'
        f'<td data-v="{users:.0f}">{_v4_num(users)}</td>'
        f'<td data-v="{ret:.0f}">{_v4_num(ret)}</td>'
        f'<td data-v="{new:.0f}">{_v4_num(new)}</td>'
        f'<td data-v="{share:.2f}"><span class="fv4-bk-share">{share:.1f}%</span></td>'
        f'<td data-v="{conv:.0f}">{_v4_num(conv)}</td>'
        f'<td data-v="{buy_rate:.3f}">{buy_rate:.2f}%</td>'
        f'<td data-v="{rev:.0f}">\u20a9{rev:,.0f}</td>'
        f'<td data-v="{aov if aov else -1:.0f}">{n(aov)}</td>'
        f'<td data-v="{cost:.0f}">\u20a9{cost:,.0f}</td>'
        f'<td data-v="{cac if cac else -1:.0f}">{n(cac)}</td>'
        f'<td data-v="{roas if roas else -1:.2f}">{roas_txt}</td>')


HEAD_NEW = ["채널", "총 방문자", "신규 방문자", "신규 비율", "회원가입", "가입률",
            "첫구매", "첫구매율", "신규 매출", "광고비", "가입 CAC", "첫구매 CAC",
            "신규 ROAS", "판정"]
HEAD_RET = ["채널", "총 방문자", "재방문자", "신규 방문자", "재방문 비율", "재구매", "재구매율",
            "재구매 매출", "객단가", "광고비", "재구매 CAC", "재구매 ROAS", "판정"]


def _v4_funnel_html(stages: list, benchmarks: dict) -> str:
    """stages: [(라벨, 값), ...] 순서대로. 앞 단계 대비 전환율을 사이에 표시하고,
    벤치마크 미달 단계는 병목(⚠)으로 하이라이트한다."""
    # 단계별 전환율과 벤치마크 대비 달성도를 먼저 구해서, '가장 많이 모자란 한 단계'만
    # 병목으로 칠한다(미달 단계를 전부 칠하면 매번 두세 칸이 빨개져서 신호가 죽는다).
    rates, ratios = {}, {}
    for i, (label, value) in enumerate(stages):
        if i == 0:
            continue
        prev_v = stages[i - 1][1]
        rates[label] = (value / prev_v * 100) if prev_v else 0
        bench = benchmarks.get(label)
        if bench:
            ratios[label] = rates[label] / bench
    worst = min(ratios, key=ratios.get) if ratios else None
    if worst is not None and ratios[worst] >= 1:
        worst = None  # 모든 단계가 벤치마크 이상이면 병목 없음

    parts = []
    for i, (label, value) in enumerate(stages):
        if i > 0:
            parts.append(
                f'<div class="fv4-conv{" low" if label == worst else ""}">{rates[label]:.1f}%</div>'
            )
        hot = label == worst
        mark = " ⚠" if hot else ""
        parts.append(
            f'<div class="fv4-stage{" hot" if hot else ""}">'
            f'<div class="fv4-stage-v">{_v4_num(value)}</div>'
            f'<div class="fv4-stage-l">{label}{mark}</div></div>'
        )
    return '<div class="fv4-funnel">' + "".join(parts) + "</div>"


def _v4_mixbar_html(plan_pct: float, act_pct: float) -> str:
    """계획 비중(검정) 위에 실제 집행 비중(라임)을 겹쳐 그린 미니 바."""
    scale = max(plan_pct, act_pct, 1e-9)
    pw = min(plan_pct / scale * 100, 100)
    aw = min(act_pct / scale * 100, 100)
    return (
        f'<div class="fv4-mix-bar"><b class="fv4-mix-plan" style="width:{pw:.1f}%;opacity:.85"></b>'
        f'<b class="fv4-mix-act" style="width:{aw:.1f}%;height:3px;top:auto;bottom:0"></b></div>'
        f'<div class="fv4-mix-lbl">{plan_pct:.1f}% / {act_pct:.1f}%</div>'
    )


# ──────────────────────────────────────────────────────────────
# 의사결정 루프 헬퍼 — "어제와 뭐가 달라졌나 / 며칠째인가 / 저번 결정은 먹혔나"
# ──────────────────────────────────────────────────────────────
# 이 블록이 하는 일은 스냅샷(오늘 숫자가 얼마다)이 아니라 '변화'를 잡아내는 것이다.
# 매일 같은 화면을 보는 사람에게 실제로 필요한 정보는 상태값이 아니라 델타라서,
# 채널별 일별 이력(GA)에서 연속 추세와 어제 대비 변화를 뽑아 문장으로 만든다.
CHANGE_MIN_USERS = 30        # 이보다 유입이 적은 채널은 변동률이 튀어서 변화 로그에서 제외
CHANGE_ALERT_PCT = 20.0      # 이 이상 움직이면 '달라진 것'으로 본다
STREAK_MIN_DAYS = 2          # 며칠 연속부터 표시할지


def _loop_daily_by_channel(gci: pd.DataFrame) -> pd.DataFrame:
    """GA 일별 데이터를 [날짜 × 매체] 단위로 합산한다(광고 매핑된 채널만)."""
    if gci is None or gci.empty:
        return pd.DataFrame(columns=["report_date", "channel", "users", "new_users", "conversions", "revenue"])
    g = gci.copy()
    g["report_date"] = pd.to_datetime(g["report_date"]).dt.date
    g = g[g["channel"].notna()]
    if g.empty:
        return pd.DataFrame(columns=["report_date", "channel", "users", "new_users", "conversions", "revenue"])
    g["channel"] = g["channel"].map(_v4_canon_channel)
    return g.groupby(["report_date", "channel"], as_index=False).agg(
        users=("users", "sum"), new_users=("new_users", "sum"),
        conversions=("conversions", "sum"), revenue=("revenue", "sum"),
    )


def _loop_streak(series: list) -> int:
    """[(날짜, 값)] 을 날짜 오름차순으로 받아, 마지막 값 기준으로 같은 방향(증가/감소)이
    며칠 연속됐는지 센다. 부호가 바뀌면 거기서 끊는다."""
    if len(series) < 2:
        return 0
    diffs = [series[i][1] - series[i - 1][1] for i in range(1, len(series))]
    if not diffs or diffs[-1] == 0:
        return 0
    sign = 1 if diffs[-1] > 0 else -1
    streak = 0
    for d in reversed(diffs):
        if (d > 0 and sign > 0) or (d < 0 and sign < 0):
            streak += 1
        else:
            break
    return streak * sign


def _loop_change_log(daily: pd.DataFrame, end: date, lookback: int = 7,
                     spend_daily: pd.DataFrame = None, universe=None) -> list:
    """마지막 날 vs 그 전날을 매체별로 한 줄씩 비교한다.

    예전에는 20% 이상 움직인 것만, 그것도 6개만 보여줬다. 그러다 보니 브랜드검색·크리테오처럼
    조용히 도는 매체는 화면에서 아예 사라져서 "오늘 얘는 어땠지"를 확인할 수가 없었다.
    운영 중인 매체는 변화가 없어도 전부 세워두고, 눈에 띄는 것만 색으로 구분한다.

    돌려주는 값: [("_meta", (오늘, 전일)), (종류, 문장), ...]
    """
    if daily is None or daily.empty:
        return []
    d = daily[daily["report_date"] <= end]
    if d.empty:
        return []
    days = sorted(d["report_date"].unique())
    if len(days) < 2:
        return []
    today_d, prev_d = days[-1], days[-2]
    cur = d[d["report_date"] == today_d].groupby("channel").sum(numeric_only=True)
    prv = d[d["report_date"] == prev_d].groupby("channel").sum(numeric_only=True)

    # 운영 매체 목록 = GA 유입이 잡힌 매체 ∪ 광고비가 나간 매체.
    # 광고비만 있고 GA 유입이 0인 매체(=UTM 누락 의심)도 반드시 보여야 해서 합집합으로 만든다.
    chans = set(cur.index) | set(prv.index)
    spend_map = {}
    if spend_daily is not None and not spend_daily.empty and "channel" in spend_daily.columns:
        sp = spend_daily.copy()
        sp["report_date"] = pd.to_datetime(sp["report_date"], errors="coerce").dt.date
        sp = sp[(sp["report_date"] >= days[max(0, len(days) - lookback)]) & (sp["report_date"] <= today_d)]
        if not sp.empty and "cost_incl_vat" in sp.columns:
            agg = sp.groupby("channel")["cost_incl_vat"].sum()
            spend_map = {_v4_canon_channel(k): float(v) for k, v in agg.items() if float(v) > 0}
            chans |= set(spend_map)
    if universe:
        chans |= {_v4_canon_channel(c) for c in universe if str(c).strip()}

    def g(df, ch, col):
        try:
            return float(df.loc[ch, col])
        except Exception:
            return 0.0

    rows = []
    for ch in chans:
        u_now, u_prev = g(cur, ch, "users"), g(prv, ch, "users")
        r_now, r_prev = g(cur, ch, "revenue"), g(prv, ch, "revenue")
        if u_now <= 0 and u_prev <= 0 and r_now <= 0 and r_prev <= 0 and ch not in spend_map:
            continue

        hist = [(dd, float(d[(d["report_date"] == dd) & (d["channel"] == ch)]["users"].sum()))
                for dd in days[-lookback:]]
        streak = _loop_streak(hist)
        streak_txt = f" · {abs(streak)}일째" if abs(streak) >= STREAK_MIN_DAYS else ""

        def delta(now, prev):
            if prev > 0:
                return (now - prev) / prev * 100
            return None if now <= 0 else float("inf")

        up_, rp_ = delta(u_now, u_prev), delta(r_now, r_prev)
        thin = max(u_now, u_prev) < CHANGE_MIN_USERS      # 표본이 작으면 변동률이 튄다

        def txt(pct):
            if pct is None:
                return "변동 없음"
            if pct == float("inf"):
                return "신규 발생"
            return f"{pct:+.0f}%"

        parts = [f'<b>{ch}</b>',
                 f'유입 {u_now:,.0f} <span class="fv4-chg-p">(전일 {u_prev:,.0f} · {txt(up_)})</span>',
                 f'매출 {r_now:,.0f}원 <span class="fv4-chg-p">(전일 {r_prev:,.0f}원 · {txt(rp_)})</span>']

        worst = min([p for p in (up_, rp_) if p is not None and p != float("inf")], default=0.0)
        best = max([p for p in (up_, rp_) if p is not None and p != float("inf")], default=0.0)
        if thin:
            kind = "flat"
            parts.append('<span class="fv4-chg-tag">표본 적음</span>')
        elif worst <= -CHANGE_ALERT_PCT:
            kind = "down"
        elif best >= CHANGE_ALERT_PCT:
            kind = "up"
        else:
            kind = "flat"

        # 광고비는 나갔는데 GA 유입이 0이면 UTM이 안 붙은 것이다 — 성과가 아니라 설정 문제다.
        if ch in spend_map and u_now <= 0 and u_prev <= 0:
            kind = "down"
            parts = [p for p in parts if "표본 적음" not in p]   # 더 중요한 경고로 덮는다
            parts.append('<span class="fv4-chg-tag warn">광고비 집행 중인데 GA 유입 0 · UTM 확인</span>')

        rows.append((kind, " · ".join(parts) + streak_txt, worst))

    order = {"down": 0, "up": 1, "flat": 2}
    rows.sort(key=lambda x: (order.get(x[0], 9), x[2]))
    return [("_meta", (today_d, prev_d))] + [(k, t) for k, t, _ in rows]


def _loop_signal_streak(daily: pd.DataFrame, channel: str, end: date, lookback: int = 7) -> str:
    """특정 채널의 유입이 며칠 연속 빠지고 있는지(또는 오르고 있는지) 짧은 문구로."""
    if daily is None or daily.empty:
        return ""
    d = daily[(daily["channel"] == channel) & (daily["report_date"] <= end)]
    if d.empty:
        return ""
    days = sorted(d["report_date"].unique())[-lookback:]
    hist = [(dd, float(d[d["report_date"] == dd]["users"].sum())) for dd in days]
    s = _loop_streak(hist)
    if abs(s) < STREAK_MIN_DAYS:
        return ""
    return f"유입 {abs(s)}일째 {'하락' if s < 0 else '상승'}"


def _loop_review_decisions(decisions: pd.DataFrame, daily: pd.DataFrame, end: date) -> list:
    """지난 결정들이 먹혔는지 되짚는다. 결정일 이전 7일 평균 vs 이후 7일 평균 유입/매출 비교.
    [(채널, 결정문구, 결정일, 판정문장)] 형태."""
    if decisions is None or decisions.empty or daily is None or daily.empty:
        return []
    out = []
    for _, row in decisions.sort_values("decided_on", ascending=False).head(5).iterrows():
        ch, when = row.get("channel"), pd.to_datetime(row.get("decided_on")).date()
        d = daily[daily["channel"] == ch]
        if d.empty:
            continue
        before = d[(d["report_date"] < when) & (d["report_date"] >= when - timedelta(days=7))]
        after = d[(d["report_date"] >= when) & (d["report_date"] <= min(end, when + timedelta(days=7)))]
        if before.empty or after.empty:
            out.append((ch, row.get("action", ""), when, "아직 결과를 볼 만큼 기간이 지나지 않았습니다."))
            continue
        b_rev, a_rev = before["revenue"].mean(), after["revenue"].mean()
        b_u, a_u = before["users"].mean(), after["users"].mean()
        rev_p = ((a_rev - b_rev) / b_rev * 100) if b_rev else 0
        usr_p = ((a_u - b_u) / b_u * 100) if b_u else 0
        verdict = "개선" if rev_p > 5 else ("악화" if rev_p < -5 else "변화 없음")
        out.append((ch, row.get("action", ""), when,
                    f"결정 후 일평균 매출 {rev_p:+.0f}%, 유입 {usr_p:+.0f}% — {verdict}"))
    return out


# ──────────────────────────────────────────────────────────────
# 광고비 소스 (신규) — API 실집행 > 대행사 주간 리포트 > 채널믹스 예산 일할
# ──────────────────────────────────────────────────────────────
# 이 대시보드의 목적은 '매일 보고 매체를 조절하는 것'인데, 매출은 GA로 어제까지 실시간인 반면
# 광고비만 주간 대행사 리포트면 서로 다른 기간을 나눈 ROAS가 나와 판단 근거가 못 된다.
# 그래서 광고비도 일 단위로 맞춘다. 매체마다 확보 난이도가 달라서 3단 폴백으로 설계했다.
#   ① API 실집행 (메타/구글 등) — 가장 정확, 인증 필요
#   ② 대행사 주간 리포트 일할 안분 — 실제 집행액이지만 주 단위
#   ③ 채널믹스 월예산 일할 — 인증 없이도 항상 값이 나오는 기본선
# 보장형 상품(네이버 맨즈탭 등)은 계약 금액이 고정이라 일별 변동이 없어서 ③이 오히려 정확하다.
AD_SPEND_SOURCE_LABEL = {
    "meta_api": "메타 API",
    "google_ads_api": "구글 API",
    "naver_api": "네이버 API",
    "kakao_api": "카카오 API",
    "criteo_api": "크리테오 API",
    "naver_gfa_api": "GFA API",
    "contract": "정액 계약(일할)",
    "manual": "직접 입력(실집행)",
    "agency_weekly": "대행사 주간(일할)",
    "budget_prorate": "예산 일할",
}
# '신규 매체'는 시기별로 실제 매체가 달랐다(사용자 확인):
#   26년 4월      = 네이버 트렌드픽 모바일
#   26년 5~8월    = 네이버 애드부스트
# 애드부스트는 GFA 계정 안에 있지만 캠페인 분류가 'ADVoost 쇼핑'으로 따로 잡히는 별도 매체다.
# GFA(웹사이트 전환)와 합쳐버리면 계획 대비 집행이 서로 섞여서 판단이 틀어지므로 분리해서 집계한다.
NEW_MEDIA_BY_MONTH = {
    4: "네이버 트렌드픽",
    5: "네이버 애드부스트", 6: "네이버 애드부스트",
    7: "네이버 애드부스트", 8: "네이버 애드부스트",
}
NEW_MEDIA_DEFAULT = "네이버 애드부스트"


def _budget_daily_spend(channel_mix: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """채널믹스 월예산을 그 달의 일수로 나눠, 선택 기간과 겹치는 날짜만큼만 채널별로 합산한다.
    보장형(맨즈탭 등)처럼 일별 변동이 없는 매체는 이 값이 실제와 가장 가깝다."""
    cols = ["channel", "cost_incl_vat"]
    if channel_mix is None or channel_mix.empty:
        return pd.DataFrame(columns=cols)
    m = channel_mix.copy()
    m["year"] = pd.to_numeric(m["year"], errors="coerce")
    m["month"] = pd.to_numeric(m["month"], errors="coerce")
    m["budget"] = pd.to_numeric(m["budget"], errors="coerce").fillna(0)
    m = m.dropna(subset=["year", "month"])

    rows = []
    for _, r in m.iterrows():
        y, mo, amt = int(r["year"]), int(r["month"]), float(r["budget"])
        if amt <= 0:
            continue
        try:
            m_start = date(y, mo, 1)
            m_end = date(y + (mo == 12), (mo % 12) + 1, 1) - timedelta(days=1)
        except ValueError:
            continue
        ov_start, ov_end = max(m_start, start), min(m_end, end)
        if ov_start > ov_end:
            continue
        days_in_month = (m_end - m_start).days + 1
        ov_days = (ov_end - ov_start).days + 1
        ch = str(r["channel"]).strip()
        if ch in ("신규 매체", "신규매체"):
            ch = NEW_MEDIA_BY_MONTH.get(mo, NEW_MEDIA_DEFAULT)
        rows.append({"channel": _v4_canon_channel(ch), "cost_incl_vat": amt * ov_days / days_in_month})
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows).groupby("channel", as_index=False).agg(cost_incl_vat=("cost_incl_vat", "sum"))


def _secrets_get(key: str, default=None):
    """st.secrets에서 최상위 값 하나를 안전하게 꺼낸다(없으면 기본값)."""
    try:
        return st.secrets[key]
    except Exception:
        return default


def _secrets_section(name: str):
    try:
        return dict(st.secrets[name])
    except Exception:
        return None


def _meta_link_clicks(row: dict) -> float:
    """광고관리자의 '링크 클릭'과 같은 값을 꺼낸다.

    insights의 clicks는 전체 클릭(좋아요·프로필 방문 등 포함)이라 화면 숫자와 안 맞는다.
    actions 배열 안의 action_type='link_click'이 화면의 링크 클릭이다.
    (계정에 따라 actions가 안 올 수도 있어 그때는 clicks로 물러난다.)
    """
    for a in row.get("actions") or []:
        if str(a.get("action_type")) == "link_click":
            try:
                return float(a.get("value") or 0)
            except Exception:
                return 0.0
    return float(row.get("clicks") or 0)


def fetch_meta_spend(start: date, end: date) -> pd.DataFrame:
    """Meta Marketing API에서 일별 광고비를 받아온다(계정 단위). 인증이 없으면 빈 결과."""
    cfg = _secrets_section("meta_ads")
    if not cfg or not cfg.get("access_token") or not cfg.get("ad_account_id"):
        return pd.DataFrame()
    import requests

    acct = str(cfg["ad_account_id"]).strip()
    if not acct.startswith("act_"):
        acct = f"act_{acct}"
    ver = str(cfg.get("api_version", "v21.0")).strip()
    url = f"https://graph.facebook.com/{ver}/{acct}/insights"
    params = {
        # clicks는 좋아요·프로필 클릭까지 포함한 '전체 클릭'이라 광고관리자 화면(링크 클릭)과
        # 숫자가 다르다. actions에서 link_click을 꺼내 쓴다.
        "fields": "spend,impressions,clicks,actions",
        "level": "account",
        "time_increment": 1,
        "time_range": json.dumps({"since": str(start), "until": str(end)}),
        "access_token": cfg["access_token"],
        "limit": 500,
    }
    rows = []
    while url:
        resp = requests.get(url, params=params, timeout=60)
        payload = resp.json()
        if "error" in payload:
            raise RuntimeError(payload["error"].get("message", str(payload["error"])))
        for d in payload.get("data", []):
            rows.append({
                "report_date": d.get("date_start"), "channel": "메타",
                # 메타도 국내는 VAT 별도 청구라 insights의 spend는 VAT 제외 금액이다.
                # 네이버·구글·크리테오·GFA와 단위를 맞추려면 1.1을 곱해야 한다.
                "cost_incl_vat": float(d.get("spend") or 0) * 1.1, "source": "meta_api",
                "impressions": float(d.get("impressions") or 0),
                "clicks": _meta_link_clicks(d),
            })
        nxt = (payload.get("paging") or {}).get("next")
        url, params = (nxt, None) if nxt else (None, None)
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce").dt.date
    return out.dropna(subset=["report_date"]).reset_index(drop=True)


# 구글애즈 API 버전 탐색 범위. 구글은 대략 분기마다 버전을 올리고 1년 남짓 지나면 폐기한다.
GOOGLE_ADS_VER_MAX = 30
GOOGLE_ADS_VER_MIN = 18


def fetch_google_ads_spend(start: date, end: date) -> pd.DataFrame:
    """Google Ads API(REST)에서 일별 광고비를 받아온다. 무거운 google-ads 라이브러리 대신
    REST + refresh token 방식을 써서 Streamlit Cloud 배포를 가볍게 유지한다."""
    cfg = _secrets_section("google_ads")
    need = ["developer_token", "client_id", "client_secret", "refresh_token", "customer_id"]
    if not cfg or any(not cfg.get(k) for k in need):
        return pd.DataFrame()
    import requests

    tok = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": cfg["client_id"], "client_secret": cfg["client_secret"],
            "refresh_token": cfg["refresh_token"], "grant_type": "refresh_token",
        },
        timeout=60,
    ).json()
    if "access_token" not in tok:
        raise RuntimeError(f"구글 토큰 발급 실패: {tok.get('error_description') or tok}")

    cid = str(cfg["customer_id"]).replace("-", "").strip()
    headers = {
        "Authorization": f"Bearer {tok['access_token']}",
        "developer-token": str(cfg["developer_token"]).strip(),
        "Content-Type": "application/json",
    }
    if cfg.get("login_customer_id"):
        headers["login-customer-id"] = str(cfg["login_customer_id"]).replace("-", "").strip()

    query = (
        "SELECT segments.date, metrics.cost_micros, metrics.impressions, metrics.clicks "
        "FROM customer "
        f"WHERE segments.date BETWEEN '{start}' AND '{end}'"
    )
    # 구글애즈 API는 버전을 URL에 박아야 하는데, 오래된 버전은 폐기되면서 404를 낸다.
    # 버전을 코드에 고정하면 몇 달마다 대시보드가 죽으므로, 최신부터 내려가며 살아있는
    # 버전을 자동으로 찾는다. Secrets에 api_version이 있으면 그것부터 시도한다.
    candidates = []
    if cfg.get("api_version"):
        candidates.append(str(cfg["api_version"]).strip())
    candidates += [f"v{n}" for n in range(GOOGLE_ADS_VER_MAX, GOOGLE_ADS_VER_MIN - 1, -1)]

    resp, used_ver, tried = None, None, []
    for ver in candidates:
        r = requests.post(
            f"https://googleads.googleapis.com/{ver}/customers/{cid}/googleAds:searchStream",
            headers=headers, json={"query": query}, timeout=90,
        )
        tried.append(ver)
        if r.status_code == 404:
            continue          # 폐기된 버전 → 다음 후보
        resp, used_ver = r, ver
        break

    if resp is None:
        raise RuntimeError(
            "구글 조회 실패: 살아있는 API 버전을 못 찾았습니다 (전부 404). "
            f"시도: {', '.join(tried[:6])}… · Secrets [google_ads]에 api_version = \"v26\" 처럼 "
            "직접 지정하면 그것부터 씁니다."
        )
    if resp.status_code >= 400:
        raise RuntimeError(f"구글 조회 실패({resp.status_code}, {used_ver}): {resp.text[:300]}")

    rows = []
    for chunk in resp.json():
        for r in chunk.get("results", []):
            d = (r.get("segments") or {}).get("date")
            m = r.get("metrics") or {}
            micros = float(m.get("costMicros") or 0)
            if d:
                # 구글은 국내에서 VAT를 별도 청구하므로 cost_micros는 VAT 제외 금액이다.
                # 다른 매체(네이버·크리테오·GFA)와 단위를 맞추려면 1.1을 곱해야 한다.
                rows.append({"report_date": d, "channel": "구글",
                             "cost_incl_vat": micros / 1_000_000 * 1.1,
                             "impressions": float(m.get("impressions") or 0),
                             "clicks": float(m.get("clicks") or 0),
                             "source": "google_ads_api"})
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce").dt.date
    out = out.dropna(subset=["report_date"])
    return out.groupby(["report_date", "channel", "source"], as_index=False).agg(
        cost_incl_vat=("cost_incl_vat", "sum"),
        impressions=("impressions", "sum"), clicks=("clicks", "sum"),
    )


# 네이버 검색광고 API — 검색광고(파워링크)/쇼핑검색/브랜드검색 일별 광고비.
# 인증은 HMAC-SHA256 서명 방식이다: "{timestamp}.{METHOD}.{path}" 를 비밀키로 서명해서 헤더에 넣는다.
# 일별 분해는 /stats 가 지원하지 않아서 날짜별로 한 번씩(하루 1콜) 호출한다.
# 캠페인 유형 → 예산 파일의 매체명 매핑:
#   WEB_SITE(파워링크) · SHOPPING(쇼핑검색) → '네이버 검색광고'  (예산 파일에 쇼핑검색이 따로 없어 합침)
#   BRAND_SEARCH                          → '네이버 브랜드검색광고'
NAVER_CAMPAIGN_TYPE_CHANNEL = {
    "WEB_SITE": "네이버 검색광고",          # 파워링크
    "SHOPPING": "네이버 쇼핑검색광고",        # 쇼핑검색 — 검색광고와 합치면 안 된다(별도 매체로 관리)
    "BRAND_SEARCH": "네이버 브랜드검색광고",
    "POWER_CONTENTS": "네이버 검색광고",
    "PLACE": "네이버 검색광고",
}
NAVER_API_BASE = "https://api.searchad.naver.com"


def _naver_headers(method: str, path: str, cfg: dict) -> dict:
    import base64, hashlib, hmac, time
    ts = str(int(time.time() * 1000))
    msg = f"{ts}.{method}.{path}"
    sig = base64.b64encode(
        hmac.new(str(cfg["secret_key"]).encode(), msg.encode(), hashlib.sha256).digest()
    ).decode()
    return {
        "X-Timestamp": ts,
        "X-API-KEY": str(cfg["api_key"]).strip(),
        "X-Customer": str(cfg["customer_id"]).strip(),
        "X-Signature": sig,
        "Content-Type": "application/json; charset=UTF-8",
    }


def _naver_campaigns(cfg: dict) -> list:
    """캠페인 목록 + 유형을 가져온다. (id, 매체명) 리스트."""
    import requests
    path = "/ncc/campaigns"
    r = requests.get(NAVER_API_BASE + path, headers=_naver_headers("GET", path, cfg), timeout=60)
    if r.status_code >= 400:
        raise RuntimeError(f"캠페인 조회 실패({r.status_code}): {r.text[:200]}")
    out = []
    for c in r.json():
        ch = NAVER_CAMPAIGN_TYPE_CHANNEL.get(str(c.get("campaignTp", "")).upper())
        if ch and c.get("nccCampaignId"):
            out.append((c["nccCampaignId"], ch))
    return out


def fetch_naver_search_spend(start: date, end: date) -> pd.DataFrame:
    """네이버 검색광고 일별 광고비.

    · salesAmt = 집행 광고비다. 이름 때문에 매출로 오해하기 쉽다.
    · salesAmt는 **VAT 포함** 금액이다 — 다른 매체처럼 1.1을 곱하면 안 된다.
    · 캠페인 유형별로 매체를 갈라 담는다(파워링크=검색광고 / SHOPPING=쇼핑검색광고 /
      BRAND_SEARCH=브랜드검색광고). 합치면 매체별 판단이 불가능해진다.
    · 브랜드검색은 정액 상품이라 salesAmt가 항상 0이고 노출·클릭만 온다.
    """
    cfg = _secrets_section("naver_searchad")
    need = ["api_key", "secret_key", "customer_id"]
    if not cfg or any(not cfg.get(k) for k in need):
        return pd.DataFrame()
    import requests, json as _json

    campaigns = _naver_campaigns(cfg)
    if not campaigns:
        return pd.DataFrame()
    ch_by_id = dict(campaigns)
    ids = [cid for cid, _ in campaigns]

    rows = []
    d = start
    while d <= end:
        path = "/stats"
        params = {
            "ids": ids,
            "fields": _json.dumps(["salesAmt", "impCnt", "clkCnt"]),
            "timeRange": _json.dumps({"since": str(d), "until": str(d)}),
        }
        r = requests.get(NAVER_API_BASE + path, headers=_naver_headers("GET", path, cfg),
                         params=params, timeout=60)
        if r.status_code >= 400:
            raise RuntimeError(f"통계 조회 실패({r.status_code}, {d}): {r.text[:200]}")
        for item in (r.json() or {}).get("data", []):
            ch = ch_by_id.get(item.get("id"))
            if not ch:
                continue
            rows.append({"report_date": d, "channel": ch,
                         # salesAmt는 이미 VAT가 포함된 금액이다(비즈머니에서 VAT 포함으로 차감됨).
                         # 예전엔 여기에 1.1을 또 곱해서 실제보다 10% 부풀려 있었다.
                         "cost_incl_vat": float(item.get("salesAmt") or 0),
                         "impressions": float(item.get("impCnt") or 0),
                         "clicks": float(item.get("clkCnt") or 0),
                         "source": "naver_api"})
        d += timedelta(days=1)

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).groupby(
        ["report_date", "channel", "source"], as_index=False
    ).agg(cost_incl_vat=("cost_incl_vat", "sum"),
          impressions=("impressions", "sum"), clicks=("clicks", "sum"))


# 카카오모먼트 API — 광고계정 보고서(일별 광고비).
#   GET https://apis.moment.kakao.com/openapi/v4/adAccounts/report
#   헤더: Authorization: Bearer {비즈니스 토큰}, adAccountId: {광고계정 ID}
#   쿼리: adAccountId, start/end(yyyyMMdd), level=AD_ACCOUNT, metricsGroup=BASIC, timeUnit=DAY
# 주의사항 (공식 문서 확인):
#   · 조회 기간은 최대 31일. 그래서 31일씩 잘라서 여러 번 호출한다.
#   · start/end 조회에서 '오늘'은 제외된다 (어차피 우리도 어제까지만 받는다).
#   · 요청 제한: 광고계정당 5초에 1회 → 호출 사이에 텀을 둔다.
#   · 비즈니스 토큰은 만료되므로 refresh_token으로 갱신해서 쓴다.
KAKAO_REPORT_URL = "https://apis.moment.kakao.com/openapi/v4/adAccounts/report"
KAKAO_TOKEN_URL = "https://kauth.kakao.com/oauth/token"
KAKAO_MAX_RANGE_DAYS = 31
KAKAO_RATE_LIMIT_SEC = 5.5


def _kakao_access_token(cfg: dict) -> str:
    """access_token이 직접 들어있으면 그대로 쓰고, refresh_token이 있으면 갱신해서 받는다."""
    import requests
    if cfg.get("refresh_token") and cfg.get("rest_api_key"):
        r = requests.post(KAKAO_TOKEN_URL, data={
            "grant_type": "refresh_token",
            "client_id": str(cfg["rest_api_key"]).strip(),
            "refresh_token": str(cfg["refresh_token"]).strip(),
            **({"client_secret": cfg["client_secret"]} if cfg.get("client_secret") else {}),
        }, timeout=60)
        j = r.json()
        if "access_token" not in j:
            raise RuntimeError(f"토큰 갱신 실패: {j.get('error_description') or j}")
        return j["access_token"]
    if cfg.get("access_token"):
        return str(cfg["access_token"]).strip()
    raise RuntimeError("access_token 또는 (rest_api_key + refresh_token)이 필요합니다")


def fetch_kakao_moment_spend(start: date, end: date) -> pd.DataFrame:
    """카카오모먼트 일별 광고비. 채널명은 예산 파일의 '카카오톡 플친'과 맞춰 '카카오'로 정규화된다."""
    cfg = _secrets_section("kakao_moment")
    if not cfg or not cfg.get("ad_account_id"):
        return pd.DataFrame()
    import requests, time

    token = _kakao_access_token(cfg)
    acct = str(cfg["ad_account_id"]).strip()
    headers = {"Authorization": f"Bearer {token}", "adAccountId": acct}

    rows = []
    chunk_start = start
    first = True
    while chunk_start <= end:
        chunk_end = min(chunk_start + timedelta(days=KAKAO_MAX_RANGE_DAYS - 1), end)
        if not first:
            time.sleep(KAKAO_RATE_LIMIT_SEC)   # 광고계정당 5초 1회 제한
        first = False
        r = requests.get(KAKAO_REPORT_URL, headers=headers, params={
            "adAccountId": acct,
            "start": chunk_start.strftime("%Y%m%d"),
            "end": chunk_end.strftime("%Y%m%d"),
            "level": "AD_ACCOUNT",
            "timeUnit": "DAY",
            "metricsGroup": "BASIC",
        }, timeout=90)
        if r.status_code >= 400:
            raise RuntimeError(f"카카오 조회 실패({r.status_code}): {r.text[:300]}")
        payload = r.json()
        if payload.get("code") not in (200, None):
            raise RuntimeError(f"카카오 오류({payload.get('code')}): {payload.get('message')}")
        for d in payload.get("data", []):
            metrics = d.get("metrics") or {}
            cost = metrics.get("cost")
            if cost is None:
                continue
            day = d.get("start")
            if not day:
                continue
            rows.append({"report_date": day, "channel": "카카오",
                         "cost_incl_vat": float(cost) * 1.1,  # 카카오 cost는 VAT 제외
                         "impressions": float(metrics.get("imp") or 0),
                         "clicks": float(metrics.get("click") or 0),
                         "source": "kakao_api"})
        chunk_start = chunk_end + timedelta(days=1)

    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce").dt.date
    out = out.dropna(subset=["report_date"])
    return out.groupby(["report_date", "channel", "source"], as_index=False).agg(
        cost_incl_vat=("cost_incl_vat", "sum"),
        impressions=("impressions", "sum"), clicks=("clicks", "sum"))


# 크리테오 Marketing Solutions API — 일별 광고비.
#   토큰: POST https://api.criteo.com/oauth2/token (grant_type=client_credentials)
#   통계: POST https://api.criteo.com/{version}/statistics/report
#         body: dimensions=[Day], metrics=[AdvertiserCost], currency, startDate/endDate
# 크리테오는 API 버전이 날짜 형식(2024-07 등)이라 바뀔 수 있어서 Secrets에서 바꿀 수 있게 뒀다.
# 버전이 안 맞으면 404가 나는데, 진단 패널에 그대로 표시되므로 거기서 바로 확인하면 된다.
CRITEO_TOKEN_URL = "https://api.criteo.com/oauth2/token"
CRITEO_BASE = "https://api.criteo.com"
# 크리테오는 API 버전이 날짜(2026-07 등)라서 몇 달마다 올라가고 옛 버전은 404가 된다.
# 구글과 같은 이유로 버전을 코드에 고정하지 않고, 최근 것부터 살아있는 버전을 찾아 쓴다.
CRITEO_VERSION_CANDIDATES = [
    "2026-10", "2026-07", "2026-04", "2026-01",
    "2025-10", "2025-07", "2025-04", "2025-01", "2024-07",
]


def _criteo_token(cfg: dict) -> str:
    import requests
    tok = requests.post(CRITEO_TOKEN_URL, data={
        "grant_type": "client_credentials",
        "client_id": str(cfg["client_id"]).strip(),
        "client_secret": str(cfg["client_secret"]).strip(),
    }, headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=60).json()
    if "access_token" not in tok:
        raise RuntimeError(f"크리테오 토큰 발급 실패: {tok.get('error_description') or tok}")
    return tok["access_token"]


def _criteo_advertiser_ids(token: str, ver: str) -> str:
    """Secrets에 advertiser_id를 안 넣어도 되게, 내 계정에 붙은 광고주 ID를 스스로 찾는다.
    (statistics/report는 advertiserIds가 필수 항목이다.)"""
    import requests
    r = requests.get(f"{CRITEO_BASE}/{ver}/advertisers/me",
                     headers={"Authorization": f"Bearer {token}"}, timeout=60)
    if r.status_code >= 400:
        return ""
    try:
        payload = r.json() or {}
    except Exception:
        return ""
    data = payload.get("data") if isinstance(payload, dict) else payload
    if isinstance(data, dict):
        data = [data]
    ids = []
    for it in data or []:
        v = it.get("id") if isinstance(it, dict) else None
        if v is None and isinstance(it, dict):
            v = (it.get("attributes") or {}).get("advertiserId")
        if v is not None:
            ids.append(str(v))
    return ",".join(ids)


def fetch_criteo_spend(start: date, end: date) -> pd.DataFrame:
    cfg = _secrets_section("criteo")
    if not cfg or not cfg.get("client_id") or not cfg.get("client_secret"):
        return pd.DataFrame()
    import requests

    token = _criteo_token(cfg)

    candidates = []
    if cfg.get("api_version"):
        candidates.append(str(cfg["api_version"]).strip())
    candidates += CRITEO_VERSION_CANDIDATES

    adv = str(cfg.get("advertiser_id", "") or "").strip()
    payload, used_ver, last_err = None, None, None
    for ver in candidates:
        adv_ids = adv or _criteo_advertiser_ids(token, ver)
        body = {
            "advertiserIds": adv_ids,
            "dimensions": ["Day"],
            "metrics": ["AdvertiserCost", "Displays", "Clicks"],
            "currency": str(cfg.get("currency", "KRW")).strip(),
            "startDate": str(start),
            "endDate": str(end),
            "format": "json",
        }
        r = requests.post(
            f"{CRITEO_BASE}/{ver}/statistics/report",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=body, timeout=90,
        )
        if r.status_code == 404:
            continue                      # 폐기된 버전 → 다음 후보
        if r.status_code >= 400:
            last_err = f"({r.status_code}, {ver}) {r.text[:250]}"
            continue
        try:
            payload = r.json()
        except Exception:
            last_err = f"({ver}) 응답 해석 실패: {r.text[:200]}"
            continue
        used_ver = ver
        break

    if used_ver is None:
        raise RuntimeError(
            "크리테오 조회 실패: " + (last_err or "살아있는 API 버전을 못 찾았습니다(전부 404)")
            + " · Secrets [criteo]에 api_version = \"2026-07\" 처럼 직접 지정할 수 있습니다."
        )

    rows_raw = payload.get("Rows") or payload.get("rows") if isinstance(payload, dict) else payload
    if not rows_raw:
        return pd.DataFrame()

    rows = []
    for it in rows_raw:
        if not isinstance(it, dict):
            continue
        day = it.get("Day") or it.get("day") or it.get("date")
        cost = it.get("AdvertiserCost", it.get("advertiserCost", it.get("advertisercost")))
        if day is None or cost is None:
            continue
        try:
            cost = float(str(cost).replace(",", ""))
        except Exception:
            continue
        def _num(*keys):
            for k in keys:
                v = it.get(k)
                if v is not None:
                    try:
                        return float(str(v).replace(",", ""))
                    except Exception:
                        pass
            return 0.0
        rows.append({"report_date": day, "channel": "크리테오",
                     "cost_incl_vat": cost * 1.1,  # 크리테오 AdvertiserCost는 VAT 제외
                     "impressions": _num("Displays", "displays"),
                     "clicks": _num("Clicks", "clicks"),
                     "source": "criteo_api"})
    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce").dt.date
    out = out.dropna(subset=["report_date"])
    return out.groupby(["report_date", "channel", "source"], as_index=False).agg(
        cost_incl_vat=("cost_incl_vat", "sum"),
        impressions=("impressions", "sum"), clicks=("clicks", "sum"))


# 네이버 성과형 디스플레이(GFA) API — 캠페인 단위 일별 광고비.
#   GET https://openapi.naver.com/v1/ad-api/{version}/adAccounts/{adAccountNo}/performance/past/campaigns
#   헤더: Authorization: Bearer {네이버 로그인 액세스 토큰}
#         AccessManagerAccountNo: {관리계정 번호}  ← 대행사(관리계정) 통해 접근할 때 필수
#   쿼리: startDate/endDate(yyyy-MM-dd), timeUnit=daily, limit, next(페이징 토큰)
#
# 공식 문서 확인 사항:
#   · 조회 기간 최대 31일 → 31일씩 잘라서 호출
#   · 전일 데이터는 금일 02시부터 조회 가능
#   · 한 번에 최대 1000행, 초과하면 응답의 next 토큰으로 페이징 (토큰 10분 유효)
#   · API 사용 신청은 네이버 공식 파트너사만 가능 → 대행사가 발급받은 토큰을 넣어 쓰는 구조
#
# GFA 계정 안에 GFA(웹사이트 전환)와 애드부스트(ADVoost 쇼핑)가 같이 있어서, 캠페인명/목적으로
# 두 매체를 갈라서 집계한다(사용자 확인 사항). 합치면 계획 대비 집행이 섞여 판단이 틀어진다.
GFA_API_BASE = "https://openapi.naver.com/v1/ad-api"
GFA_MAX_RANGE_DAYS = 31
GFA_TOKEN_URL = "https://nid.naver.com/oauth2.0/token"


def _gfa_access_token(cfg: dict) -> str:
    """access_token이 직접 있으면 그대로 쓰고, refresh_token이 있으면 갱신해서 받는다.
    네이버 액세스 토큰은 만료되므로 refresh_token 방식을 권장."""
    import requests
    if cfg.get("refresh_token") and cfg.get("client_id") and cfg.get("client_secret"):
        r = requests.get(GFA_TOKEN_URL, params={
            "grant_type": "refresh_token",
            "client_id": str(cfg["client_id"]).strip(),
            "client_secret": str(cfg["client_secret"]).strip(),
            "refresh_token": str(cfg["refresh_token"]).strip(),
        }, timeout=60)
        j = r.json()
        if "access_token" not in j:
            raise RuntimeError(f"네이버 토큰 갱신 실패: {j.get('error_description') or j}")
        return j["access_token"]
    if cfg.get("access_token"):
        return str(cfg["access_token"]).strip()
    raise RuntimeError("access_token 또는 (client_id + client_secret + refresh_token)이 필요합니다")


def _gfa_channel_of(row: dict) -> str:
    """캠페인 정보로 'GFA'인지 '애드부스트'인지 가른다.
    애드부스트는 캠페인 목적이 ADVoost 쇼핑 계열이라 이름/목적 어디든 걸리게 본다."""
    blob = " ".join(
        str(row.get(k, "")) for k in
        ("campaignName", "campaignObjective", "campaignObjectiveType", "campaignTp", "name")
    ).lower()
    if any(kw in blob for kw in ["advoost", "adboost", "애드부스트", "쇼핑"]):
        return "네이버 애드부스트"
    return "네이버 GFA"


def fetch_naver_gfa_spend(start: date, end: date) -> pd.DataFrame:
    """GFA 일별 광고비. 관리계정 번호는 여러 개를 콤마로 넣을 수 있고, 되는 걸 자동으로 찾는다.

    GFA 계정 구조가 '대표 관리계정 → 담당 관리계정 → 광고계정' 3단이라, AccessManagerAccountNo에
    어느 번호를 넣어야 하는지는 내 네이버 계정이 어느 관리계정의 멤버냐에 따라 달라진다.
    문서상으로는 '광고계정을 직접 품고 있는 관리계정'이 맞지만, 권한 부여 방식에 따라
    대표 관리계정이어야 하는 경우도 있어서 순서대로 시도하고 성공한 것을 쓴다.
    """
    cfg = _secrets_section("naver_gfa")
    if not cfg or not cfg.get("ad_account_no"):
        return pd.DataFrame()
    import requests

    token = _gfa_access_token(cfg)
    acct = str(cfg["ad_account_no"]).strip()
    ver = str(cfg.get("api_version", "1.0")).strip()
    url = f"{GFA_API_BASE}/{ver}/adAccounts/{acct}/performance/past/campaigns"

    mgr_raw = str(cfg.get("manager_account_no", "") or "").replace("/", ",").replace(" ", ",")
    candidates = [m.strip() for m in mgr_raw.split(",") if m.strip()]
    candidates.append(None)  # 헤더 없이(광고계정 직접 멤버인 경우)도 마지막에 시도

    def _pull(mgr):
        headers = {"Authorization": f"Bearer {token}"}
        if mgr:
            headers["AccessManagerAccountNo"] = mgr
        rows = []
        chunk_start = start
        while chunk_start <= end:
            chunk_end = min(chunk_start + timedelta(days=GFA_MAX_RANGE_DAYS - 1), end)
            next_token = None
            while True:
                params = {"startDate": str(chunk_start), "endDate": str(chunk_end),
                          "timeUnit": "daily", "limit": 1000}
                if next_token:
                    params["next"] = next_token
                r = requests.get(url, headers=headers, params=params, timeout=90)
                if r.status_code >= 400:
                    raise RuntimeError(f"GFA 조회 실패({r.status_code}): {r.text[:300]}")
                payload = r.json() or {}
                for it in payload.get("rows", []):
                    day = it.get("date") or it.get("statDate") or it.get("day")
                    cost = it.get("cost", it.get("salesAmt", it.get("spend")))
                    if day is None or cost is None:
                        continue
                    rows.append({"report_date": day, "channel": _gfa_channel_of(it),
                                 # GFA 리포트의 '총비용'이 VAT 포함 금액인 것을 실데이터로 확인했다
                                 # (검색광고 salesAmt와 동일). 1.1을 곱하면 10% 부풀려진다.
                                 "cost_incl_vat": float(cost),
                                 "impressions": float(it.get("imp") or it.get("impression") or 0),
                                 "clicks": float(it.get("click") or it.get("clk") or 0),
                                 "source": "naver_gfa_api"})
                next_token = payload.get("next")
                if not next_token:
                    break
            chunk_start = chunk_end + timedelta(days=1)
        return rows

    rows, last_err, used = [], None, None
    for mgr in candidates:
        try:
            rows = _pull(mgr)
            used = mgr
            break
        except Exception as ex:
            last_err = ex
            continue
    if used is None and last_err is not None:
        raise RuntimeError(
            f"{last_err} · 시도한 관리계정 번호: "
            + ", ".join(str(c) for c in candidates if c) + " (모두 실패)"
        )

    if not rows:
        return pd.DataFrame()
    out = pd.DataFrame(rows)
    out["report_date"] = pd.to_datetime(out["report_date"], errors="coerce").dt.date
    out = out.dropna(subset=["report_date"])
    out = out.groupby(["report_date", "channel", "source"], as_index=False).agg(
        cost_incl_vat=("cost_incl_vat", "sum"))
    out.attrs["gfa_manager_used"] = used
    return out


NAVER_AUTH_URL = "https://nid.naver.com/oauth2.0/authorize"


def render_gfa_token_helper():
    """네이버 GFA용 refresh token을 대시보드 안에서 발급받는 도우미.

    네이버 로그인은 외부 사이트로 나갔다 돌아오기 때문에 그 순간 Streamlit 세션이 새로
    시작된다(입력칸이 비워짐). 그래서 client_id/secret/callback은 Secrets에서 읽어
    자동으로 채워지게 하고, 돌아왔을 때 URL에 붙은 code만 있으면 바로 교환할 수 있게 한다.
    """
    from urllib.parse import urlencode

    cfg = _secrets_section("naver_gfa") or {}

    # 구글 토큰 발급도 같은 주소로 code를 돌려주기 때문에, state로 내 것인지 확인해야
    # 서로의 코드를 잡아채지 않는다.
    code = None
    qstate = "stcogfa"
    try:
        qstate = st.query_params.get("state") or ""
        code = st.query_params.get("code") if qstate == "stcogfa" else None
    except Exception:
        code = None

    has_secrets = bool(cfg.get("client_id") and cfg.get("client_secret") and cfg.get("callback_url"))
    if not has_secrets:
        st.warning(
            "**먼저 Secrets에 아래 3줄을 넣어주세요.** 네이버 로그인은 외부 사이트로 나갔다 "
            "돌아오는 방식이라, 그 사이 화면에 직접 입력한 값은 지워집니다. "
            "Secrets에 있으면 돌아왔을 때 자동으로 채워집니다."
        )
        st.code(
            '[naver_gfa]\n'
            'client_id = "네이버 개발자센터 Client ID"\n'
            'client_secret = "네이버 개발자센터 Client Secret"\n'
            'callback_url = "개발자센터에 등록한 Callback URL"',
            language="toml",
        )

    c1, c2 = st.columns(2)
    with c1:
        cid = st.text_input("Client ID", value=str(cfg.get("client_id", "")), key="gfa_cid")
    with c2:
        csec = st.text_input("Client Secret", value=str(cfg.get("client_secret", "")),
                             type="password", key="gfa_csec")
    cb = st.text_input(
        "Callback URL (개발자센터에 등록한 값과 글자 하나까지 같아야 합니다)",
        value=str(cfg.get("callback_url", "")), key="gfa_cb",
        placeholder="https://xxxx.streamlit.app",
    )

    cid, csec, cb = str(cid).strip(), str(csec).strip(), str(cb).strip()

    # ── 돌아온 상태: URL에 code가 있으면 교환부터 처리한다 ────────────────────
    if code:
        st.success("네이버 로그인 코드가 확인됐습니다.")
        if not (cid and csec):
            st.error(
                "Client ID / Secret이 비어 있어서 교환을 못 합니다. "
                "Secrets에 넣고 다시 시도하시거나, 위 칸에 직접 채운 뒤 아래 버튼을 눌러주세요."
            )
        elif st.button("② refresh token 발급", key="gfa_token_btn", type="primary"):
            import requests
            try:
                r = requests.get(GFA_TOKEN_URL, params={
                    "grant_type": "authorization_code",
                    "client_id": cid, "client_secret": csec,
                    "code": str(code).strip(), "state": str(qstate).strip(),
                }, timeout=60)
                j = r.json() or {}
            except Exception as e:
                st.error(f"요청 실패: {e}")
                return
            if not j.get("refresh_token"):
                st.error(
                    "발급 실패: " + str(j.get("error_description") or j.get("error") or j)[:300]
                    + " · 인증 코드는 1회용이라, 이미 썼거나 시간이 지났으면 아래 '처음부터 다시'를 "
                    "누르고 ①부터 하세요."
                )
            else:
                st.success("발급 완료 — 아래 블록을 Streamlit Secrets의 [naver_gfa] 부분에 반영하세요.")
                st.code(
                    "[naver_gfa]\n"
                    f'client_id = "{cid}"\n'
                    f'client_secret = "{csec}"\n'
                    f'callback_url = "{cb}"\n'
                    f'refresh_token = "{j["refresh_token"]}"\n'
                    'ad_account_no = "대행사에서 받은 STCO 광고계정 번호"\n'
                    'manager_account_no = "대행사 관리계정 번호"',
                    language="toml",
                )
                st.caption("⚠️ 실제 비밀값이 들어 있습니다. 복사만 하고 캡처는 피해주세요.")
        if st.button("처음부터 다시", key="gfa_reset_btn"):
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()
        return

    # ── 시작 상태: 로그인 링크 만들기 ────────────────────────────────────────
    if not (cid and csec and cb):
        st.info("Client ID · Client Secret · Callback URL 3개가 모두 채워지면 로그인 링크가 나타납니다.")
        return

    auth_url = NAVER_AUTH_URL + "?" + urlencode({
        "response_type": "code", "client_id": cid,
        "redirect_uri": cb, "state": "stcogfa",
    })
    st.markdown(
        f"### ① [네이버로 로그인하고 코드 받기]({auth_url})\n"
        "위 링크를 **같은 탭에서** 클릭해 로그인·동의하면 대시보드로 돌아옵니다. "
        "돌아온 뒤 다시 이 패널을 펼치면 ② 발급 버튼이 보입니다."
    )
    st.caption(
        "돌아왔는데 ② 버튼이 안 보이면, 주소창에 `?code=` 가 붙어 있는지 확인해주세요. "
        "없으면 Callback URL이 개발자센터 등록값과 다른 경우입니다."
    )


AD_SPEND_FETCHERS = [
    ("메타", fetch_meta_spend),
    ("구글", fetch_google_ads_spend),
    ("네이버 검색광고", fetch_naver_search_spend),
    ("카카오", fetch_kakao_moment_spend),
    ("크리테오", fetch_criteo_spend),
    ("네이버 GFA", fetch_naver_gfa_spend),
]
AD_SPEND_LOOKBACK_DAYS = 30
# 매체 수치는 나중에 보정된다. 크리테오는 '클릭 후 7일' 어트리뷰션이라 최소 일주일은 값이 계속
# 움직이고, 다른 매체도 무효 트래픽을 걸러내며 비용이 소폭 바뀐다. 예전엔 3일만 다시 받아서
# 나흘 전 데이터가 낡은 채로 굳어 있었다.
AD_SPEND_RESYNC_TAIL_DAYS = 8


def sync_ad_spend(existing: pd.DataFrame):
    """매체별로 일별 광고비를 받아 ad_spend_daily에 upsert한다. 한 매체가 실패해도 나머지는
    계속 진행한다 — 하나 막히면 전체가 서는 구조를 만들지 않기 위함."""
    today = date.today()
    end = today - timedelta(days=1)
    if existing is None or existing.empty or "report_date" not in existing.columns:
        start = end - timedelta(days=AD_SPEND_LOOKBACK_DAYS - 1)
    else:
        last = pd.to_datetime(existing["report_date"]).max().date()
        start = min(last - timedelta(days=AD_SPEND_RESYNC_TAIL_DAYS - 1), end)
    if start > end:
        return 0, {}, {}

    saved, errors = {}, {}
    total = 0
    for label, fn in AD_SPEND_FETCHERS:
        try:
            df = fn(start, end)
        except Exception as e:
            errors[label] = str(e)[:250]
            continue
        if df is None or df.empty:
            continue
        n = save_table("ad_spend_daily", df, "report_date,channel,source", f"{label} API")
        saved[label] = n
        total += n
    return total, saved, errors


KAKAO_AUTH_URL = "https://kauth.kakao.com/oauth/authorize"
KAKAO_MOMENT_SCOPE = "moment_management"


def render_kakao_token_helper():
    """카카오모먼트 refresh token 발급 도우미. GFA·구글과 같은 구조.

    카카오는 client_id 자리에 'REST API 키'를 넣고, 비즈니스 동의항목(moment_management)을
    scope로 요청해야 광고 리포트를 읽을 수 있는 비즈니스 토큰이 나온다.
    """
    from urllib.parse import urlencode

    cfg = _secrets_section("kakao_moment") or {}
    st.caption(
        "카카오 개발자센터의 REST API 키로 refresh token을 받습니다. "
        "카카오모먼트 API 권한(moment_management)이 부여된 뒤에 해야 정상 발급됩니다."
    )

    code = None
    try:
        qstate = st.query_params.get("state") or ""
        code = st.query_params.get("code") if qstate == "stcokakao" else None
    except Exception:
        code = None

    if not (cfg.get("rest_api_key") and cfg.get("callback_url")):
        st.warning("**먼저 Secrets에 아래를 넣어주세요.** 카카오 로그인도 외부로 나갔다 돌아오는 방식입니다.")
        st.code(
            '[kakao_moment]\n'
            'rest_api_key = "카카오 개발자센터 → 앱 키 → REST API 키"\n'
            'client_secret = "카카오 로그인 → 보안 → Client Secret (안 켰으면 이 줄 생략)"\n'
            'callback_url = "https://stco-performance-dashboard.streamlit.app"\n'
            'ad_account_id = "카카오모먼트 광고계정 ID"',
            language="toml",
        )

    c1, c2 = st.columns(2)
    with c1:
        rk = st.text_input("REST API 키", value=str(cfg.get("rest_api_key", "")), key="kko_rk")
    with c2:
        csec = st.text_input("Client Secret (선택)", value=str(cfg.get("client_secret", "")),
                             type="password", key="kko_cs")
    cb = st.text_input("Redirect URI (카카오 로그인에 등록한 값과 같아야 합니다)",
                       value=str(cfg.get("callback_url", "")), key="kko_cb",
                       placeholder="https://xxxx.streamlit.app")
    rk, csec, cb = str(rk).strip(), str(csec).strip(), str(cb).strip()

    if code:
        st.success("카카오 인증 코드가 확인됐습니다.")
        if not (rk and cb):
            st.error("REST API 키와 Redirect URI가 있어야 교환할 수 있습니다.")
        elif st.button("② refresh token 발급", key="kko_token_btn", type="primary"):
            import requests
            data = {
                "grant_type": "authorization_code",
                "client_id": rk, "redirect_uri": cb, "code": str(code).strip(),
            }
            if csec:
                data["client_secret"] = csec
            try:
                j = requests.post(KAKAO_TOKEN_URL, data=data, timeout=60).json() or {}
            except Exception as e:
                st.error(f"요청 실패: {e}")
                return
            if not j.get("refresh_token"):
                st.error("발급 실패: " + str(j.get("error_description") or j.get("error") or j)[:300])
            else:
                st.success("발급 완료 — 아래를 Secrets의 [kakao_moment]에 반영하세요.")
                lines = ["[kakao_moment]", f'rest_api_key = "{rk}"']
                if csec:
                    lines.append(f'client_secret = "{csec}"')
                lines += [
                    f'callback_url = "{cb}"',
                    f'refresh_token = "{j["refresh_token"]}"',
                    f'ad_account_id = "{str(cfg.get("ad_account_id", "카카오모먼트 광고계정 ID"))}"',
                ]
                st.code("\n".join(lines), language="toml")
                st.caption("⚠️ 실제 비밀값이 들어 있습니다. 복사만 하고 캡처는 피해주세요.")
        if st.button("처음부터 다시", key="kko_reset_btn"):
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()
        return

    if not (rk and cb):
        st.info("REST API 키와 Redirect URI가 채워지면 로그인 링크가 나타납니다.")
        return

    auth_url = KAKAO_AUTH_URL + "?" + urlencode({
        "response_type": "code", "client_id": rk, "redirect_uri": cb,
        "scope": KAKAO_MOMENT_SCOPE, "state": "stcokakao",
    })
    st.markdown(
        f"### ① [카카오로 로그인하고 코드 받기]({auth_url})\n"
        "**카카오모먼트 광고계정에 권한이 있는 카카오 계정**으로 로그인하세요."
    )
    st.caption(
        "`KOE205`(등록되지 않은 scope) 오류가 나면 아직 moment_management 권한이 부여되지 않은 상태입니다."
    )


GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_ADS_SCOPE = "https://www.googleapis.com/auth/adwords"


def render_google_token_helper():
    """구글애즈 refresh token 발급 도우미. 구조는 GFA 쪽과 같다.

    구글은 access_type=offline + prompt=consent를 같이 줘야 refresh_token을 돌려준다.
    (한 번 동의한 계정에 prompt=consent 없이 다시 요청하면 access_token만 오고
     refresh_token이 빠져서, '왜 안 나오지' 하고 헤매기 쉬운 지점이다.)
    """
    from urllib.parse import urlencode

    cfg = _secrets_section("google_ads") or {}
    st.caption(
        "구글 클라우드에서 만든 OAuth 클라이언트 정보로 refresh token을 받습니다. "
        "개발자 토큰이 '탐색기 액세스'면 이 토큰만 있으면 바로 실계정 조회가 됩니다."
    )

    code = None
    try:
        qstate = st.query_params.get("state") or ""
        code = st.query_params.get("code") if qstate == "stcogads" else None
    except Exception:
        code = None

    has_secrets = bool(cfg.get("client_id") and cfg.get("client_secret") and cfg.get("callback_url"))
    if not has_secrets:
        st.warning(
            "**먼저 Secrets에 아래를 넣어주세요.** 구글 로그인은 외부로 나갔다 돌아오는 방식이라, "
            "화면에 직접 입력한 값은 그 사이 지워집니다."
        )
        st.code(
            '[google_ads]\n'
            'developer_token = "API 센터의 개발자 토큰"\n'
            'client_id = "OAuth 클라이언트 ID"\n'
            'client_secret = "OAuth 클라이언트 보안 비밀번호"\n'
            'callback_url = "https://stco-performance-dashboard.streamlit.app"\n'
            'customer_id = "하이픈 뺀 광고계정 번호"\n'
            'login_customer_id = "하이픈 뺀 MCC 번호"',
            language="toml",
        )

    c1, c2 = st.columns(2)
    with c1:
        cid = st.text_input("OAuth 클라이언트 ID", value=str(cfg.get("client_id", "")), key="gads_cid")
    with c2:
        csec = st.text_input("클라이언트 보안 비밀번호", value=str(cfg.get("client_secret", "")),
                             type="password", key="gads_csec")
    cb = st.text_input("승인된 리디렉션 URI (구글 클라우드에 등록한 값과 정확히 같아야 합니다)",
                       value=str(cfg.get("callback_url", "")), key="gads_cb",
                       placeholder="https://xxxx.streamlit.app")
    cid, csec, cb = str(cid).strip(), str(csec).strip(), str(cb).strip()

    if code:
        st.success("구글 인증 코드가 확인됐습니다.")
        if not (cid and csec and cb):
            st.error("클라이언트 ID/Secret/리디렉션 URI가 있어야 교환할 수 있습니다.")
        elif st.button("② refresh token 발급", key="gads_token_btn", type="primary"):
            import requests
            try:
                r = requests.post(GOOGLE_TOKEN_URL, data={
                    "grant_type": "authorization_code",
                    "code": str(code).strip(),
                    "client_id": cid, "client_secret": csec,
                    "redirect_uri": cb,
                }, timeout=60)
                j = r.json() or {}
            except Exception as e:
                st.error(f"요청 실패: {e}")
                return
            if not j.get("refresh_token"):
                st.error(
                    "발급 실패: " + str(j.get("error_description") or j.get("error") or j)[:300]
                    + " · refresh_token이 안 왔다면 이미 동의한 계정이라 그렇습니다. "
                    "'처음부터 다시' 후 ①을 다시 눌러주세요(동의 화면을 강제로 띄웁니다)."
                )
            else:
                st.success("발급 완료 — 아래를 Secrets의 [google_ads]에 반영하세요.")
                st.code(
                    "[google_ads]\n"
                    f'developer_token = "{str(cfg.get("developer_token", "API 센터의 개발자 토큰"))}"\n'
                    f'client_id = "{cid}"\n'
                    f'client_secret = "{csec}"\n'
                    f'callback_url = "{cb}"\n'
                    f'refresh_token = "{j["refresh_token"]}"\n'
                    f'customer_id = "{str(cfg.get("customer_id", "하이픈 뺀 광고계정 번호"))}"\n'
                    f'login_customer_id = "{str(cfg.get("login_customer_id", "하이픈 뺀 MCC 번호"))}"',
                    language="toml",
                )
                st.caption("⚠️ 실제 비밀값이 들어 있습니다. 복사만 하고 캡처는 피해주세요.")
        if st.button("처음부터 다시", key="gads_reset_btn"):
            try:
                st.query_params.clear()
            except Exception:
                pass
            st.rerun()
        return

    if not (cid and csec and cb):
        st.info("클라이언트 ID · 보안 비밀번호 · 리디렉션 URI 3개가 채워지면 로그인 링크가 나타납니다.")
        return

    auth_url = GOOGLE_AUTH_URL + "?" + urlencode({
        "response_type": "code",
        "client_id": cid,
        "redirect_uri": cb,
        "scope": GOOGLE_ADS_SCOPE,
        "access_type": "offline",
        "prompt": "consent",
        "state": "stcogads",
    })
    st.markdown(
        f"### ① [구글로 로그인하고 코드 받기]({auth_url})\n"
        "**구글애즈 관리자 권한이 있는 계정**으로 로그인하세요. "
        "'확인되지 않은 앱' 경고가 나오면 **고급 → (앱 이름)(으)로 이동**을 누르시면 됩니다."
    )
    st.caption("돌아온 뒤 이 패널을 다시 펼치면 ② 발급 버튼이 보입니다.")


MANUAL_SPEND_CHANNELS = [
    "네이버 GFA", "네이버 애드부스트", "네이버 맨즈탭", "카카오톡 플친",
    "카카오", "크리테오", "구글", "메타", "네이버 검색광고",
    "네이버 브랜드검색광고", "네이버 쇼핑검색광고", "모비온", "AEDI",
]


def _parse_manual_spend_file(file) -> pd.DataFrame:
    """네이버 광고주센터 '결제 > 기간별 내역' 다운로드 파일처럼 '날짜 + 금액' 두 열만 있으면
    형태가 조금 달라도 읽어낸다. 헤더 위치가 제각각이라 컬럼명을 넓게 매칭한다."""
    name = getattr(file, "name", "")
    if name.lower().endswith(".csv"):
        raw = pd.read_csv(file, header=None, dtype=str)
    else:
        raw = pd.read_excel(file, header=None, dtype=str)

    DATE_KEYS = ["기간", "날짜", "일자", "date", "day"]
    COST_KEYS = ["소진액", "소진", "광고비", "총비용", "비용", "지출", "cost", "spend"]

    hdr = None
    for i in range(min(len(raw), 30)):
        cells = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
        if any(any(k in c for k in DATE_KEYS) for c in cells) and \
           any(any(k in c for k in COST_KEYS) for c in cells):
            hdr = i
            break
    if hdr is None:
        raise RuntimeError("날짜 열과 금액 열을 못 찾았습니다 — 파일 안에 '기간/날짜'와 '소진액/광고비' 같은 머리글이 있어야 합니다.")

    df = raw.iloc[hdr + 1:].copy()
    df.columns = [str(v).strip() for v in raw.iloc[hdr].tolist()]
    lower = {c: str(c).lower() for c in df.columns}
    dcol = next((c for c in df.columns if any(k in lower[c] for k in DATE_KEYS)), None)
    ccol = next((c for c in df.columns if any(k in lower[c] for k in COST_KEYS)), None)

    out = pd.DataFrame({
        "report_date": pd.to_datetime(df[dcol], errors="coerce").dt.date,
        "cost_incl_vat": pd.to_numeric(
            df[ccol].astype(str).str.replace(r"[^0-9.\-]", "", regex=True), errors="coerce"),
    })
    out = out.dropna(subset=["report_date", "cost_incl_vat"])
    out = out[out["cost_incl_vat"] > 0]
    if out.empty:
        raise RuntimeError("읽을 수 있는 행이 없습니다 — 날짜/금액 형식을 확인해주세요.")
    return out.groupby("report_date", as_index=False).agg(cost_incl_vat=("cost_incl_vat", "sum"))


def render_manual_spend_panel(ad_spend: pd.DataFrame):
    """API가 막힌 매체(GFA·애드부스트·맨즈탭 등)의 광고비를 직접 넣는 곳.

    API로 못 가져오는 매체가 있다는 이유로 대시보드 전체를 계획값으로 두면 판단이 흐려진다.
    직접 넣은 값은 '실집행'으로 취급해서 예산 일할보다 위에 놓되, 출처를 '직접 입력'으로
    화면에 남겨 API 수치와 섞이지 않게 한다.
    """
    st.caption(
        "네이버 GFA·애드부스트처럼 API가 막힌 매체의 실집행액을 넣는 곳입니다. "
        "여기 넣은 값은 예산 일할값을 덮어쓰고, 화면에는 '직접 입력(실집행)'으로 표시됩니다."
    )

    existing = pd.DataFrame(columns=["report_date", "channel", "cost_incl_vat"])
    if ad_spend is not None and not ad_spend.empty and "source" in ad_spend.columns:
        m = ad_spend[ad_spend["source"] == "manual"].copy()
        if not m.empty:
            m["report_date"] = pd.to_datetime(m["report_date"], errors="coerce").dt.date
            existing = m[["report_date", "channel", "cost_incl_vat"]].sort_values(
                ["report_date", "channel"], ascending=[False, True]).reset_index(drop=True)

    tab_edit, tab_file = st.tabs(["표에 직접 입력", "파일로 올리기"])

    with tab_edit:
        seed = existing.copy()
        if seed.empty:
            seed = pd.DataFrame([{"report_date": date.today() - timedelta(days=1),
                                  "channel": "네이버 GFA", "cost_incl_vat": 0.0}])
        edited = st.data_editor(
            seed, num_rows="dynamic", use_container_width=True, key="manual_spend_editor",
            column_config={
                "report_date": st.column_config.DateColumn("날짜", format="YYYY-MM-DD", required=True),
                "channel": st.column_config.SelectboxColumn("매체", options=MANUAL_SPEND_CHANNELS, required=True),
                "cost_incl_vat": st.column_config.NumberColumn("광고비(VAT 포함)", format="%d", min_value=0, required=True),
            },
        )
        st.caption("행 추가는 표 맨 아래 빈 줄에 입력하면 됩니다. 금액 0으로 저장하면 그 날짜는 예산 일할값으로 되돌아갑니다.")
        if st.button("저장", key="manual_spend_save", type="primary"):
            df = edited.copy()
            df = df.dropna(subset=["report_date", "channel"])
            if df.empty:
                st.warning("저장할 행이 없습니다.")
            else:
                df["cost_incl_vat"] = pd.to_numeric(df["cost_incl_vat"], errors="coerce").fillna(0)
                df["source"] = "manual"
                n = save_table("ad_spend_daily", df[["report_date", "channel", "cost_incl_vat", "source"]],
                               "report_date,channel,source", "직접 입력")
                st.success(f"{n}행 저장했습니다. 화면을 새로고침하면 반영됩니다.")
                st.cache_data.clear()

    with tab_file:
        st.caption(
            "네이버 광고주센터 → 결제 → 비즈머니 관리 → **기간별 내역**에서 '일별'로 놓고 "
            "다운로드한 파일을 그대로 올리시면 됩니다. 날짜와 소진액 열만 있으면 형식이 조금 달라도 읽습니다."
        )
        ch = st.selectbox("이 파일의 매체", MANUAL_SPEND_CHANNELS, key="manual_spend_file_ch")
        up = st.file_uploader("파일 선택 (xlsx / csv)", type=["xlsx", "xls", "csv"], key="manual_spend_file")
        if up is not None:
            try:
                parsed = _parse_manual_spend_file(up)
            except Exception as e:
                st.error(str(e))
            else:
                parsed["channel"] = ch
                st.write(f"읽은 행 {len(parsed)}개 · {parsed['report_date'].min()} ~ {parsed['report_date'].max()} · "
                         f"합계 {parsed['cost_incl_vat'].sum():,.0f}원")
                st.dataframe(parsed.head(10), use_container_width=True)
                st.warning(
                    "비즈머니 소진액은 계정 전체 금액입니다. GFA와 애드부스트가 같은 계정에서 같이 돌고 있으면 "
                    "둘이 합쳐진 값이니, 매체를 갈라 보시려면 표 입력 쪽을 쓰세요."
                )
                if st.button("이대로 저장", key="manual_spend_file_save", type="primary"):
                    parsed["source"] = "manual"
                    n = save_table("ad_spend_daily", parsed[["report_date", "channel", "cost_incl_vat", "source"]],
                                   "report_date,channel,source", f"직접 업로드: {getattr(up, 'name', '')}")
                    st.success(f"{n}행 저장했습니다.")
                    st.cache_data.clear()


# ══════════════════════════════════════════════════════════════
# 채널 성과 (CHANNEL PERFORMANCE) — 광고비·ROAS 한 장
#
# 표에 들어갈 값의 출처가 서로 다르다는 게 이 화면의 핵심이다.
#   · 노출/클릭/광고비 → 매체 API (ad_spend_daily)
#   · 구매/매출        → GA4 (ga_channel_daily)  ← 매체 자체 전환수는 일부러 안 쓴다
#   · 월예산           → 채널믹스 파일 (channel_mix_budget)
# 매체 자체 전환은 매체마다 기준(어트리뷰션 윈도우)이 달라 서로 못 더한다. GA 기준으로 통일해야
# 매체끼리 비교가 성립한다 — 사용자의 기존 KPI 체계(GA-ROAS)와도 같은 이유다.
# ══════════════════════════════════════════════════════════════

MEDIA_MASTER_DEFAULT = [
    # (매체명, 구분, 순서, 광고비 출처, 예산 라인, 예산 배분, UTM 매칭, 예산 직접 지정)
    # UTM 매칭은 'STCO UTM 리스트(260813)'의 소스/매체를 그대로 옮긴 것.
    # 예산 직접 지정은 예산 파일에 줄이 없는 매체용 — 값이 있으면 그게 우선한다.
    ("네이버 검색광고",      "자사몰",  10, "네이버 검색광고",     "네이버 검색광고",      1.0,
     "naver/cpc", 0),
    ("네이버 브랜드검색광고", "자사몰",  20, "네이버 브랜드검색광고", "네이버 브랜드검색광고",  1.0,
     "naver/brand_pc,naver/cpm", 0),
    ("메타",               "자사몰",  30, "메타",              "메타",               1.0,
     "facebook/facebook_feed,meta/cpc,sns/cpc", 0),
    # 예산 파일의 '신규 매체' 줄이 곧 애드부스트다(5~8월).
    ("GFA_애드부스트",      "자사몰",  40, "네이버 애드부스트",    "신규 매체",           1.0,
     "naver/gfa_애드부스트", 0),
    ("GFA_자사몰",         "자사몰",  50, "네이버 GFA",         "네이버 GFA",          1.0,
     "naver/gfa,naver/da,naver/display", 0),
    ("크리테오",            "자사몰",  60, "크리테오",           "크리테오",            1.0,
     "criteo/display,criteo/cpc", 0),
    ("구글_실적최대화",      "자사몰",  70, "구글",              "구글",               1.0,
     "google/cpc", 0),
    # 맨즈탭 8월 예산 500만원은 전액 자사몰이다(외부몰은 미집행).
    ("네이버 맨즈탭_자사몰",  "자사몰",  80, "네이버 맨즈탭_자사몰", "네이버 맨즈탭",        1.0,
     "naver/n.box,naversb/cpp,naver_sb/cpp", 0),
    ("카카오톡 플친",       "자사몰",  90, "카카오",             "카카오톡 플친",        1.0,
     "kakao_msg/email,kakaotalk/display", 0),
    # 법인카드 정산분. 매체는 아니지만 월 예산 합계를 맞추려면 '기타'로 들고 있어야 한다.
    ("기타",               "자사몰", 100, "",                 "법인카드정산",         1.0, "", 0),
    # ── 외부몰: 매출이 GA4에 안 잡힌다(스마트스토어). 예산도 파일에 없어 직접 지정한다. ──
    ("네이버 쇼핑검색광고",   "외부몰", 110, "네이버 쇼핑검색광고",  "",                  0.0, "", 2_200_000),
    ("네이버 맨즈탭_외부몰",  "외부몰", 120, "네이버 맨즈탭_외부몰", "",                  0.0, "", 0),
]
MEDIA_MASTER_COLS = ["media", "scope", "sort_order", "spend_channel",
                     "budget_line", "budget_share", "utm_match", "budget_override", "note"]


def media_master_frame(saved: pd.DataFrame = None) -> pd.DataFrame:
    """저장된 매체 정의가 있으면 그걸 쓰고, 없으면 기본값으로 시작한다."""
    if saved is not None and not saved.empty and "media" in saved.columns:
        df = saved.copy()
        defaults = {"sort_order": 100, "budget_share": 1.0, "budget_override": 0}
        for c in MEDIA_MASTER_COLS:
            if c not in df.columns:
                df[c] = defaults.get(c, "")
        return df[MEDIA_MASTER_COLS].sort_values("sort_order").reset_index(drop=True)
    rows = [dict(zip(MEDIA_MASTER_COLS[:-1], r)) for r in MEDIA_MASTER_DEFAULT]
    df = pd.DataFrame(rows)
    df["note"] = ""
    return df[MEDIA_MASTER_COLS]


def _cp_ga_by_media(ga_daily: pd.DataFrame, master: pd.DataFrame,
                    start: date, end: date) -> dict:
    """GA4 원본(source/medium 단위)을 매체 정의의 utm_match로 묶는다.

    매칭은 두 단계다.

    1) **완전일치** — UTM 리스트가 'Naver / GFA_애드부스트'처럼 소스/매체 쌍을 그대로 주므로,
       공백만 지우고 통째로 비교하면 대부분 여기서 끝난다. 가장 안전하다.
    2) **부분일치** — 리스트에 없는 새 UTM이 생겼을 때의 안전망. 패턴이 긴 것부터 본다.

    이 순서가 중요한 이유: 'naver/gfa_애드부스트'는 'naver/gfa'를 문자열로 포함한다.
    부분일치만 쓰면 애드부스트 매출이 통째로 GFA_자사몰로 넘어간다. 완전일치를 먼저 보면
    이 사고가 원천적으로 안 난다.

    어디에도 안 걸린 유입은 '_미매칭'으로 따로 모아, 화면에서 누락을 확인할 수 있게 한다.
    """
    out = {m: {"conv": 0.0, "rev": 0.0} for m in master["media"]}
    out["_미매칭"] = {"conv": 0.0, "rev": 0.0, "keys": {}}
    if ga_daily is None or ga_daily.empty or "source_medium" not in ga_daily.columns:
        return out
    g = ga_daily.copy()
    g["report_date"] = pd.to_datetime(g["report_date"], errors="coerce").dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end)]
    if g.empty:
        return out

    def norm(s):
        return "".join(str(s or "").lower().split())

    exact, partial = {}, []
    for _, m in master.sort_values("sort_order").iterrows():
        for kw in str(m.get("utm_match") or "").split(","):
            kw = norm(kw)
            if not kw:
                continue
            exact.setdefault(kw, m["media"])
            partial.append((len(kw), kw, m["media"]))
    partial.sort(key=lambda x: -x[0])

    for _, row in g.iterrows():
        sm = norm(row.get("source_medium"))
        conv = float(row.get("conversions") or 0)
        rev = float(row.get("revenue") or 0)
        media = exact.get(sm)
        if media is None:
            for _, kw, mm in partial:
                if kw in sm:
                    media = mm
                    break
        if media is None:
            out["_미매칭"]["conv"] += conv
            out["_미매칭"]["rev"] += rev
            k = str(row.get("source_medium") or "(없음)")
            out["_미매칭"]["keys"][k] = out["_미매칭"]["keys"].get(k, 0) + rev
            continue
        out[media]["conv"] += conv
        out[media]["rev"] += rev
    return out


def _cp_ga_breakdown(ga_daily, master, start: date, end: date) -> pd.DataFrame:
    """매체별로 '어떤 소스/매체를 합쳐서 만든 숫자인지' 펼쳐 보여준다.

    GA4 화면에서 소스/매체 하나만 보고 대시보드와 비교하면 숫자가 안 맞는다.
    대시보드는 UTM 매칭 규칙에 따라 여러 소스/매체를 한 매체로 묶기 때문이다
    (예: 메타 = facebook/facebook_feed + meta/cpc + sns/cpc).
    그 구성을 눈으로 확인할 수 있어야 '왜 다르지'로 시간을 안 뺏긴다.
    """
    cols = ["매체", "소스/매체", "구매", "매출"]
    if ga_daily is None or ga_daily.empty or "source_medium" not in ga_daily.columns:
        return pd.DataFrame(columns=cols)
    g = ga_daily.copy()
    g["report_date"] = pd.to_datetime(g["report_date"], errors="coerce").dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end)]
    if g.empty:
        return pd.DataFrame(columns=cols)
    for c in ("conversions", "revenue"):
        g[c] = pd.to_numeric(g.get(c), errors="coerce").fillna(0)

    def norm(s):
        return "".join(str(s or "").lower().split())

    exact, partial = {}, []
    for _, m in master.sort_values("sort_order").iterrows():
        for kw in str(m.get("utm_match") or "").split(","):
            kw = norm(kw)
            if kw:
                exact.setdefault(kw, m["media"])
                partial.append((len(kw), kw, m["media"]))
    partial.sort(key=lambda x: -x[0])

    def which(sm):
        n = norm(sm)
        if n in exact:
            return exact[n]
        for _, kw, media in partial:
            if kw in n:
                return media
        return "(미매칭)"

    agg = g.groupby("source_medium", as_index=False).agg(
        구매=("conversions", "sum"), 매출=("revenue", "sum"))
    agg["매체"] = agg["source_medium"].map(which)
    agg = agg.rename(columns={"source_medium": "소스/매체"})
    return agg[cols].sort_values(["매체", "매출"], ascending=[True, False])


def _cp_spend_by_channel(ad_spend: pd.DataFrame, start: date, end: date) -> pd.DataFrame:
    """기간 내 매체별 노출·클릭·광고비.

    지표마다 출처를 따로 고른다. 네이버 브랜드검색이 대표적인 이유인데, 정액(보장형) 상품이라
    API가 노출·클릭은 주지만 집행액은 안 준다(관리자 화면에도 '-'로 뜬다). 출처를 행 단위로
    하나만 고르면 API 행이 선택되면서 손으로 넣은 광고비가 통째로 날아간다.
    그래서 '값이 있는 것 중 우선순위가 높은 출처'를 지표별로 뽑는다.
    """
    cols = ["channel", "impressions", "clicks", "cost_incl_vat", "source"]
    if ad_spend is None or ad_spend.empty:
        return pd.DataFrame(columns=cols)
    a = ad_spend.copy()
    a["report_date"] = pd.to_datetime(a["report_date"], errors="coerce").dt.date
    a = a[(a["report_date"] >= start) & (a["report_date"] <= end)]
    if a.empty:
        return pd.DataFrame(columns=cols)
    for c in ("impressions", "clicks", "cost_incl_vat"):
        a[c] = pd.to_numeric(a.get(c), errors="coerce").fillna(0)
    g = a.groupby(["channel", "source"], as_index=False).agg(
        impressions=("impressions", "sum"), clicks=("clicks", "sum"),
        cost_incl_vat=("cost_incl_vat", "sum"))
    prio = {"meta_api": 0, "google_ads_api": 0, "naver_api": 0, "kakao_api": 0,
            "criteo_api": 0, "naver_gfa_api": 0,
            "contract": 1, "manual": 2, "agency_weekly": 3, "budget_prorate": 4}
    g["_p"] = g["source"].map(prio).fillna(9)
    g = g.sort_values("_p")

    out = []
    for ch, sub in g.groupby("channel"):
        def pick(col):
            hit = sub[sub[col] > 0]
            if hit.empty:
                return 0.0, ""
            r = hit.iloc[0]
            return float(r[col]), str(r["source"])
        imp, _ = pick("impressions")
        clk, _ = pick("clicks")
        cost, csrc = pick("cost_incl_vat")
        out.append({"channel": ch, "impressions": imp, "clicks": clk,
                    "cost_incl_vat": cost, "source": csrc})
    return pd.DataFrame(out, columns=cols)


def _budget_key(s) -> str:
    """예산 라인 이름을 비교하기 좋게 다듬는다.

    같은 줄인데 파일 버전마다 표기가 조금씩 다르다
    (예: '법인카드정산' ↔ '법인카드정산 (기타)'). 괄호 안 설명과 공백을 떼고 비교한다.
    """
    t = re.sub(r"\([^)]*\)", "", str(s or ""))
    return "".join(t.split()).lower()


def _channel_mix_panel(channel_mix) -> str:
    """연간 채널 믹스 패널(매체별 금액 + 비중).

    성과를 보는 화면보다 '예산을 어떻게 나눠 쓸까'를 정하는 화면에 있어야 쓸모가 있어서
    예산 관리 탭으로 옮겼다. 비중만으로는 감이 안 잡혀 금액을 원 단위로 같이 적는다.
    """
    if channel_mix is None or channel_mix.empty:
        return ""
    m = channel_mix.copy()
    m["year"] = pd.to_numeric(m["year"], errors="coerce")
    year = m["year"].max()
    m = m[m["year"] == year]
    agg = m.groupby("channel", as_index=False).agg(budget=("budget", "sum"))
    agg = agg[agg["budget"] > 0]
    if agg.empty:
        return ""
    total = float(agg["budget"].sum())
    agg["ratio"] = agg["budget"] / total * 100
    agg = agg.sort_values("budget", ascending=False).head(10)
    top = float(agg["ratio"].max()) or 1.0
    rows = "".join(
        f'<div class="fv4-mix-row"><div class="fv4-mix-row-top">'
        f'<span class="fv4-mix-ch">{r["channel"]}</span>'
        f'<span class="fv4-mix-val">\u20a9{r["budget"]:,.0f}<b>{r["ratio"]:.1f}%</b></span></div>'
        f'<div class="fv4-mix-track"><i style="width:{r["ratio"] / top * 100:.1f}%"></i></div></div>'
        for _, r in agg.iterrows())
    return ('<div class="fv4-wrap"><div class="fv4-mixpanel-wide">'
            f'<div class="fv4-eyebrow">{int(year)} CHANNEL MIX</div>'
            f'<div class="fv4-mix-amt-big">\u20a9{total:,.0f}</div>'
            '<div class="fv4-kpi-sub">연간 매체 예산</div>'
            f'<div class="fv4-mix-list">{rows}</div></div></div>')


def _media_budget(r, budget_map: dict) -> float:
    """매체 한 줄의 월 예산. 직접 지정값이 있으면 그게 우선이다.

    예산 파일(자사몰)에 줄이 없는 매체가 있다 — 네이버 쇼핑검색광고 같은 외부몰 매체.
    그런 건 '매체 정의' 패널의 '예산 직접 지정'에 넣어두면 그 값을 쓴다.
    """
    try:
        ov = float(r.get("budget_override") or 0)
    except Exception:
        ov = 0.0
    if ov > 0:
        return ov
    bl = str(r.get("budget_line") or "").strip()
    if not bl:
        return 0.0
    amt = budget_map.get(bl)
    if amt is None:                       # 표기 차이(괄호 설명 등)까지 감안해 다시 찾는다
        want = _budget_key(bl)
        amt = next((v for k, v in budget_map.items() if _budget_key(k) == want), 0)
    return float(amt or 0) * float(r.get("budget_share") or 0)


def _cp_month_budget(channel_mix: pd.DataFrame, ref: date,
                     budget: pd.DataFrame = None, scope: str = "자사몰") -> dict:
    """해당 월의 매체별 예산을 꺼낸다.

    '◆26년 월별 예산 정리'(channel_budget)가 회사 공식 계획이라 이걸 먼저 본다.
    없을 때만 '채널 믹스' 파일로 물러난다. 두 파일의 8월 합계가 다른데
    (25,020,000 vs 23,250,000), 차이는 신규 매체·법인카드정산이라 공식 파일이 맞다.
    """
    if budget is not None and not budget.empty and "budget_cost" in budget.columns:
        b = budget.copy()
        for c in ("year", "month"):
            b[c] = pd.to_numeric(b.get(c), errors="coerce")
        b = b[(b.get("scope") == scope) & (b["year"] == ref.year) & (b["month"] == ref.month)]
        _ch = b["channel"].astype(str)
        b = b[(~_ch.str.startswith("__")) & (_ch != "TOTAL")
              & (~_ch.str.contains(BUDGET_HARD_EXCLUDE_RE))]
        if not b.empty:
            return {str(r["channel"]): float(r["budget_cost"] or 0) for _, r in b.iterrows()}

    out = {}
    if channel_mix is None or channel_mix.empty:
        return out
    m = channel_mix.copy()
    for c in ("year", "month"):
        m[c] = pd.to_numeric(m.get(c), errors="coerce")
    m = m[(m["year"] == ref.year) & (m["month"] == ref.month)]
    for _, r in m.iterrows():
        out[str(r.get("channel"))] = float(r.get("budget") or 0)
    return out

CP_CSS = """
<style>
.cp-wrap{font-family:-apple-system,BlinkMacSystemFont,'Pretendard','Segoe UI',sans-serif;color:#14181F}
.cp-eyebrow{font-size:13.5px;letter-spacing:.14em;font-weight:700;color:#7A8088;text-transform:uppercase;margin-bottom:5px}
.cp-title{font-size:26px;font-weight:800;letter-spacing:-.02em;margin:0 0 14px}
/* 칸 개수가 화면마다 달라서 grid를 고정 칸수로 잡으면 빈 칸이 생기고 칸막이가 끊긴다.
   flex로 두고 칸막이를 왼쪽 테두리로 그리면 몇 개든 알아서 맞는다. */
.cp-kpis{display:flex;background:#FFF;border:1px solid #E3E1DC;border-radius:10px;
  overflow:hidden;margin-bottom:14px}
.cp-kpi{flex:1 1 0;min-width:0;padding:14px 16px;border-left:1px solid #E3E1DC}
.cp-kpi:first-child{border-left:none}
.cp-kpi .k{font-size:14px;color:#6E747C;font-weight:600;margin-bottom:6px}
.cp-kpi .v{font-size:22px;font-weight:800;letter-spacing:-.02em}
.cp-kpi .s{font-size:13.5px;color:#7A8088;margin-top:5px}
.cp-tbl{width:100%;border-collapse:collapse;background:#FFF;border:1px solid #E3E1DC;border-radius:10px;overflow:hidden;font-size:14.5px}
.cp-tbl th{background:#F4F2ED;color:#5E646C;font-size:14px;font-weight:700;text-align:right;padding:11px 12px;border-bottom:1px solid #E3E1DC;white-space:nowrap}
.cp-tbl th.l,.cp-tbl td.l{text-align:left}
.cp-ar{margin-left:6px;color:#B9BEC5;font-size:13px}
.cp-tbl th:hover{background:#EDEAE3}
.cp-tbl td{padding:11px 12px;text-align:right;border-bottom:1px solid #F0EEE9;white-space:nowrap}
.cp-tbl tr:last-child td{border-bottom:none}
.cp-tbl tr:hover td{background:#FAFAF8}
.cp-tbl td.m{font-weight:700;color:#14181F}
.cp-tbl td.sub{font-size:13.5px;color:#7A8088;font-weight:400}
.cp-badge{display:inline-block;padding:3px 8px;border-radius:5px;font-size:13px;font-weight:700}
.cp-own{background:#14181F;color:#EFEDE7}
.cp-ext{background:#E7E4F7;color:#4B3FA8}
.cp-pill{display:inline-block;padding:3px 8px;border-radius:5px;font-size:13px;font-weight:700;background:#EAF7D9;color:#3F6B12}
.cp-pill.warn{background:#FDEDE3;color:#9A4B14}
.cp-pill.zero{background:#F1F1EF;color:#9AA0A8}
.cp-mute{color:#C2C6CC}
.cp-scope{display:inline-block;margin-left:10px;padding:4px 11px;border-radius:6px;
  background:#EFEDE8;color:#6E747C;font-size:13.5px;font-weight:700;vertical-align:middle}
.cp-pace{height:5px;background:#EFEDE8;border-radius:3px;margin-top:7px;overflow:hidden;position:relative}
.cp-pace i{display:block;height:100%;background:#14181F;border-radius:3px}
.cp-pace u{position:absolute;top:-3px;width:2px;height:11px;background:#E0654A;text-decoration:none}
.cp-tot td{background:#F4F2ED;font-weight:800;border-top:2px solid #E3E1DC}
.cp-note{font-size:14px;color:#6E747C;margin-top:12px;line-height:1.85}
.cp-arrow{text-align:center;color:#9AA0A8;font-size:15px;padding-top:34px}
.cp-rec{background:#D9F27E;border-radius:12px;padding:22px 26px;margin-top:16px}
.cp-rec-eyebrow{font-size:13px;letter-spacing:.14em;font-weight:800;color:#4A5A16;margin-bottom:7px}
.cp-rec-head{font-size:23px;font-weight:800;letter-spacing:-.02em;color:#14181F;margin-bottom:5px}
.cp-rec-sub{font-size:14.5px;color:#4A5A16;margin-bottom:14px}
.cp-rec-list{margin:0;padding-left:0;list-style:none}
.cp-rec-list li{font-size:15px;color:#1E2530;padding:11px 0;border-top:1px solid rgba(0,0,0,.09);line-height:1.65}
.cp-tag{display:inline-block;padding:3px 9px;border-radius:5px;font-size:13.5px;font-weight:800;margin-right:8px}
.cp-tag.up{background:#14181F;color:#D9F27E}
.cp-tag.down{background:#FDEDE3;color:#9A4B14}
.cp-tag.keep{background:#FFF;color:#4A5A16}
.cp-tag.hold{background:#EFEFEA;color:#6E747C}
.cp-tag.skip{background:#E7E4F7;color:#4B3FA8}
.cp-rec-foot{font-size:13.5px;color:#5E6B2A;margin-top:14px;line-height:1.65}
.cp-btnpad{height:26px}
</style>
<style>
/* 기간 바: 컨트롤들을 한 덩어리 카드처럼 보이게 한다 */
.st-key-cpbar{background:#FFF;border:1px solid #E3E1DC;border-radius:12px;
  padding:10px 14px 4px;margin-bottom:14px}
.st-key-cpbar label{font-size:13px !important;color:#8A9099 !important;font-weight:600 !important}
.st-key-cpbar [data-baseweb="select"]>div,
.st-key-cpbar [data-testid="stDateInput"] input{border-radius:8px !important;font-size:14px !important}
.st-key-cpbar .stButton>button{border-radius:8px;background:#14181F;color:#F3F1EC;
  border:none;font-weight:700;font-size:14px;padding:7px 0}
.st-key-cpbar .stButton>button:hover{background:#2A3038;color:#FFF}
</style>
"""


def _cp_won(v):
    try:
        v = float(v)
    except Exception:
        return '<span class="cp-mute">—</span>'
    return f"₩{v:,.0f}"


def _cp_int(v):
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "0"


CP_MIN_SPEND_FOR_JUDGE = 100_000   # 이보다 적게 쓴 매체는 성과를 단정하지 않는다
CP_STEP = 0.10                     # 증액·감액 제안은 한 번에 10%씩 (급격한 이동 금지)


def _cp_recommendations(df: pd.DataFrame, start: date, end: date) -> str:
    """표 아래에 붙는 '그래서 뭘 하라는 건지' 코멘트를 만든다.

    판단에서 빼는 경우를 먼저 정리하는 게 핵심이다. 안 그러면 틀린 지시가 나간다.
      · 외부몰(쇼핑검색·맨즈탭 외부몰) — 매출이 GA4에 안 잡혀 ROAS가 0으로 보인다.
        이걸 그대로 읽으면 "줄여라"가 되는데 실제로는 판단 근거 자체가 없는 것이다.
      · 광고비가 API로 안 들어온 매체 — 분모가 없으니 ROAS를 만들 수 없다.
      · 광고비가 너무 적은 매체 — 표본이 작아 우연이 성과처럼 보인다.
    남은 매체만 ROAS로 줄 세우고, 이동은 한 번에 10%씩만 제안한다.
    """
    up, keep, down, hold, skip = [], [], [], [], []
    for _, r in df.iterrows():
        m, cost, roas = r["매체"], float(r["비용"] or 0), r["GA ROAS"]
        rev = float(r["GA매출"] or 0)
        if r["구분"] == "외부몰":
            skip.append((m, "외부몰 — 매출이 GA에 안 잡혀 판단 불가 (스마트스토어 연동 필요)"))
        elif cost <= 0 and rev > 0:
            skip.append((m, f"광고비 미연동 — 매출 {rev:,.0f}원은 잡히는데 비용이 없어 ROAS 계산 불가"))
        elif cost <= 0:
            continue
        elif cost < CP_MIN_SPEND_FOR_JUDGE:
            hold.append((m, cost, roas))
        elif roas is None or pd.isna(roas):
            skip.append((m, "매출 데이터 없음"))
        elif roas >= OPS_KPI_ROAS_HIGH:
            up.append((m, cost, roas))
        elif roas >= OPS_KPI_ROAS_LOW:
            keep.append((m, cost, roas))
        else:
            down.append((m, cost, roas))

    up.sort(key=lambda x: -x[2])
    down.sort(key=lambda x: x[2])
    days = (end - start).days + 1

    # 헤드라인
    if up and down:
        head = (f"<b>{up[0][0]}</b>를 늘리고 <b>{down[0][0]}</b>를 줄이세요.")
        sub = (f"{up[0][0]} ROAS {up[0][2]:,.0f}% (목표 상단 {OPS_KPI_ROAS_HIGH:.0f}% 초과) · "
               f"{down[0][0]} ROAS {down[0][2]:,.0f}% (목표 하단 {OPS_KPI_ROAS_LOW:.0f}% 미달)")
    elif up:
        head = f"<b>{up[0][0]}</b>에 예산을 더 태울 여지가 있습니다."
        sub = f"ROAS {up[0][2]:,.0f}%로 목표 상단({OPS_KPI_ROAS_HIGH:.0f}%)을 넘었습니다."
    elif down:
        head = f"<b>{down[0][0]}</b>를 손봐야 합니다."
        sub = f"ROAS {down[0][2]:,.0f}%로 목표 하단({OPS_KPI_ROAS_LOW:.0f}%)에 못 미칩니다."
    elif keep:
        head = "지금은 손댈 매체가 없습니다."
        sub = f"판단 가능한 매체가 모두 목표 구간({OPS_KPI_ROAS_LOW:.0f}~{OPS_KPI_ROAS_HIGH:.0f}%) 안에 있습니다."
    else:
        head = "아직 판단할 만한 데이터가 없습니다."
        sub = "광고비가 연동된 매체가 없거나 집행액이 너무 적습니다."

    def line(m, cost, roas, kind):
        move = cost * CP_STEP
        daily = move / days if days else move
        if kind == "up":
            act = (f"예산 <b>+{CP_STEP*100:.0f}%</b> (기간 기준 +{move:,.0f}원 · 일 +{daily:,.0f}원) 테스트 후 "
                   f"48시간 관찰")
            tag, cls = "증액 검토", "up"
        elif kind == "down":
            act = (f"예산 <b>-{CP_STEP*100:.0f}%</b> (기간 기준 -{move:,.0f}원) 또는 소재·타겟 점검 먼저")
            tag, cls = "감액·점검", "down"
        else:
            act = "현 수준 유지"
            tag, cls = "유지", "keep"
        return (f'<li><span class="cp-tag {cls}">{tag}</span> <b>{m}</b> — '
                f'ROAS {roas:,.0f}% · 집행 {cost:,.0f}원 → {act}</li>')

    items = []
    for m, c, r in up:
        items.append(line(m, c, r, "up"))
    for m, c, r in down:
        items.append(line(m, c, r, "down"))
    for m, c, r in keep:
        items.append(line(m, c, r, "keep"))
    for m, c, r in hold:
        rr = "—" if (r is None or pd.isna(r)) else f"{r:,.0f}%"
        items.append(f'<li><span class="cp-tag hold">판단 보류</span> <b>{m}</b> — '
                     f'집행 {c:,.0f}원으로 표본이 작습니다 (ROAS {rr}). '
                     f'{CP_MIN_SPEND_FOR_JUDGE:,.0f}원 넘을 때까지 판단 유보</li>')
    for m, why in skip:
        items.append(f'<li><span class="cp-tag skip">판단 제외</span> <b>{m}</b> — {why}</li>')

    return (
        '<div class="cp-wrap"><div class="cp-rec">'
        '<div class="cp-rec-eyebrow">NEXT BEST ACTION</div>'
        f'<div class="cp-rec-head">{head}</div>'
        f'<div class="cp-rec-sub">{sub}</div>'
        f'<ul class="cp-rec-list">{"".join(items)}</ul>'
        f'<div class="cp-rec-foot">기준: GA-ROAS 목표 {OPS_KPI_ROAS_LOW:.0f}~{OPS_KPI_ROAS_HIGH:.0f}% · '
        f'집행 {CP_MIN_SPEND_FOR_JUDGE:,.0f}원 미만은 판단 보류 · '
        f'예산 이동은 한 번에 {CP_STEP*100:.0f}%씩만 제안합니다 (조회 기간 {start} ~ {end}, {days}일)</div>'
        '</div></div>'
    )


CONTRACT_CHANNELS = ["네이버 브랜드검색광고", "네이버 맨즈탭_자사몰", "네이버 맨즈탭_외부몰",
                     "카카오톡 플친", "네이버 트렌드픽"]
CONTRACT_COLS = ["channel", "start_date", "end_date", "amount_incl_vat", "note"]


def contract_seed() -> pd.DataFrame:
    """저장된 계약이 없을 때 보여줄 기본 행 (실제 브랜드검색 계약 기준)."""
    return pd.DataFrame([
        {"channel": "네이버 브랜드검색광고", "start_date": date(2026, 7, 8),
         "end_date": date(2026, 8, 6), "amount_incl_vat": 1_540_000, "note": "직전 30일 계약"},
        {"channel": "네이버 브랜드검색광고", "start_date": date(2026, 8, 7),
         "end_date": date(2026, 9, 5), "amount_incl_vat": 1_540_000, "note": "현재 계약"},
    ], columns=CONTRACT_COLS)


def contracts_to_daily(contracts: pd.DataFrame) -> pd.DataFrame:
    """계약 총액을 계약 일수로 나눠 일별 광고비 행으로 펼친다.

    정액 상품은 '오늘 얼마 썼나'가 존재하지 않으므로, 계약 기간으로 균등 배분하는 게
    가장 사실에 가깝다. 월이 걸쳐 있어도(8/7~9/5) 날짜 단위로 나누므로 월 경계가 알아서 맞는다.
    """
    if contracts is None or contracts.empty:
        return pd.DataFrame(columns=["report_date", "channel", "cost_incl_vat",
                                     "impressions", "clicks", "source"])
    rows = []
    for _, c in contracts.iterrows():
        try:
            s = pd.to_datetime(c["start_date"]).date()
            e = pd.to_datetime(c["end_date"]).date()
            amt = float(c["amount_incl_vat"] or 0)
        except Exception:
            continue
        ch = str(c.get("channel") or "").strip()
        if not ch or e < s or amt <= 0:
            continue
        days = (e - s).days + 1
        per = amt / days
        for i in range(days):
            rows.append({"report_date": s + timedelta(days=i), "channel": ch,
                         "cost_incl_vat": per, "impressions": 0, "clicks": 0,
                         # 리포트 업로드(manual)와 같은 키를 쓰면 서로 덮어쓰므로 출처를 나눈다.
                         # 보장형 매체는 계약 금액이 사실이므로 우선순위도 리포트보다 위에 둔다.
                         "source": "contract"})
    return pd.DataFrame(rows)


def render_channel_performance_page(ad_spend, ga_daily, channel_mix, master=None, budget=None):
    """채널 성과 — 매체별 노출·클릭·광고비(API) × 구매·매출(GA4) × 월예산(채널믹스)."""
    st.markdown(CP_CSS, unsafe_allow_html=True)

    mst = media_master_frame(master)

    # ── 기간 바 ──────────────────────────────────────────
    # 프리셋을 고르면 날짜가 그에 맞게 채워지고, 채워진 날짜를 손으로 고쳐도 된다.
    # (date_input의 key에 preset을 넣어야 프리셋을 바꿨을 때 값이 새로 잡힌다)
    today = date.today()

    def _range_for(name):
        if name == "이번 달":
            return today.replace(day=1), today - timedelta(days=1)
        if name == "최근 7일":
            e = today - timedelta(days=1); return e - timedelta(days=6), e
        if name == "최근 30일":
            e = today - timedelta(days=1); return e - timedelta(days=29), e
        if name == "지난 달":
            f = today.replace(day=1); e = f - timedelta(days=1); return e.replace(day=1), e
        return today.replace(day=1), today - timedelta(days=1)

    try:
        bar = st.container(key="cpbar")
    except TypeError:
        bar = st.container()
    with bar:
        b1, b2, b3, b4, b5, b6 = st.columns([1.05, 1.05, 0.12, 1.05, 0.5, 1.15])
        with b1:
            preset = st.selectbox("조회 기간",
                                  ["이번 달", "최근 7일", "최근 30일", "지난 달", "직접 지정"],
                                  key="cp_preset")
        d0, d1 = _range_for(preset)
        with b2:
            start = st.date_input("시작일", d0, key=f"cp_s_{preset}", format="YYYY-MM-DD")
        with b3:
            st.markdown('<div class="cp-arrow">→</div>', unsafe_allow_html=True)
        with b4:
            end = st.date_input("종료일", d1, key=f"cp_e_{preset}", format="YYYY-MM-DD")
        with b5:
            st.markdown('<div class="cp-btnpad"></div>', unsafe_allow_html=True)
            if st.button("조회", key="cp_apply", use_container_width=True):
                st.cache_data.clear()
        with b6:
            src_label = "GA4 · 실제 데이터" if (ga_daily is not None and not ga_daily.empty) \
                else "엑셀 업로드 데이터"
            st.selectbox("데이터 소스", [src_label], key="cp_src", disabled=True)
    if start > end:
        start, end = end, start

    # ── 정액(보장형) 계약 광고비 ─────────────────────────
    # 브랜드검색은 30일 단위로 선지불하고 연장하는 상품이라 '일별 집행액'이 없다.
    # 계약 기간과 총액을 넣어두면 일할로 나눠 매일의 광고비를 만든다.
    with st.expander("📄 정액 계약 광고비 (브랜드검색 등) — 갱신할 때마다 한 줄 추가"):
        st.caption(
            "계약 기간과 총액(VAT 포함)을 넣으면 기간으로 나눠 일별 광고비로 저장합니다. "
            "월이 걸쳐 있어도(예: 8/7~9/5) 날짜 단위로 나누므로 월별 집계가 알아서 맞습니다."
        )
        saved_ct = load_table("ad_contract")
        ct = saved_ct[CONTRACT_COLS].copy() if (saved_ct is not None and not saved_ct.empty
                                                and "channel" in saved_ct.columns) else contract_seed()
        for c in ("start_date", "end_date"):
            ct[c] = pd.to_datetime(ct[c], errors="coerce").dt.date
        ct = ct.sort_values(["channel", "start_date"]).reset_index(drop=True)

        ct_edit = st.data_editor(
            ct, num_rows="dynamic", use_container_width=True, key="cp_contract_editor",
            column_config={
                "channel": st.column_config.SelectboxColumn("매체", options=CONTRACT_CHANNELS, required=True),
                "start_date": st.column_config.DateColumn("계약 시작", format="YYYY-MM-DD", required=True),
                "end_date": st.column_config.DateColumn("계약 종료", format="YYYY-MM-DD", required=True),
                "amount_incl_vat": st.column_config.NumberColumn("계약 총액(VAT 포함)", format="%d", min_value=0),
                "note": st.column_config.TextColumn("메모"),
            },
        )
        prev = contracts_to_daily(ct_edit)
        if not prev.empty:
            g = prev.groupby("channel", as_index=False).agg(
                일수=("report_date", "count"), 합계=("cost_incl_vat", "sum"))
            g["일 광고비"] = (g["합계"] / g["일수"]).round(0)
            st.caption("저장하면 이렇게 펼쳐집니다 (일할)")
            st.dataframe(g, use_container_width=True, hide_index=True)
        if st.button("저장하고 반영", key="cp_contract_save", type="primary"):
            e = ct_edit.dropna(subset=["channel", "start_date", "end_date"])
            if e.empty:
                st.warning("저장할 계약이 없습니다.")
            else:
                save_table("ad_contract", e[CONTRACT_COLS], "channel,start_date", "정액 계약")
                daily = contracts_to_daily(e)
                n = save_table("ad_spend_daily", daily,
                               "report_date,channel,source", "정액 계약 일할")
                st.success(f"계약 {len(e)}건 → 일별 {n}행 저장했습니다.")
                st.cache_data.clear()

    # ── 집계 ─────────────────────────────────────────────
    spend = _cp_spend_by_channel(ad_spend, start, end)
    spend_map = {r["channel"]: r for _, r in spend.iterrows()} if not spend.empty else {}
    ga_map = _cp_ga_by_media(ga_daily, mst, start, end)
    budget_map = _cp_month_budget(channel_mix, start, budget)

    recs = []
    for _, r in mst.iterrows():
        sc = str(r.get("spend_channel") or "").strip()
        sp = spend_map.get(sc, {})
        budget = _media_budget(r, budget_map)
        g = ga_map.get(r["media"], {"conv": 0.0, "rev": 0.0})
        cost = float(sp.get("cost_incl_vat", 0) or 0)
        recs.append({
            "구분": r.get("scope", "자사몰"), "매체": r["media"],
            "노출": float(sp.get("impressions", 0) or 0),
            "클릭": float(sp.get("clicks", 0) or 0),
            "비용": cost,
            "GA구매": g["conv"], "GA매출": g["rev"],
            "GA ROAS": (g["rev"] / cost * 100) if cost > 0 else None,
            "월예산": budget,
            "예산 소진율": (cost / budget * 100) if budget > 0 else None,
            "_src": sp.get("source", ""), "_order": int(r.get("sort_order") or 100),
        })
    df = pd.DataFrame(recs)

    # ── 구분 필터 (KPI보다 위) ────────────────────────────
    # 필터를 KPI 아래에 두면 '자사몰'을 골라도 위 숫자는 전체라서 예산·ROAS를 잘못 읽게 된다.
    # 자사몰과 외부몰은 재원도 매출 집계 방식도 달라 반드시 같이 걸러야 한다.
    scope_f = st.radio("구분", ["전체", "자사몰", "외부몰"], horizontal=True, key="cp_scope")
    view = (df if scope_f == "전체" else df[df["구분"] == scope_f]).sort_values("_order")

    # ── 상단 KPI (선택한 구분 기준) ────────────────────────
    tot_cost = view["비용"].sum()
    tot_budget = view["월예산"].sum()
    tot_rev = view["GA매출"].sum()
    tot_conv = view["GA구매"].sum()
    ref_m = start.replace(day=1)
    last_day = (ref_m.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    elapsed = (min(end, last_day).day / last_day.day * 100) if last_day.day else 0
    roas = (tot_rev / tot_cost * 100) if tot_cost > 0 else 0
    pace = (tot_cost / tot_budget * 100) if tot_budget > 0 else 0
    pace_word = "과속" if pace > elapsed + 5 else ("저속" if pace < elapsed - 5 else "정상")
    scope_note = "자사몰 + 외부몰" if scope_f == "전체" else scope_f
    ext_note = ('<div class="cp-note" style="margin:-6px 0 12px">'
                '· 외부몰 매출은 GA4에 안 잡힙니다 — 스마트스토어 연동 전까지 ROAS는 0으로 나옵니다.'
                '</div>') if scope_f == "외부몰" else ""
    st.markdown(
        '<div class="cp-wrap">'
        '<div class="cp-eyebrow">CHANNEL PERFORMANCE · LIVE MEDIA SPEND</div>'
        f'<div class="cp-title">채널별 성과 <span class="cp-scope">{scope_note}</span></div>'
        + ext_note +
        '<div class="cp-kpis">'
        f'<div class="cp-kpi"><div class="k">실제 집행 광고비</div><div class="v">{_cp_won(tot_cost)}</div>'
        f'<div class="s">VAT 포함 · 공급가 {_cp_won(tot_cost / 1.1)}</div></div>'
        f'<div class="cp-kpi"><div class="k">월 예산</div><div class="v">{_cp_won(tot_budget)}</div>'
        f'<div class="s">{ref_m.year}년 {ref_m.month}월 · {scope_note}</div></div>'
        f'<div class="cp-kpi"><div class="k">예산 소진율</div>'
        f'<div class="v">{pace:.1f}%</div>'
        f'<div class="cp-pace"><i style="width:{min(pace,100):.0f}%"></i>'
        f'<u style="left:{min(elapsed,100):.0f}%"></u></div>'
        f'<div class="s">쓴 돈 ÷ 월 예산 · <b>{pace_word}</b></div></div>'
        f'<div class="cp-kpi"><div class="k">기간 경과</div>'
        f'<div class="v">{elapsed:.0f}%</div>'
        f'<div class="s">{start.month}/{start.day}~{end.month}/{end.day} · '
        f'{last_day.day}일 중 {min(end, last_day).day}일</div></div>'
        f'<div class="cp-kpi"><div class="k">GA 매출 ROAS</div><div class="v">{roas:,.1f}%</div>'
        f'<div class="s">목표 {OPS_KPI_ROAS_LOW:.0f}~{OPS_KPI_ROAS_HIGH:.0f}%</div></div>'
        f'<div class="cp-kpi"><div class="k">GA 구매매출</div><div class="v">{_cp_won(tot_rev)}</div>'
        f'<div class="s">구매 {_cp_int(tot_conv)}건</div></div>'
        '</div></div>', unsafe_allow_html=True)

    # ── 표 ───────────────────────────────────────────────
    heads = ["구분", "매체", "노출", "클릭", "총비용(VAT 포함)", "GA 구매", "GA 매출", "GA ROAS", "월 예산", "예산 소진율"]
    th = "".join(
        f'<th class="{"l" if h in ("구분", "매체") else ""}">{h}<span class="cp-ar">&#8645;</span></th>'
        for h in heads)
    body = []
    for _, r in view.iterrows():
        badge = "cp-ext" if r["구분"] == "외부몰" else "cp-own"
        roas_v = ('<span class="cp-mute">—</span>'
                  if r["GA ROAS"] is None or pd.isna(r["GA ROAS"])
                  else f'{r["GA ROAS"]:,.1f}%')
        bud = _cp_won(r["월예산"]) if r["월예산"] > 0 else '<span class="cp-mute">미배정</span>'
        if r["예산 소진율"] is None or pd.isna(r["예산 소진율"]):
            pace_v = '<span class="cp-mute">—</span>'
        else:
            cls = "cp-pill" if r["예산 소진율"] <= 110 else "cp-pill warn"
            if r["예산 소진율"] == 0:
                cls = "cp-pill zero"
            pace_v = f'<span class="{cls}">{r["예산 소진율"]:.1f}%</span>'
        src = AD_SPEND_SOURCE_LABEL.get(r["_src"], "")
        sub = f'<div class="sub">{src}</div>' if src else ""
        nn = lambda v: -1 if v is None or (isinstance(v, float) and pd.isna(v)) else v
        body.append(
            f'<tr><td class="l"><span class="cp-badge {badge}">{r["구분"]}</span></td>'
            f'<td class="l m">{r["매체"]}{sub}</td>'
            f'<td data-v="{r["노출"]:.0f}">{_cp_int(r["노출"])}</td>'
            f'<td data-v="{r["클릭"]:.0f}">{_cp_int(r["클릭"])}</td>'
            f'<td data-v="{r["비용"]:.0f}">{_cp_won(r["비용"])}</td>'
            f'<td data-v="{r["GA구매"]:.0f}">{_cp_int(r["GA구매"])}</td>'
            f'<td data-v="{r["GA매출"]:.0f}">{_cp_won(r["GA매출"])}</td>'
            f'<td data-v="{nn(r["GA ROAS"]):.2f}">{roas_v}</td>'
            f'<td data-v="{r["월예산"]:.0f}">{bud}</td>'
            f'<td data-v="{nn(r["예산 소진율"]):.2f}">{pace_v}</td></tr>')
    s_cost, s_bud = view["비용"].sum(), view["월예산"].sum()
    s_roas = f'{view["GA매출"].sum() / s_cost * 100:,.1f}%' if s_cost > 0 else '<span class="cp-mute">—</span>'
    s_pace = f'{s_cost / s_bud * 100:,.1f}%' if s_bud > 0 else '<span class="cp-mute">—</span>'
    body.append(
        f'<tr class="cp-tot"><td class="l" colspan="2">합계</td>'
        f'<td>{_cp_int(view["노출"].sum())}</td><td>{_cp_int(view["클릭"].sum())}</td>'
        f'<td>{_cp_won(s_cost)}</td><td>{_cp_int(view["GA구매"].sum())}</td>'
        f'<td>{_cp_won(view["GA매출"].sum())}</td><td>{s_roas}</td>'
        f'<td>{_cp_won(s_bud)}</td><td>{s_pace}</td></tr>')

    note = (
        '<div class="cp-note">'
        '· 표 머리글을 누르면 그 열로 정렬됩니다(한 번 더 누르면 반대 방향). 합계행은 항상 맨 아래입니다.<br>'
        '· 노출·클릭·광고비는 <b>매체 API 실집행</b>, 구매·매출은 <b>GA4</b> 기준입니다. '
        '매체 자체 전환수는 매체마다 집계 기준이 달라 서로 더할 수 없어 쓰지 않습니다.<br>'
        '· 네이버 브랜드검색은 정액(보장형) 상품이라 API가 집행액을 안 줍니다 — 노출·클릭만 API에서 오고 '
        '광고비는 위 칸에서 직접 넣습니다.<br>'
        '· 외부몰(쇼핑검색·맨즈탭 외부몰) 매출은 GA4에 안 잡힙니다 — 스마트스토어 연동 전까지 0으로 표시됩니다.'
        '</div>'
    )
    html = (
        CP_CSS
        + '<div class="cp-wrap"><table class="cp-tbl" id="cptbl"><thead><tr>' + th + '</tr></thead>'
        + '<tbody>' + "".join(body) + '</tbody></table>' + note + '</div>'
        + """
<script>
(function(){
  var tbl = document.getElementById('cptbl');
  var ths = tbl.tHead.rows[0].cells;
  var state = {i:-1, asc:false};
  function val(row, i){
    var c = row.cells[i];
    var raw = (c.getAttribute('data-v'));
    if(raw !== null) return parseFloat(raw);
    return c.innerText.trim();
  }
  for(var i=0;i<ths.length;i++){
    (function(i){
      ths[i].style.cursor='pointer';
      ths[i].addEventListener('click', function(){
        var asc = (state.i===i) ? !state.asc : false;
        state = {i:i, asc:asc};
        var tb = tbl.tBodies[0];
        var rows = Array.prototype.slice.call(tb.rows);
        var total = rows.filter(function(r){return r.classList.contains('cp-tot');});
        rows = rows.filter(function(r){return !r.classList.contains('cp-tot');});
        rows.sort(function(a,b){
          var x=val(a,i), y=val(b,i);
          if(typeof x==='number' && typeof y==='number'){
            if(isNaN(x)) x=-Infinity; if(isNaN(y)) y=-Infinity;
            return asc ? x-y : y-x;
          }
          return asc ? String(x).localeCompare(String(y),'ko')
                     : String(y).localeCompare(String(x),'ko');
        });
        rows.concat(total).forEach(function(r){tb.appendChild(r);});
        for(var k=0;k<ths.length;k++){
          ths[k].querySelector('.cp-ar').textContent = (k===i) ? (asc?'\u2191':'\u2193') : '\u21C5';
        }
      });
    })(i);
  }
})();
</script>"""
    )
    st.components.v1.html(html, height=260 + 46 * len(view), scrolling=True)

    # 어느 매체에도 안 잡힌 GA 유입을 드러낸다. 조용히 사라지면 합계가 안 맞는 걸 눈치채기 어렵다.
    un = ga_map.get("_미매칭", {})
    if un.get("rev", 0) > 0 or un.get("conv", 0) > 0:
        with st.expander(
            f"⚠️ 매체 미매칭 GA 유입 — 구매 {un.get('conv', 0):,.0f}건 · 매출 {un.get('rev', 0):,.0f}원"
        ):
            st.caption(
                "UTM 매칭 규칙에 안 걸린 소스/매체입니다. 광고가 아니라 자연유입·직접유입이면 정상이고, "
                "광고인데 여기 있으면 '매체 정의' 패널의 UTM 매칭에 추가해주세요."
            )
            keys = sorted(un.get("keys", {}).items(), key=lambda x: -x[1])[:30]
            st.dataframe(
                pd.DataFrame(keys, columns=["소스 / 매체", "GA 매출"]),
                use_container_width=True, hide_index=True,
            )

    st.markdown(CP_CSS + _cp_recommendations(view, start, end), unsafe_allow_html=True)

    with st.expander("🔎 GA 매출이 안 맞을 때 — 매체별 UTM 구성 보기"):
        st.caption(
            "GA4 화면에서 소스/매체 하나만 보고 비교하면 숫자가 안 맞습니다. "
            "대시보드는 UTM 매칭 규칙대로 여러 소스/매체를 한 매체로 묶기 때문입니다. "
            "아래에서 각 매체가 무엇으로 구성됐는지 확인하세요."
        )
        bd = _cp_ga_breakdown(ga_daily, mst, start, end)
        if bd.empty:
            st.info("이 기간에 GA 데이터가 없습니다.")
        else:
            st.dataframe(
                bd, use_container_width=True, hide_index=True,
                column_config={
                    "구매": st.column_config.NumberColumn("GA 구매", format="%d"),
                    "매출": st.column_config.NumberColumn("GA 매출", format="₩%d"),
                },
            )
            un = bd[bd["매체"] == "(미매칭)"]
            if not un.empty:
                st.warning(
                    f"어느 매체에도 안 걸린 소스/매체가 {len(un)}개 있습니다 "
                    f"(매출 {un['매출'].sum():,.0f}원). 광고인데 여기 있으면 "
                    "'⚙️ 매체 정의'의 UTM 매칭에 추가해주세요."
                )

    with st.expander("🔄 광고비 다시 받기 (수치가 안 맞을 때)"):
        st.caption(
            "평소 동기화는 최근 3일만 다시 받습니다. 컬럼을 새로 추가했거나(노출·클릭), "
            "매체 분류를 바꿨거나, 데이터를 지운 뒤에는 과거 날짜가 낡은 채로 남으니 여기서 기간을 "
            "지정해 통째로 다시 받으세요."
        )
        r1, r2, r3 = st.columns([1, 1, 1])
        with r1:
            rs = st.date_input("시작", start, key="cp_rs")
        with r2:
            re_ = st.date_input("종료", end, key="cp_re")
        with r3:
            st.write("")
            go = st.button("다시 받기", key="cp_resync", type="primary", use_container_width=True)
        if go:
            with st.spinner("매체별로 받아오는 중..."):
                saved, errors = force_resync_ad_spend(min(rs, re_), max(rs, re_))
            for k, v in saved.items():
                (st.success if v else st.info)(f"{k}: {v}행")
            for k, v in errors.items():
                st.warning(f"{k} 실패: {v}")
            st.cache_data.clear()
            st.info("위 표를 갱신하려면 화면을 새로고침하세요.")

    # 표를 좁게 만들지 않으려고 매체 정의는 화면 맨 아래 접힌 패널로 뺀다.
    with st.expander("⚙️ 매체 정의 (구분 · 광고비 출처 · 예산 배분 · UTM 매칭)"):
        st.caption(
            "이 표가 위 성과표의 행을 결정합니다. 매체를 추가하거나 이름을 바꾸려면 여기서 고치세요. "
            "코드 수정 없이 반영됩니다."
        )
        st.markdown(
            "- **광고비 출처**: `ad_spend_daily`의 채널명. 비워두면 광고비 0으로 잡히고, "
            "'광고비 직접 입력' 패널에서 넣은 값이 있으면 그게 쓰입니다.\n"
            "- **예산 라인 / 배분**: 채널믹스 파일의 매체명과 그 예산을 몇 %로 나눠 쓸지. "
            "예) `네이버 GFA` 예산을 자사몰 0.5 / 애드부스트 0.5로 쪼갬\n"
            "- **UTM 매칭**: GA 구매·매출을 이 매체로 묶을 기준. `source/medium`에 이 문자열이 들어 있으면 "
            "해당 매체로 집계합니다. 쉼표로 여러 개, **긴 패턴이 우선**입니다."
        )
        edited = st.data_editor(
            mst, num_rows="dynamic", use_container_width=True, key="cp_master_editor",
            column_config={
                "media": st.column_config.TextColumn("매체명", required=True),
                "scope": st.column_config.SelectboxColumn("구분", options=["자사몰", "외부몰"], required=True),
                "sort_order": st.column_config.NumberColumn("순서", format="%d", width="small"),
                "spend_channel": st.column_config.TextColumn("광고비 출처(채널명)"),
                "budget_line": st.column_config.TextColumn("예산 라인(채널믹스)"),
                "budget_share": st.column_config.NumberColumn("예산 배분", format="%.2f", min_value=0.0, max_value=1.0),
                "utm_match": st.column_config.TextColumn("UTM 매칭 (쉼표 구분)", width="large"),
                "budget_override": st.column_config.NumberColumn(
                    "예산 직접 지정", format="%d", min_value=0,
                    help="예산 파일에 줄이 없는 매체용. 0이면 예산 라인×배분을 씁니다."),
                "note": st.column_config.TextColumn("메모"),
            },
        )
        mc1, mc2 = st.columns([1, 4])
        with mc1:
            if st.button("저장", key="cp_master_save", type="primary", use_container_width=True):
                e = edited.dropna(subset=["media"])
                e = e[e["media"].astype(str).str.strip() != ""]
                if e.empty:
                    st.warning("저장할 행이 없습니다.")
                else:
                    n = save_table("media_master", e[MEDIA_MASTER_COLS], "media", "매체 정의")
                    st.success(f"{n}개 매체 저장했습니다.")
                    st.cache_data.clear()
        with mc2:
            st.caption("저장하면 위 표가 바로 이 정의대로 다시 그려집니다.")

def force_resync_ad_spend(start: date, end: date):
    """기간을 지정해 광고비를 처음부터 다시 받아온다.

    평소 동기화는 '마지막 저장일 - 2일'부터만 다시 받는다(매체 수치 보정 반영용).
    그래서 컬럼을 새로 추가했거나(노출·클릭), 잘못된 데이터를 지웠거나, 매체 분류 규칙을
    바꿨을 때는 과거 날짜가 낡은 채로 남는다. 그럴 때 쓰는 버튼이다.
    """
    saved, errors = {}, {}
    for label, fn in AD_SPEND_FETCHERS:
        try:
            df = fn(start, end)
        except Exception as e:
            errors[label] = str(e)[:250]
            continue
        if df is None or df.empty:
            saved[label] = 0
            continue
        saved[label] = save_table("ad_spend_daily", df,
                                  "report_date,channel,source", f"{label} API(재수집)")
    return saved, errors


# ══════════════════════════════════════════════════════════════
# 예산 관리 — "다음 주 예산을 어디로 옮길까"
#
# 채널 성과 탭이 '무엇이 잘 되고 있나'라면, 여기는 '그래서 돈을 어떻게 옮길까'다.
# 두 가지를 강제한다.
#   ① 증액분과 감액분의 합이 0 — 총예산을 넘기지 않는다.
#   ② 정액 계약 매체는 조정 대상에서 제외 — 계약 금액은 못 줄인다.
# 판단 기준은 GA-ROAS와 '집행 페이스'(기간 경과 대비 예산 소진).
# ══════════════════════════════════════════════════════════════

BR_AD_RATIO_TARGET = 0.14   # 자사몰 광고비 비율 목표(예산 파일의 '광고비 비율' 기준)


def _br_ratio_actual(budget: pd.DataFrame, year: int, scope: str = "자사몰"):
    """연간 예산 파일에서 누적 실매출과 누적 광고선전비를 꺼내 광고비 비율을 만든다."""
    if budget is None or budget.empty or "budget_cost" not in budget.columns:
        return None, None, None
    b = budget.copy()
    for c in ("year", "month"):
        b[c] = pd.to_numeric(b.get(c), errors="coerce")
    b = b[(b.get("scope") == scope) & (b["year"] == year) & (b["month"].between(1, 12))]
    if b.empty:
        return None, None, None
    rev = b[b["channel"] == BUDGET_SENTINEL_ACTUAL_REVENUE]["budget_cost"].sum()
    _c = b["channel"].astype(str)
    spend = b[(~_c.str.startswith("__")) & (_c != "TOTAL")
              & (~_c.str.contains(BUDGET_HARD_EXCLUDE_RE))]["budget_cost"].sum()
    # 실매출이 입력된 달까지만 비교해야 비율이 안 왜곡된다
    months = sorted(b[(b["channel"] == BUDGET_SENTINEL_ACTUAL_REVENUE)
                      & (b["budget_cost"] > 0)]["month"].unique())
    if months:
        spend = b[(b["month"].isin(months)) & (~_c.str.startswith("__"))
                  & (_c != "TOTAL")
                  & (~_c.str.contains(BUDGET_HARD_EXCLUDE_RE))]["budget_cost"].sum()
    return (spend / rev if rev else None), spend, rev


def _br_year_matrix(budget, year: int, ref_month: int, scope: str = "자사몰") -> str:
    """연도별 월×매체 예산표. 예산 파일의 '매체 세부내역' 블록을 그대로 옮긴 화면.

    같은 (매체, 월)이 두 번 들어오면 합쳐져서 금액이 배로 부푼다. 파일을 여러 번 올리거나
    예전 업로드분이 남아 있으면 실제로 그런 일이 생겨서, 그릴 때 중복을 먼저 걷어낸다.
    """
    if budget is None or budget.empty or "budget_cost" not in budget.columns:
        return '<div class="cp-note">연간 예산 파일을 올리면 표시됩니다.</div>'
    b = budget.copy()
    for c in ("year", "month"):
        b[c] = pd.to_numeric(b.get(c), errors="coerce")
    # month 0 = 파일의 'TOTAL' 열, -1 = '당월 누계' 열. 촬영샘플·잔여비용은 월별로 안 나뉘고
    # TOTAL에만 값이 있어서, 월 합계만 보면 이 둘이 통째로 사라진다.
    b = b[(b.get("scope") == scope) & (b["year"] == year) & (b["month"].between(0, 12))]
    if b.empty:
        return '<div class="cp-note">해당 연도 예산 데이터가 없습니다.</div>'

    dup_note = ""
    n_before = len(b)
    b = b.drop_duplicates(subset=["channel", "month"], keep="last")
    if len(b) < n_before:
        dup_note = (f'<div class="br-warn">같은 매체·월 데이터가 {n_before - len(b)}건 중복돼 있어 '
                    '최신 것만 썼습니다. 예산 파일을 다시 올리면 깔끔해집니다.</div>')

    rev = b[b["channel"] == BUDGET_SENTINEL_ACTUAL_REVENUE].set_index("month")["budget_cost"]
    # 예전 파서로 저장된 데이터에는 '잔여 광고비 (7월~12월)', '매출 달성율' 같은 요약 줄이
    # 매체로 섞여 있다. 그대로 더하면 연간 총액이 실제의 1.5배가 된다.
    # 파일을 다시 안 올려도 되도록 그릴 때 한 번 더 걸러낸다.
    ch = b["channel"].astype(str)
    media = b[(~ch.str.startswith("__")) & (ch != "TOTAL")
              & (~ch.str.contains(BUDGET_HARD_EXCLUDE_RE))]
    if media.empty:
        return '<div class="cp-note">매체별 예산 데이터가 없습니다.</div>'

    piv = media.pivot_table(index="channel", columns="month", values="budget_cost",
                            aggfunc="sum", fill_value=0)
    for m in range(0, 13):
        if m not in piv.columns:
            piv[m] = 0
    piv = piv[sorted(piv.columns)]
    # TOTAL은 파일에 적힌 값(0열)을 그대로 쓴다. 월 합계로 계산하면 월별 배분이 없는
    # 항목(촬영샘플·잔여비용)이 0이 되어 연간 총액이 안 맞는다.
    month_sum = piv[list(range(1, 13))].sum(axis=1)
    piv["TOTAL"] = np.where(piv[0] > 0, piv[0], month_sum)
    piv["누계"] = piv[[m for m in range(1, ref_month + 1)]].sum(axis=1)
    piv = piv[piv["TOTAL"] > 0].sort_values("TOTAL", ascending=False)

    def w(v):
        return f"{v:,.0f}" if v else '<span class="cp-mute">-</span>'

    heads = (['<th class="l">매체</th>', "<th>TOTAL</th>", f"<th>~{ref_month}월 누계</th>"]
             + [f'<th class="{"br-now" if m == ref_month else ""}">{m}월</th>'
                for m in range(1, 13)])
    tot = piv.sum()

    # 광고비 비율(광고비 ÷ 실매출)은 예산 파일의 회사 목표(14%)와 직접 비교되는 줄이라 맨 위에 둔다.
    ratio_cells = ""
    for m in range(1, 13):
        r = float(rev.get(m, 0) or 0)
        val = (tot[m] / r * 100) if r > 0 else None
        cls = "br-now" if m == ref_month else ""
        if val is None:
            ratio_cells += f'<td class="{cls}"><span class="cp-mute">-</span></td>'
        else:
            over = " br-over" if val > BR_AD_RATIO_TARGET * 100 else ""
            ratio_cells += f'<td class="{cls}{over}">{val:.1f}%</td>'
    rev_done = float(rev[rev > 0].sum()) if len(rev) else 0
    spend_done = float(sum(tot[m] for m in range(1, 13) if float(rev.get(m, 0) or 0) > 0))
    ratio_all = (spend_done / rev_done * 100) if rev_done > 0 else None
    rows = [f'<tr class="br-ratio-row"><td class="l">광고비 비율 '
            f'<span class="br-tiny">목표 {BR_AD_RATIO_TARGET*100:.0f}%</span></td>'
            f'<td class="m">{f"{ratio_all:.1f}%" if ratio_all else "-"}</td>'
            f'<td><span class="cp-mute">-</span></td>{ratio_cells}</tr>']

    for ch, r in piv.iterrows():
        cells = "".join(
            f'<td class="{"br-now" if m == ref_month else ""}">{w(r[m])}</td>'
            for m in range(1, 13))
        rows.append(f'<tr><td class="l m">{ch}</td>'
                    f'<td class="m">{w(r["TOTAL"])}</td><td>{w(r["누계"])}</td>{cells}</tr>')
    rows.append(
        '<tr class="cp-tot"><td class="l">합계</td>'
        f'<td>{w(tot["TOTAL"])}</td><td>{w(tot["누계"])}</td>'
        + "".join(f'<td class="{"br-now" if m == ref_month else ""}">{w(tot[m])}</td>'
                  for m in range(1, 13)) + "</tr>")

    return (dup_note
            + '<div class="br-scroll">'
            f'<table class="cp-tbl br-matrix"><thead><tr>{"".join(heads)}</tr></thead>'
            f'<tbody>{"".join(rows)}</tbody></table></div>'
            '<div class="cp-note">단위: 원 (VAT 포함) · 파란 열이 이번 달입니다 · '
            f'{year}년 {scope} 기준 · 광고비 비율이 목표({BR_AD_RATIO_TARGET*100:.0f}%)를 '
            '넘는 달은 붉게 표시됩니다.</div>')



def render_budget_realloc_page(ad_spend, ga_daily, channel_mix, budget=None,
                               master=None, decisions=None):
    st.markdown(CP_CSS + BR_CSS, unsafe_allow_html=True)
    mst = media_master_frame(master)
    today = date.today()

    c1, c2 = st.columns([1, 3])
    with c1:
        ym = st.date_input("기준 월", today.replace(day=1), key="br_month", format="YYYY-MM-DD")
    ref = pd.to_datetime(ym).date().replace(day=1)
    last_day = (ref.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    end = min(today - timedelta(days=1), last_day)
    days_total = last_day.day
    days_done = max((end - ref).days + 1, 0)
    days_left = days_total - days_done

    spend = _cp_spend_by_channel(ad_spend, ref, end)
    spend_map = {r["channel"]: r for _, r in spend.iterrows()} if not spend.empty else {}
    ga_map = _cp_ga_by_media(ga_daily, mst, ref, end)
    bmap = _cp_month_budget(channel_mix, ref, budget)

    rows = []
    for _, r in mst.iterrows():
        sc = str(r.get("spend_channel") or "").strip()
        sp = spend_map.get(sc, {})
        bud = _media_budget(r, bmap)
        g = ga_map.get(r["media"], {"conv": 0.0, "rev": 0.0})
        cost = float(sp.get("cost_incl_vat", 0) or 0)
        src = str(sp.get("source", "") or "")
        rows.append({
            "매체": r["media"], "구분": r.get("scope", "자사몰"),
            "집행": cost, "예산": bud,
            "예산 소진율": (cost / bud * 100) if bud > 0 else None,
            "ROAS": (g["rev"] / cost * 100) if cost > 0 else None,
            "출처": src,
            # 계약 저장 전이라 출처가 아직 안 붙었어도, 정액 상품이면 조정 대상이 아니다.
            "고정": (src == "contract") or (r["media"] in CONTRACT_CHANNELS),
        })
    df = pd.DataFrame(rows)
    df = df[(df["예산"] > 0) | (df["집행"] > 0)].reset_index(drop=True)

    tot_b, tot_c = df["예산"].sum(), df["집행"].sum()
    pace = (tot_c / tot_b * 100) if tot_b > 0 else 0
    elapsed = (days_done / days_total * 100) if days_total else 0
    ratio, r_spend, r_rev = _br_ratio_actual(budget, ref.year)

    st.markdown(
        '<div class="cp-wrap"><div class="cp-eyebrow">BUDGET REALLOCATION · '
        f'{ref.year}년 {ref.month}월</div>'
        '<div class="cp-title">다음 주 예산을 어디로 옮길까요?</div>'
        '<div class="cp-kpis">'
        f'<div class="cp-kpi"><div class="k">월 예산</div><div class="v">{_cp_won(tot_b)}</div>'
        f'<div class="s">연간 예산 파일 기준</div></div>'
        f'<div class="cp-kpi"><div class="k">현재 집행</div><div class="v">{_cp_won(tot_c)}</div>'
        f'<div class="s">{ref} ~ {end} · {pace:.1f}%</div></div>'
        f'<div class="cp-kpi"><div class="k">남은 예산</div><div class="v">{_cp_won(tot_b - tot_c)}</div>'
        f'<div class="s">남은 {days_left}일 · 일 {_cp_won((tot_b - tot_c) / days_left) if days_left else "-"}</div></div>'
        f'<div class="cp-kpi"><div class="k">예산 소진율</div>'
        f'<div class="v">{pace:.1f}%</div>'
        f'<div class="br-pace"><i style="width:{min(pace,100):.0f}%"></i>'
        f'<u style="left:{min(elapsed,100):.0f}%"></u></div>'
        f'<div class="s">쓴 돈 ÷ 월 예산 · '
        f'<b>{"과속" if pace > elapsed + 5 else ("저속" if pace < elapsed - 5 else "정상")}</b></div></div>'
        f'<div class="cp-kpi"><div class="k">기간 경과</div>'
        f'<div class="v">{elapsed:.0f}%</div>'
        f'<div class="s">{ref.month}/1~{end.month}/{end.day} · '
        f'{last_day.day}일 중 {days_done}일</div></div>'
        + (f'<div class="cp-kpi"><div class="k">누적 광고비 비율</div>'
           f'<div class="v">{ratio*100:.1f}%</div>'
           f'<div class="s">목표 {BR_AD_RATIO_TARGET*100:.0f}% · '
           f'{"초과" if ratio > BR_AD_RATIO_TARGET else "여유"} '
           f'{abs(ratio - BR_AD_RATIO_TARGET)*100:.1f}%p</div></div>'
           if ratio is not None else '')
        + '</div></div>', unsafe_allow_html=True)


    def verdict(r):
        """조정 대상인지, 어느 방향인지 정한다.

        '판단할 수 없는 경우'를 먼저 걸러내는 게 중요하다. 외부몰은 매출이 GA4에 안 잡혀
        ROAS가 0으로 보이는데, 그대로 읽으면 '감액'이라는 틀린 지시가 나간다.
        정액 계약 매체는 금액이 계약으로 묶여 있어 애초에 조정할 수 없다.
        """
        if r["구분"] == "외부몰":
            return "판단 제외"
        if r["고정"]:
            return "계약 고정"
        if r["예산"] <= 0:
            return "예산 없음"
        if r["집행"] < CP_MIN_SPEND_FOR_JUDGE:
            return "판단 보류"
        roas = r["ROAS"]
        if roas is None or pd.isna(roas):
            return "판단 보류"
        if roas >= OPS_KPI_ROAS_HIGH:
            return "증액"
        if roas < OPS_KPI_ROAS_LOW:
            return "감액"
        return "유지"

    df["판정"] = df.apply(verdict, axis=1)

    # ── 표 ① 연간 월별 예산 ──────────────────────────────
    st.markdown(f"#### {ref.year}년 월별 · 매체별 예산")
    st.markdown(CP_CSS + BR_CSS
                + '<div class="cp-wrap">'
                + _br_year_matrix(budget, ref.year, ref.month)
                + '</div>', unsafe_allow_html=True)

    # ── 표 ② 이번 달 매체별 집행 현황 ─────────────────────
    st.markdown(f"#### {ref.month}월 매체별 집행 현황")
    st.markdown(
        f'<div class="stco-daterange">📆 {ref.year}-{ref.month:02d}-01 ~ {end:%Y-%m-%d}'
        f' · 예산 소진율 {pace:.1f}% · 기간 경과 {elapsed:.0f}%</div>',
        unsafe_allow_html=True,
    )
    VTAG = {"증액": "up", "감액": "down", "유지": "keep",
            "계약 고정": "hold", "판단 보류": "hold",
            "판단 제외": "skip", "예산 없음": "skip"}
    heads = ["매체", "판정", "월 예산", "집행", "예산 소진율", "남은 예산", "GA ROAS"]
    th = "".join(f'<th class="{"l" if h in ("매체", "판정") else ""}">{h}</th>' for h in heads)
    body = []
    for scope_name in ["자사몰", "외부몰"]:
        sub = df[df["구분"] == scope_name]
        if sub.empty:
            continue
        s_b, s_c = sub["예산"].sum(), sub["집행"].sum()
        body.append(
            f'<tr class="br-sub"><td class="l" colspan="2">{scope_name}</td>'
            f'<td>{_cp_won(s_b)}</td><td>{_cp_won(s_c)}</td>'
            f'<td>{(s_c / s_b * 100) if s_b else 0:.1f}%</td>'
            f'<td>{_cp_won(s_b - s_c)}</td><td></td></tr>')
        for _, r in sub.sort_values("집행", ascending=False).iterrows():
            pct = ('<span class="cp-mute">—</span>'
                   if r["예산 소진율"] is None or pd.isna(r["예산 소진율"]) else f'{r["예산 소진율"]:.1f}%')
            roas = ('<span class="cp-mute">—</span>'
                    if r["ROAS"] is None or pd.isna(r["ROAS"]) else f'{r["ROAS"]:,.0f}%')
            body.append(
                f'<tr><td class="l m"><span class="br-ind">{r["매체"]}</span></td>'
                f'<td class="l"><span class="cp-tag {VTAG.get(r["판정"], "hold")}">{r["판정"]}</span></td>'
                f'<td>{_cp_won(r["예산"])}</td><td>{_cp_won(r["집행"])}</td>'
                f'<td>{pct}</td><td>{_cp_won(r["예산"] - r["집행"])}</td>'
                f'<td>{roas}</td></tr>')
    t_b, t_c = df["예산"].sum(), df["집행"].sum()
    body.append(
        f'<tr class="cp-tot"><td class="l" colspan="2">전체 합계</td>'
        f'<td>{_cp_won(t_b)}</td><td>{_cp_won(t_c)}</td>'
        f'<td>{(t_c / t_b * 100) if t_b else 0:.1f}%</td>'
        f'<td>{_cp_won(t_b - t_c)}</td><td></td></tr>')
    st.markdown(
        f'<div class="cp-wrap"><table class="cp-tbl"><thead><tr>{th}</tr></thead>'
        f'<tbody>{"".join(body)}</tbody></table></div>', unsafe_allow_html=True)

    with st.expander("📊 연간 채널 믹스 (매체별 연간 예산 비중)"):
        st.markdown(FUNNEL_V4_CSS + _channel_mix_panel(channel_mix), unsafe_allow_html=True)

    st.markdown(
        '<div class="cp-wrap"><div class="cp-note">'
        '· 월 예산은 <b>◆26년 월별 예산 정리</b> 파일 기준입니다 (없으면 채널 믹스로 대체).<br>'
        '· <b>집행 페이스</b>는 기간 경과율과 비교합니다 — 기간 42%인데 예산 58%를 썼다면 남은 날을 못 버팁니다.<br>'
        f'· <b>광고비 비율</b>은 예산 파일의 목표({BR_AD_RATIO_TARGET*100:.0f}%)와 누적 실적을 비교합니다. '
        '실매출이 입력된 달까지만 계산해 비율이 왜곡되지 않게 했습니다.<br>'
        '· 확정 내역은 <b>의사결정 기록</b>에 저장되고, 결정일 전후 7일 성과가 자동 비교됩니다.'
        '</div></div>', unsafe_allow_html=True)


BR_CSS = """
<style>
.br-pace{height:6px;background:#EFEDE8;border-radius:3px;margin-top:8px;overflow:hidden;position:relative}
.br-pace i{display:block;height:100%;background:#14181F;border-radius:3px}
.br-pace u{position:absolute;top:-3px;width:2px;height:12px;background:#E0654A;text-decoration:none}
.br-ratio{border-radius:10px;padding:11px 15px;font-size:14px;margin-bottom:14px}
.br-ratio.ok{background:#EAF7D9;color:#3F6B12}
.br-ratio.over{background:#FDEDE3;color:#9A4B14}
.br-bal{border-radius:10px;padding:14px 18px;margin:12px 0;display:flex;
  justify-content:space-between;align-items:center}
.br-bal.ok{background:#D9F27E}
.br-bal.bad{background:#F1EFEA}
.br-bal .l1{font-size:13px;font-weight:800;letter-spacing:.12em;color:#4A5A16}
.br-bal .l2{font-size:15px;font-weight:800;color:#14181F;margin-top:2px}
.br-bal .l3{font-size:13.5px;font-weight:700;color:#4A5A16}
.br-bal.bad .l1,.br-bal.bad .l3{color:#8A9099}
.br-chip{display:inline-block;padding:3px 9px;border-radius:6px;font-size:13px;font-weight:800}
.br-chip.up{background:#EAF7D9;color:#3F6B12;border:1px solid #B7DE7E}
.br-chip.dn{background:#FDEDE3;color:#9A4B14;border:1px solid #E9C0A0}
.br-sec{font-size:14px;font-weight:800;color:#14181F;margin:18px 0 7px;letter-spacing:-.01em}
.br-sub td{background:#EFEDE8;font-weight:800;font-size:14px;border-top:1px solid #DDD9D1}
.br-scroll{overflow-x:auto;-webkit-overflow-scrolling:touch}
.br-matrix{font-size:13.5px;min-width:1260px}
.br-matrix th,.br-matrix td{padding:6px 7px;white-space:nowrap}
.br-now{background:#EAF1FD !important;font-weight:800}
th.br-now{background:#DCE8FA !important;color:#2C4E86}
.br-ratio-row td{background:#FAF8F2;font-weight:800;border-bottom:2px solid #E3E1DC}
.br-ratio-row td.br-over{color:#B0431A}
.br-tiny{font-size:13px;color:#7A8088;font-weight:600;margin-left:6px}
.br-warn{background:#FDEDE3;color:#9A4B14;border-radius:8px;padding:9px 13px;
  font-size:13.5px;margin-bottom:10px}
.br-ind{padding-left:12px;display:inline-block}
</style>
"""


def resolve_channel_spend(ad_actual: pd.DataFrame, channels_weekly: pd.DataFrame,
                          channel_mix: pd.DataFrame, start: date, end: date):
    """채널별 광고비를 3단 폴백으로 확정한다. (DataFrame[channel, cost_incl_vat, source], 요약문자열)
    같은 채널에 여러 소스가 있으면 API > 대행사 주간 > 예산 일할 순으로 하나만 채택한다."""
    api_df = pd.DataFrame(columns=["channel", "cost_incl_vat", "source"])
    if ad_actual is not None and not ad_actual.empty:
        a = ad_actual.copy()
        a["report_date"] = pd.to_datetime(a["report_date"]).dt.date
        a = a[(a["report_date"] >= start) & (a["report_date"] <= end)]
        if not a.empty:
            a["channel"] = a["channel"].map(_v4_canon_channel)
            api_df = a.groupby(["channel", "source"], as_index=False).agg(
                cost_incl_vat=("cost_incl_vat", "sum"))

    weekly_df = _channel_spend_total(channels_weekly, start, end)
    if weekly_df is not None and not weekly_df.empty:
        weekly_df = weekly_df.copy()
        weekly_df["channel"] = weekly_df["channel"].map(_v4_canon_channel)
        weekly_df = weekly_df.groupby("channel", as_index=False).agg(cost_incl_vat=("cost_incl_vat", "sum"))
        weekly_df["source"] = "agency_weekly"
    else:
        weekly_df = pd.DataFrame(columns=["channel", "cost_incl_vat", "source"])

    budget_df = _budget_daily_spend(channel_mix, start, end)
    if budget_df is not None and not budget_df.empty:
        budget_df = budget_df.copy()
        budget_df["source"] = "budget_prorate"
    else:
        budget_df = pd.DataFrame(columns=["channel", "cost_incl_vat", "source"])

    # 직접 입력은 광고 관리자 화면에서 눈으로 확인한 '실집행액'이라 대행사 주간·예산 일할보다 정확하다.
    # 다만 API가 붙은 매체는 사람이 손댈 필요 없이 API가 이기게 둔다.
    prio = {"meta_api": 0, "google_ads_api": 0, "naver_api": 0, "kakao_api": 0, "criteo_api": 0, "naver_gfa_api": 0,
            "contract": 1, "manual": 2, "agency_weekly": 3, "budget_prorate": 4}
    # 빈 DataFrame을 concat하면 pandas가 향후 동작 변경 경고를 내므로 값이 있는 것만 합친다.
    parts = [d for d in (api_df, weekly_df, budget_df) if d is not None and not d.empty]
    stacked = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame(columns=["channel", "cost_incl_vat", "source"])
    if stacked.empty:
        return pd.DataFrame(columns=["channel", "cost_incl_vat", "source"]), "광고비 데이터 없음"
    stacked["_p"] = stacked["source"].map(prio).fillna(9)
    stacked = stacked.sort_values("_p").drop_duplicates(subset=["channel"], keep="first").drop(columns="_p")

    counts = stacked["source"].value_counts().to_dict()
    summary = " · ".join(
        f"{AD_SPEND_SOURCE_LABEL.get(k, k)} {v}개" for k, v in counts.items()
    )
    return stacked.reset_index(drop=True), summary


def diagnose_ad_spend_setup() -> str:
    """광고비 매체별 연동 상태 진단. 비밀값은 출력하지 않는다."""
    L = []
    ok = lambda s: L.append(f"✅ {s}")
    ng = lambda s: L.append(f"❌ {s}")
    info = lambda s: L.append(f"   ↳ {s}")
    try:
        import requests  # noqa: F401
        ok("0. requests 설치됨")
    except Exception:
        ng("0. requests 없음 → requirements.txt에 requests 추가 필요")
        return "\n".join(L)

    test_end = date.today() - timedelta(days=1)
    test_start = test_end - timedelta(days=6)
    L.append(f"   ↳ 진단 조회 기간: {test_start} ~ {test_end} (어제까지 7일)")

    meta = _secrets_section("meta_ads")
    if not meta:
        ng("1. 메타: Secrets에 [meta_ads] 없음 (예산 일할값으로 대체 중)")
    else:
        miss = [k for k in ["access_token", "ad_account_id"] if not meta.get(k)]
        if miss:
            ng(f"1. 메타: 빠진 항목 {', '.join(miss)}")
        else:
            info(f"메타 광고계정: {meta.get('ad_account_id')}")
            try:
                df = fetch_meta_spend(test_start, test_end)
                if df.empty:
                    ng(f"1. 메타: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    ok(f"1. 메타 연동 성공: {len(df)}일, 합계 {df['cost_incl_vat'].sum():,.0f}원")
            except Exception as e:
                ng(f"1. 메타 조회 실패: {str(e)[:220]}")

    g = _secrets_section("google_ads")
    if not g:
        ng("2. 구글: Secrets에 [google_ads] 없음 (예산 일할값으로 대체 중)")
    else:
        need = ["developer_token", "client_id", "client_secret", "refresh_token", "customer_id"]
        miss = [k for k in need if not g.get(k)]
        if miss:
            ng(f"2. 구글: 빠진 항목 {', '.join(miss)}")
        else:
            info(f"구글 고객ID: {str(g.get('customer_id'))[:3]}****")
            try:
                df = fetch_google_ads_spend(test_start, test_end)
                if df.empty:
                    ng(f"2. 구글: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    ok(f"2. 구글 연동 성공: {len(df)}일, 합계 {df['cost_incl_vat'].sum():,.0f}원")
            except Exception as e:
                msg = str(e)[:250]
                ng(f"2. 구글 조회 실패: {msg}")
                if "DEVELOPER_TOKEN" in msg or "developer" in msg.lower():
                    info("→ 개발자 토큰이 아직 승인 안 됐을 수 있습니다(구글 심사 며칠 소요)")

    nv = _secrets_section("naver_searchad")
    if not nv:
        ng("2-1. 네이버 검색광고: Secrets에 [naver_searchad] 없음 (예산 일할값으로 대체 중)")
    else:
        miss = [k for k in ["api_key", "secret_key", "customer_id"] if not nv.get(k)]
        if miss:
            ng(f"2-1. 네이버: 빠진 항목 {', '.join(miss)}")
        else:
            info(f"네이버 CUSTOMER_ID: {nv.get('customer_id')}")
            try:
                df = fetch_naver_search_spend(test_start, test_end)
                if df.empty:
                    ng(f"2-1. 네이버: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    ok(f"2-1. 네이버 연동 성공: {len(df)}행, 합계 {df['cost_incl_vat'].sum():,.0f}원")
            except Exception as e:
                ng(f"2-1. 네이버 조회 실패: {str(e)[:220]}")

    kk = _secrets_section("kakao_moment")
    if not kk:
        ng("2-2. 카카오: Secrets에 [kakao_moment] 없음 (예산 일할값으로 대체 중)")
    else:
        if not kk.get("ad_account_id"):
            ng("2-2. 카카오: ad_account_id 없음")
        else:
            info(f"카카오 광고계정: {kk.get('ad_account_id')}")
            try:
                df = fetch_kakao_moment_spend(test_start, test_end)
                if df.empty:
                    ng(f"2-2. 카카오: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    ok(f"2-2. 카카오 연동 성공: {len(df)}일, 합계 {df['cost_incl_vat'].sum():,.0f}원")
            except Exception as e:
                msg = str(e)[:220]
                ng(f"2-2. 카카오 조회 실패: {msg}")
                if "권한" in msg or "401" in msg or "403" in msg:
                    info("→ 카카오모먼트 API 사용 권한 신청/승인이 필요합니다")

    cr = _secrets_section("criteo")
    if not cr:
        ng("2-3. 크리테오: Secrets에 [criteo] 없음 (예산 일할값으로 대체 중)")
    else:
        miss = [k for k in ["client_id", "client_secret"] if not cr.get(k)]
        if miss:
            ng(f"2-3. 크리테오: 빠진 항목 {', '.join(miss)}")
        else:
            info(f"크리테오 API 버전: {cr.get('api_version', CRITEO_DEFAULT_VERSION)}")
            try:
                df = fetch_criteo_spend(test_start, test_end)
                if df.empty:
                    ng(f"2-3. 크리테오: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    ok(f"2-3. 크리테오 연동 성공: {len(df)}일, 합계 {df['cost_incl_vat'].sum():,.0f}원")
            except Exception as e:
                msg = str(e)[:220]
                ng(f"2-3. 크리테오 조회 실패: {msg}")
                if "404" in msg:
                    info("→ api_version이 안 맞을 수 있습니다. Secrets에서 api_version을 바꿔보세요(예: 2025-01)")
                elif "consent" in msg.lower() or "403" in msg:
                    info("→ 광고주 자산에 대한 앱 동의(consent)가 필요합니다")

    gf = _secrets_section("naver_gfa")
    if not gf:
        ng("2-4. 네이버 GFA: Secrets에 [naver_gfa] 없음 (예산 일할값으로 대체 중)")
        info("→ API 신청은 네이버 공식 파트너사만 가능합니다. 대행사에서 발급받은 토큰을 넣어야 합니다")
    else:
        if not gf.get("ad_account_no"):
            ng("2-4. GFA: ad_account_no 없음")
        else:
            info(f"GFA 광고계정: {gf.get('ad_account_no')}"
                 + (f" / 관리계정: {gf.get('manager_account_no')}" if gf.get("manager_account_no") else ""))
            try:
                df = fetch_naver_gfa_spend(test_start, test_end)
                if df.empty:
                    ng(f"2-4. GFA: 응답은 왔지만 데이터 0행 ({test_start}~{test_end})")
                else:
                    by = df.groupby("channel")["cost_incl_vat"].sum().to_dict()
                    ok(f"2-4. GFA 연동 성공: {len(df)}행 · " +
                       " / ".join(f"{k} {v:,.0f}원" for k, v in by.items()))
            except Exception as e:
                msg = str(e)[:220]
                ng(f"2-4. GFA 조회 실패: {msg}")
                if "401" in msg or "token" in msg.lower():
                    info("→ 액세스 토큰이 만료됐을 수 있습니다(refresh_token 방식 권장)")
                elif "403" in msg or "권한" in msg:
                    info("→ 관리계정 번호(manager_account_no)가 필요하거나 권한이 없습니다")
    ng("2-5. 네이버 맨즈탭(보장형) / 카카오톡 플친(채널 월렛): 광고비 조회 API 자체가 없음 → 예산 일할값 사용")

    sb = get_supabase_client()
    if sb is None:
        ng("3. Supabase 연결 없음 (로컬 세션 모드)")
    else:
        try:
            sb.table("ad_spend_daily").select("report_date").limit(1).execute()
            ok("3. Supabase ad_spend_daily 테이블 접근 가능")
        except Exception as e:
            ng(f"3. ad_spend_daily 테이블 없음 → ad_spend_table.sql 실행 필요 ({str(e)[:150]})")
    return "\n".join(L)



# ──────────────────────────────────────────────────────────────
# GA 소재별 성과 — utm_campaign / utm_content 단위
# ──────────────────────────────────────────────────────────────
GC_HEAD = ["구분", "방문", "신규 방문", "회원가입", "가입률",
           "구매", "구매율", "매출", "객단가", "추정 광고비", "추정 ROAS", "판정"]

# utm_content 작명 규칙: 그룹명(타겟팅명)_날짜_소재이름
#   예) 패션관심타겟_260807_수피마티셔츠
# 소재 이름에도 '_'가 들어갈 수 있어서 앞에서부터 자르면 안 된다. 가운데의 날짜 토큰
# (6자리 YYMMDD 또는 8자리 YYYYMMDD)을 기준으로 앞=타겟팅, 뒤=소재로 가른다.
GC_DATE_TOKEN = re.compile(r"^(?:\d{6}|\d{8})$")

# 소재별 화면에서 기본으로 빼는 매체.
# 맨즈탭은 소재 운영·정산을 별도 시트로 관리해서, 여기 섞이면 오히려 헷갈린다.
# 화면의 '제외할 매체'에서 언제든 넣었다 뺐다 할 수 있다.
GC_DEFAULT_EXCLUDE = ["네이버 맨즈탭"]


def _gc_parse_content(v):
    """utm_content → (타겟팅, 등록일, 소재명). 규칙에 안 맞으면 통째로 소재명으로 둔다."""
    raw = str(v or "").strip()
    if not raw or raw == "(미설정)":
        return "(미설정)", "", "(미설정)"
    # GA4가 퍼센트 인코딩된 채로 주는 경우가 있다(%ED%8C%A8...). 사람이 읽게 풀어준다.
    if "%" in raw:
        try:
            from urllib.parse import unquote
            raw = unquote(raw)
        except Exception:
            pass
    parts = raw.split("_")
    idx = next((i for i, t in enumerate(parts) if GC_DATE_TOKEN.match(t)), None)
    if idx is None or idx == 0 or idx == len(parts) - 1:
        # 날짜가 없거나 맨 앞/맨 뒤면 규칙 밖이다 — 억지로 쪼개면 엉뚱한 타겟팅이 생긴다
        return "(규칙 외)", "", raw
    target = "_".join(parts[:idx])
    ymd = parts[idx]
    name = "_".join(parts[idx + 1:])
    if len(ymd) == 6:
        ymd = f"20{ymd[:2]}-{ymd[2:4]}-{ymd[4:]}"
    else:
        ymd = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:]}"
    return target or "(미설정)", ymd, name or "(미설정)"


def _gc_rows(cre: pd.DataFrame, start: date, end: date, level: str,
             exclude=None) -> pd.DataFrame:
    """소재 데이터를 원하는 단위(매체/캠페인/소재)로 접는다."""
    cols = ["key", "channel", "sessions", "new", "signup", "conv", "rev"]
    if cre is None or cre.empty:
        return pd.DataFrame(columns=cols)
    g = cre.copy()
    g["report_date"] = pd.to_datetime(g["report_date"], errors="coerce").dt.date
    g = g[(g["report_date"] >= start) & (g["report_date"] <= end)]
    if g.empty:
        return pd.DataFrame(columns=cols)
    for c in ("sessions", "conversions", "revenue", "signups"):
        g[c] = pd.to_numeric(g.get(c), errors="coerce").fillna(0).astype(float)
    # 채널 퍼널·채널 성과와 같은 기준으로 광고 매체만 남긴다.
    # UTM 매핑(utm_channel_map)에 있는 소스/매체 = 우리가 돈 내고 돌리는 광고다.
    # 매핑이 안 된 건 organic·referral·direct·(not set)이라 소재 성과로 볼 게 없다.
    g = g[g.apply(classify_ga_bucket, axis=1) == "광고"]
    if g.empty:
        return pd.DataFrame(columns=cols)
    g["channel"] = g["channel"].map(_v4_canon_channel)
    if exclude:
        g = g[~g["channel"].isin(set(exclude))]
        if g.empty:
            return pd.DataFrame(columns=cols)
    is_new = g["user_type"] == "신규"
    g["new"] = np.where(is_new, g["sessions"], 0.0)
    # 회원가입은 신규 유입의 성과로 본다 (재방문자의 가입은 사실상 없다)
    g["signup"] = np.where(is_new, g["signups"], 0.0)

    parsed = g["creative"].map(_gc_parse_content)
    g["target"] = parsed.map(lambda t: t[0])
    g["cre_date"] = parsed.map(lambda t: t[1])
    g["cre_name"] = parsed.map(lambda t: t[2])

    if level == "매체":
        g["key"] = g["channel"]
    elif level == "캠페인":
        g["key"] = g["channel"] + " · " + g["campaign"]
    elif level == "타겟팅":
        g["key"] = g["channel"] + " · " + g["target"]
    else:
        # 소재명은 같은데 타겟팅이 다르면 성과가 다르다 — 둘을 붙여야 한 줄이 한 소재가 된다
        g["key"] = (g["cre_name"] + '<span class="gc-sub">'
                    + g["channel"] + " · " + g["target"]
                    + np.where(g["cre_date"] != "", " · " + g["cre_date"], "")
                    + "</span>")

    out = g.groupby(["key", "channel"], as_index=False).agg(
        sessions=("sessions", "sum"), new=("new", "sum"), signup=("signup", "sum"),
        conv=("conversions", "sum"), rev=("revenue", "sum"))
    return out.sort_values("rev", ascending=False)[cols]


def _gc_row_html(r, cost, extra_cls="") -> str:
    """소재 표의 한 줄. 광고비가 소재 단위로 없어서 방문 비중으로 나눈 '추정'이다."""
    ses = float(r["sessions"] or 0)
    new = float(r["new"] or 0)
    sign = float(r["signup"] or 0)
    conv = float(r["conv"] or 0)
    rev = float(r["rev"] or 0)
    s_rate = (sign / new * 100) if new else 0.0
    b_rate = (conv / ses * 100) if ses else 0.0
    aov = (rev / conv) if conv else 0.0
    roas = (rev / cost * 100) if cost > 0 else 0.0
    roas_txt = f"{roas:,.0f}%" if cost > 0 else "-"
    label, cls = _v4_verdict(roas, cost)
    name = r["key"] if isinstance(r, dict) or "key" in r else ""
    return (
        f'<tr class="{extra_cls}">'
        f'<td class="l">{name}</td>'
        f'<td data-v="{ses:.0f}">{_v4_num(ses)}</td>'
        f'<td data-v="{new:.0f}">{_v4_num(new)}</td>'
        f'<td data-v="{sign:.0f}">{_v4_num(sign)}</td>'
        f'<td data-v="{s_rate:.3f}">{s_rate:.2f}%</td>'
        f'<td data-v="{conv:.0f}">{_v4_num(conv)}</td>'
        f'<td data-v="{b_rate:.3f}">{b_rate:.2f}%</td>'
        f'<td data-v="{rev:.0f}">{_v4_num(rev, "원")}</td>'
        f'<td data-v="{aov:.0f}">{_v4_num(aov, "원")}</td>'
        f'<td data-v="{cost:.0f}">{_v4_num(cost, "원")}</td>'
        f'<td data-v="{roas:.2f}">{roas_txt}</td>'
        f'<td><span class="fv4-chip {cls}">{label}</span></td>'
        f'</tr>'
    )


def render_ga_creative_page(cre: pd.DataFrame, ad_spend: pd.DataFrame = None,
                            utm_map: pd.DataFrame = None):
    """GA 소재별 성과 — 대행사 리포트가 아니라 GA4의 utm_campaign/utm_content로 본다.

    매체 리포트(소재별 성과 탭)와 나눠둔 이유: 매체가 신고하는 전환은 매체마다 기준이 달라
    서로 못 더한다. GA는 한 기준이라 소재끼리 비교가 된다. 대신 GA에는 광고비가 없어서
    ROAS는 방문 비중으로 나눈 '추정치'다 — 순위 판단용이지 정산용이 아니다.
    """
    st.markdown(FUNNEL_V4_CSS, unsafe_allow_html=True)
    st.markdown(
        '<div class="fv4-wrap"><div class="fv4-eyebrow">CREATIVE · GA4</div>'
        '<div class="fv4-h2">GA 소재별 성과</div></div>', unsafe_allow_html=True)

    c1, c2 = st.columns([3, 1])
    with c2:
        if st.button("🔄 소재 데이터 동기화", use_container_width=True, key="gc_sync"):
            cmap = (build_utm_channel_lookup(utm_map)
                    if utm_map is not None and not utm_map.empty else {})
            with st.spinner("GA4에서 소재별 데이터를 받는 중..."):
                n, s_, e_, err = sync_ga4_creative_daily(cre, cmap, force_full=cre is None or cre.empty)
            if err:
                st.error(err)
            else:
                st.success(f"{s_} ~ {e_} · {n}행 저장했습니다")
                st.cache_data.clear()
                st.rerun()

    if cre is None or cre.empty:
        st.info(
            "아직 소재별 데이터가 없습니다. 오른쪽 **🔄 소재 데이터 동기화**를 눌러주세요.\n\n"
            "먼저 Supabase에 `ga_creative_daily` 테이블이 있어야 합니다 (배포 SQL 참고)."
        )
        return

    d = cre.copy()
    d["report_date"] = pd.to_datetime(d["report_date"], errors="coerce").dt.date
    d = d.dropna(subset=["report_date"])
    if d.empty:
        st.info("날짜를 읽을 수 있는 행이 없습니다.")
        return

    # 매체 판정은 저장된 channel을 믿지 않고 지금의 UTM 매핑으로 다시 입힌다.
    # 동기화할 때 매핑이 비어 있었거나 그 뒤에 매핑을 고쳤을 수 있어서, 화면 기준으로 맞춘다.
    lookup = (build_utm_channel_lookup(utm_map)
              if utm_map is not None and not utm_map.empty else {})
    if lookup:
        d["channel"] = d["source_medium"].map(
            lambda sm: lookup.get(str(sm).strip().lower()))
    elif "channel" not in d.columns:
        d["channel"] = None
    with c1:
        start, end = period_filter(d["report_date"].min(), d["report_date"].max(),
                                   key="ga_creative", default_preset="이번달")

    level = st.radio("보기 단위", ["소재", "타겟팅", "캠페인", "매체"],
                     horizontal=True, key="gc_level")
    # 볼 매체 고르기.
    # 맨즈탭은 소재 운영·정산을 별도 시트로 관리해서 평소에는 빼두는데, 가끔 맨즈탭만 따로
    # 봐야 할 때가 있다. 매번 나머지를 다 빼는 건 번거로우니 셋 중에 고르게 한다.
    all_ch = sorted({_v4_canon_channel(c) for c in d["channel"].dropna().unique()})
    sep_ch = [c for c in GC_DEFAULT_EXCLUDE if c in all_ch]      # 따로 보는 매체
    sep_label = " · ".join(sep_ch) if sep_ch else "별도 관리 매체"
    scope = st.radio(
        "보기 대상",
        [f"{sep_label} 제외", f"{sep_label}만", "직접 고르기"],
        horizontal=True, key="gc_scope",
        help=f"{sep_label}은 소재를 별도 시트로 관리하셔서 평소에는 빼둡니다.",
    ) if sep_ch else "직접 고르기"

    if scope.endswith("만"):
        exclude = [c for c in all_ch if c not in sep_ch]
    elif scope.endswith("제외"):
        exclude = list(sep_ch)
    else:
        exclude = st.multiselect(
            "제외할 매체", all_ch, default=[], key="gc_exclude",
            help="여기 담긴 매체가 표에서 빠집니다.",
        )
    rows = _gc_rows(d, start, end, level, exclude=exclude)
    if rows.empty:
        if not lookup:
            st.warning(
                "UTM 매핑이 비어 있어 광고 매체를 가려낼 수 없습니다. "
                "사이드바에서 **UTM 리스트 파일**을 먼저 올려주세요 "
                "(소스/매체 → 매체명 대응표)."
            )
        else:
            st.info("선택한 기간에 광고 유입 데이터가 없습니다.")
        return

    # ── UTM 설정 상태 진단 ── 소재가 (미설정)이면 표가 뭉개진다. 비율을 먼저 보여준다.
    per = d[(d["report_date"] >= start) & (d["report_date"] <= end)].copy()
    per = per[per.apply(classify_ga_bucket, axis=1) == "광고"]
    if exclude:
        per = per[~per["channel"].map(_v4_canon_channel).isin(set(exclude))]
    per["sessions"] = pd.to_numeric(per["sessions"], errors="coerce").fillna(0)
    tot_ses = float(per["sessions"].sum())
    unset = float(per[per["creative"] == "(미설정)"]["sessions"].sum())
    if tot_ses > 0 and unset / tot_ses > 0.2:
        st.warning(
            f"소재(utm_content)가 안 붙은 방문이 {unset:,.0f} / {tot_ses:,.0f}건 "
            f"({unset / tot_ses * 100:.0f}%)입니다. 광고 랜딩 URL에 `utm_content`를 넣어야 "
            "소재별 비교가 됩니다 — 운영 도구 › UTM 빌더에서 만들 수 있습니다."
        )

    # ── 광고비 배분 ── 매체 광고비를 그 매체 안에서 방문 비중대로 나눈다.
    spend = _cp_spend_by_channel(ad_spend, start, end) if ad_spend is not None else pd.DataFrame()
    spend_by_ch = {}
    if spend is not None and not spend.empty:
        for _, r in spend.iterrows():
            spend_by_ch[_v4_canon_channel(r["channel"])] = float(r.get("cost_incl_vat", 0) or 0)
    ch_ses = rows.groupby("channel")["sessions"].sum().to_dict()

    def cost_of(r):
        ch = r["channel"]
        tot = float(ch_ses.get(ch, 0) or 0)
        if tot <= 0:
            return 0.0
        return spend_by_ch.get(ch, 0.0) * float(r["sessions"]) / tot

    body = [_gc_row_html(r, cost_of(r)) for _, r in rows.iterrows()]
    tot_r = rows[["sessions", "new", "signup", "conv", "rev"]].sum()
    tot_r["key"] = "TOTAL"
    sum_html = _gc_row_html(tot_r, sum(spend_by_ch.get(c, 0.0) for c in ch_ses),
                            "fv4-sum-row nosort")

    head = list(GC_HEAD)
    head[0] = {"소재": "소재 (타겟팅 · 등록일)", "타겟팅": "매체 · 타겟팅",
               "캠페인": "매체 · 캠페인", "매체": "매체"}[level]
    th = "".join(f'<th class="{"l" if i == 0 else ""}">{h}'
                 f'<span class="fv4-ar">&#8645;</span></th>'
                 for i, h in enumerate(head))
    table = (f'<table class="fv4-tbl" id="gctbl"><thead><tr>{th}</tr></thead>'
             f'<tbody>{sum_html}{"".join(body)}</tbody></table>')

    card = (
        FUNNEL_V4_CSS
        + '<div class="fv4-wrap"><div class="fv4-card">'
        f'<span class="fv4-badge-dark">{level}별</span>'
        f'<div class="fv4-card-title">{level}별 GA 성과</div>'
        '<div class="fv4-card-sub">GA4의 utm_campaign / utm_content 기준입니다. '
        '매체가 신고하는 전환은 매체마다 기준이 달라 서로 못 더하지만, GA는 한 기준이라 '
        '소재끼리 비교가 됩니다. <b>UTM 매핑된 광고 매체만</b> 셉니다 — '
        '자연유입·레퍼럴·(not set)은 소재가 없어서 제외합니다.</div>'
        '<div class="fv4-bk-cap">머리글을 누르면 정렬됩니다. TOTAL은 맨 위 고정입니다.<br>'
        '소재 이름은 <b>utm_content</b>를 <code>타겟팅_날짜_소재명</code> 규칙으로 쪼갠 것입니다 — '
        '같은 소재라도 타겟팅이 다르면 따로 셉니다. 규칙에 안 맞는 값은 '
        '<b>(규칙 외)</b>로 모아 보여주니, 그게 많으면 UTM 작명을 맞춰주세요.<br>'
        '<b>추정 광고비</b>는 매체 광고비를 그 매체 안에서 방문 비중대로 나눈 값입니다 — '
        'GA에는 소재별 광고비가 없습니다. <b>순위를 볼 때만</b> 쓰시고 정산에는 쓰지 마세요. '
        f'광고비가 {FUNNEL_MIN_SPEND:,.0f}원 미만이면 표본이 작아 <b>판단 보류</b>로 둡니다.</div>'
        + table + '</div></div>'
        + """
<script>
(function(){
  var t=document.getElementById('gctbl'); if(!t) return;
  var ths=t.tHead.rows[0].cells, st={i:-1,asc:false};
  function v(row,i){var c=row.cells[i]; if(!c) return '';
    var d=c.getAttribute('data-v'); return d!==null?parseFloat(d):c.innerText.trim();}
  for(var i=0;i<ths.length;i++){(function(i){
    ths[i].style.cursor='pointer';
    ths[i].onclick=function(){
      var asc=(st.i===i)?!st.asc:false; st={i:i,asc:asc};
      var tb=t.tBodies[0], all=Array.prototype.slice.call(tb.rows);
      var fixed=all.filter(function(r){return r.classList.contains('nosort');});
      var mov=all.filter(function(r){return !r.classList.contains('nosort');});
      mov.sort(function(a,b){var x=v(a,i),y=v(b,i);
        if(typeof x==='number'&&typeof y==='number'){return asc?x-y:y-x;}
        return asc?String(x).localeCompare(String(y),'ko')
                  :String(y).localeCompare(String(x),'ko');});
      fixed.concat(mov).forEach(function(r){tb.appendChild(r);});
    };})(i);}
})();
</script>"""
    )
    st.components.v1.html(card, height=min(1400, 420 + 40 * len(body)), scrolling=True)


def render_ga_channel_funnel_page(
    audience: pd.DataFrame,
    ga_channel_inflow: pd.DataFrame,
    inflow_revenue: pd.DataFrame,
    channels_weekly: pd.DataFrame,
    channel_mix: pd.DataFrame,
    ga_daily: pd.DataFrame = None,
    utm_map: pd.DataFrame = None,
    decisions: pd.DataFrame = None,
    ad_spend: pd.DataFrame = None,
    master: pd.DataFrame = None,
):
    """채널 퍼널 리포트 V4 — '신규 고객 발굴' / '매출 확보' 두 목적축을 하나의 화면에서 본다.
    노출·클릭·가입·구매·매출은 channel_audience_snapshot(대행사 캠페인 신규/리타겟 태그 기준),
    방문·재방문은 ga_channel_inflow(GA 신규/재방문 사용자 기준), 계획 예산은 channel_mix_budget에서
    가져와 채널명을 정규화(_v4_canon_channel)해서 한 줄에 놓는다."""
    st.markdown(FUNNEL_V4_CSS, unsafe_allow_html=True)

    # ── GA4 자동 연동 ── 엑셀 업로드 없이, 페이지를 열 때 저장된 마지막 날짜 이후만 받아 채운다.
    # (Streamlit Community Cloud는 크론을 못 돌리므로 '열 때 자동 동기화'가 가장 현실적이다.)
    ga_daily = ga_daily if ga_daily is not None else pd.DataFrame()
    lookup = build_utm_channel_lookup(utm_map) if utm_map is not None and not utm_map.empty else {}
    sync_note, sync_err = "", None
    if "fv4_synced" not in st.session_state:
        n, s, e, sync_err = sync_ga4_channel_daily(ga_daily, lookup)
        st.session_state["fv4_synced"] = True
        if n:
            st.cache_data.clear()
            ga_daily = load_table("ga_channel_daily")
            sync_note = f"GA4 자동 동기화 {n:,}행 ({s} ~ {e})"

    # 광고비도 같은 시점에 동기화한다(매체별로 실패해도 나머지는 계속 진행).
    ad_spend = ad_spend if ad_spend is not None else pd.DataFrame()
    spend_sync_note, spend_errors = "", {}
    if "fv4_spend_synced" not in st.session_state:
        n_sp, saved, spend_errors = sync_ad_spend(ad_spend)
        st.session_state["fv4_spend_synced"] = True
        if n_sp:
            st.cache_data.clear()
            ad_spend = load_table("ad_spend_daily")
            spend_sync_note = "광고비 동기화 " + ", ".join(f"{k} {v}행" for k, v in saved.items())

    # GA4 API로 받은 데이터가 있으면 그걸 우선 쓰고, 없으면 기존 엑셀 업로드분으로 폴백한다.
    ga_source_label = "GA4 API(자동)"
    if ga_daily is not None and not ga_daily.empty:
        ga_channel_inflow = ga_inflow_source(ga_daily, ga_channel_inflow)
    else:
        ga_source_label = "GA 엑셀 업로드(수동)"

    last_sync = None
    if ga_daily is not None and not ga_daily.empty and "uploaded_at" in ga_daily.columns:
        try:
            last_sync = pd.to_datetime(ga_daily["uploaded_at"]).max()
        except Exception:
            last_sync = None
    ga_last_date = None
    if ga_channel_inflow is not None and not ga_channel_inflow.empty:
        ga_last_date = pd.to_datetime(ga_channel_inflow["report_date"]).max().date()

    c_fresh, c_btn = st.columns([5, 1])
    with c_fresh:
        bits = [f"데이터 출처: {ga_source_label}"]
        if ga_last_date:
            bits.append(f"GA 최신 일자 {ga_last_date}")
        if last_sync is not None and pd.notna(last_sync):
            bits.append(f"마지막 동기화 {last_sync:%m/%d %H:%M}")
        if sync_note:
            bits.append(sync_note)
        if spend_sync_note:
            bits.append(spend_sync_note)
        st.caption(" · ".join(bits))
    with c_btn:
        if st.button("🔄 지금 동기화", key="fv4_sync_btn", use_container_width=True):
            st.session_state.pop("fv4_synced", None)
            st.cache_data.clear()
            st.rerun()
    if sync_err:
        st.warning(
            f"GA4 자동 연동이 아직 안 됐습니다 — {sync_err} "
            "(설정 전까지는 기존 엑셀 업로드 데이터로 표시됩니다.)"
        )
    if ga_last_date and (date.today() - ga_last_date).days > 2:
        st.warning(f"GA 데이터가 {(date.today() - ga_last_date).days}일 지연돼 있습니다 — 최신 수치가 아닐 수 있습니다.")

    if ga_source_label != "GA4 API(자동)" or sync_err:
        with st.expander("🔍 GA4 연동 진단 (연동이 안 되면 이걸 펼쳐서 캡쳐해주세요)", expanded=bool(sync_err)):
            st.caption("비밀값(private_key 등)은 출력하지 않습니다 — 형식이 맞는지만 검사합니다.")
            if st.button("진단 실행", key="fv4_diag_btn"):
                with st.spinner("확인 중..."):
                    st.code(diagnose_ga4_setup(), language=None)

    with st.expander("💰 광고비 연동 진단 (매체별 연동 상태)"):
        st.caption("연동 안 된 매체는 채널믹스 예산 일할값으로 자동 대체됩니다 — 대시보드는 항상 동작합니다.")
        if spend_errors:
            for k, v in spend_errors.items():
                st.warning(f"{k} 광고비 조회 실패: {v}")
        if st.button("광고비 진단 실행", key="fv4_spend_diag_btn"):
            with st.spinner("확인 중..."):
                st.code(diagnose_ad_spend_setup(), language=None)

    # OAuth에서 돌아오면(?code=...) 해당 패널이 저절로 펼쳐지게 한다.
    # 접힌 채로 돌아오면 "눌렀는데 아무 반응이 없다"로 보이기 때문.
    try:
        _oauth_state = st.query_params.get("state") or ""
        _oauth_code = st.query_params.get("code") or ""
    except Exception:
        _oauth_state, _oauth_code = "", ""
    _back_gfa = bool(_oauth_code) and _oauth_state == "stcogfa"
    _back_gads = bool(_oauth_code) and _oauth_state == "stcogads"
    _back_kko = bool(_oauth_code) and _oauth_state == "stcokakao"
    if _back_gfa or _back_gads or _back_kko:
        _who = "네이버" if _back_gfa else ("구글" if _back_gads else "카카오")
        st.success(f"{_who} 로그인에서 돌아왔습니다 — 아래 토큰 발급 패널에서 ② 버튼을 눌러주세요.")

    with st.expander("🔑 네이버 GFA 토큰 발급 (refresh token 만들기)", expanded=_back_gfa):
        render_gfa_token_helper()

    with st.expander("🔑 구글애즈 토큰 발급 (refresh token 만들기)", expanded=_back_gads):
        render_google_token_helper()

    with st.expander("🔑 카카오모먼트 토큰 발급 (refresh token 만들기)", expanded=_back_kko):
        render_kakao_token_helper()

    with st.expander("✍️ 광고비 직접 입력 (API가 막힌 매체용)"):
        render_manual_spend_panel(ad_spend)

    if audience.empty and ga_channel_inflow.empty and inflow_revenue.empty:
        st.info(
            "아직 데이터가 없습니다. 사이드바 '① 주간 리포트 업로드'(캠페인 신규/리타겟 분류) · "
            "'② GA 유입 데이터 업로드'를 먼저 진행해주세요."
        )
        return

    date_pool = []
    if not audience.empty:
        date_pool.append(pd.to_datetime(audience["as_of_date"]))
    if not ga_channel_inflow.empty:
        date_pool.append(pd.to_datetime(ga_channel_inflow["report_date"]))
    if not inflow_revenue.empty:
        date_pool.append(pd.to_datetime(inflow_revenue["report_date"]))
    if not date_pool:
        st.info("선택 가능한 기간 데이터가 없습니다.")
        return
    all_dates = pd.concat(date_pool)
    min_d, max_d = all_dates.min().date(), all_dates.max().date()
    st.subheader("🔎 기간 필터")
    start, end = period_filter(min_d, max_d, key="channel_funnel", default_preset="이번달")

    span = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span - 1)

    # ── 집계 ──
    gci = pd.DataFrame()
    if not ga_channel_inflow.empty:
        gci = ga_channel_inflow.copy()
        gci["report_date"] = pd.to_datetime(gci["report_date"]).dt.date
    fg = gci[(gci["report_date"] >= start) & (gci["report_date"] <= end)] if not gci.empty else pd.DataFrame()
    pg = gci[(gci["report_date"] >= prev_start) & (gci["report_date"] <= prev_end)] if not gci.empty else pd.DataFrame()

    def _sum(df, col):
        return float(df[col].sum()) if (df is not None and not df.empty and col in df.columns) else 0.0

    users_now, users_prev = _sum(fg, "users"), _sum(pg, "users")
    new_now, new_prev = _sum(fg, "new_users"), _sum(pg, "new_users")
    ret_now, ret_prev = _sum(fg, "returning_users"), _sum(pg, "returning_users")
    conv_now, conv_prev = _sum(fg, "conversions"), _sum(pg, "conversions")
    rev_now, rev_prev = _sum(fg, "revenue"), _sum(pg, "revenue")
    # 목적축별 — 신규 발굴은 가입·첫구매·신규매출, 매출 확보는 재구매·재구매매출을 본다
    sign_now, sign_prev = _sum(fg, "new_signups"), _sum(pg, "new_signups")
    nconv_now, nconv_prev = _sum(fg, "new_conv"), _sum(pg, "new_conv")
    nrev_now, nrev_prev = _sum(fg, "new_rev"), _sum(pg, "new_rev")
    rconv_now, rconv_prev = _sum(fg, "ret_conv"), _sum(pg, "ret_conv")
    rrev_now, rrev_prev = _sum(fg, "ret_rev"), _sum(pg, "ret_rev")

    bucket_share = {"광고": 0.0, "자연유입": 0.0, "기타": 0.0}
    if not fg.empty:
        tmp = fg.copy()
        tmp["_b"] = tmp.apply(classify_ga_bucket, axis=1)
        for b, v in tmp.groupby("_b")["users"].sum().items():
            bucket_share[b] = float(v)
    bucket_total = sum(bucket_share.values()) or 1.0

    # 광고비 출처 ①: '유입·매출 비교' 파일. 최근 날짜가 비어 있는 경우가 많아서, 비어 있으면
    # ②: 대행사 주간 매체 리포트(channel_weekly)의 일할 안분값으로 자동 대체한다.
    # (①이 비어서 광고비가 0에 가깝게 잡히는 바람에 ROAS가 2000%대로 뻥튀기된 적 있음)
    ad_spend_period = 0.0
    spend_src = "유입·매출 비교 파일"
    if not inflow_revenue.empty:
        ir = inflow_revenue.copy()
        ir["report_date"] = pd.to_datetime(ir["report_date"]).dt.date
        fir = ir[(ir["report_date"] >= start) & (ir["report_date"] <= end)]
        ad_spend_period = float(fir["cost_incl_vat"].sum()) if not fir.empty else 0.0
    # GA-매출 전체(자연유입 포함)를 광고비로 나누면 범위가 안 맞으므로, ROAS는 '광고' 버킷
    # 매출만 광고비로 나눈다(유입·매출 비교 페이지에서 이미 정리된 원칙과 동일).
    ad_revenue_period = 0.0
    if not fg.empty:
        _t = fg.copy()
        _t["_b"] = _t.apply(classify_ga_bucket, axis=1)
        ad_revenue_period = float(_t.loc[_t["_b"] == "광고", "revenue"].sum())
    site_roas = 0.0  # 실제 계산은 아래에서 spend_sum(주간 리포트 합계)을 구한 뒤에 한다

    # 채널 단위 — 대표 채널명으로 정규화해서 합산
    aud_new = _funnel_from_audience(audience, start, end, "신규")
    aud_re = _funnel_from_audience(audience, start, end, "리타겟팅")
    visits = _ga_visits_by_channel(ga_channel_inflow, start, end)
    # 광고비: API 실집행 > 대행사 주간(일할) > 채널믹스 예산(일할) 순으로 채널마다 확정한다.
    spend_all, spend_src_summary = resolve_channel_spend(ad_spend, channels_weekly, channel_mix, start, end)
    spend_by_channel_src = dict(zip(spend_all["channel"], spend_all["source"])) if not spend_all.empty else {}
    # 계획 비중은 '연간' 채널 믹스 기준으로 잡는다 — 표 헤더가 '연간 계획 / 실제 집행'이고,
    # 기간을 좁힐 때마다 계획 비중까지 흔들리면 계획 대비 이탈을 판단할 수 없기 때문이다.
    mix_ratio = pd.DataFrame(columns=["channel", "budget", "budget_ratio"])
    if channel_mix is not None and not channel_mix.empty:
        _m = channel_mix.copy()
        _m["year"] = pd.to_numeric(_m["year"], errors="coerce")
        _m = _m[_m["year"] == _m["year"].max()]
        mix_ratio = _m.groupby("channel", as_index=False).agg(budget=("budget", "sum"))
        _yt = mix_ratio["budget"].sum()
        mix_ratio["budget_ratio"] = np.where(_yt > 0, mix_ratio["budget"] / _yt * 100, 0)

    for _df in (aud_new, aud_re, visits, mix_ratio):
        if _df is not None and not _df.empty:
            _df["channel"] = _df["channel"].map(_v4_canon_channel)

    def _regroup(df, how):
        if df is None or df.empty:
            return df
        return df.groupby("channel", as_index=False).agg(how)

    aud_new = _regroup(aud_new, {c: "sum" for c in ["impressions", "clicks", "cost_incl_vat", "signups", "conversions", "revenue"]})
    aud_re = _regroup(aud_re, {c: "sum" for c in ["impressions", "clicks", "cost_incl_vat", "signups", "conversions", "revenue"]})
    visits = _regroup(visits, {"new_users": "sum", "returning_users": "sum"})

    if mix_ratio is not None and not mix_ratio.empty:
        mix_ratio = mix_ratio.groupby("channel", as_index=False).agg(budget=("budget", "sum"))
        _tot = mix_ratio["budget"].sum()
        mix_ratio["budget_ratio"] = np.where(_tot > 0, mix_ratio["budget"] / _tot * 100, 0)

    spend_sum = float(spend_all["cost_incl_vat"].sum()) if (spend_all is not None and not spend_all.empty) else 0.0

    # 이제 광고비는 항상 resolve_channel_spend에서 나오므로, 그 합계를 정식 광고비로 쓴다.
    if spend_sum > 0:
        ad_spend_period = spend_sum
        spend_src = spend_src_summary
    site_roas = (ad_revenue_period / ad_spend_period * 100) if ad_spend_period > 0 else 0.0

    def _build(base, visit_col):
        if base is None or base.empty:
            return pd.DataFrame()
        df = base.copy()
        if visits is not None and not visits.empty:
            df = df.merge(visits[["channel", visit_col]], on="channel", how="left")
        else:
            df[visit_col] = np.nan
        df = add_kpis(df)
        df["signup_cac"] = np.where(df["signups"] > 0, df["cost_incl_vat"] / df["signups"], 0)
        if mix_ratio is not None and not mix_ratio.empty:
            df = df.merge(mix_ratio[["channel", "budget_ratio"]], on="channel", how="left")
        else:
            df["budget_ratio"] = np.nan
        if spend_all is not None and not spend_all.empty:
            # source 컬럼까지 붙으면 뒤쪽 표 로직과 컬럼이 꼬이므로 필요한 두 개만 가져온다.
            _sp = spend_all[["channel", "cost_incl_vat"]].rename(columns={"cost_incl_vat": "_spend_all"})
            df = df.merge(_sp, on="channel", how="left")
        else:
            df["_spend_all"] = df["cost_incl_vat"]
        df["_spend_all"] = df["_spend_all"].fillna(df["cost_incl_vat"])
        df["act_ratio"] = np.where(spend_sum > 0, df["_spend_all"] / spend_sum * 100, 0)
        df["drift"] = df["act_ratio"] - df["budget_ratio"].fillna(df["act_ratio"])
        return df.sort_values("budget_ratio", ascending=False, na_position="last")

    new_df = _build(aud_new, "new_users")
    re_df = _build(aud_re, "returning_users")

    # ── 액션 신호 ──
    signals = []
    pool = new_df if not new_df.empty else re_df
    if not pool.empty:
        scored = pool[pool["cost_incl_vat"] >= FUNNEL_MIN_SPEND]
        if not scored.empty:
            worst = scored.sort_values("roas").iloc[0]
            if worst["roas"] < OPS_KPI_ROAS_LOW:
                signals.append(("bad", "효율 미달", worst["channel"],
                                f"{worst['channel']} 신규 ROAS {worst['roas']:.0f}%",
                                f"목표 {OPS_KPI_ROAS_LOW:.0f}% 미만 · 소재/타겟팅 점검 필요"))
            best = scored.sort_values("roas", ascending=False).iloc[0]
            if best["roas"] >= OPS_KPI_ROAS_HIGH:
                signals.append(("good", "증액 후보", best["channel"],
                                f"{best['channel']} 신규 ROAS {best['roas']:.0f}%",
                                f"목표 하단({OPS_KPI_ROAS_LOW:.0f}%) 대비 {best['roas'] / OPS_KPI_ROAS_LOW:.1f}배 · 단계적 증액 검토"))
        # 이미 다른 신호로 잡힌 채널은 건너뛴다 — 배너 3칸이 같은 매체로 채워지면
        # "볼 곳이 세 군데"라는 신호의 의미가 없어진다.
        used = {ch for _, _, ch, _, _ in signals}
        drifted = pool[(pool["drift"].abs() >= FUNNEL_MIX_DRIFT_PP) & (~pool["channel"].isin(used))]
        if not drifted.empty:
            d = drifted.reindex(drifted["drift"].abs().sort_values(ascending=False).index).iloc[0]
            signals.append(("warn", "예산 이탈", d["channel"],
                            f"{d['channel']} 집행 {d['drift']:+.1f}%p",
                            "연간 계획 비중 대비 · 예산 소진 속도 점검"))
    if not signals:
        signals.append(("hold", "이상 없음", "", "긴급 신호 없음", "목표 구간 안에서 운영 중입니다"))

    # 헤드라인은 "A 효율을 점검하고, B 증액 여지를 확인하고, C 예산 속도를 조정하세요." 처럼
    # 마지막 절만 종결어미가 되도록 '하고'로 이어 붙였다가 끝을 '하세요.'로 바꾼다.
    verb = {"bad": "효율을 점검하고", "good": "증액 여지를 확인하고", "warn": "예산 속도를 조정하고"}
    head_bits = [f"{ch} {verb[k]}" for k, _, ch, _, _ in signals if k in verb]
    if head_bits:
        headline = ",<br>".join(head_bits)
        headline = headline[: -len("하고")] + "하세요." if headline.endswith("하고") else headline + "."
    else:
        headline = "오늘은 큰 이상 신호가 없습니다."

    # 신호마다 '며칠째'를 덧붙인다 — 오늘 처음 빨개진 것과 5일째 빨간 것은 액션이 다르다.
    _daily_for_streak = _loop_daily_by_channel(gci)
    sig_html = ""
    for k, chip, _ch, t, s in signals:
        streak_txt = _loop_signal_streak(_daily_for_streak, _ch, end) if _ch else ""
        extra = f' · {streak_txt}' if streak_txt else ""
        sig_html += (
            f'<div class="fv4-signal"><span class="fv4-chip {k}">{chip}</span>'
            f'<div class="fv4-signal-title">{t}</div><div class="fv4-signal-sub">{s}{extra}</div></div>'
        )
    st.markdown(
        '<div class="fv4-wrap"><div class="fv4-banner">'
        '<div class="fv4-banner-lead">'
        f'<div class="fv4-count">{len(signals)}</div>'
        f'<div><div class="fv4-banner-eyebrow">오늘의 액션 신호</div>'
        f'<div class="fv4-banner-title">{headline}</div></div></div>'
        f'<div class="fv4-signals">{sig_html}</div></div></div>',
        unsafe_allow_html=True,
    )

    # ── 어제 대비 달라진 것 ── 스냅샷이 아니라 '변화'를 먼저 보여준다.
    daily_ch = _loop_daily_by_channel(gci)
    change_logs = _loop_change_log(
        daily_ch, end, spend_daily=ad_spend,
        universe=list(master["media"]) if (master is not None and not master.empty
                                           and "media" in master.columns) else None)
    if change_logs and len(change_logs) > 1:
        (_, (_cmp_new, _cmp_old)), *_body = change_logs
        items = "".join(
            f'<div class="fv4-chg-item"><span class="fv4-chg-dot {k}"></span>{t}</div>'
            for k, t in _body
        )
        # '어제 대비'라고만 적으면 실제로 언제와 언제를 비교했는지 알 수 없다. GA4 동기화가
        # 하루 밀려 있으면 어제가 아니라 그제와 그그제를 비교하고 있을 수도 있어서 날짜를 박아둔다.
        _hdr = f"{_cmp_new:%-m/%d}(vs {_cmp_old:%-m/%d}) 달라진 것"
        _warn = ""
        if _cmp_new >= date.today() - timedelta(days=1):
            # GA4는 구매·매출을 최대 이틀까지 늦게 채운다. 그래서 가장 최근 날짜의 매출이
            # 0으로 보이는 건 '안 팔린 것'이 아니라 '아직 안 들어온 것'일 때가 많다.
            _warn = ('<div class="fv4-chg-note">매출은 GA4 집계가 최대 이틀 늦습니다 — '
                     f'{_cmp_new:%-m/%d} 매출 0원은 아직 안 들어온 것일 수 있습니다. '
                     '확정 판단은 하루 뒤에 하세요.</div>')
        st.markdown(
            f'<div class="fv4-wrap"><div class="fv4-chg"><div class="fv4-chg-h">{_hdr}</div>'
            f'{items}{_warn}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown(
        '<div class="fv4-wrap"><div class="fv4-eyebrow">TWO FUNNELS, ONE DECISION</div>'
        '<div class="fv4-h2">목적별 퍼널 성과</div></div>',
        unsafe_allow_html=True,
    )
    if "fv4_mode" not in st.session_state:
        st.session_state["fv4_mode"] = "신규 고객 발굴"
    _sp, c_new, c_re = st.columns([6, 2, 2])
    with c_new:
        if st.button("신규 고객 발굴", key="fv4_btn_new", use_container_width=True,
                     type="primary" if st.session_state["fv4_mode"] == "신규 고객 발굴" else "secondary"):
            st.session_state["fv4_mode"] = "신규 고객 발굴"
            st.rerun()
    with c_re:
        if st.button("매출 확보", key="fv4_btn_re", use_container_width=True,
                     type="primary" if st.session_state["fv4_mode"] == "매출 확보" else "secondary"):
            st.session_state["fv4_mode"] = "매출 확보"
            st.rerun()

    mode = st.session_state["fv4_mode"]
    is_new = mode == "신규 고객 발굴"

    # ── KPI 스트립 ── 목적축(신규 발굴 / 매출 확보)에 따라 세우는 숫자가 달라진다.
    # 두 탭이 같은 KPI를 보여주면 탭을 나눈 의미가 없다 — 신규는 '얼마나 데려왔나',
    # 매출 확보는 '있던 사람이 얼마나 다시 샀나'가 판단 기준이다.
    axis_users = new_now if is_new else ret_now
    axis_users_prev = new_prev if is_new else ret_prev
    axis_ratio = (axis_users / users_now * 100) if users_now else 0
    axis_conv, axis_conv_prev = (nconv_now, nconv_prev) if is_new else (rconv_now, rconv_prev)
    axis_rev, axis_rev_prev = (nrev_now, nrev_prev) if is_new else (rrev_now, rrev_prev)
    axis_roas = (axis_rev / ad_spend_period * 100) if ad_spend_period > 0 else None
    roas_label = "신규 ROAS" if is_new else "재구매 ROAS"
    # 광고비를 못 구하면 ROAS를 억지로 계산하지 않고 '-'로 둔다(틀린 숫자로 판단하는 게 제일 위험).
    roas_sub = (
        f"광고 ROAS {site_roas:.0f}% · 광고비 {spend_src}"
        if ad_spend_period > 0 else "광고 ROAS - (광고비 데이터 없음)"
    )
    # 광고비는 신규·재방문으로 나눌 수 없어 전액 기준이다. 그래서 신규 ROAS + 재구매 ROAS가
    # 전체 ROAS가 된다 — 각각만 보면 낮아 보이는 게 정상이라 그 점을 같이 적어준다.
    axis_roas_sub = (
        f"{roas_label} {axis_roas:.0f}% · 전체 매출 {rev_now:,.0f}원"
        if axis_roas is not None else f"{roas_label} - (광고비 데이터 없음)"
    )
    ad_p = bucket_share["광고"] / bucket_total * 100
    org_p = bucket_share["자연유입"] / bucket_total * 100
    etc_p = bucket_share["기타"] / bucket_total * 100
    st.markdown(
        '<div class="fv4-wrap"><div class="fv4-kpis">'
        f'<div class="fv4-kpi"><div class="fv4-kpi-label">총 방문자</div>'
        f'<div class="fv4-kpi-value">{users_now:,.0f}{_v4_delta_html(users_now, users_prev)}</div>'
        f'<div class="fv4-kpi-sub">신규 {new_now:,.0f} + 재방문 {ret_now:,.0f}</div></div>'
        f'<div class="fv4-kpi"><div class="fv4-kpi-label">{"신규 방문자" if is_new else "재방문자"}</div>'
        f'<div class="fv4-kpi-value">{axis_users:,.0f}{_v4_delta_html(axis_users, axis_users_prev)}</div>'
        f'<div class="fv4-kpi-sub">전체 방문의 {axis_ratio:.1f}%</div></div>'
        + (
            f'<div class="fv4-kpi"><div class="fv4-kpi-label">회원가입</div>'
            f'<div class="fv4-kpi-value">{sign_now:,.0f}{_v4_delta_html(sign_now, sign_prev)}</div>'
            f'<div class="fv4-kpi-sub">가입률 {(sign_now / new_now * 100) if new_now else 0:.2f}% · '
            f'첫구매 {nconv_now:,.0f}건</div></div>'
            if is_new else
            f'<div class="fv4-kpi"><div class="fv4-kpi-label">재구매</div>'
            f'<div class="fv4-kpi-value">{rconv_now:,.0f}{_v4_delta_html(rconv_now, rconv_prev)}</div>'
            f'<div class="fv4-kpi-sub">재구매율 {(rconv_now / ret_now * 100) if ret_now else 0:.2f}% · '
            f'객단가 {(rrev_now / rconv_now) if rconv_now else 0:,.0f}원</div></div>'
        ) +
        f'<div class="fv4-kpi"><div class="fv4-kpi-label">{"신규 매출" if is_new else "재구매 매출"}</div>'
        f'<div class="fv4-kpi-value money">{axis_rev:,.0f}원{_v4_delta_html(axis_rev, axis_rev_prev)}</div>'
        f'<div class="fv4-kpi-sub">{axis_roas_sub}</div></div>'
        f'<div class="fv4-kpi"><div class="fv4-kpi-label">유입 구성</div>'
        f'<div class="fv4-stack"><i style="width:{ad_p:.1f}%;background:#17170f"></i>'
        f'<i style="width:{org_p:.1f}%;background:#6b5ce7"></i>'
        f'<i style="width:{etc_p:.1f}%;background:#cfcdbf"></i></div>'
        f'<div class="fv4-kpi-sub">광고 {ad_p:.0f}% · 자연 {org_p:.0f}% · 기타 {etc_p:.0f}%</div></div>'
        "</div></div>",
        unsafe_allow_html=True,
    )

    # ── 목적별 퍼널 (토글) ──
    # 이 화면은 대행사 리포트를 쓰지 않는다 — 매체 API(노출·클릭·광고비)와 GA4(사용자·가입·구매·매출)만
    # 본다. 둘 다 매일 자동으로 들어오므로 사람이 파일을 올리지 않아도 최신 상태가 유지된다.
    buckets, media = _v4_ga_rows(ga_channel_inflow, start, end)
    spend_now = _cp_spend_by_channel(ad_spend, start, end) if ad_spend is not None \
        else pd.DataFrame(columns=["channel", "impressions", "clicks", "cost_incl_vat", "source"])
    spend_map = {r["channel"]: r for _, r in spend_now.iterrows()}

    if media.empty and buckets.empty:
        st.info("선택한 기간에 GA4 유입 데이터가 없습니다. 상단 '🔄 지금 동기화'를 눌러주세요.")
    else:
        ad_row = buckets[buckets["key"] == "광고"]
        ad_row = ad_row.iloc[0] if not ad_row.empty else None
        ad_media = list(media["key"])
        agg_sp = lambda col: sum(float(spend_map.get(c, {}).get(col, 0) or 0) for c in ad_media)
        ad_imp, ad_clk, ad_cost = agg_sp("impressions"), agg_sp("clicks"), agg_sp("cost_incl_vat")

        if is_new:
            stages = [
                ("노출", ad_imp), ("클릭", ad_clk),
                ("신규 방문", float(ad_row["new"]) if ad_row is not None else 0),
                ("회원가입", float(ad_row["signup"]) if ad_row is not None else 0),
                ("첫구매", float(ad_row["new_conv"]) if ad_row is not None else 0),
            ]
            bench = FUNNEL_BENCHMARK_NEW
            badge, title = "ACQUISITION", "노출에서 첫구매까지"
            sub = ("신규 사용자가 가입하고 첫 구매까지 오는 구간입니다. "
                   "노출·클릭이 광고에만 있는 숫자라 <b>이 퍼널은 광고 유입만</b> 봅니다 — "
                   "위 KPI·아래 TOTAL(자연유입·기타 포함)보다 작은 게 정상입니다.")
            head, rowfn = HEAD_NEW, _v4_row_new
        else:
            stages = [
                ("재노출", ad_imp), ("재클릭", ad_clk),
                ("재방문", float(ad_row["ret"]) if ad_row is not None else 0),
                ("재구매", float(ad_row["ret_conv"]) if ad_row is not None else 0),
            ]
            bench = FUNNEL_BENCHMARK_RETURN
            badge, title = "RETENTION", "재노출에서 재구매까지"
            sub = ("이미 방문한 고객이 다시 사는 구간입니다. 재구매 건수·매출·ROAS를 봅니다. "
                   "노출·클릭이 광고에만 있는 숫자라 <b>이 퍼널은 광고 유입만</b> 봅니다 — "
                   "위 KPI·아래 TOTAL(자연유입·기타 포함)보다 작은 게 정상입니다.")
            head, rowfn = HEAD_RET, _v4_row_ret

        def verdict_of(r, cost):
            rev = float(r.get("new_rev" if is_new else "ret_rev", 0) or 0)
            roas = (rev / cost * 100) if cost > 0 else 0.0
            return _v4_verdict(roas, cost)

        # ── ① 채널 대분류 ──
        brows = []
        for b in ["광고", "자연유입", "기타"]:
            s_ = buckets[buckets["key"] == b]
            if s_.empty:
                continue
            s_ = s_.iloc[0]
            if b == "광고":
                nm = f'<b>{b}</b><span class="fv4-bk-sub">{len(media)}개 매체</span>'
                brows.append(rowfn(nm, s_, ad_cost, "fv4-bk-row") + "<td></td></tr>")
            else:
                nm = f'<b>{b}</b>'
                brows.append(rowfn(nm, s_, 0.0, "fv4-bk-row") + "<td></td></tr>")
        brows_sum = ""
        if not buckets.empty:
            tot = buckets[["users", "new", "ret", "signup", "new_conv", "new_rev",
                           "ret_conv", "ret_rev", "conv", "rev"]].sum()
            brows_sum = rowfn("TOTAL", tot, ad_cost, "fv4-sum-row nosort") + "<td></td></tr>"

        # ── ② 광고 매체별 상세 ──
        mrows = []
        for _, m in media.iterrows():
            ch = m["key"]
            cost = float(spend_map.get(ch, {}).get("cost_incl_vat", 0) or 0)
            label, cls = verdict_of(m, cost)
            mrows.append(rowfn(f'<span class="fv4-ind">{ch}</span>', m, cost)
                         + f'<td><span class="fv4-chip {cls}">{label}</span></td></tr>')
        mrows_sum = ""
        if not media.empty:
            mt = media[["users", "new", "ret", "signup", "new_conv", "new_rev",
                        "ret_conv", "ret_rev", "conv", "rev"]].sum()
            mrows_sum = rowfn("TOTAL", mt, ad_cost, "fv4-sum-row nosort") + "<td></td></tr>"

        def build_table(tid, head_labels, sum_row, data_rows):
            th = "".join(
                f'<th class="{"l" if h == "채널" else ""}">{h}'
                f'<span class="fv4-ar">&#8645;</span></th>' for h in head_labels)
            return (f'<table class="fv4-tbl" id="{tid}"><thead><tr>{th}</tr></thead>'
                    f'<tbody>{sum_row}{"".join(data_rows)}</tbody></table>')

        card = (
            FUNNEL_V4_CSS
            + '<div class="fv4-wrap"><div class="fv4-card">'
            f'<span class="fv4-badge-dark">{badge}</span>'
            f'<div class="fv4-card-title">{title}</div><div class="fv4-card-sub">{sub}</div>'
            + _v4_funnel_html(stages, bench)
            + '<div class="fv4-sec">채널 대분류</div>'
            + '<div class="fv4-bk-cap">머리글을 누르면 그 표만 정렬됩니다. TOTAL 줄은 항상 맨 위 고정입니다. '
              '광고비는 신규·재방문으로 나눌 수 없어 <b>전액 기준</b>이라, '
              '신규 ROAS와 재구매 ROAS를 더하면 전체 ROAS가 됩니다.<br>'
              '<b>총 방문자 = 신규 방문자 + 재방문자</b>로 딱 맞습니다(GA4 방문수 기준). '
              'GA4 보고서의 <b>총 사용자</b>는 중복을 뺀 사람 수라 이보다 작고 신규+재방문과도 '
              '안 맞는데, 그 값은 매체별로 나눌 수가 없어 쓰지 않습니다.</div>'
            + build_table("fvtblA", head, brows_sum, brows)
            + '<div class="fv4-sec">광고 매체별 상세</div>'
            + '<div class="fv4-bk-cap">위 표의 <b>광고</b> 줄을 매체로 쪼갠 것입니다 — '
              '아래 TOTAL이 위 <b>광고</b> 줄과 같아야 정상입니다.</div>'
            + build_table("fvtblB", head, mrows_sum, mrows) + '</div></div>'
            + """
<script>
(function(){
  function mk(id){
    var t=document.getElementById(id); if(!t) return;
    var ths=t.tHead.rows[0].cells, st={i:-1,asc:false};
    function v(row,i){var c=row.cells[i]; if(!c) return '';
      var d=c.getAttribute('data-v'); return d!==null?parseFloat(d):c.innerText.trim();}
    for(var i=0;i<ths.length;i++){(function(i){
      ths[i].style.cursor='pointer';
      ths[i].onclick=function(){
        var asc=(st.i===i)?!st.asc:false; st={i:i,asc:asc};
        var tb=t.tBodies[0], all=Array.prototype.slice.call(tb.rows);
        var fixed=all.filter(function(r){return r.classList.contains('nosort');});
        var mov=all.filter(function(r){return !r.classList.contains('nosort');});
        mov.sort(function(a,b){var x=v(a,i),y=v(b,i);
          if(typeof x==='number'&&typeof y==='number'){return asc?x-y:y-x;}
          return asc?String(x).localeCompare(String(y),'ko')
                    :String(y).localeCompare(String(x),'ko');});
        fixed.concat(mov).forEach(function(r){tb.appendChild(r);});
        for(var k=0;k<ths.length;k++){
          ths[k].querySelector('.fv4-ar').textContent=(k===i)?(asc?'\u2191':'\u2193'):'\u21C5';}
      };})(i);}
  }
  mk('fvtblA'); mk('fvtblB');
})();
</script>"""
        )
        st.components.v1.html(
            card, height=430 + 44 * (len(brows) + len(mrows) + 2), scrolling=True)
        st.caption(
            f"⚠ 표시는 해당 단계 전환율이 벤치마크 미만이라는 뜻입니다 · "
            f"위 퍼널의 노출·클릭·가입 출처: 대행사 리포트(캠페인 {'신규' if is_new else '리타겟'} 태그) · "
            f"표의 사용자·가입·구매·매출 출처: {ga_source_label} · "
            f"광고비 출처: 매체 API·계약·업로드 · {FUNNEL_MIN_SPEND:,}원 미만은 '판단 보류'"
        )
        with st.expander("🔍 판정 근거 보기 (왜 이 판정인지)"):
            for _, m in media.iterrows():
                ch = m["key"]
                sp = spend_map.get(ch, {})
                cost = float(sp.get("cost_incl_vat", 0) or 0)
                rev = float(m["new_rev"] if is_new else m["ret_rev"])
                conv = float(m["new_conv"] if is_new else m["ret_conv"])
                base = float(m["new"] if is_new else m["ret"])
                roas = (rev / cost * 100) if cost > 0 else 0.0
                label, _cls = _v4_verdict(roas, cost)
                kind = "신규" if is_new else "재구매"
                reason = (
                    f"{kind} ROAS {roas:.0f}% (목표 {OPS_KPI_ROAS_LOW:.0f}~{OPS_KPI_ROAS_HIGH:.0f}%), "
                    f"광고비 {cost:,.0f}원, {kind} 매출 {rev:,.0f}원, "
                    f"{'신규' if is_new else '재방문'} {base:,.0f}명 → {kind} {conv:,.0f}건"
                )
                if is_new and m.get("signup", 0):
                    reason += f" · 가입 {m['signup']:,.0f}명 (가입 CAC {cost / m['signup']:,.0f}원)"
                if conv:
                    reason += f" · {kind} CAC {cost / conv:,.0f}원"
                _src = str(sp.get("source", "") or "")
                if _src:
                    reason += f" · 광고비 출처: {AD_SPEND_SOURCE_LABEL.get(_src, _src)}"
                if cost < FUNNEL_MIN_SPEND:
                    reason += f" · 광고비가 {FUNNEL_MIN_SPEND:,}원 미만이라 판단 보류"
                st.markdown(f"**{ch} → {label}**  \n{reason}")

    # ── NEXT BEST ACTION (목적축별 · 매체 전부) ──
    st.markdown(CP_CSS + _v4_actions_html(media, spend_map, is_new, start, end),
                unsafe_allow_html=True)

    st.markdown(
        '<div class="fv4-wrap"><div class="fv4-foot">'
        '<span>STCO PERFORMANCE · 내부 의사결정용</span>'
        f"<span>{start:%Y-%m-%d} ~ {end:%Y-%m-%d}</span></div></div>",
        unsafe_allow_html=True,
    )

    # ── 의사결정 기록 & 회고 ──
    # 대시보드를 보고 내린 결정을 남겨두면, 다음에 열었을 때 그 결정이 먹혔는지 자동으로
    # 되짚어준다. 이게 있어야 '보기만 하는 대시보드'가 아니라 루프가 닫힌다.
    st.markdown("---")
    st.markdown("#### 🧾 의사결정 기록 & 회고")
    decisions = decisions if decisions is not None else pd.DataFrame()

    reviews = _loop_review_decisions(decisions, daily_ch, end)
    if reviews:
        items = "".join(
            f'<div class="fv4-rev-item"><div class="fv4-rev-h">{ch} · {act}</div>'
            f'<div class="fv4-rev-s">{when} 결정 → {verdict}</div></div>'
            for ch, act, when, verdict in reviews
        )
        st.markdown(f'<div class="fv4-wrap"><div class="fv4-rev">{items}</div></div>', unsafe_allow_html=True)
    else:
        st.caption("아직 기록된 결정이 없습니다. 아래에서 오늘 내린 결정을 남겨두면 다음 주에 결과를 자동으로 되짚어드립니다.")

    with st.expander("➕ 오늘 내린 결정 기록하기"):
        ch_options = sorted(set(pd.concat([new_df, re_df])["channel"].dropna())) if not (new_df.empty and re_df.empty) else []
        with st.form("fv4_decision_form", clear_on_submit=True):
            d1, d2 = st.columns([1, 2])
            with d1:
                dch = st.selectbox("매체", ch_options if ch_options else ["(없음)"], key="fv4_dec_ch")
                dwhen = st.date_input("결정일", value=date.today(), key="fv4_dec_when")
            with d2:
                dact = st.text_input("결정 내용", placeholder="예: 메타 신규발굴 예산 10% 축소", key="fv4_dec_act")
                dnote = st.text_input("근거 / 메모", placeholder="예: ROAS 192%로 3일 연속 목표 미달", key="fv4_dec_note")
            if st.form_submit_button("기록 저장", type="primary"):
                if not dact.strip():
                    st.warning("결정 내용을 입력해주세요.")
                else:
                    row = pd.DataFrame([{
                        "decided_on": str(dwhen), "channel": dch, "action": dact.strip(),
                        "note": dnote.strip() or None,
                    }])
                    n = save_table("decision_log", row, "decided_on,channel,action", "대시보드 입력")
                    if n:
                        st.cache_data.clear()
                        st.success("기록했습니다. 다음 기간에 결과를 자동으로 되짚어드립니다.")
                        st.rerun()


def render_ga4_page():
    looker_view_url = (
        "https://lookerstudio.google.com/u/0/reporting/"
        "7177b0a5-7d7e-4f07-af76-17f2436b317e/page/p_bbwwb7lo4c"
    )
    looker_embed_url = (
        "https://lookerstudio.google.com/embed/reporting/"
        "7177b0a5-7d7e-4f07-af76-17f2436b317e/page/p_bbwwb7lo4c"
    )
    st.markdown("### 구글 애널리틱스(GA4) 라이브 리포트")
    st.caption(
        "대행사가 만든 리포트라 일반 공개(링크가 있는 모든 사용자)로 바꾸기 어려우면, "
        "대시보드 안에 그대로 넣는(임베드) 대신 아래 버튼으로 본인 구글 계정 권한으로 새 창에서 열어 보세요."
    )
    st.link_button("📊 GA4 리포트 새 창에서 열기", looker_view_url, use_container_width=True)

    with st.expander("대시보드 안에 직접 띄워보기 (권한 있으면 아래에 표시됨)"):
        st.caption(
            "대행사에게 Looker Studio에서 파일 > 삽입 보고서(Embed report)를 켜달라고 요청하면 "
            "이 안에 화면이 그대로 뜹니다. 권한이 없으면 로그인 요청이나 빈 화면이 보일 수 있어요."
        )
        scale = st.slider(
            "리포트 축소 비율", min_value=0.4, max_value=1.0, value=1.0, step=0.05, key="looker_scale",
        )
        # 폭을 줄이면 리포트 내부 표가 반응형으로 찌그러지며 컬럼이 잘리므로,
        # 원본 크기(1400x1000)로 그린 뒤 CSS로 통째로 축소(zoom-out)해서
        # 컬럼이 안 잘리고 더 많은 내용이 한 화면에 들어오게 한다.
        native_w, native_h = 1400, 1000
        disp_w, disp_h = int(native_w * scale), int(native_h * scale)
        st.markdown(
            f'<div style="width:{disp_w}px; height:{disp_h}px; overflow:hidden; margin:0 auto; border:1px solid #e5e8eb; border-radius:8px;">'
            f'<iframe src="{looker_embed_url}" width="{native_w}" height="{native_h}" '
            f'style="border:0; transform:scale({scale}); transform-origin:0 0;" allowfullscreen></iframe>'
            f'</div>',
            unsafe_allow_html=True,
        )


# ──────────────────────────────────────────────────────────────
# 운영 코멘트 (신규) — 데이터를 바탕으로 자동 생성하는 실무 코멘트/제안
# performance-marketing-analysis 스킬(코멘트 작성 순서·소재 필터링·예산재배분 원칙)과
# 프로젝트 KPI 기준(판단 지표 GA-ROAS, 목표 200~300%)을 그대로 반영한 룰 기반 코멘트.
# ──────────────────────────────────────────────────────────────
OPS_KPI_ROAS_LOW = 200    # KPI 목표 하단(%) — GA 기준
OPS_KPI_ROAS_HIGH = 300   # KPI 목표 상단(%), 300%+가 목표
OPS_MIN_CHANNEL_SPEND = 100_000   # 매체 판단 보류 기준: 광고비 10만원 미만
OPS_CREATIVE_MIN_SPEND = 50_000   # 소재 판단 보류 기준(performance-marketing-analysis 스킬 기본값과 동일)


def _ops_kpi_status(roas: float) -> str:
    if roas >= OPS_KPI_ROAS_HIGH:
        return "목표 초과 달성"
    if roas >= OPS_KPI_ROAS_LOW:
        return "목표 구간 내"
    return "목표 미달"


def _ops_fmt_pp(v: float) -> str:
    # 표(render_html_table)와 같은 색 규칙: ▲(상승)는 빨강, ▼(하락)는 파랑.
    # HTML 문자열이라 이 값을 쓰는 st.markdown 호출엔 반드시 unsafe_allow_html=True를 줘야 한다.
    color = "#d93025" if v >= 0 else "#1a73e8"
    arrow = "▲" if v >= 0 else "▼"
    return f'<span style="color:{color}">{arrow}{abs(v):,.0f}%p</span>'


def _ops_fmt_pct(v: float) -> str:
    color = "#d93025" if v >= 0 else "#1a73e8"
    arrow = "▲" if v >= 0 else "▼"
    return f'<span style="color:{color}">{arrow}{abs(v):,.1f}%</span>'


def _ops_next_action(text: str) -> str:
    """'다음 액션' 라벨부터 그 뒤 제안 문장까지 전부 굵은 빨간 텍스트로 강조해서 눈에 띄게 한다.
    이 함수가 붙은 문장을 st.markdown()으로 렌더링할 땐 unsafe_allow_html=True가 필요하다."""
    return f'<b style="color:#d93025;">다음 액션: {text}</b>'


def _short_week_label(row) -> str:
    """weekly 행에서 'N월 N주차' 짧은 라벨을 뽑는다. 원본 엑셀 라벨에 '주차' 표기가 없으면
    week_start~week_end 날짜(MM/DD)로 대체한다."""
    m = re.search(r"(\d{1,2})월\s*(\d+)\s*주차", str(row.get("label", "")))
    if m:
        return f"{m.group(1)}월 {m.group(2)}주차"
    return f"{row['week_start']:%m/%d}~{row['week_end']:%m/%d}"


def _ops_driver_sentence(cur, prev, cur_label: str, prev_label: str, roas_col: str = "roas", roas_label: str = "ROAS") -> str:
    """단순히 'ROAS가 올랐다/내렸다'만 말하지 않고, 광고비·노출·클릭·매출이 어느 방향으로
    같이 움직였는지(볼륨 측) + CVR·객단가가 어떻게 바뀌어서 ROAS가 그렇게 움직였는지(효율 측)를
    묶어서 설명하는 문장을 만든다. 월별/주간별 코멘트에서 공용으로 쓴다.
    반환값에 <span>/<b> 태그가 섞여 있어 st.markdown()에 unsafe_allow_html=True가 필요하다."""

    def pct(a, b):
        return (a - b) / abs(b) * 100 if b else 0

    d_cost = cur.get("cost_incl_vat", 0) - prev.get("cost_incl_vat", 0)
    d_impr_pct = pct(cur.get("impressions", 0), prev.get("impressions", 0))
    d_clicks_pct = pct(cur.get("clicks", 0), prev.get("clicks", 0))
    d_rev_pct = pct(cur.get("revenue", 0), prev.get("revenue", 0))
    d_cvr_pct = pct(cur.get("cvr", 0), prev.get("cvr", 0))
    d_aov = cur.get("aov", 0) - prev.get("aov", 0)
    d_aov_pct = pct(cur.get("aov", 0), prev.get("aov", 0))
    d_roas_pp = cur.get(roas_col, 0) - prev.get(roas_col, 0)

    cost_word = "증가" if d_cost >= 0 else "감소"
    metrics = [("노출수", d_impr_pct), ("클릭수", d_clicks_pct), ("매출", d_rev_pct)]
    same_dir = [n for n, v in metrics if (v >= 0) == (d_cost >= 0)]
    diff_dir = [n for n, v in metrics if (v >= 0) != (d_cost >= 0)]

    if len(same_dir) == 3:
        vol_sentence = f"{cur_label}는 {prev_label} 대비 광고비가 {cost_word}하면서 노출수·클릭수·매출도 동반 {cost_word}했습니다."
    elif same_dir:
        vol_sentence = (
            f"{cur_label}는 {prev_label} 대비 광고비가 {cost_word}했고 {'·'.join(same_dir)}도 함께 {cost_word}했지만, "
            f"{'·'.join(diff_dir)}는 반대로 움직였습니다."
        )
    else:
        vol_sentence = f"{cur_label}는 {prev_label} 대비 광고비가 {cost_word}했지만, 노출수·클릭수·매출은 반대로 움직였습니다."

    eff_bits = []
    if abs(d_cvr_pct) >= 0.5:
        eff_bits.append(f"CVR이 {_ops_fmt_pct(d_cvr_pct)} {'상승' if d_cvr_pct >= 0 else '하락'}")
    if abs(d_aov_pct) >= 0.5:
        eff_bits.append(f"객단가가 {_ops_fmt_pct(d_aov_pct)}({d_aov:+,.0f}원) {'상승' if d_aov_pct >= 0 else '하락'}")

    roas_verb = "상승" if d_roas_pp >= 0 else "하락"
    if eff_bits:
        eff_sentence = "다만 " + ", ".join(eff_bits) + f"한 영향으로 {roas_label}는 {_ops_fmt_pp(d_roas_pp)} {roas_verb}했습니다."
    else:
        eff_sentence = f"{roas_label}는 {_ops_fmt_pp(d_roas_pp)} {roas_verb}했습니다."

    return vol_sentence + " " + eff_sentence


def _ops_top_mover_sentence(
    df: pd.DataFrame, key_col: str, period_col: str, cur_period, prev_period,
    cost_col: str = "cost_incl_vat", rev_col: str = "revenue", min_abs_cost: float = 0, min_abs_rev: float = 0,
) -> str:
    """전체(또는 그룹) 수치가 왜 그렇게 움직였는지, 어떤 매체/소재가 광고비·매출 변화를
    가장 크게 이끌었는지 짚어주는 문장 하나를 만든다. 매체별/소재별 코멘트에서 공용으로 쓴다.
    df엔 [key_col, period_col, cost_col, rev_col]가 있어야 한다. cur/prev 중 한쪽에만 있는
    key(이번 주 신규 집행/운영 중지된 매체 등)도 그 기간엔 0원으로 보고 비교 대상에 포함한다 —
    실제로 매체 운영을 중지/재개하는 것 자체가 흔한 원인이라 빠뜨리면 안 된다."""
    if df is None or df.empty:
        return ""
    cur_df = df[df[period_col] == cur_period].groupby(key_col)[[cost_col, rev_col]].sum()
    prev_df = df[df[period_col] == prev_period].groupby(key_col)[[cost_col, rev_col]].sum()
    all_keys = cur_df.index.union(prev_df.index)
    if len(all_keys) == 0:
        return ""
    cur_df = cur_df.reindex(all_keys, fill_value=0)
    prev_df = prev_df.reindex(all_keys, fill_value=0)
    d_cost = cur_df[cost_col] - prev_df[cost_col]
    d_rev = cur_df[rev_col] - prev_df[rev_col]

    bits = []
    cost_driver = None
    if len(d_cost) and d_cost.abs().max() > min_abs_cost:
        cost_driver = d_cost.abs().idxmax()
        v = d_cost[cost_driver]
        stop_note = "(운영 중지)" if cur_df.loc[cost_driver, cost_col] == 0 and v < 0 else (
            "(신규 집행)" if prev_df.loc[cost_driver, cost_col] == 0 and v > 0 else ""
        )
        bits.append(f"{cost_driver}에서 광고비가 {abs(v):,.0f}원 {'증가' if v >= 0 else '감소'}{stop_note}한 영향이 컸고")
    if len(d_rev) and d_rev.abs().max() > min_abs_rev:
        rev_driver = d_rev.abs().idxmax()
        v = d_rev[rev_driver]
        if rev_driver == cost_driver:
            bits.append(f"같은 곳에서 매출도 {abs(v):,.0f}원 {'증가' if v >= 0 else '감소'}하며 영향을 키웠습니다")
        else:
            connector = "대신 " if bits else ""
            bits.append(f"{connector}{rev_driver}에서 매출이 {abs(v):,.0f}원 {'증가' if v >= 0 else '감소'}한 영향이 컸습니다")
    if not bits:
        return ""
    if len(bits) == 1:
        return bits[0] + "."
    return bits[0] + ", " + bits[1] + "."


def _ops_channel_efficiency_driver_sentence(
    df: pd.DataFrame, key_col: str, period_col: str, cur_period, prev_period,
    cost_col: str = "cost_incl_vat", conv_col: str = "conversions", clicks_col: str = "clicks",
    rev_col: str = "revenue",
) -> str:
    """광고비·매출 증감(위 _ops_top_mover_sentence)만으로는 설명이 안 되는 ROAS 변화 —
    즉 광고비·매출이 둘 다 줄었는데도 ROAS는 오히려 오른 경우처럼, '효율' 쪽에서 원인을 찾아야
    할 때 쓴다. 두 기간 모두 광고비가 있었던 매체 중 ROAS(%)가 가장 많이 오른/내린 매체를 찾아
    CVR·객단가 변화로 설명한다."""
    if df is None or df.empty:
        return ""
    cur_g = df[df[period_col] == cur_period].groupby(key_col)[[cost_col, conv_col, clicks_col, rev_col]].sum()
    prev_g = df[df[period_col] == prev_period].groupby(key_col)[[cost_col, conv_col, clicks_col, rev_col]].sum()
    common = [k for k in cur_g.index.intersection(prev_g.index)
              if cur_g.loc[k, cost_col] > 0 and prev_g.loc[k, cost_col] > 0]
    if not common:
        return ""
    best_k, best_d_roas, best_stats = None, 0, None
    for k in common:
        cur_roas = cur_g.loc[k, rev_col] / cur_g.loc[k, cost_col] * 100
        prev_roas = prev_g.loc[k, rev_col] / prev_g.loc[k, cost_col] * 100
        d_roas = cur_roas - prev_roas
        if best_k is None or abs(d_roas) > abs(best_d_roas):
            cur_cvr = cur_g.loc[k, conv_col] / cur_g.loc[k, clicks_col] * 100 if cur_g.loc[k, clicks_col] else 0
            prev_cvr = prev_g.loc[k, conv_col] / prev_g.loc[k, clicks_col] * 100 if prev_g.loc[k, clicks_col] else 0
            cur_aov = cur_g.loc[k, rev_col] / cur_g.loc[k, conv_col] if cur_g.loc[k, conv_col] else 0
            prev_aov = prev_g.loc[k, rev_col] / prev_g.loc[k, conv_col] if prev_g.loc[k, conv_col] else 0
            best_k, best_d_roas = k, d_roas
            best_stats = {"d_cvr": cur_cvr - prev_cvr, "d_aov": cur_aov - prev_aov}
    if best_k is None or abs(best_d_roas) < 1:
        return ""
    bits = []
    if abs(best_stats["d_cvr"]) >= 0.01:
        bits.append(f"CVR이 {abs(best_stats['d_cvr']):.1f}%p {'상승' if best_stats['d_cvr'] >= 0 else '하락'}")
    if abs(best_stats["d_aov"]) >= 1:
        bits.append(f"객단가가 {abs(best_stats['d_aov']):,.0f}원 {'상승' if best_stats['d_aov'] >= 0 else '하락'}")
    eff = ", ".join(bits) if bits else "효율 지표가 개선"
    verb = "오른" if best_d_roas >= 0 else "내린"
    return f"이 중 {best_k}에서 {eff}하며 ROAS가 {abs(best_d_roas):.0f}%p {verb} 것이 전체 ROAS 변동의 주요 요인으로 보입니다."


WEEKDAY_KOR = ["월", "화", "수", "목", "금", "토", "일"]

_WEEKLY_ZERO_CHECK_COLS = ["impressions", "clicks", "cost_excl_vat", "cost_incl_vat", "conversions", "revenue"]


def _drop_trailing_zero_weeks(weekly: pd.DataFrame) -> pd.DataFrame:
    """parse_weekly()에서 이미 걸러내지만, Supabase에 예전 코드로 저장된 빈(전부 0) 미래 주
    placeholder 행이 그대로 남아있을 수도 있어서 화면에 쓸 때도 한 번 더 방어적으로 걸러준다.
    week_start 기준 정렬 후, 값이 하나도 없는 말미 행들을 제거한다."""
    if weekly is None or weekly.empty:
        return weekly
    w = weekly.sort_values("week_start").reset_index(drop=True)
    cols = [c for c in _WEEKLY_ZERO_CHECK_COLS if c in w.columns]
    if not cols:
        return w
    metric_sum = w[cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
    nonzero_idx = metric_sum[metric_sum > 0].index
    if len(nonzero_idx):
        w = w.loc[: nonzero_idx.max()]
    return w.reset_index(drop=True)


def _weekly_period_label(weekly: pd.DataFrame, label_prefix: str = "주간 코멘트 정리") -> str:
    """weekly(주간 통합 데이터)의 가장 최근 주를 '{prefix} - 26년 7월 5주차 : 07/27(월) ~08/02(일)'
    형식으로 만든다. 원본 엑셀의 'N월 N주차' 라벨 텍스트에서 'N주차' 숫자만 재사용하고,
    날짜 범위와 요일은 week_start/week_end로 직접 계산해 매주 자동으로 갱신되게 한다."""
    weekly = _drop_trailing_zero_weeks(weekly)
    if weekly is None or weekly.empty:
        return label_prefix
    w = weekly.sort_values("week_end").iloc[-1]
    week_start, week_end = w["week_start"], w["week_end"]
    if isinstance(week_start, str):
        week_start = pd.to_datetime(week_start).date()
    if isinstance(week_end, str):
        week_end = pd.to_datetime(week_end).date()
    year2 = f"{week_start.year % 100:02d}"
    month = week_start.month
    m = re.search(r"(\d+)\s*주차", str(w.get("label", "")))
    week_part = f"{m.group(1)}주차" if m else ""
    start_s = f"{week_start.month:02d}/{week_start.day:02d}({WEEKDAY_KOR[week_start.weekday()]})"
    end_s = f"{week_end.month:02d}/{week_end.day:02d}({WEEKDAY_KOR[week_end.weekday()]})"
    period = f"{year2}년 {month}월 {week_part} : {start_s} ~{end_s}".replace("  ", " ")
    return f"{label_prefix} - {period}"


def render_ops_comment_monthly(monthly: pd.DataFrame, heading: str = "#### 💬 월별 코멘트", channels: pd.DataFrame = None):
    """월별 총평 코멘트 — 종합 대시보드의 '1) 월별 누적' 표 아래와, 운영 코멘트 탭의 ①에서
    공용으로 쓴다. GA-ROAS를 KPI 판단 기준(200~300%)으로 쓰고, 자체 ROAS는 참고로 같이 보여준다.
    channels(매체별 월간 데이터)가 주어지면, '전체'만 보고 끝내지 않고 어느 매체가 광고비/매출
    변화를 가장 크게 이끌었는지 매체명까지 짚어주는 문장을 한 줄 더 붙인다."""
    st.markdown(heading)
    if monthly.empty:
        st.caption("월별 데이터가 아직 없습니다.")
        return
    m = monthly.copy()
    m["report_month"] = pd.to_datetime(m["report_month"]).dt.date
    m = add_kpis(m.sort_values("report_month")).reset_index(drop=True)
    cur = m.iloc[-1]
    prev = m.iloc[-2] if len(m) >= 2 else None
    cur_label = f"{cur['report_month'].year}년 {cur['report_month'].month}월"
    cur_roas, cur_garoas = cur.get("roas", 0), cur.get("ga_roas", 0)
    status = _ops_kpi_status(cur_garoas)

    st.markdown(f"**{cur_label} GA-ROAS {cur_garoas:,.0f}%로 {status}** (자체 ROAS {cur_roas:,.0f}%, KPI {OPS_KPI_ROAS_LOW}~{OPS_KPI_ROAS_HIGH}%)")

    body = []
    if cur_roas and cur_garoas and abs(cur_roas - cur_garoas) / max(cur_garoas, 1) > 0.3:
        body.append(
            f"자체 ROAS와 GA-ROAS 격차가 {_ops_fmt_pp(cur_roas - cur_garoas)}로 큰 편이라, "
            "두 지표 중 어느 쪽을 메인으로 볼지 다시 짚어볼 필요가 있습니다."
        )
    trend_down = None
    if prev is not None:
        prev_label = f"{prev['report_month'].year}년 {prev['report_month'].month}월"
        d_garoas = cur_garoas - prev.get("ga_roas", 0)
        trend_down = d_garoas < 0
        body.append(_ops_driver_sentence(cur, prev, cur_label, prev_label, roas_col="ga_roas", roas_label="GA-ROAS"))
        if channels is not None and not channels.empty and "report_month" in channels.columns:
            ch = channels.copy()
            ch["report_month"] = pd.to_datetime(ch["report_month"]).dt.date
            mover = _ops_top_mover_sentence(
                ch, key_col="channel", period_col="report_month",
                cur_period=cur["report_month"], prev_period=prev["report_month"],
            )
            if mover:
                body.append(mover)
            eff_mover = _ops_channel_efficiency_driver_sentence(
                ch, key_col="channel", period_col="report_month",
                cur_period=cur["report_month"], prev_period=prev["report_month"],
            )
            if eff_mover:
                body.append(eff_mover)
    if body:
        st.markdown(" ".join(body), unsafe_allow_html=True)

    if status == "목표 미달":
        if trend_down:
            body_next = (
                "목표 미달에 하락세까지 겹쳤습니다. 이번 달 남은 기간엔 ②매체별·③소재별 코멘트를 먼저 "
                "확인해 원인(매체/소재/타겟팅)을 좁히고, 다음 달 예산 조정 여부를 결정하는 것을 권장합니다."
            )
        else:
            body_next = (
                "목표 미달이지만 반등 조짐이 있어, 이번 달은 매체·소재 단위 원인 점검만 진행하고 "
                "다음 달 초까지 개선 여부를 지켜본 뒤 예산 조정을 검토하세요."
            )
    elif status == "목표 초과 달성":
        body_next = (
            "목표를 안정적으로 상회하고 있어, 다음 달엔 ②매체별 코멘트에서 목표 초과로 분류된 매체 위주로 "
            "10~20% 증액을 검토하는 것을 권장합니다."
        )
    else:
        body_next = "목표 구간 내에서 안정적으로 운영되고 있어, 다음 달에도 현재 운영 기조(매체 비중·소재 구성)를 유지하는 것을 권장합니다."
    st.markdown(_ops_next_action(body_next), unsafe_allow_html=True)


def render_ops_comment_weekly(
    weekly: pd.DataFrame, heading: str = "#### 💬 주간별 코멘트", channels_weekly: pd.DataFrame = None,
):
    """주간별 코멘트 — 종합 대시보드의 '2) 주간별 누적' 표 아래에 쓴다.
    주간 데이터엔 GA-ROAS가 없어(주간 리포트 시트 구조상 GA 비교 컬럼 없음) 자체 ROAS만 쓴다.
    channels_weekly(매체 개별 시트의 주간 데이터)가 주어지면, 어느 매체가 이번 주 광고비/매출
    변화를 가장 크게 이끌었는지 매체명까지 짚어주는 문장을 한 줄 더 붙인다."""
    st.markdown(heading)
    weekly = _drop_trailing_zero_weeks(weekly)
    if weekly is None or weekly.empty:
        st.caption("주간 데이터가 아직 없습니다.")
        return
    w = add_kpis(weekly.sort_values("week_start")).reset_index(drop=True)
    cur = w.iloc[-1]
    prev = w.iloc[-2] if len(w) >= 2 else None
    cur_label = f"{cur['week_start']:%Y-%m-%d}~{cur['week_end']:%Y-%m-%d}"
    cur_roas = cur.get("roas", 0)
    status = _ops_kpi_status(cur_roas)

    st.markdown(f"**{cur_label} 자체 ROAS {cur_roas:,.0f}%로 {status}** (KPI {OPS_KPI_ROAS_LOW}~{OPS_KPI_ROAS_HIGH}%, GA 비교는 월별 코멘트 참고)")

    trend_down = None
    if prev is not None:
        prev_label = f"{prev['week_start']:%Y-%m-%d}~{prev['week_end']:%Y-%m-%d}"
        d_roas = cur_roas - prev.get("roas", 0)
        trend_down = d_roas < 0
        st.markdown(
            _ops_driver_sentence(
                cur, prev, _short_week_label(cur), _short_week_label(prev), roas_col="roas", roas_label="ROAS"
            ),
            unsafe_allow_html=True,
        )
        if channels_weekly is not None and not channels_weekly.empty and "week_start" in channels_weekly.columns:
            cwk = channels_weekly.copy()
            cwk["week_start"] = pd.to_datetime(cwk["week_start"]).dt.date
            mover = _ops_top_mover_sentence(
                cwk, key_col="channel", period_col="week_start",
                cur_period=cur["week_start"], prev_period=prev["week_start"],
            )
            if mover:
                st.markdown(mover)
            eff_mover = _ops_channel_efficiency_driver_sentence(
                cwk, key_col="channel", period_col="week_start",
                cur_period=cur["week_start"], prev_period=prev["week_start"],
            )
            if eff_mover:
                st.markdown(eff_mover)
            cur_ch = sorted(cwk.loc[cwk["week_start"] == cur["week_start"], "channel"].unique())
            prev_ch = sorted(cwk.loc[cwk["week_start"] == prev["week_start"], "channel"].unique())
            st.caption(
                f"※ 매체별 주간 데이터 인식: 전체 {cwk['channel'].nunique()}개 매체 · "
                f"{_short_week_label(cur)} {len(cur_ch)}개({', '.join(cur_ch) or '없음'}) · "
                f"{_short_week_label(prev)} {len(prev_ch)}개({', '.join(prev_ch) or '없음'})"
            )
            # 매체마다 가장 최근에 잡힌 주가 언제인지(뒤처진 매체 탭이 있는지) 바로 보이게.
            latest_per_channel = (
                cwk.groupby("channel")["week_start"].max().sort_values(ascending=False)
            )
            latest_lines = "; ".join(f"{ch}: {wk}" for ch, wk in latest_per_channel.items())
            st.caption(f"※ 매체별 최신 주간 데이터 시점 — {latest_lines}")
        elif channels_weekly is not None and channels_weekly.empty:
            st.caption("※ 매체별 주간 데이터가 아직 없습니다 — 매체별 코멘트는 전체 합산 기준으로만 표시됩니다.")

    if trend_down is None:
        next_action = "아직 비교할 전주 데이터가 없어 다음 주 데이터가 쌓이면 추세를 판단할 수 있습니다."
    elif status == "목표 미달" and trend_down:
        next_action = "목표 미달+하락세가 겹쳤습니다. 이번 주 안에 소재/타겟팅을 먼저 점검하고, 다음 주에도 개선이 없으면 예산 축소를 검토하세요."
    elif status == "목표 미달":
        next_action = "목표는 미달이지만 반등 조짐이 있어, 다음 주까지는 현재 설정을 유지하며 지켜보는 것을 권장합니다."
    elif not trend_down and status != "목표 미달":
        next_action = "상승 추세라 다음 주도 현재 운영을 유지하며 지켜보는 것을 권장합니다."
    else:
        next_action = "하락 추세이니 다음 주엔 원인(소재 소진·시즌성 등)을 점검해보는 것을 권장합니다."
    st.markdown(_ops_next_action(next_action), unsafe_allow_html=True)


def _ops_channel_bucket_lines(df: pd.DataFrame, roas_col: str) -> list:
    """매체별 df('channel','cost_incl_vat', roas_col 컬럼 필요)를 목표초과/근접/미달로 나눠
    코멘트 문장 리스트를 만든다. 광고비가 OPS_MIN_CHANNEL_SPEND 미만인 매체는 판단 보류로 뺀다.
    그룹 전체의 min~max 범위 대신, 매체별 실제 수치를 각각 이름과 함께 보여준다."""
    if df.empty or roas_col not in df.columns:
        return []
    judged = df[df["cost_incl_vat"] >= OPS_MIN_CHANNEL_SPEND].copy()
    held = df[df["cost_incl_vat"] < OPS_MIN_CHANNEL_SPEND].copy()
    lines = []
    if not judged.empty:
        def _bucket(v):
            if v >= OPS_KPI_ROAS_HIGH:
                return "목표 초과"
            if v >= OPS_KPI_ROAS_LOW:
                return "목표 근접"
            return "목표 미달"

        judged["구분"] = judged[roas_col].map(_bucket)
        judged = judged.sort_values(roas_col, ascending=False)
        over = judged[judged["구분"] == "목표 초과"]
        near = judged[judged["구분"] == "목표 근접"]
        under = judged[judged["구분"] == "목표 미달"]

        def _names_with_roas(sub):
            return ", ".join(f"{r.channel}({getattr(r, roas_col):,.0f}%)" for r in sub.itertuples())

        if not over.empty:
            lines.append(
                f"**목표 초과 ({len(over)}개 매체)**: {_names_with_roas(over)} — 목표를 상회합니다. "
                + _ops_next_action("이번 주 중 10~20% 증액을 적용해보고, 다음 주 재측정 후 유지·추가 증액 여부를 결정하는 것을 권장합니다.")
            )
        if not near.empty:
            lines.append(
                f"**목표 근접 ({len(near)}개 매체)**: {_names_with_roas(near)} — 목표 구간 안쪽입니다. "
                + _ops_next_action("현재 예산·소재를 유지하며 다음 주까지 추세를 지켜보는 것을 권장합니다.")
            )
        if not under.empty:
            lines.append(
                f"**목표 미달 ({len(under)}개 매체)**: {_names_with_roas(under)} — 목표에 못 미칩니다. "
                + _ops_next_action(
                    "이번 주엔 소재 교체를 먼저 시도하고, 다음 주에도 개선이 없으면 총 예산의 10~15% "
                    "내에서 목표 초과 매체 쪽으로 단계적으로 이동하는 것을 권장합니다."
                )
            )
    if not held.empty:
        lines.append(f"판단 보류(광고비 {OPS_MIN_CHANNEL_SPEND:,}원 미만, 표본 부족): {', '.join(held['channel'])}")
    return lines


def render_ops_comment_channel_narrative(
    df: pd.DataFrame, roas_col: str, heading: str = "#### 💬 매체별 코멘트", footnote: str = "",
):
    """매체별 표 바로 아래 붙이는 짧은 서술형 코멘트(목표초과/근접/미달 분류 + 예산재배분 제안만).
    표는 이미 위에 있으므로 여기서는 표를 다시 그리지 않는다."""
    st.markdown(heading)
    lines = _ops_channel_bucket_lines(df, roas_col)
    if not lines:
        st.caption("판정 가능한 매체 데이터가 아직 없습니다.")
        return
    for line in lines:
        st.markdown(line, unsafe_allow_html=True)
    if footnote:
        st.caption(footnote)


def render_operation_comment_page(
    weekly: pd.DataFrame, monthly: pd.DataFrame, daily: pd.DataFrame,
    channels: pd.DataFrame, snapshot: pd.DataFrame, creatives: pd.DataFrame,
    audience: pd.DataFrame, inflow_revenue: pd.DataFrame, ga_channel_inflow: pd.DataFrame,
    agency_notes: pd.DataFrame = None,
):
    """⓪ 대행사 운영 메모 → ① 전체 총평 → ② 매체별 코멘트+예산재배분 제안 →
    ③ 소재별 제안 → ④ GA 유입 코멘트 순서로, 업로드된 데이터를 바탕으로 룰 기반 코멘트를
    자동 생성한다. 어디까지나 1차 초안이라 최종 운영 판단은 직접 검토가 필요하다는 점을 상단에 명시한다."""
    st.info(
        f"아래 코멘트는 업로드된 데이터를 기준으로 자동 생성한 1차 초안입니다. "
        f"판단 지표는 GA-ROAS, KPI 목표는 {OPS_KPI_ROAS_LOW}~{OPS_KPI_ROAS_HIGH}%입니다. "
        "최종 운영 판단은 직접 검토 후 반영해주세요."
    )

    # ── ⓪ 대행사 운영 메모 (매체통합 시트 하단 자유 텍스트) ──
    # 제목에 그 주의 연도/월/N주차/날짜범위(요일 포함)를 자동으로 붙인다 — 매주 업로드할 때마다
    # weekly(통합 주간별)의 가장 최근 주 기준으로 갱신된다.
    st.markdown(f"## {_weekly_period_label(weekly)}")
    if agency_notes is None or agency_notes.empty:
        st.caption("주간 리포트의 '매체통합' 시트 하단에서 인식된 운영 메모가 아직 없습니다.")
    else:
        latest_note = agency_notes.sort_values("as_of_date").iloc[-1]
        st.caption(f"기준일: {latest_note['as_of_date']} (주간 리포트 하단 코멘트를 그대로 정리한 것입니다)")
        st.markdown(latest_note["note_text"])
    st.caption("※ 실제 파일 구조 검증 전 휴리스틱 추출입니다 — 다음 업로드에서 누락/오인식이 있으면 알려주세요.")

    st.markdown("---")

    # ── ① 전체 총평 ──
    render_ops_comment_monthly(monthly, heading="## ① 전체 총평", channels=channels)

    st.markdown("---")

    # ── ② 매체별 코멘트 + 예산재배분 제안 ──
    st.markdown("## ② 매체별 코멘트 + 예산재배분 제안")
    if snapshot.empty:
        st.caption("매체별 당월 스냅샷 데이터가 아직 없습니다 (① 주간 리포트 업로드 필요).")
    else:
        snap = snapshot.copy()
        latest_month = snap["as_of_month"].max()
        snap = add_kpis(snap[snap["as_of_month"] == latest_month]).reset_index(drop=True)
        st.caption(f"기준월: {latest_month}")

        for line in _ops_channel_bucket_lines(snap, "ga_roas"):
            st.markdown(line, unsafe_allow_html=True)

        judged = snap[snap["cost_incl_vat"] >= OPS_MIN_CHANNEL_SPEND]
        gap_flag = judged[
            (judged["roas"] > 0) & (judged["ga_roas"] > 0)
            & ((judged["roas"] - judged["ga_roas"]).abs() / judged["ga_roas"] > 0.3)
        ]
        if not gap_flag.empty:
            st.caption(
                "⚠️ 자체 ROAS와 GA-ROAS 격차가 큰 매체: "
                + ", ".join(
                    f"{r.channel}(자체 {r.roas:,.0f}% vs GA {r.ga_roas:,.0f}%)"
                    for r in gap_flag.itertuples()
                )
            )

        show_cols = ["channel", "cost_incl_vat", "roas", "ga_roas"]
        show_cols = [c for c in show_cols if c in snap.columns]
        table = format_display(snap[show_cols].sort_values("cost_incl_vat", ascending=False))
        render_html_table(korify(table), raw=snap[show_cols])

    st.markdown("---")

    # ── ③ 소재별 제안 ──
    st.markdown("## ③ 소재별 제안")
    if creatives.empty:
        st.caption("소재별 성과 데이터가 아직 없습니다.")
    else:
        cr = creatives.copy()
        cr["as_of_date"] = pd.to_datetime(cr["as_of_date"]).dt.date
        cr_latest = cr.sort_values("as_of_date").drop_duplicates(subset=["channel", "creative"], keep="last")
        agg = cr_latest.groupby(["channel", "creative"], as_index=False).agg(
            impressions=("impressions", "sum"), clicks=("clicks", "sum"),
            cost_incl_vat=("cost_incl_vat", "sum"), conversions=("conversions", "sum"),
            revenue=("revenue", "sum"),
        )
        agg = add_kpis(agg)
        total_cost, total_rev = agg["cost_incl_vat"].sum(), agg["revenue"].sum()
        account_avg_roas = (total_rev / total_cost * 100) if total_cost else 0

        def _judge(row):
            if row["cost_incl_vat"] < OPS_CREATIVE_MIN_SPEND:
                return "판단 보류"
            if account_avg_roas <= 0:
                return "판단 보류"
            ratio = row["roas"] / account_avg_roas
            if ratio >= 1.2:
                return "우수"
            if ratio <= 0.7:
                return "부진"
            return "평균 수준"

        agg["판정"] = agg.apply(_judge, axis=1)
        good = agg[agg["판정"] == "우수"].sort_values("roas", ascending=False)
        bad = agg[agg["판정"] == "부진"].sort_values("roas")
        held_n = int((agg["판정"] == "판단 보류").sum())

        st.caption(f"계정 평균 ROAS: {account_avg_roas:,.0f}% · 광고비 {OPS_CREATIVE_MIN_SPEND:,}원 미만 {held_n}개 소재는 판단 보류")

        # 전체 소재 수치가 왜 그렇게 움직였는지, 어떤 소재가 광고비/매출 변화를 가장 크게
        # 이끌었는지 이전 스냅샷과 비교해 짚어준다(직전 저장 시점 대비).
        cr_dates = sorted(cr["as_of_date"].unique())
        if len(cr_dates) >= 2:
            cur_snap_date, prev_snap_date = cr_dates[-1], cr_dates[-2]
            cr_key = cr.copy()
            cr_key["key"] = cr_key["channel"].astype(str) + " " + cr_key["creative"].astype(str)
            creative_mover = _ops_top_mover_sentence(
                cr_key, key_col="key", period_col="as_of_date",
                cur_period=cur_snap_date, prev_period=prev_snap_date,
            )
            if creative_mover:
                st.markdown(f"직전 저장({prev_snap_date}) 대비 이번({cur_snap_date}) 소재 수치는 {creative_mover}")

        if not good.empty:
            top_good = ", ".join(f"{r.creative}({r.roas:,.0f}%)" for r in good.head(3).itertuples())
            st.markdown(
                f"**효율 우수 소재 ({len(good)}개)**: 계정 평균 대비 1.2배 이상 — 상위로는 {top_good} 등입니다. "
                + _ops_next_action("이번 주 중 예산을 증액하거나 동일 소재를 타 캠페인/타겟팅으로 확장 적용해보는 것을 권장합니다."),
                unsafe_allow_html=True,
            )
            st.dataframe(
                korify(format_display(good[["channel", "creative", "cost_incl_vat", "roas"]])),
                use_container_width=True, hide_index=True,
            )
        if not bad.empty:
            worst_bad = ", ".join(f"{r.creative}({r.roas:,.0f}%)" for r in bad.head(3).itertuples())
            st.markdown(
                f"**효율 부진 소재 ({len(bad)}개)**: 계정 평균 대비 0.7배 이하 — 하위로는 {worst_bad} 등입니다. "
                "다만 막 시작한 소재는 학습 기간(3~7일)을 감안해 성급히 끄지 않는 것을 권장합니다. "
                + _ops_next_action("이번 주엔 문구/이미지 교체를 먼저 시도하고, 다음 주까지 개선이 없으면 축소/OFF를 검토하세요."),
                unsafe_allow_html=True,
            )
            st.dataframe(
                korify(format_display(bad[["channel", "creative", "cost_incl_vat", "roas"]])),
                use_container_width=True, hide_index=True,
            )
        if good.empty and bad.empty:
            st.caption("우수/부진으로 구분될 만큼 표본이 확보된 소재가 아직 없습니다.")

    st.markdown("---")

    # ── ④ GA 유입 코멘트 ──
    st.markdown("## ④ GA 유입 코멘트")
    if inflow_revenue.empty:
        st.caption("유입·매출 비교 데이터가 아직 없습니다.")
    else:
        iv = inflow_revenue.copy()
        iv["report_date"] = pd.to_datetime(iv["report_date"]).dt.date
        iv["ym"] = iv["report_date"].map(lambda d: (d.year, d.month))
        monthly_iv = iv.groupby("ym", as_index=False).agg(users=("users", "sum"), new_users=("new_users", "sum"))
        monthly_iv = monthly_iv.sort_values("ym")
        if len(monthly_iv):
            cur_iv = monthly_iv.iloc[-1]
            cur_y, cur_mo = cur_iv["ym"]
            body = [f"{cur_y}년 {cur_mo}월 총 방문자 {cur_iv['users']:,.0f}명, 신규 방문자 {cur_iv['new_users']:,.0f}명입니다."]
            if len(monthly_iv) >= 2:
                prev_iv = monthly_iv.iloc[-2]
                d_users = cur_iv["users"] - prev_iv["users"]
                d_users_pct = (d_users / prev_iv["users"] * 100) if prev_iv["users"] else 0
                py, pmo = prev_iv["ym"]
                body.append(
                    f"전월({py}년 {pmo}월) 대비 방문자수는 {_ops_fmt_pct(d_users_pct)} "
                    f"({d_users:+,.0f}명) {'증가' if d_users >= 0 else '감소'}했습니다."
                )
            st.markdown(" ".join(body), unsafe_allow_html=True)

        if not ga_channel_inflow.empty:
            gc = ga_channel_inflow.copy()
            gc["report_date"] = pd.to_datetime(gc["report_date"]).dt.date
            max_d = gc["report_date"].max()
            min_d = max_d - timedelta(days=29)
            recent = gc[gc["report_date"] >= min_d]
            top_src = (
                recent.groupby("source_medium", as_index=False)["users"].sum()
                .sort_values("users", ascending=False).head(3)
            )
            if not top_src.empty:
                names = ", ".join(f"{r.source_medium}({r.users:,.0f}명)" for r in top_src.itertuples())
                st.markdown(f"최근 30일 기준 유입 상위 소스/매체는 {names} 순입니다.")
                st.markdown(
                    _ops_next_action(
                        f"상위 소스인 {top_src.iloc[0]['source_medium']} 유입을 다음 주에도 "
                        "유지되는지 확인하고, 유입은 있는데 구매로 안 이어지는 소스/매체가 있으면 랜딩페이지·"
                        "타겟팅을 점검하는 것을 권장합니다."
                    ),
                    unsafe_allow_html=True,
                )
            if recent["channel"].isna().all():
                st.caption(
                    "※ '매체'(채널 그룹핑) 매핑이 아직 안 되어 있어 소스/매체 단위로만 코멘트했습니다. "
                    "매핑 완료되면 채널 단위 코멘트로 보강할 수 있습니다."
                )


# ──────────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────────
def main():
    # st.title()은 raw HTML(이미지)을 못 받아서, 이모지(📊) 대신 커스텀 아이콘을 쓰려면
    # 직접 <h1>을 그려야 한다. inject_theme()의 전역 h1 스타일(글자크기/굵기/색)은
    # 태그 자체를 그대로 쓰기 때문에 동일하게 적용된다.
    st.markdown(
        f'<h1 style="display:flex;align-items:center;gap:10px;margin:0 0 0.5rem 0;">'
        f'<img src="data:image/png;base64,{PAGE_TITLE_ICON_B64}" style="height:34px;width:auto;" />'
        f'STCO 온라인팀 광고/마케팅 성과 대시보드</h1>',
        unsafe_allow_html=True,
    )
    render_upload_panel()

    # 어느 페이지를 보든 모든 테이블을 다 불러오면 화면 전환·정렬·날짜 변경 같은 사소한 조작마다
    # 수십 번의 DB 왕복이 생긴다(Streamlit은 조작할 때마다 스크립트를 처음부터 다시 실행한다).
    # 그래서 '지금 보는 페이지가 실제로 쓰는 테이블'만 그때그때 불러온다.
    T = load_table

    # 데이터가 아예 없을 때의 안내는 가벼운 두 테이블로만 판단한다.
    weekly, monthly = T("weekly_overview"), T("monthly_overview")
    if weekly.empty and monthly.empty:
        st.info("아직 저장된 데이터가 없습니다. 왼쪽 사이드바에서 주간 리포트 파일을 업로드하고 '전체 저장하기'를 눌러주세요.")
        return

    def to_date(df, col):
        if df is not None and not df.empty and col in df.columns:
            df[col] = pd.to_datetime(df[col]).dt.date
        return df

    to_date(weekly, "week_start"); to_date(weekly, "week_end")
    to_date(monthly, "report_month")

    page = render_nav()  # ← st.tabs() 대신 사이드바 그룹 네비게이션

    if page == "종합 대시보드":
        render_overview_page(weekly, monthly, to_date(T("daily_overview"), "report_date"),
                             T("channel_monthly"), T("channel_weekly"), T("ga_channel_inflow"))
    elif page == "매체별 성과":
        render_channel_page(T("channel_monthly"), T("channel_snapshot"), T("ga_channel_inflow"))
    elif page == "타겟팅별 성과":
        render_targeting_performance_page(T("channel_audience_snapshot"),
                                          creatives_fallback=T("creative_performance"))
    elif page == "소재별 성과":
        render_creative_performance(T("creative_performance"))
    elif page == "예산 현황":
        render_budget_page(monthly, T("channel_budget"))
    elif page == "운영 코멘트":
        render_operation_comment_page(
            weekly, monthly, to_date(T("daily_overview"), "report_date"),
            T("channel_monthly"), T("channel_snapshot"), T("creative_performance"),
            T("channel_audience_snapshot"), T("inflow_revenue_daily"),
            T("ga_channel_inflow"), T("agency_notes"),
        )
    elif page == "채널 퍼널 리포트":
        render_ga_channel_funnel_page(
            T("channel_audience_snapshot"), T("ga_channel_inflow"),
            T("inflow_revenue_daily"), T("channel_weekly"), T("channel_mix_budget"),
            ga_daily=T("ga_channel_daily"), utm_map=T("utm_channel_map"),
            decisions=T("decision_log"), ad_spend=T("ad_spend_daily"),
            master=T("media_master"),
        )
    elif page == "채널 성과":
        render_channel_performance_page(
            T("ad_spend_daily"), T("ga_channel_daily"), T("channel_mix_budget"),
            master=T("media_master"), budget=T("channel_budget"),
        )
    elif page == "예산 관리":
        render_budget_realloc_page(
            T("ad_spend_daily"), T("ga_channel_daily"), T("channel_mix_budget"),
            budget=T("channel_budget"), master=T("media_master"),
            decisions=T("decision_log"),
        )
    elif page == "GA 소재별 성과":
        render_ga_creative_page(T("ga_creative_daily"), ad_spend=T("ad_spend_daily"),
                                utm_map=T("utm_channel_map"))
    elif page == "GA 매체별 유입 경로":
        render_ga_channel_inflow_page(
            ga_inflow_source(T("ga_channel_daily"), T("ga_channel_inflow")))
    elif page == "GA4 라이브 리포트":
        render_ga4_page()
    elif page == "유입·매출 비교":
        render_inflow_revenue_page(
            T("inflow_revenue_daily"),
            ga_inflow_source(T("ga_channel_daily"), T("ga_channel_inflow")))
    elif page in NAV_PAGES_COMING_SOON:
        render_coming_soon(page)


if __name__ == "__main__":
    main()
